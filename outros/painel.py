import sqlite3
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB      = os.path.join(os.path.dirname(os.path.abspath(__file__)), "financeiro.db")
OUTPUT  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "outputs", "painel_qualidade.xlsx")

VERDE   = PatternFill("solid", fgColor="C6EFCE")
VERMELHO= PatternFill("solid", fgColor="FFC7CE")
AMARELO = PatternFill("solid", fgColor="FFEB9C")
CINZA   = PatternFill("solid", fgColor="D9D9D9")
AZUL    = PatternFill("solid", fgColor="BDD7EE")

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def gerar_painel():
    conn = sqlite3.connect(DB)

    # Dados financeiros
    df_fin = pd.read_sql("""
        SELECT ticker, ano,
            MAX(receita_liquida)    AS tem_dre,
            MAX(ativo_total)        AS tem_balanco,
            MAX(fco)                AS tem_fc
        FROM financeiros_anuais
        GROUP BY ticker, ano
    """, conn)

    # Dividendos
    df_div = pd.read_sql("""
        SELECT ticker, ano, MAX(dividendo_por_acao) AS tem_div
        FROM dividendos_anuais
        GROUP BY ticker, ano
    """, conn)

    # Preços
    df_pre = pd.read_sql("""
        SELECT ticker, ano, MAX(preco_medio) AS tem_preco
        FROM precos_anuais
        GROUP BY ticker, ano
    """, conn)

    conn.close()

    # Junta tudo
    df = df_fin.merge(df_div, on=["ticker", "ano"], how="outer")
    df = df.merge(df_pre, on=["ticker", "ano"], how="outer")
    df = df.sort_values(["ticker", "ano"])

    tickers = sorted(df["ticker"].unique())
    anos    = sorted(df["ano"].unique())

    wb = Workbook()

    # ── Aba 1: Checklist completo ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Checklist"

    colunas = ["DRE", "Balanço", "FC", "Dividendos", "Preços"]

    # Cabeçalho anos
    ws.cell(1, 1, "Ticker").font = Font(bold=True)
    ws.cell(1, 1).fill = CINZA
    col = 2
    for ano in anos:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+4)
        c = ws.cell(1, col, str(ano))
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.fill = AZUL
        for sub_i, sub in enumerate(colunas):
            ws.cell(2, col + sub_i, sub).font = Font(bold=True, size=8)
            ws.cell(2, col + sub_i).alignment = Alignment(horizontal="center")
            ws.cell(2, col + sub_i).fill = CINZA
        col += 5

    # Dados
    for row_i, ticker in enumerate(tickers, start=3):
        ws.cell(row_i, 1, ticker).font = Font(bold=True)
        col = 2
        for ano in anos:
            sub = df[(df["ticker"] == ticker) & (df["ano"] == ano)]
            vals = [
                sub["tem_dre"].values[0]     if len(sub) else None,
                sub["tem_balanco"].values[0] if len(sub) else None,
                sub["tem_fc"].values[0]      if len(sub) else None,
                sub["tem_div"].values[0]     if len(sub) else None,
                sub["tem_preco"].values[0]   if len(sub) else None,
            ]
            for v in vals:
                c = ws.cell(row_i, col)
                if v is not None and v == v:  # not NaN
                    c.value = "✓"
                    c.fill = VERDE
                else:
                    c.value = "✗"
                    c.fill = VERMELHO
                c.alignment = Alignment(horizontal="center")
                col += 1

    ws.column_dimensions["A"].width = 10
    for i in range(2, col):
        ws.column_dimensions[get_column_letter(i)].width = 5

    # ── Aba 2: Resumo por empresa ──────────────────────────────────────────
    ws2 = wb.create_sheet("Resumo")
    ws2.append(["Ticker", "Anos com DRE", "Anos c/ Balanço", "Anos c/ FC", "Anos c/ Dividendos", "Anos c/ Preços", "Cobertura %"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = AZUL

    for ticker in tickers:
        sub = df[df["ticker"] == ticker]
        total = len(anos)
        dre  = sub["tem_dre"].notna().sum()
        bal  = sub["tem_balanco"].notna().sum()
        fc   = sub["tem_fc"].notna().sum()
        div  = sub["tem_div"].notna().sum()
        pre  = sub["tem_preco"].notna().sum()
        cobert = round((dre + bal + fc + div + pre) / (total * 5) * 100, 1)
        ws2.append([ticker, dre, bal, fc, div, pre, f"{cobert}%"])

    for col in range(1, 8):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    wb.save(OUTPUT)
    print(f"✅ Painel salvo em: outputs/painel_qualidade.xlsx")

if __name__ == "__main__":
    gerar_painel()
