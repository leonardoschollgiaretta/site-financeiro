import sqlite3
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "financeiro.db")
SAIDA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "outputs", "ranking.xlsx")

AZUL_ESC = "1F3864"
AZUL_MED = "2E75B6"
AZUL_CLR = "BDD7EE"
VERDE    = "C6EFCE"
AMARELO  = "FFEB9C"
VERMELHO = "FFC7CE"
CINZA    = "F2F2F2"
BRANCO   = "FFFFFF"

# ── Definição dos indicadores ───────────────────────────────────────────────
# (nome, inverso, min_ref, max_ref, peso, categoria)
INDICADORES = [
    ("ROE (%)",            False,  0,  40, 0.30, "Rentabilidade"),
    ("Margem Liq (%)",     False,  0,  35, 0.25, "Rentabilidade"),
    ("Margem EBITDA (%)",  False,  0,  50, 0.10, "Rentabilidade"),
    ("FCO/Receita (%)",    False,  0,  40, 0.20, "Fluxo de Caixa"),
    ("Div.Liq/EBITDA",     True,   0,   6, 0.15, "Endividamento"),
    ("DY (%)",             False,  0,  10, 0.10, "Mercado"),
]

CORES_CAT = {
    "Rentabilidade":  "DAEEF3",
    "Fluxo de Caixa": "E2EFDA",
    "Endividamento":  "FCE4D6",
    "Mercado":        "FFF2CC",
}

def borda():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def cor_score(s):
    if s is None: return BRANCO
    if s >= 70:   return VERDE
    if s >= 40:   return AMARELO
    return VERMELHO

PRIORIDADE_FONTE = {"investsite": 0, "statusinvest": 1, "yfinance": 2, "fundamentus": 3, "manual": 4}

def carregar_dados(ano, moeda):
    conn = sqlite3.connect(DB)
    fin_raw = pd.read_sql(f"""
        SELECT f.ticker, e.nome, e.setor, f.fonte,
               f.receita_liquida, f.lucro_liquido, f.ebitda,
               f.patrimonio_liquido, f.divida_liquida, f.fco
        FROM financeiros_anuais f
        JOIN empresas e ON e.ticker = f.ticker
        WHERE f.ano = {ano} AND e.moeda = '{moeda}'
    """, conn)
    divs = pd.read_sql(f"""
        SELECT ticker, SUM(dividendo_por_acao) AS dividendo
        FROM dividendos_anuais WHERE ano = {ano} GROUP BY ticker
    """, conn)
    precos = pd.read_sql(f"""
        SELECT ticker, preco_medio FROM precos_anuais WHERE ano = {ano}
    """, conn)
    conn.close()

    # Mescla múltiplas fontes: investsite > statusinvest > yfinance > ...
    if not fin_raw.empty:
        fin_raw["_prio"] = fin_raw["fonte"].map(lambda f: PRIORIDADE_FONTE.get(f, 99))
        fin_raw = fin_raw.sort_values(["ticker", "_prio"])
        cols_dados = [c for c in fin_raw.columns if c not in ["ticker","nome","setor","fonte","_prio"]]
        # groupby ticker e pega primeiro não-nulo por coluna
        fin = fin_raw.groupby("ticker")[cols_dados].first().reset_index()
        # recupera nome e setor (sempre do mesmo ticker)
        meta = fin_raw.groupby("ticker")[["nome","setor"]].first().reset_index()
        fin = fin.merge(meta, on="ticker")
    else:
        fin = fin_raw

    df = fin.merge(divs, on="ticker", how="left")
    df = df.merge(precos, on="ticker", how="left")
    return df

def safe_div(a, b, allow_neg_b=False):
    try:
        fa, fb = float(a), float(b)
        if pd.isna(fa) or pd.isna(fb) or fb == 0: return None
        if not allow_neg_b and fb < 0: return None
        return round(fa / fb, 2)
    except:
        return None

def normalizar(val, inv, minv, maxv):
    """Converte valor bruto em score 0-100. Negativo = 0 sempre."""
    if val is None or pd.isna(val): return None
    if val < 0: return 0
    if inv:
        if val <= minv: return 100.0
        if val >= maxv: return 0.0
        return round(100 * (maxv - val) / (maxv - minv), 1)
    else:
        if val <= minv: return 0.0
        if val >= maxv: return 100.0
        return round(100 * (val - minv) / (maxv - minv), 1)

def calcular_indicadores(df):
    # Valores brutos
    df["ROE (%)"]           = df.apply(lambda r: safe_div(r.lucro_liquido * 100, r.patrimonio_liquido), axis=1)
    df["Margem Liq (%)"]    = df.apply(lambda r: safe_div(r.lucro_liquido * 100, r.receita_liquida), axis=1)
    df["Margem EBITDA (%)"] = df.apply(lambda r: safe_div(r.ebitda * 100, r.receita_liquida), axis=1)
    df["FCO/Receita (%)"]   = df.apply(lambda r: safe_div(r.fco * 100, r.receita_liquida), axis=1)
    df["DY (%)"]            = df.apply(lambda r: safe_div((r.get("dividendo") or 0) * 100, r.preco_medio), axis=1)
    df["Div.Liq/EBITDA"]    = df.apply(
        lambda r: safe_div(r.divida_liquida, r.ebitda) if pd.notna(r.ebitda) and r.ebitda > 0 else None, axis=1)

    # Score por indicador (0-100)
    for nome, inv, minv, maxv, _, _ in INDICADORES:
        df[f"S:{nome}"] = df[nome].apply(lambda v: normalizar(v, inv, minv, maxv))

    # Score final ponderado
    def score_final(row):
        pts, total_peso = 0, 0
        for nome, inv, minv, maxv, peso, _ in INDICADORES:
            s = row.get(f"S:{nome}")
            if s is None: continue
            pts        += peso * s
            total_peso += peso
        return round(pts / total_peso, 1) if total_peso > 0 else 0

    df["Score Final"] = df.apply(score_final, axis=1)
    df = df.sort_values("Score Final", ascending=False).reset_index(drop=True)
    df.index += 1
    return df

def escrever_tabela(ws, df, moeda, ano):
    simbolo = "USD" if moeda == "USD" else "R$"
    brd = borda()

    # ── Colunas fixas + par (valor bruto | score) por indicador ────────────
    fixas = ["#", "Ticker", "Empresa", "Setor",
             f"Receita\n({simbolo} B)", f"Lucro\n({simbolo} B)"]
    n_fixas = len(fixas)

    total_cols = n_fixas + len(INDICADORES) * 2 + 1  # +1 = Score Final

    # Linha 1: Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws.cell(1, 1, f"RANKING {moeda} — ANO {ano}  |  Score 0-100 por indicador")
    t.font = Font(bold=True, color="FFFFFF", size=13)
    t.fill = PatternFill("solid", fgColor=AZUL_ESC)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Linha 2: cabeçalhos fixos + agrupamento por indicador
    for ci, cn in enumerate(fixas, 1):
        ws.merge_cells(start_row=2, start_column=ci, end_row=3, end_column=ci)
        c = ws.cell(2, ci, cn)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=AZUL_ESC)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd

    col = n_fixas + 1
    for nome, inv, _, _, peso, cat in INDICADORES:
        cor_cat = CORES_CAT.get(cat, AZUL_CLR)
        # Merge pra agrupar "Valor | Score"
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
        cab = ws.cell(2, col, f"{nome}\n(peso {int(peso*100)}%)")
        cab.font = Font(bold=True, size=9, color="000000")
        cab.fill = PatternFill("solid", fgColor=cor_cat)
        cab.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cab.border = brd
        # Subcabeçalhos Valor | Score
        for sub_ci, sub_txt in enumerate(["Valor", "Score\n(0-100)"], col):
            c = ws.cell(3, sub_ci, sub_txt)
            c.font = Font(bold=True, size=8, color="000000")
            c.fill = PatternFill("solid", fgColor=cor_cat)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = brd
        col += 2

    # Score Final
    ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
    sf = ws.cell(2, col, "SCORE\nFINAL")
    sf.font = Font(bold=True, color="FFFFFF", size=10)
    sf.fill = PatternFill("solid", fgColor=AZUL_ESC)
    sf.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sf.border = brd

    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 28

    # ── Dados ──────────────────────────────────────────────────────────────
    for rank, row in df.iterrows():
        bg = CINZA if rank % 2 == 0 else BRANCO
        ri = rank + 3  # linha real (1=título, 2-3=cabeçalhos)

        # Fixas
        fixas_vals = [
            rank,
            row["ticker"],
            row["nome"],
            row.get("setor", ""),
            round(row["receita_liquida"] / 1e9, 2) if pd.notna(row.get("receita_liquida")) else "-",
            round(row["lucro_liquido"] / 1e9, 2)   if pd.notna(row.get("lucro_liquido"))   else "-",
        ]
        for ci, val in enumerate(fixas_vals, 1):
            c = ws.cell(ri, ci, val if val == val else "-")
            c.font = Font(size=10, bold=(ci == 1))
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left" if ci in [3,4] else "center", vertical="center")
            c.border = brd

        # Indicadores: valor bruto + score
        col = n_fixas + 1
        for nome, _, _, _, _, cat in INDICADORES:
            cor_cat = CORES_CAT.get(cat, AZUL_CLR)
            val  = row.get(nome)
            sc   = row.get(f"S:{nome}")

            # Valor bruto
            exibir = round(val, 2) if (val is not None and pd.notna(val)) else "-"
            c = ws.cell(ri, col, exibir)
            c.font = Font(size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = brd
            col += 1

            # Score 0-100
            exibir_s = sc if (sc is not None and pd.notna(sc)) else "-"
            c = ws.cell(ri, col, exibir_s)
            c.font = Font(size=10, bold=True)
            c.fill = PatternFill("solid", fgColor=cor_score(sc))
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = brd
            col += 1

        # Score Final
        sf_val = row.get("Score Final")
        c = ws.cell(ri, col, sf_val if sf_val is not None else "-")
        c.font = Font(size=11, bold=True)
        c.fill = PatternFill("solid", fgColor=cor_score(sf_val))
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd

        ws.row_dimensions[ri].height = 20

    # Larguras
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    for i in range(n_fixas + 1, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 9

def gerar_ranking(ano=2024):
    wb = Workbook()
    wb.remove(wb.active)

    for moeda, label in [("BRL", "Brasil"), ("USD", "EUA")]:
        df = carregar_dados(ano, moeda)
        if df.empty:
            print(f"  ⚠️ Sem dados {moeda} para {ano}")
            continue
        df = calcular_indicadores(df)
        ws = wb.create_sheet(label)
        escrever_tabela(ws, df, moeda, ano)

        print(f"\n  📊 Ranking {moeda} ({ano})")
        print(f"  {'#':<3} {'Ticker':<7} {'ROE':>7} {'Margem':>8} {'FCO/Rec':>8} {'D/EBITDA':>9} {'Score':>6}")
        print(f"  {'-'*50}")
        for i, row in df.iterrows():
            print(f"  #{i:<2} {row['ticker']:<7}"
                  f" {str(row.get('ROE (%)', '-')):>7}%"
                  f" {str(row.get('Margem Liq (%)', '-')):>8}%"
                  f" {str(row.get('FCO/Receita (%)', '-')):>8}%"
                  f" {str(row.get('Div.Liq/EBITDA', '-')):>9}"
                  f" {str(row.get('Score Final', '-')):>6}")

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    print(f"\n✅ Ranking salvo em outputs/ranking.xlsx")

if __name__ == "__main__":
    ano = input("Ano para o ranking [Enter = 2024]: ").strip()
    ano = int(ano) if ano else 2024
    print(f"\n📊 Gerando ranking {ano}...")
    gerar_ranking(ano)
