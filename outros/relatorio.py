import sqlite3
import os
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "financeiro.db")
SAIDA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "outputs", "relatorio.xlsx")

def carregar_empresa(ticker):
    conn = sqlite3.connect(DB)

    # Carrega todas as fontes e mescla: statusinvest tem prioridade nos dados financeiros
    fin_raw = pd.read_sql(f"SELECT * FROM financeiros_anuais WHERE ticker='{ticker}' ORDER BY ano DESC, fonte ASC", conn)

    # Mescla por ano: para cada campo, usa o primeiro valor não-nulo
    # prioridade: statusinvest > yfinance > manual
    PRIORIDADE = {"statusinvest": 0, "yfinance": 1, "fundamentus": 2, "manual": 3}
    if not fin_raw.empty:
        fin_raw["_prio"] = fin_raw["fonte"].map(lambda f: PRIORIDADE.get(f, 99))
        fin_raw = fin_raw.sort_values(["ano", "_prio"])
        colunas_dados = [c for c in fin_raw.columns if c not in ["ticker","ano","fonte","moeda","_prio"]]
        fin = fin_raw.groupby("ano")[colunas_dados].first().reset_index()
        fin = fin.sort_values("ano", ascending=False)
    else:
        fin = fin_raw

    precos = pd.read_sql(f"SELECT ano, preco_min, preco_max, preco_medio FROM precos_anuais WHERE ticker='{ticker}' ORDER BY ano DESC", conn)
    divs   = pd.read_sql(f"SELECT ano, dividendo_por_acao FROM dividendos_anuais WHERE ticker='{ticker}' ORDER BY ano DESC", conn)
    emp    = pd.read_sql(f"SELECT * FROM empresas WHERE ticker='{ticker}'", conn)
    conn.close()
    return fin, precos, divs, emp

def m(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "-"
    return round(float(val) / 1_000_000, 2)

def adicionar_aba(wb, ticker):
    ticker = ticker.upper()
    fin, precos, divs, empresa = carregar_empresa(ticker)

    if fin.empty:
        print(f"   ⚠️  {ticker} — sem dados no banco. Rode o coletor primeiro.")
        return

    anos = sorted(fin["ano"].unique(), reverse=True)
    nome = empresa["nome"].values[0] if not empresa.empty else ticker

    # Busca dados 2026 ao vivo
    print(f"   🔄 Buscando dados 2026 de {ticker}...")
    try:
        acao = yf.Ticker(f"{ticker}.SA")
        info = acao.info
        preco_hoje = info.get("currentPrice") or info.get("regularMarketPrice")

        hist2026 = acao.history(start="2026-01-01")
        if not hist2026.empty:
            hist2026.index = hist2026.index.tz_localize(None)
            preco_min_2026 = round(hist2026["Close"].min(), 2)
            preco_max_2026 = round(hist2026["Close"].max(), 2)
            preco_med_2026 = round(hist2026["Close"].mean(), 2)
        else:
            preco_min_2026 = preco_max_2026 = preco_med_2026 = preco_hoje

        div2026 = acao.dividends
        if not div2026.empty:
            div2026.index = div2026.index.tz_localize(None)
            div2026 = div2026[div2026.index.year == 2026].sum()
            div2026 = round(float(div2026), 4) if div2026 else "-"
        else:
            div2026 = "-"

        tem_2026 = True
    except Exception as e:
        print(f"   ⚠️  Não foi possível buscar 2026: {e}")
        tem_2026 = False
        preco_hoje = preco_min_2026 = preco_max_2026 = preco_med_2026 = "-"
        div2026 = "-"

    anos_exibir = [2026] + list(anos) if tem_2026 else list(anos)

    ws = wb.create_sheet(title=ticker)

    AZUL_ESC  = "1F3864"
    AZUL_MED  = "2E75B6"
    AZUL_CLR  = "BDD7EE"
    VERDE_CLR = "E2EFDA"

    lado  = Side(style="thin", color="CCCCCC")
    borda = Border(left=lado, right=lado, top=lado, bottom=lado)

    def hd(cell, texto, verde=False):
        cell.value = texto
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1E8449" if verde else AZUL_ESC)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda

    def secao(row, texto):
        ws.merge_cells(f"A{row}:{get_column_letter(1+len(anos_exibir))}{row}")
        c = ws.cell(row=row, column=1, value=f"  {texto}")
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=AZUL_MED)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20

    def linha(row, label, valores, negrito=False, cor=None):
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(bold=negrito, size=10)
        c.fill = PatternFill("solid", fgColor=cor if cor else "FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = borda
        for i, val in enumerate(valores):
            is_2026 = (anos_exibir[i] == 2026)
            cell = ws.cell(row=row, column=2+i, value=val)
            cell.font = Font(bold=negrito, size=10, italic=is_2026)
            bg = VERDE_CLR if is_2026 else (cor if cor else "FFFFFF")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = borda
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
        ws.row_dimensions[row].height = 18

    # Cabeçalho
    ws.merge_cells(f"A1:{get_column_letter(1+len(anos_exibir))}1")
    c = ws["A1"]
    c.value = f"{nome} ({ticker})  |  Valores em R$ milhões"
    c.font = Font(bold=True, color="FFFFFF", size=13)
    c.fill = PatternFill("solid", fgColor=AZUL_ESC)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    hd(ws.cell(row=2, column=1), "INDICADOR")
    for i, ano in enumerate(anos_exibir):
        hd(ws.cell(row=2, column=2+i),
           "2026 (atual)" if ano == 2026 else str(ano),
           verde=(ano == 2026))
    ws.row_dimensions[2].height = 22

    def vals(campo):
        result = []
        for ano in anos_exibir:
            if ano == 2026:
                row = fin[fin["ano"] == anos[0]] if anos else pd.DataFrame()
            else:
                row = fin[fin["ano"] == ano]
            result.append(m(row[campo].values[0]) if not row.empty else "-")
        return result

    def vals_raw(campo):
        result = []
        for ano in anos_exibir:
            if ano == 2026:
                row = fin[fin["ano"] == anos[0]] if anos else pd.DataFrame()
            else:
                row = fin[fin["ano"] == ano]
            if row.empty:
                result.append("-")
            else:
                v = row[campo].values[0]
                result.append(round(float(v), 4) if v and not pd.isna(v) else "-")
        return result

    def vals_precos(campo):
        result = []
        for ano in anos_exibir:
            if ano == 2026:
                if campo == "preco_min":   result.append(preco_min_2026)
                elif campo == "preco_max": result.append(preco_max_2026)
                else:                      result.append(preco_med_2026)
            else:
                row = precos[precos["ano"] == ano]
                if row.empty:
                    result.append("-")
                else:
                    v = row[campo].values[0]
                    result.append(round(float(v), 2) if v else "-")
        return result

    def vals_divs():
        result = []
        for ano in anos_exibir:
            if ano == 2026:
                result.append(div2026)
            else:
                row = divs[divs["ano"] == ano]
                if row.empty:
                    result.append("-")
                else:
                    v = row["dividendo_por_acao"].values[0]
                    result.append(round(float(v), 4) if v else "-")
        return result

    r = [3]
    def add_secao(n):
        secao(r[0], n); r[0] += 1
    def add(label, valores, negrito=False, cor=None):
        linha(r[0], label, valores, negrito=negrito, cor=cor); r[0] += 1
    def espaco():
        r[0] += 1

    add_secao("📈  DADOS DE MERCADO")
    add("Preço Atual (R$)",           [preco_hoje if a == 2026 else "-" for a in anos_exibir])
    add("Preço Mínimo do Ano (R$)",   vals_precos("preco_min"))
    add("Preço Máximo do Ano (R$)",   vals_precos("preco_max"))
    add("Preço Médio do Ano (R$)",    vals_precos("preco_medio"))
    add("Dividendo por Ação (R$)",    vals_divs())
    espaco()

    add_secao("📊  DRE  |  R$ milhões  |  2026 = último resultado 2025")
    add("Receita Líquida",            vals("receita_liquida"),    negrito=True, cor=AZUL_CLR)
    add("Custo dos Produtos (CPV)",   vals("custo_receita"))
    add("Lucro Bruto",                vals("lucro_bruto"),        negrito=True, cor=AZUL_CLR)
    add("EBITDA",                     vals("ebitda"),             negrito=True, cor=AZUL_CLR)
    add("EBIT",                       vals("ebit"),               negrito=True, cor=AZUL_CLR)
    add("Receitas Financeiras",       vals("rec_financeiras"))
    add("Despesas Financeiras",       vals("desp_financeiras"))
    add("IR & CSLL",                  vals("ir_csll"))
    add("Lucro Líquido",              vals("lucro_liquido"),      negrito=True, cor=AZUL_CLR)
    add("EPS — Lucro por Ação",       vals_raw("eps"))
    espaco()

    add_secao("🏦  BALANÇO  |  R$ milhões  |  2026 = último resultado 2025")
    add("Caixa & Equivalentes",       vals("caixa"))
    add("Contas a Receber",           vals("contas_receber"))
    add("Estoques",                   vals("estoques"))
    add("TOTAL ATIVO CIRCULANTE",     vals("ativo_circulante"),   negrito=True, cor=AZUL_CLR)
    add("TOTAL DO ATIVO",             vals("ativo_total"),        negrito=True, cor=AZUL_CLR)
    add("Empréstimos CP",             vals("divida_cp"))
    add("Empréstimos LP",             vals("divida_lp"))
    add("TOTAL PATRIMÔNIO LÍQUIDO",   vals("patrimonio_liquido"), negrito=True, cor=AZUL_CLR)
    add("Dívida Bruta",               vals("divida_bruta"))
    add("Dívida Líquida",             vals("divida_liquida"))
    espaco()

    add_secao("💰  FLUXO DE CAIXA  |  R$ milhões  |  2026 = último resultado 2025")
    add("Fluxo Caixa Operacional",    vals("fco"),                negrito=True, cor=AZUL_CLR)
    add("CAPEX",                      vals("capex"))
    add("Free Cash Flow",             vals("fcf"),                negrito=True, cor=AZUL_CLR)
    add("Fluxo Caixa Investimentos",  vals("fci"))
    add("Fluxo Caixa Financiamentos", vals("fcf_financiamento"))
    add("Dividendos Pagos",           vals("dividendos_pagos"))

    ws.column_dimensions["A"].width = 32
    for i in range(len(anos_exibir)):
        ws.column_dimensions[get_column_letter(2+i)].width = 15

    print(f"   ✅ {ticker} — aba criada ({len(anos_exibir)} colunas incluindo 2026)")

# ── MAIN ─────────────────────────────────────────────────────────────
entrada = input("Digite os tickers separados por vírgula: ")
tickers = [t.strip().upper() for t in entrada.split(",")]

wb = Workbook()
wb.remove(wb.active)

print(f"\n📊 Gerando relatorio.xlsx com {len(tickers)} empresa(s)...")

for ticker in tickers:
    adicionar_aba(wb, ticker)

wb.save(SAIDA)
print(f"\n✅ Salvo em outputs/relatorio.xlsx")
print(f"   Abas: {' | '.join(tickers)}")