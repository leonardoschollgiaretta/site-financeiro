"""
Gera o arquivo config_ranking.xlsx com todos os indicadores e pesos padrão.
Rode uma vez. Depois edite o Excel diretamente pra ajustar os pesos.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SAIDA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "outputs", "config_ranking.xlsx")

INDICADORES = [
    # (categoria, nome, coluna_db, inverso, peso_padrao, descricao)
    # ── RENTABILIDADE ──────────────────────────────────────────────
    ("Rentabilidade", "ROE (%)",            "calc",  False, 8, "Lucro Líquido / Patrimônio Líquido × 100"),
    ("Rentabilidade", "ROA (%)",            "calc",  False, 5, "Lucro Líquido / Ativo Total × 100"),
    ("Rentabilidade", "Margem Líquida (%)", "calc",  False, 7, "Lucro Líquido / Receita × 100"),
    ("Rentabilidade", "Margem EBITDA (%)",  "calc",  False, 6, "EBITDA / Receita × 100"),
    ("Rentabilidade", "Margem EBIT (%)",    "calc",  False, 4, "EBIT / Receita × 100"),
    ("Rentabilidade", "Margem Bruta (%)",   "calc",  False, 3, "Lucro Bruto / Receita × 100"),
    # ── CRESCIMENTO ────────────────────────────────────────────────
    ("Crescimento",   "Cresc. Receita (%)", "calc",  False, 6, "Variação da Receita vs ano anterior"),
    ("Crescimento",   "Cresc. Lucro (%)",   "calc",  False, 6, "Variação do Lucro vs ano anterior"),
    # ── ENDIVIDAMENTO ──────────────────────────────────────────────
    ("Endividamento", "Dív.Líq/EBITDA",     "calc",  True,  7, "Dívida Líquida / EBITDA (menor = melhor)"),
    ("Endividamento", "Dív.Líq/PL",         "calc",  True,  5, "Dívida Líquida / Patrimônio Líquido (menor = melhor)"),
    ("Endividamento", "Endividamento (%)",   "calc",  True,  4, "Dívida Bruta / Ativo Total × 100 (menor = melhor)"),
    ("Endividamento", "Liquidez Corrente",   "calc",  False, 4, "Ativo Circulante / Passivo Circulante (maior = melhor)"),
    # ── FLUXO DE CAIXA ────────────────────────────────────────────
    ("Fluxo de Caixa","FCO/Receita (%)",    "calc",  False, 6, "Fluxo de Caixa Operacional / Receita × 100"),
    ("Fluxo de Caixa","Qualidade Lucro",    "calc",  False, 5, "FCO / Lucro Líquido (FCO > lucro = bom sinal)"),
    ("Fluxo de Caixa","FCL/Receita (%)",    "calc",  False, 5, "Free Cash Flow / Receita × 100"),
    ("Fluxo de Caixa","CAPEX/Receita (%)",  "calc",  True,  3, "CAPEX / Receita × 100 (menor = menos intensivo)"),
    # ── MERCADO ───────────────────────────────────────────────────
    ("Mercado",       "Dividend Yield (%)", "calc",  False, 5, "Dividendo por Ação / Preço Médio × 100"),
    ("Mercado",       "P/L",                "calc",  True,  4, "Preço Médio / Lucro por Ação (menor = mais barato)"),
    ("Mercado",       "P/VP",               "calc",  True,  3, "Preço / Valor Patrimonial por Ação (menor = mais barato)"),
    ("Mercado",       "EV/EBITDA",          "calc",  True,  4, "Enterprise Value / EBITDA (menor = mais barato)"),
]

wb = Workbook()
ws = wb.active
ws.title = "Pesos"

# ── Estilos ────────────────────────────────────────────────────────────────
def cel(row, col, val, bold=False, cor_fundo=None, cor_fonte="000000", center=False):
    c = ws.cell(row, col, val)
    c.font = Font(bold=bold, color=cor_fonte, size=10)
    if cor_fundo:
        c.fill = PatternFill("solid", fgColor=cor_fundo)
    c.alignment = Alignment(horizontal="center" if center else "left",
                            vertical="center", wrap_text=True)
    s = Side(style="thin", color="CCCCCC")
    c.border = Border(left=s, right=s, top=s, bottom=s)
    return c

CORES_CAT = {
    "Rentabilidade": "DAEEF3",
    "Crescimento":   "E2EFDA",
    "Endividamento": "FCE4D6",
    "Fluxo de Caixa":"EBF1DE",
    "Mercado":       "FFF2CC",
}

# ── Cabeçalho ─────────────────────────────────────────────────────────────
ws.merge_cells("A1:F1")
t = ws.cell(1, 1, "⚙️  CONFIGURAÇÃO DO RANKING — Defina os pesos de 0 a 10  (0 = ignorar indicador)")
t.font = Font(bold=True, color="FFFFFF", size=12)
t.fill = PatternFill("solid", fgColor="1F3864")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

cabecalhos = ["Categoria", "Indicador", "Maior = Melhor?", "PESO (0-10)", "Peso %", "Descrição"]
for ci, cab in enumerate(cabecalhos, 1):
    cel(2, ci, cab, bold=True, cor_fundo="2E75B6", cor_fonte="FFFFFF", center=True)
ws.row_dimensions[2].height = 22

# ── Dados ─────────────────────────────────────────────────────────────────
for ri, (cat, nome, _, inv, peso, desc) in enumerate(INDICADORES, start=3):
    cor = CORES_CAT.get(cat, "FFFFFF")
    cel(ri, 1, cat,                    cor_fundo=cor)
    cel(ri, 2, nome,                   cor_fundo=cor, bold=True)
    cel(ri, 3, "NÃO" if inv else "SIM",cor_fundo=cor, center=True)
    c = cel(ri, 4, peso,               cor_fundo="FFFFC0", bold=True, center=True)  # ← EDITAR AQUI
    # Fórmula % automática
    ws.cell(ri, 5).value = f"=D{ri}/SUM($D$3:$D${2+len(INDICADORES)})"
    ws.cell(ri, 5).number_format = "0.0%"
    ws.cell(ri, 5).alignment = Alignment(horizontal="center", vertical="center")
    s = Side(style="thin", color="CCCCCC")
    ws.cell(ri, 5).border = Border(left=s, right=s, top=s, bottom=s)
    cel(ri, 6, desc,                   cor_fundo="FFFFFF")
    ws.row_dimensions[ri].height = 20

# ── Validação: peso só de 0 a 10 ──────────────────────────────────────────
dv = DataValidation(type="whole", operator="between", formula1=0, formula2=10,
                    showErrorMessage=True,
                    errorTitle="Valor inválido",
                    error="Digite um número inteiro entre 0 e 10")
dv.sqref = f"D3:D{2+len(INDICADORES)}"
ws.add_data_validation(dv)

# ── Instrução extra ────────────────────────────────────────────────────────
linha_inst = 3 + len(INDICADORES) + 1
ws.merge_cells(f"A{linha_inst}:F{linha_inst}")
inst = ws.cell(linha_inst, 1,
    "💡  Edite apenas a coluna PESO (D). 0 = ignorar. O ranking recalcula automaticamente ao rodar ranking.py")
inst.font = Font(italic=True, color="595959", size=9)
inst.fill = PatternFill("solid", fgColor="F2F2F2")
inst.alignment = Alignment(horizontal="left", vertical="center")

# ── Larguras ──────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 10
ws.column_dimensions["F"].width = 55

os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
wb.save(SAIDA)
print(f"✅ Config gerado em outputs/config_ranking.xlsx")
print(f"   Edite a coluna PESO (amarela) e rode ranking.py")
