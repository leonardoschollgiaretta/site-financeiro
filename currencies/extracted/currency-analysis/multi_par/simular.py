"""
simular.py -- Simulador multi-par.

Fluxo:
  1. Le parametros.xlsx (cria com defaults se nao existir)
  2. Verifica se ha cotacoes do par escolhido (avisa pra rodar coletar.py)
  3. Recalcula janelas se necessario (automatico)
  4. Filtra janelas similares no SQLite
  5. Gera resultado.xlsx com tabelas + graficos

Uso: python multi_par/simular.py
"""
import sys
import os
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent))
from banco import criar_banco, conectar, DB_PATH
from coletar import PARES
from janelas import (
    calcular_janelas, janelas_atualizadas,
    TAMANHOS_BASE, HORIZONTES,
)

BASE_DIR = Path(__file__).resolve().parent.parent
PARAM_PATH = Path(__file__).resolve().parent / "parametros.xlsx"
OUT_DIR    = BASE_DIR / "outputs" / "multi_par"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Defaults ───────────────────────────────────────────────────────────────────
DEFAULTS = {
    "par":              "USDBRL",
    "amostra":          2500,
    "periodo_seguinte": 10,            # 5, 10, 21 ou 63 (precisa estar em HORIZONTES)
    "tam_base_atual":   20,
    "tam_base_alvo":    "AUTO",        # AUTO = usa tam_base_atual
    "faixa_tempo":      5,
    "var_alvo":         "AUTO",        # AUTO = calcula da serie
    "faixa_baixo":      1.0,
    "faixa_cima":       1.0,
    "limite_caiu":      -5.0,
    "limite_lateral":   2.0,
    "limite_subiu":     5.0,
}

# ── Estilos ────────────────────────────────────────────────────────────────────
P_HDR  = PatternFill("solid", fgColor="1F3864")
P_SUB  = PatternFill("solid", fgColor="2E75B6")
P_INPUT= PatternFill("solid", fgColor="FFFF00")
P_CALC = PatternFill("solid", fgColor="DDEBF7")
P_OUT  = PatternFill("solid", fgColor="E2EFDA")
P_OUT2 = PatternFill("solid", fgColor="FFF2CC")
P_LBL  = PatternFill("solid", fgColor="F2F2F2")
P_DEST = PatternFill("solid", fgColor="00D4AA")
P_NEG  = PatternFill("solid", fgColor="FF6B6B")

F_HDR   = Font(name="Arial", color="FFFFFF", bold=True, size=12)
F_SUB   = Font(name="Arial", color="FFFFFF", bold=True, size=10)
F_LBL   = Font(name="Arial", bold=True, size=10)
F_VAL   = Font(name="Arial", size=10)
F_INPUT = Font(name="Arial", color="0000FF", bold=True, size=11)
F_CALC  = Font(name="Arial", color="333333", bold=True, size=10)
F_OUT   = Font(name="Arial", bold=True, size=11)
F_BIG   = Font(name="Arial", bold=True, size=14)
F_NOTE  = Font(name="Arial", italic=True, color="666666", size=9)

CTR = Alignment(horizontal="center", vertical="center")
LFT = Alignment(horizontal="left",   vertical="center", indent=1)

def brd():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ─────────────────────────────────────────────────────────────────────────────
#  Parametros (Excel)
# ─────────────────────────────────────────────────────────────────────────────

def gerar_parametros_default():
    """Cria parametros.xlsx com valores default."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Parametros"

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "  PARAMETROS DO SIMULADOR  --  edite as celulas amarelas"
    c.font = F_HDR; c.fill = P_HDR; c.alignment = LFT
    ws.row_dimensions[1].height = 26

    linhas = [
        ("Par",                       "par",              "Pares disponiveis: " + ", ".join(PARES.keys())),
        ("Tamanho da amostra (dias)", "amostra",          "Quantos dias do historico considerar (max ~2500)"),
        ("Periodo seguinte (dias)",   "periodo_seguinte", f"Horizonte a projetar. Opcoes: {HORIZONTES}"),
        ("Tamanho base ATUAL (dias)", "tam_base_atual",   "Quantos dias atras eu olho pra calcular a variacao atual"),
        ("Tamanho base ALVO",         "tam_base_alvo",    "Default: AUTO (=tam_base_atual). Sobrescreva pra simular"),
        ("Faixa de tempo (+/- dias)", "faixa_tempo",      "Aceita janelas com tamanho base entre [alvo - X, alvo + X]"),
        ("Variacao alvo (%)",         "var_alvo",         "Default: AUTO (calculado). Sobrescreva pra simular cenario"),
        ("Faixa pra BAIXO (%)",       "faixa_baixo",      "Quanto pra menos do alvo eu aceito"),
        ("Faixa pra CIMA (%)",        "faixa_cima",       "Quanto pra mais do alvo eu aceito"),
        ("Limite Caiu Forte (%)",     "limite_caiu",      "Var <= este valor = CAIU FORTE (negativo)"),
        ("Limite Lateral (+/- %)",    "limite_lateral",   "Faixa lateral em torno de zero"),
        ("Limite Subiu Forte (%)",    "limite_subiu",     "Var >= este valor = SUBIU FORTE"),
    ]
    for i, (label, chave, ajuda) in enumerate(linhas, start=3):
        c = ws.cell(i, 1, label)
        c.font = F_LBL; c.fill = P_LBL; c.alignment = LFT; c.border = brd()
        c = ws.cell(i, 2, DEFAULTS[chave])
        c.font = F_INPUT; c.fill = P_INPUT; c.alignment = CTR; c.border = brd()
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=4)
        ca = ws.cell(i, 3, "  " + ajuda)
        ca.font = F_NOTE; ca.alignment = LFT
        ws.row_dimensions[i].height = 20

    # Dropdown do par
    dv = DataValidation(type="list", formula1=f'"{",".join(PARES.keys())}"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws.cell(3, 2))

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30

    wb.save(PARAM_PATH)
    print(f"parametros.xlsx criado em {PARAM_PATH}")
    print("Edite os valores e rode novamente.")


def ler_parametros():
    """Le parametros do Excel. Retorna dict com valores tipados."""
    if not PARAM_PATH.exists():
        gerar_parametros_default()
        return None

    wb = load_workbook(PARAM_PATH, data_only=True)
    ws = wb["Parametros"]

    ordem = [
        "par", "amostra", "periodo_seguinte", "tam_base_atual",
        "tam_base_alvo", "faixa_tempo", "var_alvo",
        "faixa_baixo", "faixa_cima",
        "limite_caiu", "limite_lateral", "limite_subiu",
    ]
    valores = {}
    for i, chave in enumerate(ordem, start=3):
        valores[chave] = ws.cell(i, 2).value

    # Tipagem
    valores["par"] = str(valores["par"]).strip().upper()
    valores["amostra"] = int(valores["amostra"])
    valores["periodo_seguinte"] = int(valores["periodo_seguinte"])
    valores["tam_base_atual"] = int(valores["tam_base_atual"])
    valores["faixa_tempo"] = int(valores["faixa_tempo"])
    valores["faixa_baixo"] = float(valores["faixa_baixo"])
    valores["faixa_cima"]  = float(valores["faixa_cima"])
    valores["limite_caiu"]    = float(valores["limite_caiu"])
    valores["limite_lateral"] = float(valores["limite_lateral"])
    valores["limite_subiu"]   = float(valores["limite_subiu"])

    # AUTO -> resolvido depois (precisa de dados)
    return valores


# ─────────────────────────────────────────────────────────────────────────────
#  Calculo do cenario atual e busca de similares
# ─────────────────────────────────────────────────────────────────────────────

def calcular_cenario_atual(conn, par, tam_base):
    """
    Retorna (preco_ini, preco_fim, var_pct, data_ini, data_fim)
    do periodo BASE atual = ultimos tam_base dias do par.
    """
    rows = conn.execute(
        "SELECT data, preco FROM cotacoes WHERE par=? ORDER BY data DESC LIMIT ?",
        (par, tam_base)
    ).fetchall()
    if len(rows) < tam_base:
        return None

    rows.reverse()  # ordem cronologica
    data_ini, preco_ini = rows[0]
    data_fim, preco_fim = rows[-1]
    var_pct = (preco_fim / preco_ini - 1) * 100
    return preco_ini, preco_fim, var_pct, data_ini, data_fim


def buscar_similares(conn, par, params):
    """
    Faz SELECT no banco filtrando janelas similares.
    Retorna DataFrame com as janelas encontradas.

    Aceita QUALQUER periodo_seguinte:
      - Se for um dos pre-calculados (HORIZONTES = [5,10,21,63]) usa direto do banco
      - Caso contrario, busca data_ini + tamanho_base no banco e calcula var_seg
        on-the-fly usando as cotacoes brutas
    """
    horizonte = params["periodo_seguinte"]

    # Calcula data_corte: amostra-N dias antes do ultimo dia
    ult = conn.execute(
        "SELECT MAX(data) FROM cotacoes WHERE par=?", (par,)
    ).fetchone()[0]
    if not ult:
        return pd.DataFrame()

    dts = [r[0] for r in conn.execute(
        "SELECT data FROM cotacoes WHERE par=? ORDER BY data DESC LIMIT ?",
        (par, params["amostra"])
    ).fetchall()]
    if not dts:
        return pd.DataFrame()
    data_corte = dts[-1]   # mais antiga dentro da amostra

    tam_min = params["tam_base_alvo_resolvido"] - params["faixa_tempo"]
    tam_max = params["tam_base_alvo_resolvido"] + params["faixa_tempo"]
    var_min = params["var_alvo_resolvido"] - params["faixa_baixo"]
    var_max = params["var_alvo_resolvido"] + params["faixa_cima"]

    # ── Caso A: horizonte pre-calculado ─────────────────────────────────────
    if horizonte in HORIZONTES:
        horizonte_col = f"var_seg_{horizonte}"
        sql = f"""
            SELECT data_ini, tamanho_base, preco_ini, preco_fim_base,
                   var_base_pct, {horizonte_col} AS var_seg
            FROM janelas
            WHERE par=?
              AND data_ini >= ?
              AND tamanho_base BETWEEN ? AND ?
              AND var_base_pct BETWEEN ? AND ?
              AND {horizonte_col} IS NOT NULL
            ORDER BY data_ini
        """
        df = pd.read_sql(sql, conn, params=(
            par, data_corte, tam_min, tam_max, var_min, var_max
        ))
        return df

    # ── Caso B: horizonte custom ────────────────────────────────────────────
    # Pega janelas que casam com tempo + variacao (ainda sem var_seg)
    sql = """
        SELECT data_ini, tamanho_base, preco_ini, preco_fim_base, var_base_pct
        FROM janelas
        WHERE par=?
          AND data_ini >= ?
          AND tamanho_base BETWEEN ? AND ?
          AND var_base_pct BETWEEN ? AND ?
        ORDER BY data_ini
    """
    df = pd.read_sql(sql, conn, params=(
        par, data_corte, tam_min, tam_max, var_min, var_max
    ))

    if df.empty:
        df["var_seg"] = []
        return df

    # Carrega TODAS as cotacoes do par como dict {data: (idx, preco)}
    rows = conn.execute(
        "SELECT data, preco FROM cotacoes WHERE par=? ORDER BY data",
        (par,)
    ).fetchall()
    data_para_idx = {data: i for i, (data, _) in enumerate(rows)}
    precos = [p for _, p in rows]
    n = len(precos)

    # Para cada janela, calcula data_fim_base e var_seg N dias depois
    var_segs = []
    for _, row in df.iterrows():
        data_ini = row["data_ini"]
        T = int(row["tamanho_base"])
        idx_ini = data_para_idx.get(data_ini)
        if idx_ini is None:
            var_segs.append(None)
            continue
        idx_fim_base = idx_ini + T - 1
        idx_fim_seg  = idx_fim_base + horizonte
        if idx_fim_seg >= n:
            var_segs.append(None)
            continue
        p_fim_base = precos[idx_fim_base]
        p_fim_seg  = precos[idx_fim_seg]
        if p_fim_base <= 0:
            var_segs.append(None)
            continue
        var_segs.append((p_fim_seg / p_fim_base - 1) * 100)

    df["var_seg"] = var_segs
    df = df.dropna(subset=["var_seg"]).reset_index(drop=True)
    return df


# ─────────────────────────────────────────────────────────────────────────────
#  Geracao do Excel de resultado
# ─────────────────────────────────────────────────────────────────────────────

def classificar(var, lim_caiu, lim_lat, lim_subiu):
    if var <= lim_caiu:               return "CAIU FORTE"
    if var <= -lim_lat:               return "CAIU MEDIO"
    if var <  lim_lat:                return "LATERAL"
    if var <  lim_subiu:              return "SUBIU MEDIO"
    return "SUBIU FORTE"


def montar_histograma(var_seg_series, lim_inf=-15, lim_sup=15, passo=0.5):
    """Retorna (centros_bins, contagens) para o histograma."""
    bins = []
    cur = lim_inf
    while cur < lim_sup:
        bins.append(cur)
        cur += passo
    bins.append(lim_sup)

    centros = []
    contagens = []
    for i in range(len(bins) - 1):
        inf, sup = bins[i], bins[i+1]
        c = ((var_seg_series >= inf) & (var_seg_series < sup)).sum()
        centros.append(round((inf + sup) / 2, 2))
        contagens.append(int(c))
    return centros, contagens


def gerar_excel_resultado(params, cenario, df_similares, out_path, total_testadas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    par = params["par"]

    # ── Titulo ───────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"  RESULTADO  --  {par}  --  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font = F_HDR; c.fill = P_HDR; c.alignment = LFT
    ws.row_dimensions[1].height = 28

    # ── Bloco: Cenario atual ──────────────────────────────────────────────────
    ws.merge_cells("A3:F3")
    c = ws["A3"]
    c.value = "  CENARIO ATUAL"
    c.font = F_SUB; c.fill = P_SUB; c.alignment = LFT
    ws.row_dimensions[3].height = 22

    p_ini, p_fim, var_atu, data_ini, data_fim = cenario
    info_cen = [
        ("Periodo BASE atual (dias)",       params["tam_base_atual"]),
        ("Data inicial",                    data_ini),
        ("Data final",                      data_fim),
        ("Preco inicial",                   round(p_ini, 4)),
        ("Preco final",                     round(p_fim, 4)),
        ("Variacao no periodo BASE",        f"{var_atu:+.2f}%"),
    ]
    for i, (label, valor) in enumerate(info_cen, start=4):
        c = ws.cell(i, 1, label); c.font = F_LBL; c.fill = P_LBL
        c.alignment = LFT; c.border = brd()
        c = ws.cell(i, 2, valor); c.font = F_CALC; c.fill = P_CALC
        c.alignment = CTR; c.border = brd()

    # ── Bloco: Filtro aplicado ────────────────────────────────────────────────
    ws.merge_cells("A11:F11")
    c = ws["A11"]
    c.value = "  FILTRO APLICADO"
    c.font = F_SUB; c.fill = P_SUB; c.alignment = LFT
    ws.row_dimensions[11].height = 22

    info_filtro = [
        ("Tamanho base ALVO",     params["tam_base_alvo_resolvido"]),
        ("Faixa de tempo (+/-)",  params["faixa_tempo"]),
        ("Variacao ALVO",         f"{params['var_alvo_resolvido']:+.2f}%"),
        ("Faixa pra baixo (%)",   params["faixa_baixo"]),
        ("Faixa pra cima (%)",    params["faixa_cima"]),
        ("Periodo seguinte (d)",  params["periodo_seguinte"]),
        ("Amostra (dias)",        params["amostra"]),
    ]
    for i, (label, valor) in enumerate(info_filtro, start=12):
        c = ws.cell(i, 1, label); c.font = F_LBL; c.fill = P_LBL
        c.alignment = LFT; c.border = brd()
        c = ws.cell(i, 2, valor); c.font = F_CALC; c.fill = P_CALC
        c.alignment = CTR; c.border = brd()

    # ── Bloco: Resultado totais ───────────────────────────────────────────────
    ws.merge_cells("A20:F20")
    c = ws["A20"]
    c.value = "  RESULTADO"
    c.font = F_SUB; c.fill = P_DEST; c.alignment = LFT
    ws.row_dimensions[20].height = 22

    n_sim = len(df_similares)

    c = ws.cell(21, 1, "Janelas testadas (na amostra)")
    c.font = F_LBL; c.fill = P_LBL; c.alignment = LFT; c.border = brd()
    c = ws.cell(21, 2, total_testadas)
    c.font = F_OUT; c.fill = P_OUT; c.alignment = CTR; c.border = brd()
    c.number_format = "#,##0"

    c = ws.cell(22, 1, "Janelas SIMILARES")
    c.font = F_LBL; c.fill = P_LBL; c.alignment = LFT; c.border = brd()
    c = ws.cell(22, 2, n_sim)
    c.font = F_BIG; c.fill = P_DEST; c.alignment = CTR; c.border = brd()
    c.number_format = "#,##0"

    if n_sim == 0:
        ws.cell(24, 1, "  Nenhuma janela similar encontrada. Relaxe os filtros.").font = F_NOTE
        wb.save(out_path)
        return

    # ── Bloco: Estatisticas (opcao C) ─────────────────────────────────────────
    var_seg = df_similares["var_seg"].values

    ws.merge_cells("A24:F24")
    c = ws["A24"]
    c.value = "  ESTATISTICAS DO PERIODO SEGUINTE"
    c.font = F_SUB; c.fill = P_OUT2; c.alignment = LFT
    c.font = Font(name="Arial", color="000000", bold=True, size=10)
    ws.row_dimensions[24].height = 22

    import numpy as np
    estats = [
        ("Pior caso (min)",   float(np.min(var_seg))),
        ("Percentil 10",      float(np.percentile(var_seg, 10))),
        ("Percentil 25",      float(np.percentile(var_seg, 25))),
        ("Mediana (50)",      float(np.percentile(var_seg, 50))),
        ("Percentil 75",      float(np.percentile(var_seg, 75))),
        ("Percentil 90",      float(np.percentile(var_seg, 90))),
        ("Melhor caso (max)", float(np.max(var_seg))),
        ("Media",             float(np.mean(var_seg))),
        ("Desvio padrao",     float(np.std(var_seg))),
    ]
    for i, (label, valor) in enumerate(estats, start=25):
        c = ws.cell(i, 1, label); c.font = F_LBL; c.fill = P_LBL
        c.alignment = LFT; c.border = brd()
        c = ws.cell(i, 2, valor); c.font = F_OUT; c.fill = P_OUT2
        c.alignment = CTR; c.border = brd()
        c.number_format = '+0.00"%";-0.00"%";0.00"%"'

    # ── Bloco: Distribuicao por faixas (opcao A) ──────────────────────────────
    ws.merge_cells("A35:F35")
    c = ws["A35"]
    c.value = "  DISTRIBUICAO POR FAIXAS"
    c.font = F_SUB; c.fill = P_OUT; c.alignment = LFT
    c.font = Font(name="Arial", color="000000", bold=True, size=10)
    ws.row_dimensions[35].height = 22

    classes = [classificar(v, params["limite_caiu"], params["limite_lateral"],
                           params["limite_subiu"]) for v in var_seg]
    contagens = {
        "CAIU FORTE":  classes.count("CAIU FORTE"),
        "CAIU MEDIO":  classes.count("CAIU MEDIO"),
        "LATERAL":     classes.count("LATERAL"),
        "SUBIU MEDIO": classes.count("SUBIU MEDIO"),
        "SUBIU FORTE": classes.count("SUBIU FORTE"),
    }
    cores_faixa = {
        "CAIU FORTE":  "FF6B6B",
        "CAIU MEDIO":  "F4B084",
        "LATERAL":     "BFBFBF",
        "SUBIU MEDIO": "A9D08E",
        "SUBIU FORTE": "00D4AA",
    }

    headers = ["Faixa", "Casos", "% das similares"]
    for i, h in enumerate(headers):
        c = ws.cell(36, 1 + i, h)
        c.font = F_LBL; c.fill = P_LBL; c.alignment = CTR; c.border = brd()

    linha_inicio_faixas = 37
    for i, (nome, n) in enumerate(contagens.items()):
        r = linha_inicio_faixas + i
        c = ws.cell(r, 1, nome)
        c.font = Font(name="Arial", bold=True, size=10,
                      color="FFFFFF" if nome in ("CAIU FORTE", "SUBIU FORTE") else "000000")
        c.fill = PatternFill("solid", fgColor=cores_faixa[nome])
        c.alignment = LFT; c.border = brd()

        c = ws.cell(r, 2, n)
        c.font = F_OUT; c.fill = P_OUT; c.alignment = CTR; c.border = brd()
        c.number_format = "0"

        c = ws.cell(r, 3, n / n_sim if n_sim else 0)
        c.font = F_BIG; c.fill = P_OUT; c.alignment = CTR; c.border = brd()
        c.number_format = "0.0%"

    # ── Grafico de faixas ─────────────────────────────────────────────────────
    chart = BarChart()
    chart.type = "col"
    chart.style = 12
    chart.title = "Distribuicao por faixas (% das similares)"
    chart.y_axis.title = "% das similares"

    dados_ref = Reference(ws, min_col=3, max_col=3,
                          min_row=linha_inicio_faixas,
                          max_row=linha_inicio_faixas + 4)
    chart.add_data(dados_ref, titles_from_data=False)

    # Labels curtos em coluna H
    labels = ["Caiu Forte", "Caiu Medio", "Lateral", "Subiu Medio", "Subiu Forte"]
    for i, lbl in enumerate(labels):
        ws.cell(linha_inicio_faixas + i, 8, lbl)
    cat_ref = Reference(ws, min_col=8, max_col=8,
                        min_row=linha_inicio_faixas,
                        max_row=linha_inicio_faixas + 4)
    chart.set_categories(cat_ref)

    # Cores
    series = chart.series[0]
    series.dPt = []
    for i, nome in enumerate(["CAIU FORTE","CAIU MEDIO","LATERAL","SUBIU MEDIO","SUBIU FORTE"]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties = GraphicalProperties(solidFill=cores_faixa[nome])
        series.dPt.append(pt)
    series.dLbls = DataLabelList(showVal=True)
    series.dLbls.numFmt = "0.0%"
    chart.legend = None
    chart.width = 18; chart.height = 9
    ws.add_chart(chart, "E20")

    # ── Histograma ────────────────────────────────────────────────────────────
    centros, conts = montar_histograma(pd.Series(var_seg))
    HIST_LINHA_INI = 50
    ws.cell(HIST_LINHA_INI - 1, 1, "bin_centro").font = F_HDR
    ws.cell(HIST_LINHA_INI - 1, 1).fill = P_HDR
    ws.cell(HIST_LINHA_INI - 1, 2, "contagem").font = F_HDR
    ws.cell(HIST_LINHA_INI - 1, 2).fill = P_HDR
    for i, (cen, n) in enumerate(zip(centros, conts)):
        ws.cell(HIST_LINHA_INI + i, 1, cen).number_format = "0.0"
        ws.cell(HIST_LINHA_INI + i, 2, n).number_format = "0"

    hist = BarChart()
    hist.type = "col"
    hist.style = 12
    hist.title = f"Histograma  --  variacao % no periodo seguinte ({params['periodo_seguinte']} dias)"
    hist.y_axis.title = "Numero de janelas"
    hist.x_axis.title = "Variacao % no periodo seguinte"

    hist_data = Reference(ws, min_col=2, max_col=2,
                          min_row=HIST_LINHA_INI, max_row=HIST_LINHA_INI + len(centros) - 1)
    hist.add_data(hist_data, titles_from_data=False)
    hist_cat = Reference(ws, min_col=1, max_col=1,
                         min_row=HIST_LINHA_INI, max_row=HIST_LINHA_INI + len(centros) - 1)
    hist.set_categories(hist_cat)
    hist.gapWidth = 0
    hist.series[0].graphicalProperties = GraphicalProperties(solidFill="00D4AA")
    hist.x_axis.delete = False
    hist.x_axis.number_format = '0.0"%"'
    hist.x_axis.tickLblSkip = 4
    hist.x_axis.tickMarkSkip = 4
    hist.x_axis.majorTickMark = "out"
    hist.y_axis.delete = False
    hist.y_axis.number_format = "0"
    hist.y_axis.majorTickMark = "out"
    hist.legend = None
    hist.width = 30; hist.height = 12
    ws.add_chart(hist, "E40")

    # ── Aba: detalhe das janelas similares ────────────────────────────────────
    ws_d = wb.create_sheet("Janelas Similares")
    headers = ["data_ini", "tamanho_base", "preco_ini", "preco_fim_base",
               "var_base_%", "var_seg_%"]
    for i, h in enumerate(headers, 1):
        c = ws_d.cell(1, i, h)
        c.font = F_HDR; c.fill = P_HDR; c.alignment = CTR; c.border = brd()
    for r, row in enumerate(df_similares.itertuples(index=False), start=2):
        ws_d.cell(r, 1, row.data_ini)
        ws_d.cell(r, 2, row.tamanho_base)
        ws_d.cell(r, 3, round(row.preco_ini, 4))
        ws_d.cell(r, 4, round(row.preco_fim_base, 4))
        ws_d.cell(r, 5, round(row.var_base_pct, 2))
        ws_d.cell(r, 6, round(row.var_seg, 2))
    for col, w in enumerate([12, 14, 12, 14, 12, 12], 1):
        ws_d.column_dimensions[get_column_letter(col)].width = w

    # Larguras
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    wb.save(out_path)


# ─────────────────────────────────────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  SIMULADOR MULTI-PAR")
    print("=" * 60)

    # Garante que o banco existe
    criar_banco()

    # Le parametros
    params = ler_parametros()
    if params is None:
        return  # criou parametros.xlsx default; usuario deve editar e rodar de novo

    par = params["par"]
    print(f"\nPar selecionado: {par}")

    if par not in PARES:
        print(f"  ERRO: par '{par}' nao esta na lista. Disponiveis: {list(PARES.keys())}")
        return

    # Valida horizonte (aceita qualquer numero positivo)
    if params["periodo_seguinte"] <= 0:
        print(f"  ERRO: periodo_seguinte deve ser positivo")
        return
    if params["periodo_seguinte"] not in HORIZONTES:
        print(f"  Periodo seguinte custom ({params['periodo_seguinte']} dias) — calculando on-the-fly")

    conn = conectar()

    # Verifica cotacoes
    n_cot = conn.execute(
        "SELECT COUNT(*) FROM cotacoes WHERE par=?", (par,)
    ).fetchone()[0]
    if n_cot == 0:
        print(f"\n  Sem cotacoes no banco para {par}.")
        print(f"  Rode primeiro: python multi_par/coletar.py")
        conn.close()
        return
    print(f"  {n_cot:,} cotacoes no banco")

    # Recalcula janelas se preciso (modo a)
    if not janelas_atualizadas(par, conn):
        print(f"  janelas desatualizadas — recalculando...")
        calcular_janelas(par, conn)
    else:
        print(f"  janelas atualizadas")

    # Resolve AUTO
    if str(params["tam_base_alvo"]).strip().upper() == "AUTO":
        params["tam_base_alvo_resolvido"] = params["tam_base_atual"]
    else:
        params["tam_base_alvo_resolvido"] = int(params["tam_base_alvo"])

    # Cenario atual
    cenario = calcular_cenario_atual(conn, par, params["tam_base_atual"])
    if cenario is None:
        print(f"  ERRO: nao tem dados suficientes para periodo_base={params['tam_base_atual']}")
        conn.close()
        return
    p_ini, p_fim, var_atu, data_ini, data_fim = cenario
    print(f"\n  Cenario atual ({data_ini} -> {data_fim}):")
    print(f"    Preco {p_ini:.4f} -> {p_fim:.4f}  ({var_atu:+.2f}%)")

    if str(params["var_alvo"]).strip().upper() == "AUTO":
        params["var_alvo_resolvido"] = var_atu
    else:
        params["var_alvo_resolvido"] = float(params["var_alvo"])

    print(f"  Variacao alvo: {params['var_alvo_resolvido']:+.2f}%")
    print(f"  Filtro: tempo {params['tam_base_alvo_resolvido']} +/- {params['faixa_tempo']}d  |  "
          f"variacao [{params['var_alvo_resolvido']-params['faixa_baixo']:+.2f}, "
          f"{params['var_alvo_resolvido']+params['faixa_cima']:+.2f}]%")

    # Total de janelas testadas (na amostra) — conta janelas onde da pra projetar
    # com o horizonte solicitado.
    dts = [r[0] for r in conn.execute(
        "SELECT data FROM cotacoes WHERE par=? ORDER BY data DESC LIMIT ?",
        (par, params["amostra"])
    ).fetchall()]
    data_corte = dts[-1] if dts else "0000-00-00"

    if params["periodo_seguinte"] in HORIZONTES:
        # Caso pre-calculado: usa a coluna do banco
        horizonte_col = f"var_seg_{params['periodo_seguinte']}"
        total_testadas = conn.execute(
            f"SELECT COUNT(*) FROM janelas "
            f"WHERE par=? AND data_ini >= ? AND {horizonte_col} IS NOT NULL",
            (par, data_corte)
        ).fetchone()[0]
    else:
        # Caso custom: precisa estimar por aproximacao
        # Pegamos o numero de janelas com data_ini >= data_corte e descontamos
        # as que nao caberiam no horizonte (proximas demais do final)
        horizonte = params["periodo_seguinte"]
        # Pega ate uma data que comporte horizonte+max_tamanho_base de folga
        rows_total = conn.execute(
            "SELECT COUNT(*) FROM janelas WHERE par=? AND data_ini >= ?",
            (par, data_corte)
        ).fetchone()[0]
        # Aproximacao: subtrai estimativa de janelas no fim que nao tem horizonte
        # (cada tamanho_base perde aprox `horizonte` janelas no final)
        from janelas import TAMANHOS_BASE as TB
        descarte_estimado = horizonte * len(TB)
        total_testadas = max(0, rows_total - descarte_estimado)

    # Busca similares
    df = buscar_similares(conn, par, params)
    print(f"\n  {total_testadas:,} janelas testadas  -->  {len(df):,} similares")

    conn.close()

    # Gera resultado
    out_path = OUT_DIR / f"resultado_{par}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    gerar_excel_resultado(params, cenario, df, out_path, total_testadas)
    print(f"\n  Resultado salvo em: {out_path}")


if __name__ == "__main__":
    main()
