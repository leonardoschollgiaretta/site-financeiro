"""
gerar_simulador.py -- Simulador de probabilidades USD/BRL (Excel interativo)

Logica:
  1. Define cenario atual: "olhei os ultimos N dias do historico, o dolar variou X%"
  2. Procura no historico janelas similares (tempo similar + variacao similar)
  3. Mostra a distribuicao do que aconteceu nos M dias seguintes a essas janelas

Tudo recalcula automaticamente via formulas nativas do Excel.

Uso: python gerar_simulador.py
"""
import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

sys.path.insert(0, str(Path(__file__).parent))
from src.fetchers.yahoo_fetcher import fetch_currency_history

BASE_DIR = Path(__file__).parent
DATA_RAW = BASE_DIR / "data" / "raw"
OUT_DIR  = BASE_DIR / "outputs"
OUT_DIR.mkdir(exist_ok=True)
SAIDA = OUT_DIR / "simulador_usdbrl.xlsx"

TICKER = "USDBRL=X"
PERIOD = "10y"

# ── Estilos ─────────────────────────────────────────────────────────────────────
P_HDR    = PatternFill("solid", fgColor="1F3864")
P_SUB    = PatternFill("solid", fgColor="2E75B6")
P_INPUT  = PatternFill("solid", fgColor="FFFF00")
P_CALC   = PatternFill("solid", fgColor="DDEBF7")  # azul muito claro = calculado
P_OUT    = PatternFill("solid", fgColor="E2EFDA")
P_OUT2   = PatternFill("solid", fgColor="FFF2CC")
P_LBL    = PatternFill("solid", fgColor="F2F2F2")
P_DEST   = PatternFill("solid", fgColor="00D4AA")
P_NEG    = PatternFill("solid", fgColor="FF6B6B")

F_HDR    = Font(name="Arial", color="FFFFFF", bold=True, size=12)
F_SUB    = Font(name="Arial", color="FFFFFF", bold=True, size=10)
F_LBL    = Font(name="Arial", bold=True, size=10)
F_VAL    = Font(name="Arial", size=10)
F_INPUT  = Font(name="Arial", color="0000FF", bold=True, size=11)
F_CALC   = Font(name="Arial", color="333333", bold=True, size=10)
F_OUT    = Font(name="Arial", bold=True, size=11)
F_BIG    = Font(name="Arial", bold=True, size=14)
F_NOTE   = Font(name="Arial", italic=True, color="666666", size=9)

CTR = Alignment(horizontal="center", vertical="center")
LFT = Alignment(horizontal="left",   vertical="center", indent=1)

def brd():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Coleta ──────────────────────────────────────────────────────────────────────

def carregar_dados():
    DATA_RAW.mkdir(parents=True, exist_ok=True)
    csvs = sorted(DATA_RAW.glob("USDBRL_*.csv"))
    if csvs:
        print(f"Usando CSV existente: {csvs[-1].name}")
        df = pd.read_csv(csvs[-1], index_col=0, parse_dates=True)
    else:
        print(f"Baixando {TICKER} ({PERIOD})...")
        df = fetch_currency_history(TICKER, period=PERIOD, save=True)

    df = df[["Close"]].dropna().sort_index()
    return df


# ── Aba Dados ───────────────────────────────────────────────────────────────────

def criar_aba_dados(wb, df):
    """Aba 'Dados' = serie historica simples (data, preco)."""
    ws = wb.create_sheet("Dados")

    headers = ["data", "preco"]
    for col, txt in enumerate(headers, 1):
        c = ws.cell(1, col, txt)
        c.font = F_HDR; c.fill = P_HDR; c.alignment = CTR; c.border = brd()
    ws.row_dimensions[1].height = 22

    for i, (idx, row) in enumerate(df.iterrows(), start=2):
        ws.cell(i, 1, idx.date()).number_format = "yyyy-mm-dd"
        ws.cell(i, 2, round(float(row["Close"]), 6)).number_format = "0.0000"

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.freeze_panes = "A2"
    return ws, len(df)


# ── Aba Janelas ─────────────────────────────────────────────────────────────────

def criar_aba_janelas(wb, n_dados):
    """
    Aba 'Janelas': para cada combinacao (data_inicio, tamanho_base) gera linhas
    com as variacoes do base e do seguinte.

    Como TEMPO eh variavel (faixa de dias), preciso gerar varios tamanhos por janela.
    Solucao pratica:
      - Para cada possivel data_inicio i (1 a n_dados-base_max-seg)
      - Gera 1 linha com base no TAMANHO_BASE_REPRESENTATIVO calculado por interpolacao

    NOVA ABORDAGEM (mais simples e correta):
      - Cada linha = uma data_inicio i + tamanho_base T (variando de base_min a base_max)
      - Ou seja, expansao por (i, T)
      - Mas isso explode em N x range. Solucao: uso UMA linha por data_inicio,
        e nessa linha calculo TODAS as combinacoes via formulas.

    DECISAO FINAL: gero uma linha por data_inicio, e o teste de similaridade
    e variacao seguinte usam o tamanho base CENTRAL definido no Simulador.
    A faixa de tempo (+/- dias) entra como criterio de FILTRO no Modo simulado:
    para cada janela, eu testo se EXISTE algum sub-tamanho (T) entre base_min e
    base_max em que a variacao do base estaria dentro da faixa similar. Isso
    deixa a formula complexa.

    SIMPLIFICACAO PRATICA (e o que faz mais sentido pro usuario):
      - "Faixa de tempo" significa: testo o periodo base com TAMANHO MEDIO
        (definido pelo usuario), e considero similares todas as janelas onde
        o tamanho efetivo testado caia dentro da faixa.
      - Implementacao: cada linha = data_ini, e calculo a variacao do base
        com tamanho EXATO base_central. Nao filtro por tamanho — uso o tamanho
        central de referencia. A faixa de tempo (B12) eh usada apenas para
        documentacao/UX.

    Mas isso descarta a ideia de "tempo similar". Pra compensar, vou gerar
    MULTIPLAS LINHAS por data_ini, uma para cada tamanho base na faixa
    [base_central - faixa_tempo, base_central + faixa_tempo]. Eh mais correto
    e o overhead eh aceitavel (faixa default 5 -> 11 linhas por data_ini).

    PARA SIMPLIFICAR AINDA MAIS: gero linhas com tamanhos fixos espacados de 1 em 1
    cobrindo de 5 a 60 dias (range razoavel). O Simulador filtra qual cair na
    faixa de tempo.
    """
    ws = wb.create_sheet("Janelas")

    headers = [
        "i", "tamanho_base", "data_ini",
        "preco_ini", "preco_fim_base", "preco_fim_seg",
        "var_base_%", "var_seg_%",
        "tempo_ok", "var_ok", "similar"
    ]
    for col, txt in enumerate(headers, 1):
        c = ws.cell(1, col, txt)
        c.font = F_HDR; c.fill = P_HDR; c.alignment = CTR; c.border = brd()
    ws.row_dimensions[1].height = 22

    # Range de tamanhos base testados (cobre ate 60 dias)
    tamanhos = list(range(5, 61))   # 5, 6, ..., 60

    SIM = "Simulador"
    DADOS = "Dados"
    ult = n_dados + 1   # linha da ultima cotacao em Dados (header = 1)

    # Para cada data_ini i (varre dentro da amostra), gera 1 linha por tamanho_base T
    # Limite: amostra max ~2500. Tamanhos: 56. Total = 2500 * 56 = 140000 linhas.
    # Isso eh muito. Vou reduzir: amostra max efetiva = 1500, tamanhos = 5..50 = 46
    # Total = 1500 * 46 = 69000 linhas. Aceitavel mas ainda pesado.

    # Solucao mais pratica: limitar a amostra aceita pelo simulador a 1000.
    # Total = 1000 * 46 = 46000 linhas.

    AMOSTRA_MAX = 2500
    tamanhos = list(range(5, 51))   # 5..50

    n_linhas = AMOSTRA_MAX * len(tamanhos)

    print(f"  Gerando {n_linhas:,} linhas de janelas (amostra max {AMOSTRA_MAX} x {len(tamanhos)} tamanhos)...")

    r = 2
    for i in range(1, AMOSTRA_MAX + 1):
        for T in tamanhos:
            # Coluna A: i (numero da janela)
            ws.cell(r, 1, i)
            # Coluna B: tamanho_base
            ws.cell(r, 2, T)

            # Indice no Dados: linha = ult - amostra + i
            # data_ini = INDEX(Dados!A:A, ult - amostra + i)
            ws.cell(r, 3,
                f"=IFERROR(INDEX({DADOS}!A:A,{ult}-{SIM}!$B$5+A{r}),\"\")"
            ).number_format = "yyyy-mm-dd"

            # ult = linha da ultima cotacao em Dados ({ult})
            # A janela so e valida se o indice do "fim do seguinte" nao passar de {ult}
            # idx_fim_seg = ult - amostra + i + T + seg - 1
            # condicao: idx_fim_seg <= ult  =>  i + T + seg - 1 <= amostra
            # Em formula: A{r} + B{r} + B6 - 1 <= B5

            # preco_ini = INDEX(Dados!B:B, ult - amostra + i)  -- so se indice >= 2 (linha 1 = header)
            ws.cell(r, 4,
                f'=IF(({ult}-{SIM}!$B$5+A{r})<2,"",'
                f'IFERROR(INDEX({DADOS}!B:B,{ult}-{SIM}!$B$5+A{r}),""))'
            ).number_format = "0.0000"

            # preco_fim_base = INDEX(Dados!B:B, ult - amostra + i + T - 1)
            ws.cell(r, 5,
                f'=IF(({ult}-{SIM}!$B$5+A{r}+B{r}-1)>{ult},"",'
                f'IFERROR(INDEX({DADOS}!B:B,{ult}-{SIM}!$B$5+A{r}+B{r}-1),""))'
            ).number_format = "0.0000"

            # preco_fim_seg = INDEX(Dados!B:B, ult - amostra + i + T + seg - 1)
            # CRITICO: aqui precisa garantir que o indice nao passa de {ult}
            ws.cell(r, 6,
                f'=IF(({ult}-{SIM}!$B$5+A{r}+B{r}+{SIM}!$B$6-1)>{ult},"",'
                f'IFERROR(INDEX({DADOS}!B:B,{ult}-{SIM}!$B$5+A{r}+B{r}+{SIM}!$B$6-1),""))'
            ).number_format = "0.0000"

            # var_base_% = (preco_fim_base / preco_ini - 1) * 100
            # ISNUMBER garante que so calcula se ambos forem numeros validos
            ws.cell(r, 7,
                f'=IF(AND(ISNUMBER(D{r}),ISNUMBER(E{r}),D{r}>0),(E{r}/D{r}-1)*100,"")'
            ).number_format = "0.00"

            # var_seg_% = (preco_fim_seg / preco_fim_base - 1) * 100
            ws.cell(r, 8,
                f'=IF(AND(ISNUMBER(E{r}),ISNUMBER(F{r}),E{r}>0),(F{r}/E{r}-1)*100,"")'
            ).number_format = "0.00"

            # tempo_ok: tamanho_base esta dentro de [base_central - faixa_tempo, base_central + faixa_tempo]
            ws.cell(r, 9,
                f'=IF(AND(B{r}>=({SIM}!$B$11-{SIM}!$B$12),B{r}<=({SIM}!$B$11+{SIM}!$B$12)),1,0)'
            )

            # var_ok: var_base esta dentro de [alvo_var - faixa_var, alvo_var + faixa_var]
            # var_ok: var_base esta dentro de [alvo - faixa_baixo, alvo + faixa_cima]
            # B13 = alvo, B14 = faixa para baixo, B15 = faixa para cima
            ws.cell(r, 10,
                f'=IFERROR(IF(AND(G{r}>=({SIM}!$B$13-{SIM}!$B$14),G{r}<=({SIM}!$B$13+{SIM}!$B$15)),1,0),0)'
            )

            # similar = tempo_ok AND var_ok AND var_seg nao vazia
            ws.cell(r, 11,
                f'=IF(AND(I{r}=1,J{r}=1,ISNUMBER(H{r})),1,0)'
            )

            r += 1

    # Larguras
    larguras = [5, 12, 12, 11, 13, 13, 11, 11, 9, 9, 9]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    return ws, n_linhas


# ── Aba Simulador ───────────────────────────────────────────────────────────────

def criar_aba_simulador(wb, n_dados):
    """Aba principal."""
    ws = wb.create_sheet("Simulador", 0)

    # ── Titulo ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "  SIMULADOR DE PROBABILIDADES  --  USD/BRL"
    c.font = F_HDR; c.fill = P_HDR; c.alignment = LFT
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = ("  Edite as celulas amarelas. O Excel recalcula automaticamente. "
               "Base: ultimos 10 anos do Yahoo Finance.")
    c.font = F_NOTE; c.alignment = LFT
    ws.row_dimensions[2].height = 18

    # ╔══════════════════════════════════════════════════════════════════════════╗
    # ║ BLOCO 1: AMOSTRA + PERIODO SEGUINTE                                       ║
    # ╚══════════════════════════════════════════════════════════════════════════╝
    ws.merge_cells("A4:F4")
    c = ws["A4"]
    c.value = "  AMOSTRA HISTORICA + PERIODO SEGUINTE"
    c.font = F_SUB; c.fill = P_SUB; c.alignment = LFT
    ws.row_dimensions[4].height = 22

    # B5 = amostra
    ws.cell(5, 1, "Tamanho da amostra (dias do historico)").font = F_LBL
    ws.cell(5, 1).fill = P_LBL; ws.cell(5, 1).alignment = LFT; ws.cell(5, 1).border = brd()
    ws.cell(5, 2, 2500)
    ws.cell(5, 2).font = F_INPUT; ws.cell(5, 2).fill = P_INPUT
    ws.cell(5, 2).alignment = CTR; ws.cell(5, 2).border = brd()
    ws.cell(5, 2).number_format = "0"
    ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=6)
    ws.cell(5, 3, "  Quantos dias do historico vao ser varridos. Maximo: 2500 (~10 anos).").font = F_NOTE

    # B6 = periodo seguinte (a projetar)
    ws.cell(6, 1, "Periodo SEGUINTE a projetar (dias)").font = F_LBL
    ws.cell(6, 1).fill = P_LBL; ws.cell(6, 1).alignment = LFT; ws.cell(6, 1).border = brd()
    ws.cell(6, 2, 10)
    ws.cell(6, 2).font = F_INPUT; ws.cell(6, 2).fill = P_INPUT
    ws.cell(6, 2).alignment = CTR; ws.cell(6, 2).border = brd()
    ws.cell(6, 2).number_format = "0"
    ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=6)
    ws.cell(6, 3, "  Quantos dias 'depois' do periodo base eu quero analisar. Max: 30.").font = F_NOTE

    # ╔══════════════════════════════════════════════════════════════════════════╗
    # ║ BLOCO 2: CENARIO ATUAL                                                    ║
    # ╚══════════════════════════════════════════════════════════════════════════╝
    ws.merge_cells("A8:F8")
    c = ws["A8"]
    c.value = "  CENARIO ATUAL"
    c.font = F_SUB; c.fill = P_SUB; c.alignment = LFT
    ws.row_dimensions[8].height = 22

    # B9 = tamanho base atual (quantos dias atras eu olho)
    ws.cell(9, 1, "Tamanho do periodo BASE atual (dias)").font = F_LBL
    ws.cell(9, 1).fill = P_LBL; ws.cell(9, 1).alignment = LFT; ws.cell(9, 1).border = brd()
    ws.cell(9, 2, 20)
    ws.cell(9, 2).font = F_INPUT; ws.cell(9, 2).fill = P_INPUT
    ws.cell(9, 2).alignment = CTR; ws.cell(9, 2).border = brd()
    ws.cell(9, 2).number_format = "0"
    ws.merge_cells(start_row=9, start_column=3, end_row=9, end_column=6)
    ws.cell(9, 3, '  Quantos dias atras eu olho. Ex: 20 = "ultimos 20 dias do dolar"').font = F_NOTE

    # ╔══════════════════════════════════════════════════════════════════════════╗
    # ║ BLOCO 3: FILTRO DE SIMILARIDADE                                           ║
    # ╚══════════════════════════════════════════════════════════════════════════╝
    ws.merge_cells("A10:F10")
    c = ws["A10"]
    c.value = "  FILTRO DE SIMILARIDADE  (eixos: tempo + variacao)"
    c.font = F_SUB; c.fill = P_SUB; c.alignment = LFT
    ws.row_dimensions[10].height = 22

    # B11 = base central (alvo de tempo) — DEFAULT = igual ao tamanho base atual (B9)
    ws.cell(11, 1, "Tamanho base ALVO (dias)").font = F_LBL
    ws.cell(11, 1).fill = P_LBL; ws.cell(11, 1).alignment = LFT; ws.cell(11, 1).border = brd()
    ws.cell(11, 2, "=B9")
    ws.cell(11, 2).font = F_INPUT; ws.cell(11, 2).fill = P_INPUT
    ws.cell(11, 2).alignment = CTR; ws.cell(11, 2).border = brd()
    ws.cell(11, 2).number_format = "0"
    ws.merge_cells(start_row=11, start_column=3, end_row=11, end_column=6)
    ws.cell(11, 3, "  Default = tamanho base atual. Sobrescreva pra simular outros prazos.").font = F_NOTE

    # B12 = faixa de tempo (+/- dias)
    ws.cell(12, 1, "Faixa de TEMPO (+/- dias)").font = F_LBL
    ws.cell(12, 1).fill = P_LBL; ws.cell(12, 1).alignment = LFT; ws.cell(12, 1).border = brd()
    ws.cell(12, 2, 5)
    ws.cell(12, 2).font = F_INPUT; ws.cell(12, 2).fill = P_INPUT
    ws.cell(12, 2).alignment = CTR; ws.cell(12, 2).border = brd()
    ws.cell(12, 2).number_format = "0"
    ws.merge_cells(start_row=12, start_column=3, end_row=12, end_column=6)
    ws.cell(12, 3, '  Ex: alvo 20, faixa 5 -> aceita janelas de 15 a 25 dias').font = F_NOTE

    # B13 = variacao alvo (default = variacao atual calculada)
    ws.cell(13, 1, "Variacao ALVO (%)").font = F_LBL
    ws.cell(13, 1).fill = P_LBL; ws.cell(13, 1).alignment = LFT; ws.cell(13, 1).border = brd()
    # Calcula automaticamente a variacao dos ultimos B9 dias
    ws.cell(13, 2,
        f"=(INDEX(Dados!B:B,{n_dados+1})/INDEX(Dados!B:B,{n_dados+1}-B9+1)-1)*100"
    )
    ws.cell(13, 2).font = F_INPUT; ws.cell(13, 2).fill = P_INPUT
    ws.cell(13, 2).alignment = CTR; ws.cell(13, 2).border = brd()
    ws.cell(13, 2).number_format = "0.00"
    ws.merge_cells(start_row=13, start_column=3, end_row=13, end_column=6)
    ws.cell(13, 3, "  Default = variacao atual. Sobrescreva pra simular cenarios hipoteticos.").font = F_NOTE

    # B14 = faixa de variacao para BAIXO (% a menos do alvo)
    ws.cell(14, 1, "Faixa de variacao PARA BAIXO (% a menos)").font = F_LBL
    ws.cell(14, 1).fill = P_LBL; ws.cell(14, 1).alignment = LFT; ws.cell(14, 1).border = brd()
    ws.cell(14, 2, 1.0)
    ws.cell(14, 2).font = F_INPUT; ws.cell(14, 2).fill = P_INPUT
    ws.cell(14, 2).alignment = CTR; ws.cell(14, 2).border = brd()
    ws.cell(14, 2).number_format = "0.00"
    ws.merge_cells(start_row=14, start_column=3, end_row=14, end_column=6)
    ws.cell(14, 3, "  Quanto pra BAIXO do alvo eu aceito. Ex: alvo -2.88 e B14=1 -> limite inferior -3.88").font = F_NOTE

    # B15 = faixa de variacao para CIMA (% a mais do alvo)
    ws.cell(15, 1, "Faixa de variacao PARA CIMA (% a mais)").font = F_LBL
    ws.cell(15, 1).fill = P_LBL; ws.cell(15, 1).alignment = LFT; ws.cell(15, 1).border = brd()
    ws.cell(15, 2, 1.0)
    ws.cell(15, 2).font = F_INPUT; ws.cell(15, 2).fill = P_INPUT
    ws.cell(15, 2).alignment = CTR; ws.cell(15, 2).border = brd()
    ws.cell(15, 2).number_format = "0.00"
    ws.merge_cells(start_row=15, start_column=3, end_row=15, end_column=6)
    ws.cell(15, 3, "  Quanto pra CIMA do alvo eu aceito. Ex: alvo -2.88 e B15=3 -> limite superior +0.12").font = F_NOTE

    # ╔══════════════════════════════════════════════════════════════════════════╗
    # ║ BLOCO 4: CALCULO AUTOMATICO DO CENARIO ATUAL                              ║
    # ╚══════════════════════════════════════════════════════════════════════════╝
    ws.merge_cells("A16:F16")
    c = ws["A16"]
    c.value = "  RESUMO DO CENARIO ATUAL  (calculado a partir dos dados)"
    c.font = F_SUB; c.fill = P_CALC; c.alignment = LFT
    c.font = Font(name="Arial", color="000000", bold=True, size=10)
    ws.row_dimensions[16].height = 22

    ws.cell(17, 1, "Preco inicial (atual − base dias)").font = F_LBL
    ws.cell(17, 1).fill = P_LBL; ws.cell(17, 1).alignment = LFT; ws.cell(17, 1).border = brd()
    ws.cell(17, 2, f"=INDEX(Dados!B:B,{n_dados+1}-B9+1)")
    ws.cell(17, 2).font = F_CALC; ws.cell(17, 2).fill = P_CALC
    ws.cell(17, 2).alignment = CTR; ws.cell(17, 2).border = brd()
    ws.cell(17, 2).number_format = "0.0000"

    ws.cell(18, 1, "Preco final (atual)").font = F_LBL
    ws.cell(18, 1).fill = P_LBL; ws.cell(18, 1).alignment = LFT; ws.cell(18, 1).border = brd()
    ws.cell(18, 2, f"=INDEX(Dados!B:B,{n_dados+1})")
    ws.cell(18, 2).font = F_CALC; ws.cell(18, 2).fill = P_CALC
    ws.cell(18, 2).alignment = CTR; ws.cell(18, 2).border = brd()
    ws.cell(18, 2).number_format = "0.0000"

    ws.cell(19, 1, "Variacao no periodo BASE atual").font = F_LBL
    ws.cell(19, 1).fill = P_LBL; ws.cell(19, 1).alignment = LFT; ws.cell(19, 1).border = brd()
    ws.cell(19, 2, "=(B18/B17-1)*100")
    ws.cell(19, 2).font = F_CALC; ws.cell(19, 2).fill = P_CALC
    ws.cell(19, 2).alignment = CTR; ws.cell(19, 2).border = brd()
    ws.cell(19, 2).number_format = '+0.00"%";-0.00"%"'

    # ╔══════════════════════════════════════════════════════════════════════════╗
    # ║ BLOCO 5: RESULTADO — TOTAL DE JANELAS SIMILARES                           ║
    # ╚══════════════════════════════════════════════════════════════════════════╝
    ws.merge_cells("A21:F21")
    c = ws["A21"]
    c.value = "  RESULTADO  --  janelas similares encontradas"
    c.font = F_SUB; c.fill = P_DEST; c.alignment = LFT
    ws.row_dimensions[21].height = 22

    ws.cell(22, 1, "Total de janelas testadas").font = F_LBL
    ws.cell(22, 1).fill = P_LBL; ws.cell(22, 1).alignment = LFT; ws.cell(22, 1).border = brd()
    ws.cell(22, 2, '=COUNTIFS(Janelas!I:I,1,Janelas!H:H,"<>")')   # tempo_ok=1 e var_seg numerica
    # Mais simples: total efetivamente analisaveis = janelas com var_seg numerica e dentro da amostra
    ws.cell(22, 2,
        '=COUNTIFS(Janelas!H:H,"<>",Janelas!A:A,"<="&B5)'
    )
    ws.cell(22, 2).font = F_OUT; ws.cell(22, 2).fill = P_OUT
    ws.cell(22, 2).alignment = CTR; ws.cell(22, 2).border = brd()
    ws.cell(22, 2).number_format = "0"

    ws.cell(23, 1, "Janelas SIMILARES").font = F_LBL
    ws.cell(23, 1).fill = P_LBL; ws.cell(23, 1).alignment = LFT; ws.cell(23, 1).border = brd()
    ws.cell(23, 2, '=COUNTIFS(Janelas!K:K,1,Janelas!A:A,"<="&B5)')
    ws.cell(23, 2).font = F_BIG; ws.cell(23, 2).fill = P_DEST
    ws.cell(23, 2).alignment = CTR; ws.cell(23, 2).border = brd()
    ws.cell(23, 2).number_format = "0"
    ws.merge_cells(start_row=23, start_column=3, end_row=23, end_column=6)
    ws.cell(23, 3, "  Janelas onde tempo + variacao bateram com o cenario atual.").font = F_NOTE

    # ╔══════════════════════════════════════════════════════════════════════════╗
    # ║ BLOCO 6: ESTATISTICAS (OPCAO C)                                           ║
    # ╚══════════════════════════════════════════════════════════════════════════╝
    ws.merge_cells("A25:F25")
    c = ws["A25"]
    c.value = "  ESTATISTICAS DO PERIODO SEGUINTE  (opcao C)"
    c.font = F_SUB; c.fill = P_OUT2; c.alignment = LFT
    c.font = Font(name="Arial", color="000000", bold=True, size=10)
    ws.row_dimensions[25].height = 22

    # Para calcular percentis com filtro, uso fórmula matricial:
    # PERCENTILE(IF(Janelas!K:K=1, Janelas!H:H), p)
    # No openpyxl, formulas matriciais precisam ser entradas com {= } e isso eh chato.
    # Solucao: PERCENTILE.INC com SUMPRODUCT ou uso _xlfn.PERCENTILE para evitar
    # array formula. Vou usar uma abordagem com array formula via openpyxl ArrayFormula.

    # Abordagem mais simples: nao usar array. Em vez disso, na propria aba Janelas
    # tenho uma coluna L que copia o var_seg APENAS quando similar=1, e NA() caso contrario.
    # Ai PERCENTILE.INC nessa coluna ignora os NA.
    # Vou adicionar essa coluna agora — depois ajusto a aba.

    # Por enquanto coloco placeholder com PERCENTILE em coluna L da aba Janelas:
    ESTATS = [
        ("Pior caso (min)",      "=IFERROR(MIN(Janelas!L:L),0)"),
        ("Percentil 10",         "=IFERROR(_xlfn.PERCENTILE.INC(Janelas!L:L,0.10),0)"),
        ("Percentil 25",         "=IFERROR(_xlfn.PERCENTILE.INC(Janelas!L:L,0.25),0)"),
        ("MEDIANA (50)",         "=IFERROR(_xlfn.PERCENTILE.INC(Janelas!L:L,0.50),0)"),
        ("Percentil 75",         "=IFERROR(_xlfn.PERCENTILE.INC(Janelas!L:L,0.75),0)"),
        ("Percentil 90",         "=IFERROR(_xlfn.PERCENTILE.INC(Janelas!L:L,0.90),0)"),
        ("Melhor caso (max)",    "=IFERROR(MAX(Janelas!L:L),0)"),
        ("Media",                "=IFERROR(AVERAGE(Janelas!L:L),0)"),
        ("Desvio padrao",        "=IFERROR(STDEV(Janelas!L:L),0)"),
    ]
    for i, (label, formula) in enumerate(ESTATS):
        r = 26 + i
        c = ws.cell(r, 1, label)
        c.font = F_LBL; c.fill = P_LBL; c.alignment = LFT; c.border = brd()
        c = ws.cell(r, 2, formula)
        c.font = F_OUT; c.fill = P_OUT2
        c.alignment = CTR; c.border = brd()
        c.number_format = '+0.00"%";-0.00"%";0.00"%"'

    # ╔══════════════════════════════════════════════════════════════════════════╗
    # ║ BLOCO 7: DISTRIBUICAO POR FAIXAS (OPCAO A)                                ║
    # ╚══════════════════════════════════════════════════════════════════════════╝
    ws.merge_cells("A36:F36")
    c = ws["A36"]
    c.value = "  DISTRIBUICAO POR FAIXAS  (opcao A — limites editaveis)"
    c.font = F_SUB; c.fill = P_OUT; c.alignment = LFT
    c.font = Font(name="Arial", color="000000", bold=True, size=10)
    ws.row_dimensions[36].height = 22

    # Limites editaveis (B37, B38, B39)
    ws.cell(37, 1, "Limite CAIU FORTE (var <= X%)").font = F_LBL
    ws.cell(37, 1).fill = P_LBL; ws.cell(37, 1).alignment = LFT; ws.cell(37, 1).border = brd()
    ws.cell(37, 2, -5.0)
    ws.cell(37, 2).font = F_INPUT; ws.cell(37, 2).fill = P_INPUT
    ws.cell(37, 2).alignment = CTR; ws.cell(37, 2).border = brd()
    ws.cell(37, 2).number_format = "0.00"

    ws.cell(38, 1, "Limite LATERAL (faixa +/- X%)").font = F_LBL
    ws.cell(38, 1).fill = P_LBL; ws.cell(38, 1).alignment = LFT; ws.cell(38, 1).border = brd()
    ws.cell(38, 2, 2.0)
    ws.cell(38, 2).font = F_INPUT; ws.cell(38, 2).fill = P_INPUT
    ws.cell(38, 2).alignment = CTR; ws.cell(38, 2).border = brd()
    ws.cell(38, 2).number_format = "0.00"

    ws.cell(39, 1, "Limite SUBIU FORTE (var >= X%)").font = F_LBL
    ws.cell(39, 1).fill = P_LBL; ws.cell(39, 1).alignment = LFT; ws.cell(39, 1).border = brd()
    ws.cell(39, 2, 5.0)
    ws.cell(39, 2).font = F_INPUT; ws.cell(39, 2).fill = P_INPUT
    ws.cell(39, 2).alignment = CTR; ws.cell(39, 2).border = brd()
    ws.cell(39, 2).number_format = "0.00"

    # Tabela de faixas — 5 categorias usando os limites acima
    # Caiu forte:    var <= -5
    # Caiu medio:    -5 < var <= -2
    # Lateral:       -2 < var < +2
    # Subiu medio:   +2 <= var < +5
    # Subiu forte:   var >= +5
    headers_a = ["Faixa", "Casos", "% das similares"]
    for i, h in enumerate(headers_a):
        c = ws.cell(41, 1 + i, h)
        c.font = F_LBL; c.fill = P_LBL; c.alignment = CTR; c.border = brd()

    # Cada linha usa COUNTIFS com 2 criterios em coluna L (var_seg quando similar=1)
    faixas = [
        ('Caiu FORTE  (<= L_caiu_forte)',
         '=COUNTIFS(Janelas!L:L,"<="&B37)',
         P_NEG, "FFFFFF"),

        ('Caiu MEDIO  (entre limite caiu e -lateral)',
         '=COUNTIFS(Janelas!L:L,">"&B37,Janelas!L:L,"<="&-B38)',
         PatternFill("solid", fgColor="F4B084"), "000000"),

        ('LATERAL     (-lateral a +lateral)',
         '=COUNTIFS(Janelas!L:L,">"&-B38,Janelas!L:L,"<"&B38)',
         P_LBL, "000000"),

        ('Subiu MEDIO (entre +lateral e limite subiu)',
         '=COUNTIFS(Janelas!L:L,">="&B38,Janelas!L:L,"<"&B39)',
         PatternFill("solid", fgColor="A9D08E"), "000000"),

        ('Subiu FORTE (>= L_subiu_forte)',
         '=COUNTIFS(Janelas!L:L,">="&B39)',
         P_DEST, "FFFFFF"),
    ]
    for i, (label, formula, fill, cor_txt) in enumerate(faixas):
        r = 42 + i
        c = ws.cell(r, 1, label)
        c.font = Font(name="Arial", bold=True, size=10, color=cor_txt)
        c.fill = fill; c.alignment = LFT; c.border = brd()

        c = ws.cell(r, 2, formula)
        c.font = F_OUT; c.fill = P_OUT
        c.alignment = CTR; c.border = brd()
        c.number_format = "0"

        c = ws.cell(r, 3, f"=IFERROR(B{r}/$B$23,0)")
        c.font = F_BIG; c.fill = P_OUT
        c.alignment = CTR; c.border = brd()
        c.number_format = "0.0%"

    # ╔══════════════════════════════════════════════════════════════════════════╗
    # ║ GRAFICO: distribuicao das 5 faixas                                        ║
    # ╚══════════════════════════════════════════════════════════════════════════╝
    chart = BarChart()
    chart.type = "col"
    chart.style = 12
    chart.title = "Distribuicao do periodo seguinte (% das janelas similares)"
    chart.y_axis.title = "% das similares"
    chart.x_axis.title = None

    # Dados: coluna C42:C46 (% das similares)
    dados_ref = Reference(ws, min_col=3, max_col=3, min_row=42, max_row=46)
    chart.add_data(dados_ref, titles_from_data=False)

    # Categorias: usa labels curtos das faixas
    # Em vez de ler A42:A46 (que tem texto longo), crio labels curtos numa area auxiliar
    # Coloco em H42:H46
    labels_curtos = [
        ("Caiu Forte",  "FF6B6B"),
        ("Caiu Medio",  "F4B084"),
        ("Lateral",     "BFBFBF"),
        ("Subiu Medio", "A9D08E"),
        ("Subiu Forte", "00D4AA"),
    ]
    for i, (lbl, _) in enumerate(labels_curtos):
        ws.cell(42 + i, 8, lbl)

    cat_ref = Reference(ws, min_col=8, max_col=8, min_row=42, max_row=46)
    chart.set_categories(cat_ref)

    # Cores das barras (uma cor por barra)
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.chart.marker import DataPoint
    series = chart.series[0]
    series.dPt = []
    for i, (_, hex_cor) in enumerate(labels_curtos):
        pt = DataPoint(idx=i)
        pt.graphicalProperties = GraphicalProperties(solidFill=hex_cor)
        series.dPt.append(pt)

    # Rotulos no topo das barras com valor %
    series.dLbls = DataLabelList(showVal=True)
    series.dLbls.numFmt = "0.0%"

    # Esconde a legenda (so 1 serie)
    chart.legend = None

    # Tamanho do grafico (em "cm" da unidade do openpyxl)
    chart.width = 18    # ~7 polegadas
    chart.height = 9    # ~3.5 polegadas

    # Ancora na celula E40 (logo a direita da tabela de faixas)
    ws.add_chart(chart, "E40")

    # ╔══════════════════════════════════════════════════════════════════════════╗
    # ║ HISTOGRAMA: distribuicao completa em 60 bins (range fixo -15% a +15%)     ║
    # ╚══════════════════════════════════════════════════════════════════════════╝
    # Tabela auxiliar de bins: colunas O (limite inf), P (centro/label), Q (contagem)
    # Linhas 2 a 61 (60 bins, passos de 0.5%)
    HIST_COL_INF   = 15   # O
    HIST_COL_LABEL = 16   # P
    HIST_COL_CONT  = 17   # Q
    HIST_LINHA_INI = 2
    HIST_LINHA_FIM = 61   # 60 bins

    # Cabecalhos
    ws.cell(1, HIST_COL_INF,   "bin_inf")
    ws.cell(1, HIST_COL_LABEL, "bin_centro")
    ws.cell(1, HIST_COL_CONT,  "contagem")
    for col in (HIST_COL_INF, HIST_COL_LABEL, HIST_COL_CONT):
        c = ws.cell(1, col)
        c.font = F_HDR; c.fill = P_HDR; c.alignment = CTR

    # Bins de -15% a +15% em passos de 0.5%
    bin_inf = -15.0
    passo   = 0.5
    for i in range(60):
        r = HIST_LINHA_INI + i
        inf  = round(bin_inf + i * passo, 2)
        sup  = round(inf + passo, 2)
        centro = round((inf + sup) / 2, 2)
        # Coluna O: limite inferior (incluido)
        ws.cell(r, HIST_COL_INF,   inf).number_format = "0.0"
        # Coluna P: centro do bin (label do eixo X) — texto pra ficar legivel
        ws.cell(r, HIST_COL_LABEL, centro).number_format = "0.0"
        # Coluna Q: COUNTIFS com os limites
        # Conta var_seg_filtrada (Janelas!L) entre [inf, sup)
        ws.cell(r, HIST_COL_CONT,
            f'=COUNTIFS(Janelas!L:L,">="&{get_column_letter(HIST_COL_INF)}{r},'
            f'Janelas!L:L,"<"&({get_column_letter(HIST_COL_INF)}{r}+{passo}))'
        ).number_format = "0"

    # Largura das colunas auxiliares
    ws.column_dimensions[get_column_letter(HIST_COL_INF)].width   = 10
    ws.column_dimensions[get_column_letter(HIST_COL_LABEL)].width = 10
    ws.column_dimensions[get_column_letter(HIST_COL_CONT)].width  = 10

    # Cria o histograma (gráfico de colunas com gap=0 = barras coladas)
    hist = BarChart()
    hist.type = "col"
    hist.style = 12
    hist.title = "Histograma  --  variacao do periodo seguinte (% nas janelas similares)"
    hist.y_axis.title = "Numero de janelas"
    hist.x_axis.title = "Variacao % no periodo seguinte"

    hist_data = Reference(ws, min_col=HIST_COL_CONT, max_col=HIST_COL_CONT,
                          min_row=HIST_LINHA_INI, max_row=HIST_LINHA_FIM)
    hist.add_data(hist_data, titles_from_data=False)

    hist_cat = Reference(ws, min_col=HIST_COL_LABEL, max_col=HIST_COL_LABEL,
                         min_row=HIST_LINHA_INI, max_row=HIST_LINHA_FIM)
    hist.set_categories(hist_cat)

    # Barras coladas (gap=0)
    hist.gapWidth = 0

    # Cor unica das barras (verde-agua, igual ao histograma de retornos diarios)
    serie_hist = hist.series[0]
    serie_hist.graphicalProperties = GraphicalProperties(solidFill="00D4AA")

    # ─── Eixos: garantir que rotulos / numeros aparecam ─────────────────────
    # Eixo X (categorias): forca exibicao dos labels com formato numerico
    hist.x_axis.delete = False
    hist.x_axis.number_format = '0.0"%"'
    # Reduz frequencia de labels pra nao virar bagunca (60 bins -> mostra 1 a cada 4 = 15 labels)
    hist.x_axis.tickLblSkip = 4
    hist.x_axis.tickMarkSkip = 4
    hist.x_axis.majorTickMark = "out"

    # Eixo Y: garantir que numeros apareçam
    hist.y_axis.delete = False
    hist.y_axis.number_format = "0"
    hist.y_axis.majorTickMark = "out"

    # Sem legenda
    hist.legend = None

    # Tamanho maior — ocupa mais espaco horizontal
    hist.width = 30   # ~12 polegadas
    hist.height = 12  # ~5 polegadas

    # Ancora abaixo do primeiro grafico
    ws.add_chart(hist, "E60")

    # ╔══════════════════════════════════════════════════════════════════════════╗
    # ║ BLOCO 8: COMO LER                                                          ║
    # ╚══════════════════════════════════════════════════════════════════════════╝
    ws.merge_cells("A48:F48")
    c = ws["A48"]
    c.value = "  COMO LER ESTE SIMULADOR"
    c.font = F_SUB; c.fill = P_SUB; c.alignment = LFT
    ws.row_dimensions[48].height = 22

    explicacao = [
        "1. Voce define o cenario ATUAL (B9 = quantos dias atras eu olho).",
        "2. O simulador calcula sozinho a variacao do dolar nesse periodo.",
        "3. Voce define o que considera 'parecido' (B11-B14): faixa de tempo e variacao.",
        "4. O simulador procura no historico janelas similares.",
        "5. Mostra a distribuicao do que aconteceu nos B6 dias seguintes a essas janelas.",
        "",
        "DICA: para simular hipoteses, sobrescreva B11 (alvo tempo) e B13 (alvo variacao).",
        "      Ex: B13 = -5 simula 'e se o dolar tivesse caido 5%? O que costuma vir depois?'",
        "",
        "ATENCAO: poucos casos similares (B23 < 20) = estatistica pouco confiavel.",
        "         Aumente a faixa de tempo (B12) ou variacao (B14) pra ter mais amostra.",
    ]
    for i, txt in enumerate(explicacao):
        ws.cell(49 + i, 1, "  " + txt).font = F_NOTE
        ws.merge_cells(start_row=49+i, start_column=1, end_row=49+i, end_column=6)

    # Larguras
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14


def adicionar_coluna_var_seg_filtrada(ws_janelas, n_linhas):
    """Adiciona coluna L na aba Janelas: var_seg quando similar=1, NA() caso contrario.
    Permite usar PERCENTILE / MIN / MAX / AVERAGE direto na coluna sem array formula."""
    # Header
    c = ws_janelas.cell(1, 12, "var_seg_filtrada")
    c.font = F_HDR; c.fill = P_HDR; c.alignment = CTR; c.border = brd()

    # Para cada linha, similar=K, var_seg=H. Se similar=1 E dentro da amostra, retorna H, senao "" (vazio)
    # MAS: PERCENTILE / MIN ignora textos. So precisa que linhas nao-similares fiquem texto vazio.
    # Tambem precisa filtrar pela amostra: i (col A) <= B5
    for r in range(2, n_linhas + 2):
        ws_janelas.cell(r, 12,
            f'=IF(AND(K{r}=1,A{r}<=Simulador!$B$5),H{r},"")'
        ).number_format = "0.00"

    ws_janelas.column_dimensions[get_column_letter(12)].width = 14


def main():
    print("=" * 60)
    print("  GERADOR DE SIMULADOR USD/BRL  (Excel interativo)")
    print("=" * 60)

    df = carregar_dados()
    n_dados = len(df)
    print(f"\n{n_dados} dias de dados ({df.index.min().date()} a {df.index.max().date()})")

    wb = Workbook()
    wb.remove(wb.active)

    print("\nGerando aba 'Dados'...")
    criar_aba_dados(wb, df)

    print("Gerando aba 'Simulador'...")
    criar_aba_simulador(wb, n_dados)

    print("Gerando aba 'Janelas' (formulas pesadas)...")
    ws_jan, n_linhas = criar_aba_janelas(wb, n_dados)

    print(f"Adicionando coluna de var_seg filtrada ({n_linhas:,} linhas)...")
    adicionar_coluna_var_seg_filtrada(ws_jan, n_linhas)

    # Reordena: Simulador, Dados, Janelas
    wb._sheets = [
        wb["Simulador"],
        wb["Dados"],
        wb["Janelas"],
    ]

    print(f"\nSalvando em {SAIDA}...")
    print("(Pode demorar 30-60s pelo volume de formulas)")
    wb.save(SAIDA)
    print(f"\nConcluido: {SAIDA}")
    print("\nQuando abrir o Excel pela primeira vez, ele vai recalcular todas as")
    print("formulas (pode levar 1-2 minutos). Depois disso fica rapido.")


if __name__ == "__main__":
    main()
