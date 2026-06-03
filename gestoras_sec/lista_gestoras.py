"""
Gera Excel mestre com todas as gestoras do banco gestoras_sec.db.

- Atribui ID sequencial ESTÁVEL a cada CIK (tabela gestora_id no banco).
- Ordenação inicial: por valor de holdings no trimestre mais recente (decrescente).
- CIKs novos recebem o próximo ID livre.

Saída: outputs/lista_gestoras_AAAAMMDD_HHMM.xlsx

Uso:
    python lista_gestoras.py
"""
import os, sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE, 'gestoras_sec.db')
OUT_DIR = os.path.join(BASE, 'outputs')
os.makedirs(OUT_DIR, exist_ok=True)

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
TITLE_FONT  = Font(bold=True, size=14, color='1F4E79')
SUB_FONT    = Font(italic=True, size=9, color='595959')
THIN        = Side(border_style='thin', color='BFBFBF')
BOX         = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left', vertical='center')
RIGHT       = Alignment(horizontal='right', vertical='center')
FMT_USD     = '"US$ "#,##0'

def garantir_tabela_id(conn):
    conn.cursor().execute('''
    CREATE TABLE IF NOT EXISTS gestora_id (
        id    INTEGER PRIMARY KEY AUTOINCREMENT,
        cik   TEXT UNIQUE NOT NULL
    )''')
    conn.commit()

def atribuir_ids(conn):
    cur = conn.cursor()
    cur.execute('SELECT MAX(trimestre) FROM filings_13f')
    ult = cur.fetchone()[0]
    if not ult: return None
    # ordena por valor agregado no último trim, desc
    cur.execute('''
        SELECT DISTINCT g.cik,
               COALESCE((SELECT valor_total FROM filings_13f f
                         WHERE f.cik=g.cik AND f.trimestre=? LIMIT 1), 0) as vt
        FROM gestoras g ORDER BY vt DESC, g.cik
    ''', (ult,))
    candidatos = [r[0] for r in cur.fetchall()]
    cur.execute('SELECT cik FROM gestora_id')
    ja_tem = {r[0] for r in cur.fetchall()}
    novos = [c for c in candidatos if c not in ja_tem]
    if novos:
        cur.executemany('INSERT INTO gestora_id (cik) VALUES (?)', [(c,) for c in novos])
        conn.commit()
        print(f'Atribuído ID para {len(novos)} CIK(s) novo(s).')
    return ult

def coletar(conn, ult):
    cur = conn.cursor()
    return cur.execute('''
        WITH stats AS (
            SELECT cik, COUNT(DISTINCT trimestre) n_trim,
                   SUM(n_holdings) total_holdings
            FROM filings_13f GROUP BY cik
        ),
        atual AS (
            SELECT cik, n_holdings, valor_total
            FROM filings_13f WHERE trimestre=?
        )
        SELECT i.id, i.cik, g.apelido, g.categoria, g.nome_sec,
               s.n_trim, s.total_holdings,
               a.n_holdings, a.valor_total
        FROM gestora_id i
        LEFT JOIN gestoras g ON g.cik=i.cik
        LEFT JOIN stats s    ON s.cik=i.cik
        LEFT JOIN atual a    ON a.cik=i.cik
        ORDER BY i.id
    ''', (ult,)).fetchall()

def main():
    if not os.path.exists(DB):
        print(f'Banco não encontrado: {DB}\nRode antes: python carga_sec_13f.py 2026Q1'); return
    conn = sqlite3.connect(DB)
    garantir_tabela_id(conn)
    ult = atribuir_ids(conn)
    if not ult: print('Banco vazio.'); return

    cur = conn.cursor()
    cur.execute('SELECT COUNT(DISTINCT trimestre) FROM filings_13f')
    n_trim = cur.fetchone()[0]

    rows = coletar(conn, ult)
    print(f'Trimestre de referência: {ult}  |  {len(rows)} gestoras')

    out = os.path.join(OUT_DIR, f'lista_gestoras_{datetime.now():%Y%m%d_%H%M}.xlsx')
    wb = Workbook(); ws = wb.active; ws.title = 'Gestoras'

    ws.cell(1,1, 'Gestoras institucionais — 13F-HR (SEC)').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.cell(2,1, f'{len(rows)} gestoras | refer. {ult} | hist. no banco: {n_trim} trimestre(s) | '
                 f'Use ID com `python historico_gestora.py <ID>`.'
            ).font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    headers = ['ID','CIK','Apelido','Categoria','Nome registrado SEC',
               'Trim. c/ filing','Total holdings','Holdings (atual)','Valor (atual)']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[4].height = 28

    for i, r in enumerate(rows, 1):
        rid, cik, apelido, cat, nome_sec, n_t, tot_h, n_h, vt = r
        rowi = 4 + i
        vals = [rid, cik, apelido, cat, nome_sec, n_t or 0, tot_h or 0, n_h, vt]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(rowi, c, v); cell.border = BOX
            if c in (1, 2, 6, 7, 8): cell.alignment = CENTER
            elif c in (3, 4, 5):     cell.alignment = LEFT
            else:                    cell.alignment = RIGHT
        ws.cell(rowi, 9).number_format = FMT_USD
        if i % 2 == 0:
            for c in range(1, 10):
                ws.cell(rowi, c).fill = PatternFill('solid', fgColor='F2F2F2')

    larg = [6, 10, 28, 16, 55, 10, 10, 12, 22]
    for i, w in enumerate(larg, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'D5'
    ws.auto_filter.ref = f'A4:{get_column_letter(len(headers))}{4+len(rows)}'

    wb.save(out)
    print(f'\nExcel salvo: {out}')
    print(f'Tamanho: {os.path.getsize(out)/1024:.1f} KB')

if __name__ == '__main__':
    main()
