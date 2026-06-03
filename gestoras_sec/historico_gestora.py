"""
Histórico completo de holdings de UMA gestora institucional, ao longo dos trimestres do banco.

Uso:
    python historico_gestora.py 5            # por ID (lista_gestoras)
    python historico_gestora.py 1067983      # por CIK
    python historico_gestora.py              # interativo
"""
import os, sys, sqlite3, re
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

BASE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE, 'gestoras_sec.db')
OUT_DIR = os.path.join(BASE, 'outputs')
os.makedirs(OUT_DIR, exist_ok=True)

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
TITLE_FONT  = Font(bold=True, size=14, color='1F4E79')
SUB_FONT    = Font(italic=True, size=9, color='595959')
TOTAL_FILL  = PatternFill('solid', fgColor='DDEBF7')
NEW_FILL    = PatternFill('solid', fgColor='E2EFDA')
OUT_FILL    = PatternFill('solid', fgColor='FCE4D6')
THIN        = Side(border_style='thin', color='BFBFBF')
BOX         = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left', vertical='center')
RIGHT       = Alignment(horizontal='right', vertical='center')
FMT_USD     = '"US$ "#,##0'
FMT_PCT     = '0.00%'

def resolver_gestora(conn, termo):
    cur = conn.cursor()
    s = str(termo).strip()
    # ID puro?
    if s.isdigit() and len(s) <= 5:
        cur.execute('SELECT cik, id FROM gestora_id WHERE id=?', (int(s),))
        r = cur.fetchone()
        if r: return r
    # CIK puro?
    cur.execute('SELECT cik, id FROM gestora_id WHERE cik=? OR cik=?',
                (s, s.lstrip('0')))
    r = cur.fetchone()
    if r: return r
    # busca por apelido/nome
    cur.execute('''SELECT g.cik, i.id, g.apelido FROM gestoras g
                   JOIN gestora_id i ON i.cik=g.cik
                   WHERE UPPER(g.apelido) LIKE ? OR UPPER(g.nome_sec) LIKE ?''',
                (f'%{s.upper()}%', f'%{s.upper()}%'))
    rs = cur.fetchall()
    if len(rs) == 1: return rs[0][:2]
    if len(rs) > 1:
        print(f'Ambíguo. {len(rs)} matches:')
        for c, i, a in rs[:10]: print(f'  ID {i}  CIK {c}  {a}')
        return None, None
    return None, None

def info_gestora(conn, cik):
    return conn.cursor().execute(
        'SELECT apelido, categoria, nome_sec FROM gestoras WHERE cik=?', (cik,)
    ).fetchone()

def trimestres(conn):
    return [r[0] for r in conn.cursor().execute(
        'SELECT DISTINCT trimestre FROM filings_13f ORDER BY trimestre').fetchall()]

def historico(conn, cik):
    return conn.cursor().execute('''
        SELECT trimestre, cusip, name_of_issuer, title_of_class,
               value_usd, shares, share_type, put_call
        FROM holdings WHERE cik=?
        ORDER BY trimestre, value_usd DESC
    ''', (cik,)).fetchall()

def aba_resumo(wb, fid, cik, info, trims, hist):
    apelido, categoria, nome_sec = info
    ws = wb.create_sheet('Resumo', 0)
    ws.cell(1,1, apelido or nome_sec or '(sem nome)').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.cell(2,1, f'CIK: {cik}  |  ID: {fid}  |  Categoria: {categoria or "—"}').font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws.cell(3,1, f'Nome registrado SEC: {nome_sec}').font = SUB_FONT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)

    headers = ['Trimestre','Nº holdings','Valor total (US$)','Top 1 posição','Concentração top 10']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(5, c, h); cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[5].height = 22

    por_trim = defaultdict(list)
    for r in hist:
        t, _cusip, name, _toc, vl, _sh, _st, _pc = r
        por_trim[t].append((name, vl or 0))

    for i, t in enumerate(trims, 1):
        rowi = 5 + i
        items = sorted(por_trim.get(t, []), key=lambda x:-x[1])
        nh = len(items)
        tot = sum(v for _,v in items)
        top1 = items[0] if items else ('—', 0)
        top10 = sum(v for _, v in items[:10])
        conc = (top10/tot) if tot else None
        ws.cell(rowi, 1, t).alignment = CENTER
        ws.cell(rowi, 2, nh).alignment = CENTER
        c3 = ws.cell(rowi, 3, tot if tot else None); c3.number_format = FMT_USD; c3.alignment = RIGHT
        ws.cell(rowi, 4, f'{(top1[0] or "")[:30]} (US$ {top1[1]:,.0f})').alignment = LEFT
        c5 = ws.cell(rowi, 5, conc); c5.number_format = FMT_PCT; c5.alignment = CENTER
        for c in range(1,6): ws.cell(rowi,c).border = BOX
        if i % 2 == 0:
            for c in range(1,6): ws.cell(rowi,c).fill = PatternFill('solid', fgColor='F7F9FC')

    for i, w in enumerate([12, 12, 22, 50, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A6'

def aba_historico(wb, hist, trims):
    """Matriz papel × trimestre (valor US$)."""
    ws = wb.create_sheet('Histórico')
    ws.cell(1,1, 'Matriz de holdings — valor (US$) por papel × trimestre').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.cell(2,1, 'Verde = nova posição (trim. anterior=0); Laranja = posição zerada (saiu).').font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    # agrupa por CUSIP (papel)
    pos = defaultdict(lambda: {t: 0.0 for t in trims})
    nomes = {}
    for r in hist:
        t, cusip, name, toc, vl, _sh, _st, _pc = r
        key = cusip or f'(sem cusip) {name}'
        pos[key][t] += (vl or 0)
        if key not in nomes: nomes[key] = (name or '', toc or '')

    t_ult = trims[-1]
    def chave(k): return (-pos[k][t_ult], -sum(pos[k].values()))
    papeis = sorted(pos.keys(), key=chave)

    headers = ['CUSIP','Nome do papel','Classe'] + [t for t in trims] + ['Total agregado']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h); cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[4].height = 24

    for i, k in enumerate(papeis, 1):
        rowi = 4 + i
        name, toc = nomes[k]
        ws.cell(rowi, 1, k[:11] if not k.startswith('(') else '—').alignment = CENTER
        ws.cell(rowi, 2, name[:50]).alignment = LEFT
        ws.cell(rowi, 3, toc[:20]).alignment = LEFT
        ant = 0.0; total = 0.0
        for j, t in enumerate(trims, 1):
            vl = pos[k][t]
            cell = ws.cell(rowi, 3+j, vl if vl else None)
            cell.number_format = FMT_USD; cell.alignment = RIGHT; cell.border = BOX
            if vl > 0 and ant == 0 and j > 1: cell.fill = NEW_FILL
            elif vl == 0 and ant > 0:          cell.fill = OUT_FILL
            ant = vl; total += vl
        c_tot = ws.cell(rowi, 3+len(trims)+1, total if total else None)
        c_tot.number_format = FMT_USD; c_tot.alignment = RIGHT; c_tot.border = BOX
        c_tot.font = Font(bold=True)
        for c in (1, 2, 3): ws.cell(rowi, c).border = BOX
        if i % 2 == 0:
            for c in range(1, 4+len(trims)+1):
                if ws.cell(rowi, c).fill.fgColor.rgb in (None, '00000000', 'FFFFFFFF'):
                    ws.cell(rowi, c).fill = PatternFill('solid', fgColor='F2F2F2')

    tot_row = 4 + len(papeis) + 1
    ws.cell(tot_row, 1, 'TOTAL').font = Font(bold=True)
    ws.cell(tot_row, 1).alignment = CENTER
    for c in (1,2,3):
        ws.cell(tot_row, c).fill = TOTAL_FILL; ws.cell(tot_row, c).border = BOX
    for j, t in enumerate(trims, 1):
        s = sum(pos[k][t] for k in papeis)
        cell = ws.cell(tot_row, 3+j, s if s else None)
        cell.number_format = FMT_USD; cell.font = Font(bold=True)
        cell.alignment = RIGHT; cell.border = BOX; cell.fill = TOTAL_FILL
    s_g = sum(sum(pos[k].values()) for k in papeis)
    cell = ws.cell(tot_row, 3+len(trims)+1, s_g)
    cell.number_format = FMT_USD; cell.font = Font(bold=True)
    cell.alignment = RIGHT; cell.border = BOX; cell.fill = TOTAL_FILL

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 18
    for j in range(len(trims)):
        ws.column_dimensions[get_column_letter(4+j)].width = 18
    ws.column_dimensions[get_column_letter(4+len(trims))].width = 20
    ws.freeze_panes = 'D5'

def aba_detalhes(wb, hist):
    ws = wb.create_sheet('Detalhes')
    ws.cell(1,1, 'Detalhes — 1 linha por posição × trimestre').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    headers = ['Trimestre','CUSIP','Nome do papel','Classe','Valor (US$)',
               'Quantidade','Tipo','Put/Call']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h); cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[3].height = 22
    for i, r in enumerate(hist, 1):
        t, cusip, name, toc, vl, sh, st, pc = r
        rowi = 3 + i
        vals = [t, cusip, (name or '')[:55], toc, vl, sh, st, pc]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(rowi, c, v); cell.border = BOX
            if c in (1,2,4,7,8): cell.alignment = CENTER
            elif c == 3:         cell.alignment = LEFT
            else:                cell.alignment = RIGHT
        ws.cell(rowi, 5).number_format = FMT_USD
        ws.cell(rowi, 6).number_format = '#,##0'
        if i % 2 == 0:
            for c in range(1, 9):
                ws.cell(rowi, c).fill = PatternFill('solid', fgColor='F2F2F2')
    for i, w in enumerate([10, 12, 45, 18, 18, 16, 8, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'

def main():
    if not os.path.exists(DB):
        print(f'Banco não encontrado: {DB}'); return
    conn = sqlite3.connect(DB)

    termo = sys.argv[1] if len(sys.argv)>1 else input('> ID, CIK ou nome da gestora: ').strip()
    if not termo: print('Nada informado.'); return

    cik, fid = resolver_gestora(conn, termo)
    if not cik:
        print(f'Não encontrei gestora para "{termo}". Rode antes: python lista_gestoras.py')
        return

    info = info_gestora(conn, cik)
    if not info: print(f'CIK {cik} sem registro.'); return
    hist = historico(conn, cik)
    if not hist: print(f'Gestora {info[0]} sem holdings no banco.'); return
    trims = trimestres(conn)

    print(f'Gestora: {info[0]} ({info[2]})')
    print(f'  CIK {cik} | ID {fid} | {len(hist)} linhas | {len({h[0] for h in hist})} trim.')

    wb = Workbook(); wb.remove(wb.active)
    aba_historico(wb, hist, trims)
    aba_detalhes(wb, hist)
    aba_resumo(wb, fid, cik, info, trims, hist)

    safe = re.sub(r'[^A-Za-z0-9._-]', '_', (info[0] or 'gestora'))[:40]
    out = os.path.join(OUT_DIR, f'historico_gest_{fid:04d}_{safe}_{datetime.now():%Y%m%d_%H%M}.xlsx')
    wb.save(out)
    print(f'Excel salvo: {out}  ({os.path.getsize(out)/1024:.1f} KB)')

if __name__ == '__main__':
    main()
