"""
Relatório Excel — quem detém um ou mais papéis americanos em um trimestre.

Uso (interativo):
    python relatorio_papel.py

Pede:
  - Papel(eis): nomes (substring) ou CUSIPs separados por vírgula
                ex: APPLE,MICROSOFT,037833100
  - Trimestre (AAAAQn) — enter = mais recente

Saída: outputs/papeis_*_AAAAQn_AAAAMMDD_HHMM.xlsx
       - aba RESUMO (se >1 papel)
       - 1 aba por papel com gestoras detentoras
"""
import os, sys, sqlite3, re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

BASE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE, 'gestoras_sec.db')
OUT_DIR = os.path.join(BASE, 'outputs')
os.makedirs(OUT_DIR, exist_ok=True)

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
TITLE_FONT  = Font(bold=True, size=14, color='1F4E79')
SUB_FONT    = Font(italic=True, size=9, color='595959')
TOTAL_FILL  = PatternFill('solid', fgColor='DDEBF7')
THIN        = Side(border_style='thin', color='BFBFBF')
BOX         = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left', vertical='center')
RIGHT       = Alignment(horizontal='right', vertical='center')
FMT_USD     = '"US$ "#,##0'
FMT_INT     = '#,##0'

def conectar():
    if not os.path.exists(DB):
        print(f'Banco não encontrado: {DB}'); sys.exit(1)
    return sqlite3.connect(DB)

def trims_disponiveis(conn):
    return [r[0] for r in conn.cursor().execute(
        'SELECT DISTINCT trimestre FROM filings_13f ORDER BY trimestre').fetchall()]

def detectar(termo):
    s = termo.strip().lstrip('﻿').replace(' ','').upper()
    # CUSIP precisa ter MISTURA de letras e dígitos (não pode ser só letras)
    if len(s) == 9 and s.isalnum() and any(c.isdigit() for c in s):
        return ('cusip', s)
    return ('name', s)

def buscar(conn, termo, trim):
    modo, valor = detectar(termo)
    cur = conn.cursor()
    if modo == 'cusip':
        sql = '''SELECT h.cik, g.apelido, g.categoria, h.name_of_issuer,
                        h.title_of_class, h.cusip,
                        h.shares, h.share_type, h.value_usd, h.put_call,
                        f.valor_total
                 FROM holdings h LEFT JOIN gestoras g ON g.cik=h.cik
                 LEFT JOIN filings_13f f ON f.accession=h.accession
                 WHERE h.cusip=? AND h.trimestre=?
                 ORDER BY h.value_usd DESC'''
        rows = cur.execute(sql, (valor, trim)).fetchall()
    else:
        sql = '''SELECT h.cik, g.apelido, g.categoria, h.name_of_issuer,
                        h.title_of_class, h.cusip,
                        h.shares, h.share_type, h.value_usd, h.put_call,
                        f.valor_total
                 FROM holdings h LEFT JOIN gestoras g ON g.cik=h.cik
                 LEFT JOIN filings_13f f ON f.accession=h.accession
                 WHERE UPPER(h.name_of_issuer) LIKE ? AND h.trimestre=?
                 ORDER BY h.value_usd DESC'''
        rows = cur.execute(sql, (f'%{valor}%', trim)).fetchall()
    return rows, modo, valor

def ajustar(ws, larguras):
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def aba_papel(wb, termo_label, trim, rows):
    nome_aba = re.sub(r'[^A-Za-z0-9]','_', termo_label)[:31]
    if nome_aba in wb.sheetnames: nome_aba = nome_aba[:28] + '_2'
    ws = wb.create_sheet(nome_aba)
    ws.cell(1,1, f'Detentores de "{termo_label}" — {trim}').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    total_val = sum((r[8] or 0) for r in rows)
    total_sh  = sum((r[6] or 0) for r in rows)
    ws.cell(2,1, f'{len(rows)} posições | Quantidade total: {total_sh:,.0f} | Valor agregado: US$ {total_val:,.0f}').font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    headers = ['#','CIK','Gestora','Categoria','Papel','Classe','CUSIP','Quant.','Valor (US$)','% AUM gestora']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h); cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[4].height = 24

    for i, r in enumerate(rows, 1):
        cik, ap, cat, nm, toc, cu, sh, st, vl, pc, vt_aum = r
        rowi = 4 + i
        pct_aum = ((vl/vt_aum) if (vl and vt_aum and vt_aum>0) else None)
        vals = [i, cik, ap, cat, nm, toc, cu, sh, vl, pct_aum]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(rowi, c, v); cell.border = BOX
            if c in (1, 2, 4, 6, 7): cell.alignment = CENTER
            elif c in (3, 5):         cell.alignment = LEFT
            else:                     cell.alignment = RIGHT
        ws.cell(rowi, 8).number_format = FMT_INT
        ws.cell(rowi, 9).number_format = FMT_USD
        ws.cell(rowi, 10).number_format = '0.00%'
        if i % 2 == 0:
            for c in range(1, 11):
                ws.cell(rowi, c).fill = PatternFill('solid', fgColor='F2F2F2')

    if rows:
        tot_row = 4 + len(rows) + 1
        ws.cell(tot_row, 1, 'TOTAL').font = Font(bold=True)
        ws.cell(tot_row, 8, total_sh).number_format = FMT_INT
        ws.cell(tot_row, 9, total_val).number_format = FMT_USD
        for c in range(1, 11):
            ws.cell(tot_row, c).font = Font(bold=True)
            ws.cell(tot_row, c).fill = TOTAL_FILL
            ws.cell(tot_row, c).border = BOX
        rng = f'J5:J{4+len(rows)}'
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type='min', start_color='FFFFFF',
            mid_type='percentile', mid_value=50, mid_color='FFE699',
            end_type='max', end_color='FF7F7F'))

    ajustar(ws, [5, 10, 32, 16, 38, 14, 12, 16, 18, 12])
    ws.freeze_panes = 'A5'

def aba_resumo(wb, termos, trim, dados):
    ws = wb.create_sheet('RESUMO', 0)
    ws.cell(1,1, f'Resumo comparativo — {trim}').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(2,1, f'Papeis: {", ".join(termos)}').font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    headers = ['Papel buscado','Nº posições','Quant. total','Valor agregado (US$)','Top 1 gestora','% do top 1']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h); cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[4].height = 24

    for i, t in enumerate(termos, 1):
        rows = dados.get(t, [])
        rowi = 4 + i
        nf = len(rows)
        sh = sum((r[6] or 0) for r in rows)
        vl = sum((r[8] or 0) for r in rows)
        top1 = ''
        pct1 = None
        if rows:
            top1 = f'{(rows[0][1] or "—")[:30]} (US$ {(rows[0][8] or 0):,.0f})'
            pct1 = (rows[0][8]/vl) if vl else None
        vals = [t, nf, sh, vl, top1, pct1]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(rowi, c, v); cell.border = BOX
            if c == 5: cell.alignment = LEFT
            elif c in (1,): cell.alignment = LEFT
            else: cell.alignment = (CENTER if c in (2,6) else RIGHT)
        ws.cell(rowi, 3).number_format = FMT_INT
        ws.cell(rowi, 4).number_format = FMT_USD
        ws.cell(rowi, 6).number_format = '0.00%'
        if i % 2 == 0:
            for c in range(1, 7):
                ws.cell(rowi, c).fill = PatternFill('solid', fgColor='F2F2F2')
    ajustar(ws, [25, 12, 16, 22, 55, 12])
    ws.freeze_panes = 'A5'

def main():
    conn = conectar()
    trims = trims_disponiveis(conn)
    if not trims:
        print('Banco vazio. Rode carga_sec_13f.py.'); return
    print(f'Trimestres disponíveis: {", ".join(trims)}\n')

    termos_raw = input('> Papel(eis) — separados por vírgula (ex: APPLE,MICROSOFT,037833100): ').strip()
    if not termos_raw: print('Nada informado.'); return
    termos = [t.strip().lstrip('﻿​').lstrip() for t in termos_raw.split(',') if t.strip()]

    trim_raw = input(f'> Trimestre AAAAQn (enter = {trims[-1]}): ').strip()
    if not trim_raw: trim = trims[-1]
    elif trim_raw in trims: trim = trim_raw
    else:
        print(f'Trimestre {trim_raw} não no banco. Disponíveis: {", ".join(trims)}'); return

    print(f'\nGerando relatório para {len(termos)} papel(eis) em {trim}...')
    dados = {}
    for t in termos:
        rows, modo, valor = buscar(conn, t, trim)
        dados[t] = rows
        print(f'  {t} (busca {modo}): {len(rows)} posições')

    wb = Workbook(); wb.remove(wb.active)
    if len(termos) > 1: aba_resumo(wb, termos, trim, dados)
    for t in termos: aba_papel(wb, t, trim, dados[t])

    label = re.sub(r'[^A-Za-z0-9_]+','_', '_'.join(termos))[:40] if len(termos)<=4 else f'{len(termos)}papeis'
    out = os.path.join(OUT_DIR, f'papeis_{label}_{trim}_{datetime.now():%Y%m%d_%H%M}.xlsx')
    wb.save(out)
    print(f'\nExcel salvo: {out}  ({os.path.getsize(out)/1024:.1f} KB)')

if __name__ == '__main__':
    main()
