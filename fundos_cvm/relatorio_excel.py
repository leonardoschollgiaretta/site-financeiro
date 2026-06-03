"""
Relatório Excel — posição de cada fundo em uma ou mais ações em um mês.

Uso (interativo):
    python relatorio_excel.py

Pede:
  - Tickers (separados por vírgula): ex. PETR4,VALE3,CMIG4
  - Mês (AAAAMM):                    ex. 202604  (vazio = último disponível)

Gera 1 Excel em outputs/ com:
  - 1 aba por ticker (fundos detentores ordenados por valor)
  - 1 aba RESUMO comparando todos os tickers
"""
import os, sys, sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

BASE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE, 'fundos_cvm.db')
OUT_DIR = os.path.join(BASE, 'outputs')
os.makedirs(OUT_DIR, exist_ok=True)

# ---- estilos ----
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
TITLE_FONT  = Font(bold=True, size=14, color='1F4E79')
SUB_FONT    = Font(italic=True, size=9, color='595959')
THIN        = Side(border_style='thin', color='BFBFBF')
BOX         = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left', vertical='center', wrap_text=False)
RIGHT       = Alignment(horizontal='right', vertical='center')

FMT_INT     = '#,##0'
FMT_BRL     = 'R$ #,##0.00;[Red]-R$ #,##0.00'
FMT_BRL_M   = 'R$ #,##0.00,," mi"'   # em milhões
FMT_PCT     = '0.00%'

def conectar():
    if not os.path.exists(DB):
        print(f'ERRO: banco não encontrado em {DB}\nRode antes: python carga_cvm_cda.py --ultimos 12')
        sys.exit(1)
    return sqlite3.connect(DB)

def periodos_disponiveis(conn):
    cur = conn.cursor()
    cur.execute('SELECT DISTINCT periodo FROM posicoes_acoes ORDER BY periodo')
    return [r[0] for r in cur.fetchall()]

def fmt_periodo_humano(p):
    """202604 -> 'abr/2026'"""
    if not p or len(p)!=6: return p
    meses = ['jan','fev','mar','abr','mai','jun','jul','ago','set','out','nov','dez']
    return f'{meses[int(p[4:6])-1]}/{p[:4]}'

def buscar_posicoes(conn, ticker, periodo):
    cur = conn.cursor()
    sql = '''
    SELECT p.cnpj_fundo, f.denominacao, f.tp_fundo_classe,
           p.tp_ativo, p.qt_pos_final, p.vl_mercado, p.vl_custo,
           f.patrimonio_liq,
           CASE WHEN f.patrimonio_liq>0 THEN p.vl_mercado*1.0/f.patrimonio_liq END,
           (p.vl_mercado - p.vl_custo),
           CASE WHEN p.vl_custo>0 THEN (p.vl_mercado - p.vl_custo)/p.vl_custo END
    FROM posicoes_acoes p
    LEFT JOIN fundos f ON f.cnpj = p.cnpj_fundo AND f.periodo = p.periodo
    WHERE p.cd_ativo = ? AND p.periodo = ?
    ORDER BY p.vl_mercado DESC
    '''
    return cur.execute(sql, (ticker.upper(), periodo)).fetchall()

def ajustar_largura(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def montar_aba_ticker(wb, ticker, periodo, rows):
    aba = ticker.upper()[:31]
    if aba in wb.sheetnames:
        aba = aba[:28] + '_2'
    ws = wb.create_sheet(aba)

    ws.cell(1,1, f'Posição em {ticker.upper()} — {fmt_periodo_humano(periodo)}').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    if rows:
        total_val = sum((r[5] or 0) for r in rows)
        total_qt  = sum((r[4] or 0) for r in rows)
        ws.cell(2,1, f'{len(rows)} fundos detentores | quantidade total: {total_qt:,.0f} ações | '
                     f'valor agregado: R$ {total_val:,.2f}').font = SUB_FONT
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    else:
        ws.cell(2,1, 'Nenhum fundo encontrado para este ticker neste período.').font = SUB_FONT

    headers = ['#','CNPJ','Denominação','Tipo classe','Tipo ativo','Quantidade',
               'Valor mercado (R$)','Valor custo (R$)','PL do fundo (R$)','% do PL',
               'Resultado (R$)','Resultado %']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[4].height = 24

    for i, r in enumerate(rows, 1):
        cnpj, denom, tipo_cl, tp_at, qt, vl, vc, pl, pct, res, res_pct = r
        rowi = 4 + i
        values = [i, cnpj, denom, tipo_cl, tp_at, qt, vl, vc, pl, pct, res, res_pct]
        for c, v in enumerate(values, 1):
            cell = ws.cell(rowi, c, v); cell.border = BOX
            if c == 1: cell.alignment = CENTER
            elif c in (2,3,4,5): cell.alignment = LEFT
            else: cell.alignment = RIGHT
        ws.cell(rowi, 6).number_format = FMT_INT
        ws.cell(rowi, 7).number_format = FMT_BRL
        ws.cell(rowi, 8).number_format = FMT_BRL
        ws.cell(rowi, 9).number_format = FMT_BRL
        ws.cell(rowi,10).number_format = FMT_PCT
        ws.cell(rowi,11).number_format = FMT_BRL
        ws.cell(rowi,12).number_format = FMT_PCT
        # zebra
        if i % 2 == 0:
            for c in range(1,13):
                ws.cell(rowi,c).fill = PatternFill('solid', fgColor='F2F2F2')

    # Totais
    if rows:
        tot_row = 4 + len(rows) + 1
        ws.cell(tot_row, 1, 'TOTAL').font = Font(bold=True)
        ws.cell(tot_row, 6, sum((r[4] or 0) for r in rows)).number_format = FMT_INT
        ws.cell(tot_row, 7, sum((r[5] or 0) for r in rows)).number_format = FMT_BRL
        ws.cell(tot_row, 8, sum((r[6] or 0) for r in rows)).number_format = FMT_BRL
        ws.cell(tot_row,11, sum((r[9] or 0) for r in rows)).number_format = FMT_BRL
        for c in range(1,13):
            ws.cell(tot_row, c).font = Font(bold=True)
            ws.cell(tot_row, c).fill = PatternFill('solid', fgColor='DDEBF7')
            ws.cell(tot_row, c).border = BOX

        # Heatmap em % do PL
        rng = f'J5:J{4+len(rows)}'
        rule = ColorScaleRule(start_type='min', start_color='FFFFFF',
                              mid_type='percentile', mid_value=50, mid_color='FFE699',
                              end_type='max', end_color='FF7F7F')
        ws.conditional_formatting.add(rng, rule)

    ajustar_largura(ws, [5, 20, 55, 17, 22, 16, 18, 18, 18, 10, 18, 12])
    ws.freeze_panes = 'A5'

def montar_resumo(wb, tickers, periodo, dados_por_ticker):
    """dados_por_ticker: {ticker: rows}"""
    ws = wb.create_sheet('RESUMO', 0)
    ws.cell(1,1, f'Resumo comparativo — {fmt_periodo_humano(periodo)}').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(2,1, f'Tickers analisados: {", ".join(tickers)}').font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    headers = ['Ticker','Nº de fundos','Quantidade total','Valor agregado (R$)',
               'Valor custo (R$)','Maior posição (fundo)']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4,c,h); cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[4].height = 24

    for i, tk in enumerate(tickers, 1):
        rows = dados_por_ticker.get(tk.upper(), [])
        rowi = 4 + i
        nf = len(rows)
        qt = sum((r[4] or 0) for r in rows)
        vl = sum((r[5] or 0) for r in rows)
        vc = sum((r[6] or 0) for r in rows)
        maior = ''
        if rows:
            denom = rows[0][1] or ''
            val_maior = rows[0][5] or 0
            maior = f'{denom[:50]} (R$ {val_maior:,.0f})'
        values = [tk.upper(), nf, qt, vl, vc, maior]
        for c, v in enumerate(values, 1):
            cell = ws.cell(rowi, c, v); cell.border = BOX
            cell.alignment = CENTER if c==1 else (LEFT if c==6 else RIGHT)
        ws.cell(rowi,3).number_format = FMT_INT
        ws.cell(rowi,4).number_format = FMT_BRL
        ws.cell(rowi,5).number_format = FMT_BRL
        if i % 2 == 0:
            for c in range(1,7):
                ws.cell(rowi,c).fill = PatternFill('solid', fgColor='F2F2F2')

    ajustar_largura(ws, [12, 14, 20, 22, 22, 75])
    ws.freeze_panes = 'A5'

def main():
    print('='*60)
    print('Relatório Excel — Fundos detentores por ação')
    print('='*60)
    conn = conectar()
    periodos = periodos_disponiveis(conn)
    if not periodos:
        print('Banco vazio. Rode antes: python carga_cvm_cda.py --ultimos 12')
        return
    print(f'\nPeríodos disponíveis no banco: {", ".join(periodos)}')

    tickers_raw = input('\n> Ticker(s) — separados por vírgula (ex: PETR4,VALE3): ').strip()
    if not tickers_raw:
        print('Nada informado. Saindo.')
        return
    tickers = [t.strip().lstrip('﻿').upper() for t in tickers_raw.split(',') if t.strip()]

    p_raw = input(f'> Mês AAAAMM (enter = {periodos[-1]}): ').strip()
    if not p_raw:
        periodo = periodos[-1]
    elif p_raw in periodos:
        periodo = p_raw
    else:
        print(f'! Período {p_raw} não está no banco. Disponíveis: {", ".join(periodos)}')
        return

    print(f'\nGerando relatório para {len(tickers)} ticker(s) em {fmt_periodo_humano(periodo)}...')

    dados = {}
    for tk in tickers:
        rows = buscar_posicoes(conn, tk, periodo)
        dados[tk] = rows
        print(f'  {tk}: {len(rows)} fundos')

    wb = Workbook(); wb.remove(wb.active)
    if len(tickers) > 1:
        montar_resumo(wb, tickers, periodo, dados)
    for tk in tickers:
        montar_aba_ticker(wb, tk, periodo, dados[tk])

    tk_label = '_'.join(tickers) if len(tickers) <= 4 else f'{len(tickers)}tickers'
    out = os.path.join(OUT_DIR, f'fundos_{tk_label}_{periodo}_{datetime.now():%Y%m%d_%H%M}.xlsx')
    wb.save(out)
    print(f'\nRelatório salvo: {out}')
    print(f'Tamanho: {os.path.getsize(out)/1024:.1f} KB')

if __name__ == '__main__':
    main()
