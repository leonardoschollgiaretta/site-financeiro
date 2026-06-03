"""
Gera Excel com o histórico completo de posições em ações de UM fundo,
desde o início do banco até o período mais recente.

Uso:
    python historico_fundo.py 42              # por ID (da lista_fundos.xlsx)
    python historico_fundo.py 36352539000157  # por CNPJ (com ou sem pontuação)
    python historico_fundo.py                 # modo interativo
"""
import os, sys, sqlite3, re
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

BASE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE, 'fundos_cvm.db')
OUT_DIR = os.path.join(BASE, 'outputs')
os.makedirs(OUT_DIR, exist_ok=True)

# estilos
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
TITLE_FONT  = Font(bold=True, size=14, color='1F4E79')
SUB_FONT    = Font(italic=True, size=9, color='595959')
TOTAL_FILL  = PatternFill('solid', fgColor='DDEBF7')
NEW_FILL    = PatternFill('solid', fgColor='E2EFDA')   # verde claro - entrada
OUT_FILL    = PatternFill('solid', fgColor='FCE4D6')   # laranja claro - saída
THIN        = Side(border_style='thin', color='BFBFBF')
BOX         = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left', vertical='center')
RIGHT       = Alignment(horizontal='right', vertical='center')
FMT_BRL     = 'R$ #,##0.00;[Red]-R$ #,##0.00'
FMT_BRL_DELTA = '"+"R$ #,##0.00;[Red]"-"R$ #,##0.00;"—"'
FMT_INT     = '#,##0'
FMT_PCT     = '0.00%'

MESES_PT = ['jan','fev','mar','abr','mai','jun','jul','ago','set','out','nov','dez']

def fmt_periodo(p): return f'{MESES_PT[int(p[4:6])-1]}/{p[:4]}'

def resolver_fundo(conn, termo):
    """Recebe ID numérico (pequeno) ou CNPJ. Retorna (cnpj, id) ou (None,None)."""
    cur = conn.cursor()
    s = str(termo).strip()
    # ID se tiver até 7 dígitos puros
    if s.isdigit() and len(s) <= 7:
        cur.execute('SELECT cnpj, id FROM fundo_id WHERE id=?', (int(s),))
        r = cur.fetchone()
        if r: return r
    # CNPJ
    cnpj_clean = re.sub(r'\D','', s)
    if len(cnpj_clean) >= 8:
        cur.execute("""SELECT cnpj, id FROM fundo_id
                       WHERE REPLACE(REPLACE(REPLACE(cnpj,'.',''),'/',''),'-','') LIKE ?""",
                    (f'%{cnpj_clean}%',))
        rs = cur.fetchall()
        if len(rs) == 1: return rs[0]
        if len(rs) > 1:
            print(f'CNPJ ambíguo. {len(rs)} matches:')
            for c, i in rs[:10]: print(f'  ID {i} | {c}')
            return None, None
    return None, None

def buscar_nome(conn, cnpj):
    """Pega o nome mais recente do fundo."""
    cur = conn.cursor()
    cur.execute('''SELECT denominacao, tp_fundo_classe, patrimonio_liq, periodo
                   FROM fundos WHERE cnpj=? ORDER BY periodo DESC LIMIT 1''', (cnpj,))
    return cur.fetchone()

def buscar_historico(conn, cnpj):
    """Retorna todas as posições do fundo, todos os períodos."""
    cur = conn.cursor()
    cur.execute('''
        SELECT p.periodo, p.dt_compt, p.cd_ativo, p.ds_ativo, p.tp_ativo,
               p.qt_pos_final, p.vl_mercado, p.vl_custo,
               f.patrimonio_liq
        FROM posicoes_acoes p
        LEFT JOIN fundos f ON f.cnpj=p.cnpj_fundo AND f.periodo=p.periodo
        WHERE p.cnpj_fundo=?
        ORDER BY p.periodo, p.vl_mercado DESC
    ''', (cnpj,))
    return cur.fetchall()

def buscar_pl_por_periodo(conn, cnpj):
    cur = conn.cursor()
    cur.execute('SELECT periodo, patrimonio_liq FROM fundos WHERE cnpj=? ORDER BY periodo', (cnpj,))
    return dict(cur.fetchall())

def todos_periodos(conn):
    cur = conn.cursor()
    cur.execute('SELECT DISTINCT periodo FROM posicoes_acoes ORDER BY periodo')
    return [r[0] for r in cur.fetchall()]

# ---------------- abas ----------------
def aba_resumo(wb, fid, cnpj, nome, tipo, pl_atual, periodo_atual, hist, pls, periodos):
    ws = wb.create_sheet('Resumo', 0)
    ws.cell(1,1, nome or '(sem nome)').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.cell(2,1, f'CNPJ: {cnpj}  |  ID: {fid}  |  Tipo: {tipo or "—"}').font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.cell(3,1, f'Histórico no banco: {len(periodos)} período(s) — '
                 f'{fmt_periodo(periodos[0])} a {fmt_periodo(periodos[-1])}').font = SUB_FONT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)

    headers = ['Período','PL do fundo (R$)','Valor em ações (R$)','Nº ativos']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(5, c, h); cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[5].height = 22

    # agrega por período
    por_periodo = defaultdict(lambda: [0.0, set()])
    for r in hist:
        p, _dt, cod, _ds, _tp, _qt, vl, _vc, _pl = r
        por_periodo[p][0] += (vl or 0)
        por_periodo[p][1].add(cod)

    for i, p in enumerate(periodos, 1):
        rowi = 5 + i
        vl_total, ativos = por_periodo.get(p, (0.0, set()))
        ws.cell(rowi, 1, fmt_periodo(p)).alignment = CENTER
        c2 = ws.cell(rowi, 2, pls.get(p)); c2.number_format = FMT_BRL; c2.alignment = RIGHT
        c3 = ws.cell(rowi, 3, vl_total if vl_total else None); c3.number_format = FMT_BRL; c3.alignment = RIGHT
        ws.cell(rowi, 4, len(ativos) if ativos else 0).alignment = CENTER
        for c in range(1,5): ws.cell(rowi,c).border = BOX
        if i % 2 == 0:
            for c in range(1,5): ws.cell(rowi,c).fill = PatternFill('solid', fgColor='F7F9FC')

    for i, w in enumerate([14, 22, 22, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A6'

def aba_historico(wb, hist, periodos):
    """Matriz tickers x meses (valores R$). Inclui Δ vs mês anterior, entrada (verde) e saída (laranja)."""
    ws = wb.create_sheet('Histórico')
    ws.cell(1,1, 'Matriz de posições — valor de mercado (R$) por ticker × mês').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.cell(2,1, 'Verde = nova posição (mês anterior=0); Laranja = posição zerada (saiu).').font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    # tabela tickers -> {periodo: (qt, vl)}
    pos = defaultdict(lambda: {p: (0.0, 0.0) for p in periodos})
    nomes = {}
    for r in hist:
        p, _dt, cod, ds, _tp, qt, vl, _vc, _pl = r
        pos[cod][p] = ((pos[cod][p][0] or 0)+(qt or 0), (pos[cod][p][1] or 0)+(vl or 0))
        if cod not in nomes: nomes[cod] = ds or ''

    # ordena tickers por valor total do período mais recente (depois soma total)
    p_ult = periodos[-1]
    def chave_ord(cod):
        return (-pos[cod][p_ult][1], -sum(v for _,v in pos[cod].values()))
    tickers = sorted(pos.keys(), key=chave_ord)

    # cabeçalho
    headers = ['Ticker','Descrição'] + [fmt_periodo(p) for p in periodos] + ['Total agregado']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h); cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[4].height = 24

    for i, cod in enumerate(tickers, 1):
        rowi = 4 + i
        ws.cell(rowi, 1, cod).font = Font(bold=True)
        ws.cell(rowi, 1).alignment = CENTER
        ws.cell(rowi, 2, nomes.get(cod,'')[:50]).alignment = LEFT
        ant = 0.0
        total = 0.0
        for j, p in enumerate(periodos, 1):
            qt, vl = pos[cod][p]
            cell = ws.cell(rowi, 2+j, vl if vl else None)
            cell.number_format = FMT_BRL; cell.alignment = RIGHT; cell.border = BOX
            # highlight entrada/saída
            if vl > 0 and ant == 0 and j > 1:
                cell.fill = NEW_FILL
            elif vl == 0 and ant > 0:
                cell.fill = OUT_FILL
            ant = vl
            total += vl
        c_tot = ws.cell(rowi, 2+len(periodos)+1, total if total else None)
        c_tot.number_format = FMT_BRL; c_tot.alignment = RIGHT; c_tot.border = BOX
        c_tot.font = Font(bold=True)
        ws.cell(rowi, 1).border = BOX; ws.cell(rowi, 2).border = BOX
        if i % 2 == 0:
            for c in range(1, 3+len(periodos)+1):
                if ws.cell(rowi,c).fill.fgColor.rgb in (None, '00000000', 'FFFFFFFF'):
                    ws.cell(rowi,c).fill = PatternFill('solid', fgColor='F2F2F2')

    # linha de TOTAL
    tot_row = 4 + len(tickers) + 1
    ws.cell(tot_row, 1, 'TOTAL').font = Font(bold=True)
    ws.cell(tot_row, 1).alignment = CENTER
    ws.cell(tot_row, 1).fill = TOTAL_FILL; ws.cell(tot_row, 1).border = BOX
    ws.cell(tot_row, 2).fill = TOTAL_FILL; ws.cell(tot_row, 2).border = BOX
    for j, p in enumerate(periodos, 1):
        s = sum(pos[c][p][1] for c in tickers)
        cell = ws.cell(tot_row, 2+j, s if s else None)
        cell.number_format = FMT_BRL; cell.font = Font(bold=True)
        cell.alignment = RIGHT; cell.border = BOX; cell.fill = TOTAL_FILL
    s_geral = sum(sum(v for _,v in pos[c].values()) for c in tickers)
    cell = ws.cell(tot_row, 2+len(periodos)+1, s_geral)
    cell.number_format = FMT_BRL; cell.font = Font(bold=True)
    cell.alignment = RIGHT; cell.border = BOX; cell.fill = TOTAL_FILL

    # larguras
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    for j in range(len(periodos)):
        ws.column_dimensions[get_column_letter(3+j)].width = 16
    ws.column_dimensions[get_column_letter(3+len(periodos))].width = 17
    ws.freeze_panes = 'C5'

def aba_detalhes(wb, hist):
    """Granular: 1 linha por ticker x período. Inclui qt, vl, custo, % PL."""
    ws = wb.create_sheet('Detalhes')
    ws.cell(1,1, 'Detalhes — posição por ticker × período').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    headers = ['Período','Ticker','Tipo','Descrição','Quantidade',
               'Valor mercado (R$)','Valor custo (R$)','Resultado (R$)','% do PL']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h); cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[3].height = 22
    for i, r in enumerate(hist, 1):
        p, _dt, cod, ds, tp, qt, vl, vc, pl = r
        rowi = 3 + i
        valores = [fmt_periodo(p), cod, tp, (ds or '')[:55], qt, vl, vc,
                   (vl or 0)-(vc or 0),
                   ((vl/pl) if (vl and pl and pl>0) else None)]
        for c, v in enumerate(valores, 1):
            cell = ws.cell(rowi, c, v); cell.border = BOX
            if c in (1,2,3): cell.alignment = CENTER
            elif c == 4:     cell.alignment = LEFT
            else:            cell.alignment = RIGHT
        ws.cell(rowi, 5).number_format = FMT_INT
        ws.cell(rowi, 6).number_format = FMT_BRL
        ws.cell(rowi, 7).number_format = FMT_BRL
        ws.cell(rowi, 8).number_format = FMT_BRL
        ws.cell(rowi, 9).number_format = FMT_PCT
        if i % 2 == 0:
            for c in range(1, 10):
                ws.cell(rowi, c).fill = PatternFill('solid', fgColor='F2F2F2')
    for i, w in enumerate([10, 10, 22, 50, 16, 18, 18, 18, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'
    if len(hist):
        rng = f'I4:I{3+len(hist)}'
        rule = ColorScaleRule(start_type='min', start_color='FFFFFF',
                              mid_type='percentile', mid_value=50, mid_color='FFE699',
                              end_type='max', end_color='FF7F7F')
        ws.conditional_formatting.add(rng, rule)

# ---------------- main ----------------
def main():
    if not os.path.exists(DB):
        print(f'Banco não encontrado: {DB}\nRode antes: python carga_cvm_cda.py --ultimos 12')
        return
    conn = sqlite3.connect(DB)

    if len(sys.argv) > 1:
        termo = sys.argv[1]
    else:
        termo = input('> ID ou CNPJ do fundo: ').strip()
    if not termo:
        print('Nada informado.'); return

    cnpj, fid = resolver_fundo(conn, termo)
    if not cnpj:
        print(f'Não encontrei fundo para "{termo}". Gere a lista mestre antes: python lista_fundos.py')
        return

    info = buscar_nome(conn, cnpj)
    if not info:
        print(f'CNPJ {cnpj} sem registro em fundos. Existe na lista mas pode não ter posição.')
        return
    nome, tipo, pl_atual, periodo_atual = info

    hist = buscar_historico(conn, cnpj)
    if not hist:
        print(f'Fundo {nome} (ID {fid}) — sem posições em ações no banco.')
        return

    periodos = todos_periodos(conn)
    pls = buscar_pl_por_periodo(conn, cnpj)

    print(f'Fundo: {nome}')
    print(f'  CNPJ {cnpj} | ID {fid} | PL atual: R$ {(pl_atual or 0):,.0f}')
    print(f'  {len(hist)} posições ao longo de {len({r[0] for r in hist})} mês(es)')

    wb = Workbook(); wb.remove(wb.active)
    aba_historico(wb, hist, periodos)
    aba_detalhes(wb, hist)
    aba_resumo(wb, fid, cnpj, nome, tipo, pl_atual, periodo_atual, hist, pls, periodos)

    # nome de arquivo seguro
    safe = re.sub(r'[^A-Za-z0-9._-]', '_', (nome or 'fundo'))[:50]
    out = os.path.join(OUT_DIR, f'historico_{fid:05d}_{safe}_{datetime.now():%Y%m%d_%H%M}.xlsx')
    wb.save(out)
    print(f'\nExcel salvo: {out}')
    print(f'Tamanho: {os.path.getsize(out)/1024:.1f} KB')

if __name__ == '__main__':
    main()
