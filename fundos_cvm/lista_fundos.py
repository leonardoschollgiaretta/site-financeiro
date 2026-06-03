"""
Gera Excel mestre com todos os fundos que tiveram posição em ações em algum período do banco.

- Atribui ID sequencial ESTÁVEL a cada CNPJ (tabela fundo_id no banco).
- Ordenação inicial dos IDs por PL decrescente no período mais recente.
- CNPJs novos (que aparecerem em cargas futuras) recebem o próximo ID livre.

Saída: outputs/lista_fundos_AAAAMMDD_HHMM.xlsx

Uso:
    python lista_fundos.py
"""
import os, sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE, 'fundos_cvm.db')
OUT_DIR = os.path.join(BASE, 'outputs')
os.makedirs(OUT_DIR, exist_ok=True)

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
TITLE_FONT  = Font(bold=True, size=14, color='1F4E79')
SUB_FONT    = Font(italic=True, size=9, color='595959')
THIN        = Side(border_style='thin', color='BFBFBF')
BOX         = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left', vertical='center')
RIGHT       = Alignment(horizontal='right', vertical='center')
FMT_BRL     = 'R$ #,##0.00;[Red]-R$ #,##0.00'

def garantir_tabela_id(conn):
    """Cria tabela fundo_id se não existir. ID estável por CNPJ."""
    cur = conn.cursor()
    cur.execute('''
    CREATE TABLE IF NOT EXISTS fundo_id (
        id    INTEGER PRIMARY KEY AUTOINCREMENT,
        cnpj  TEXT UNIQUE NOT NULL
    )''')
    conn.commit()

def atribuir_ids(conn):
    """Atribui IDs para CNPJs novos. Ordem: PL decrescente no último período."""
    cur = conn.cursor()
    cur.execute('SELECT MAX(periodo) FROM posicoes_acoes')
    ult = cur.fetchone()[0]
    if not ult:
        print('Banco vazio. Rode carga_cvm_cda.py primeiro.')
        return None
    # Lista todos os CNPJs que já tiveram posição em ações, ordenados por PL no último período
    cur.execute('''
        SELECT DISTINCT p.cnpj_fundo,
               COALESCE(f.patrimonio_liq, 0) as pl
        FROM posicoes_acoes p
        LEFT JOIN fundos f ON f.cnpj=p.cnpj_fundo AND f.periodo=?
        ORDER BY pl DESC, p.cnpj_fundo
    ''', (ult,))
    candidatos = [r[0] for r in cur.fetchall()]
    cur.execute('SELECT cnpj FROM fundo_id')
    ja_tem = {r[0] for r in cur.fetchall()}
    novos = [c for c in candidatos if c not in ja_tem]
    if novos:
        cur.executemany('INSERT INTO fundo_id (cnpj) VALUES (?)', [(c,) for c in novos])
        conn.commit()
        print(f'Atribuído ID para {len(novos)} CNPJ(s) novo(s).')
    return ult

def coletar_dados(conn, periodo_atual):
    """Para cada fundo, coleta: nome (último período), PL, nº meses com posição,
    valor total em ações no período atual, soma quantidades."""
    cur = conn.cursor()
    sql = '''
    WITH stats AS (
        SELECT cnpj_fundo,
               COUNT(DISTINCT periodo) as n_meses,
               COUNT(DISTINCT cd_ativo) as n_ativos_distintos_total
        FROM posicoes_acoes
        GROUP BY cnpj_fundo
    ),
    atual AS (
        SELECT cnpj_fundo,
               SUM(vl_mercado) as vl_acoes_atual,
               COUNT(DISTINCT cd_ativo) as n_ativos_atual
        FROM posicoes_acoes
        WHERE periodo = ?
        GROUP BY cnpj_fundo
    )
    SELECT i.id, i.cnpj,
           COALESCE(f.denominacao, '(sem nome no período atual)'),
           f.tp_fundo_classe,
           f.patrimonio_liq,
           s.n_meses,
           s.n_ativos_distintos_total,
           a.vl_acoes_atual,
           a.n_ativos_atual,
           CASE WHEN f.patrimonio_liq>0 AND a.vl_acoes_atual IS NOT NULL
                THEN a.vl_acoes_atual/f.patrimonio_liq END as pct_pl_acoes
    FROM fundo_id i
    LEFT JOIN stats s ON s.cnpj_fundo = i.cnpj
    LEFT JOIN fundos f ON f.cnpj = i.cnpj AND f.periodo = ?
    LEFT JOIN atual  a ON a.cnpj_fundo = i.cnpj
    ORDER BY i.id
    '''
    return cur.execute(sql, (periodo_atual, periodo_atual)).fetchall()

def fmt_periodo_humano(p):
    if not p: return ''
    meses = ['jan','fev','mar','abr','mai','jun','jul','ago','set','out','nov','dez']
    return f'{meses[int(p[4:6])-1]}/{p[:4]}'

def gerar_excel(rows, periodo_atual, periodos_total):
    out = os.path.join(OUT_DIR, f'lista_fundos_{datetime.now():%Y%m%d_%H%M}.xlsx')
    wb = Workbook(); ws = wb.active; ws.title = 'Fundos'

    ws.cell(1,1, 'Lista mestre de fundos com posição em ações').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.cell(2,1, f'{len(rows)} fundos | dados do período {fmt_periodo_humano(periodo_atual)} | '
                 f'histórico do banco: {periodos_total} meses | '
                 f'Use a coluna ID com `python historico_fundo.py <ID>` para gerar histórico individual.'
            ).font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    headers = ['ID','CNPJ','Denominação','Tipo classe','PL (R$)',
               'Nº meses c/ pos.','Nº ativos (total)','Valor em ações (R$)',
               'Nº ativos (atual)','% PL em ações']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[4].height = 28

    for i, r in enumerate(rows, 1):
        rid, cnpj, denom, tipo, pl, n_meses, n_at_tot, vl_at, n_at_at, pct = r
        rowi = 4 + i
        valores = [rid, cnpj, denom, tipo, pl, n_meses or 0, n_at_tot or 0, vl_at, n_at_at, pct]
        for c, v in enumerate(valores, 1):
            cell = ws.cell(rowi, c, v); cell.border = BOX
            if c in (1, 6, 7, 9): cell.alignment = CENTER
            elif c in (2, 3, 4):  cell.alignment = LEFT
            else:                 cell.alignment = RIGHT
        ws.cell(rowi, 5).number_format = FMT_BRL
        ws.cell(rowi, 8).number_format = FMT_BRL
        ws.cell(rowi, 10).number_format = '0.00%'
        if i % 2 == 0:
            for c in range(1,11):
                ws.cell(rowi,c).fill = PatternFill('solid', fgColor='F2F2F2')

    largs = [6, 20, 60, 22, 18, 8, 9, 20, 8, 10]
    for i, w in enumerate(largs, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'D5'
    ws.auto_filter.ref = f'A4:{get_column_letter(len(headers))}{4+len(rows)}'

    wb.save(out)
    return out

def main():
    if not os.path.exists(DB):
        print(f'Banco não encontrado: {DB}\nRode antes: python carga_cvm_cda.py --ultimos 12')
        return
    conn = sqlite3.connect(DB)
    garantir_tabela_id(conn)
    periodo_atual = atribuir_ids(conn)
    if not periodo_atual: return

    # Conta períodos totais no banco
    cur = conn.cursor()
    cur.execute('SELECT COUNT(DISTINCT periodo) FROM posicoes_acoes')
    n_periodos = cur.fetchone()[0]

    print(f'Coletando dados (período de referência: {periodo_atual})...')
    rows = coletar_dados(conn, periodo_atual)
    print(f'  {len(rows)} fundos na lista')

    out = gerar_excel(rows, periodo_atual, n_periodos)
    print(f'\nExcel gerado: {out}')
    print(f'Tamanho: {os.path.getsize(out)/1024:.1f} KB')
    print(f'\nUse a coluna ID para gerar histórico individual:')
    print(f'  python historico_fundo.py <ID>')

if __name__ == '__main__':
    main()
