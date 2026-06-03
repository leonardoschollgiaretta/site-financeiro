"""
Pivot temporal: para um ticker, mostra TODOS os fundos que tiveram posição
em CADA mês disponível no banco. Linhas = fundos, colunas = meses.

Uso:
    python pivot_acao.py RECV3                  # imprime tabela no console
    python pivot_acao.py RECV3 --excel          # gera .xlsx formatado
    python pivot_acao.py RECV3 --excel --abrir  # gera e abre o Excel
"""
from __future__ import annotations

import argparse
import os
import sqlite3
import sys
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE, 'fundos_cvm.db')


def fmt_periodo(p: str) -> str:
    """'202509' -> 'set/25'."""
    meses = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun',
             'jul', 'ago', 'set', 'out', 'nov', 'dez']
    ano, mes = int(p[:4]), int(p[4:])
    return f'{meses[mes - 1]}/{str(ano)[2:]}'


def fmt_mi(v: float | None) -> str:
    if v is None or v == 0:
        return '       -'
    return f'{v / 1e6:8.1f}'


def buscar(ticker: str):
    ticker = ticker.upper().strip()
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row

    # Quais meses tem dado pra esse ticker?
    periodos = [r[0] for r in conn.execute("""
        SELECT DISTINCT periodo FROM posicoes_acoes
        WHERE cd_ativo = ? ORDER BY periodo
    """, (ticker,)).fetchall()]

    if not periodos:
        print(f'Sem dados para {ticker} no banco.')
        sys.exit(0)

    # Posições por fundo × mês (em R$)
    sql = """
        SELECT p.cnpj_fundo,
               COALESCE(f.denominacao, '(sem nome)') AS denominacao,
               p.periodo,
               SUM(p.vl_mercado) AS vl
        FROM posicoes_acoes p
        LEFT JOIN fundos f
               ON f.cnpj = p.cnpj_fundo AND f.periodo = p.periodo
        WHERE p.cd_ativo = ?
        GROUP BY p.cnpj_fundo, denominacao, p.periodo
    """
    rows = conn.execute(sql, (ticker,)).fetchall()
    conn.close()

    # Indexar: { cnpj: { 'nome': str, 'pos': {periodo: vl} } }
    fundos: dict[str, dict] = {}
    for r in rows:
        f = fundos.setdefault(r['cnpj_fundo'], {'nome': r['denominacao'], 'pos': {}})
        f['pos'][r['periodo']] = r['vl']
        # nome mais recente prevalece se vier null
        if r['denominacao'] and r['denominacao'] != '(sem nome)':
            f['nome'] = r['denominacao']

    # Ordenar fundos pelo último mês onde tiveram posição (decrescente)
    ult = periodos[-1]
    fundos_ord = sorted(
        fundos.items(),
        key=lambda kv: (kv[1]['pos'].get(ult, 0) or 0),
        reverse=True,
    )

    return ticker, periodos, fundos_ord


def imprimir_console(ticker: str, periodos: list[str], fundos_ord: list):
    cab_meses = ' '.join(f'{fmt_periodo(p):>8}' for p in periodos)
    print()
    print(f'{ticker} - fundos com posicao (R$ milhoes)')
    print(f'Banco: {len(periodos)} períodos, {len(fundos_ord)} fundos no total')
    print()
    print(f'{"Fundo":<55} {cab_meses}')
    print('-' * (55 + len(cab_meses) + 1))

    for cnpj, dados in fundos_ord:
        nome = dados['nome'][:54]
        celulas = ' '.join(fmt_mi(dados['pos'].get(p)) for p in periodos)
        print(f'{nome:<55} {celulas}')

    # Linhas de totais
    print('-' * (55 + len(cab_meses) + 1))
    totais = [sum((d['pos'].get(p, 0) or 0) for _, d in fundos_ord) for p in periodos]
    print(f'{"TOTAL":<55} {" ".join(fmt_mi(t) for t in totais)}')
    n_fundos = [sum(1 for _, d in fundos_ord if (d['pos'].get(p, 0) or 0) > 0) for p in periodos]
    print(f'{"N fundos com posicao":<55} {" ".join(f"{n:>8}" for n in n_fundos)}')


def exportar_excel(ticker: str, periodos: list[str], fundos_ord: list, abrir: bool):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print('openpyxl não instalado. Rode: pip install openpyxl')
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ticker

    azul = PatternFill(start_color='1d3557', end_color='1d3557', fill_type='solid')
    cinza = PatternFill(start_color='f3f6fb', end_color='f3f6fb', fill_type='solid')
    bold_white = Font(bold=True, color='FFFFFF')
    bold = Font(bold=True)

    # Cabeçalho
    ws.cell(1, 1, f'{ticker} — posições por fundo (R$ milhões)').font = Font(bold=True, size=14)
    ws.cell(2, 1, f'{len(periodos)} períodos · {len(fundos_ord)} fundos · gerado em {datetime.now():%d/%m/%Y %H:%M}').font = Font(italic=True, color='6b7280')

    # Linha de cabeçalho da tabela
    linha_cab = 4
    ws.cell(linha_cab, 1, 'CNPJ').font = bold_white
    ws.cell(linha_cab, 2, 'Fundo').font = bold_white
    for j, p in enumerate(periodos):
        c = ws.cell(linha_cab, 3 + j, fmt_periodo(p))
        c.font = bold_white
        c.alignment = Alignment(horizontal='right')
    for col in range(1, 3 + len(periodos)):
        ws.cell(linha_cab, col).fill = azul

    # Linhas de dados
    for i, (cnpj, dados) in enumerate(fundos_ord):
        r = linha_cab + 1 + i
        ws.cell(r, 1, cnpj)
        ws.cell(r, 2, dados['nome'])
        for j, p in enumerate(periodos):
            v = dados['pos'].get(p)
            cell = ws.cell(r, 3 + j, v / 1e6 if v else None)
            cell.number_format = '#,##0.0;[Red]-#,##0.0;—'

    # Totais
    r_tot = linha_cab + 1 + len(fundos_ord)
    ws.cell(r_tot, 2, 'TOTAL').font = bold
    ws.cell(r_tot + 1, 2, 'Nº fundos').font = bold
    for j, p in enumerate(periodos):
        total = sum((d['pos'].get(p, 0) or 0) for _, d in fundos_ord)
        n = sum(1 for _, d in fundos_ord if (d['pos'].get(p, 0) or 0) > 0)
        ws.cell(r_tot, 3 + j, total / 1e6).number_format = '#,##0.0'
        ws.cell(r_tot, 3 + j).font = bold
        ws.cell(r_tot + 1, 3 + j, n).font = bold
    for col in range(1, 3 + len(periodos)):
        ws.cell(r_tot, col).fill = cinza
        ws.cell(r_tot + 1, col).fill = cinza

    # Larguras
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 55
    for j in range(len(periodos)):
        ws.column_dimensions[get_column_letter(3 + j)].width = 11
    ws.freeze_panes = 'C5'

    out = os.path.join(BASE, f'pivot_{ticker}_{datetime.now():%Y%m%d}.xlsx')
    wb.save(out)
    print(f'\nExcel gerado: {out}')

    if abrir:
        os.startfile(out)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('ticker', nargs='?', help='Ticker da ação (ex: RECV3). Se omitido, pergunta.')
    ap.add_argument('--excel', action='store_true', help='Exporta para Excel')
    ap.add_argument('--abrir', action='store_true', help='Abre o Excel após gerar')
    args = ap.parse_args()

    ticker_in = args.ticker
    excel = args.excel
    abrir = args.abrir

    # Modo interativo se rodar sem argumentos
    if not ticker_in:
        ticker_in = input('Ticker da ação (ex: RECV3): ').strip()
        if not ticker_in:
            print('Ticker vazio. Saindo.')
            return
        resp = input('Exportar para Excel? [S/n]: ').strip().lower()
        excel = resp != 'n'
        if excel:
            resp = input('Abrir o Excel ao terminar? [S/n]: ').strip().lower()
            abrir = resp != 'n'

    ticker, periodos, fundos_ord = buscar(ticker_in)

    if excel:
        exportar_excel(ticker, periodos, fundos_ord, abrir=abrir)
    else:
        imprimir_console(ticker, periodos, fundos_ord)


if __name__ == '__main__':
    main()
