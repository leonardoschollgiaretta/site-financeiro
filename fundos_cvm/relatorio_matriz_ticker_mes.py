"""
Relatório Excel — MATRIZ ticker × mês.

Cada linha = um ticker (ação). Cada coluna = um período (AAAAMM).
Célula = soma do valor de mercado (R$) que TODOS os fundos têm naquele
ticker, naquele mês (totalizador agregado).

Inclui todos os tickers presentes no banco.
Ordena por valor de mercado do último mês disponível (maiores no topo).

Uso:
    python relatorio_matriz_ticker_mes.py
"""
import os, sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

BASE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE, 'fundos_cvm.db')
OUT_DIR = os.path.join(BASE, 'outputs')
os.makedirs(OUT_DIR, exist_ok=True)

# ---- estilos (mesmo padrão do relatorio_excel.py) ----
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
TITLE_FONT  = Font(bold=True, size=14, color='1F4E79')
SUB_FONT    = Font(italic=True, size=9, color='595959')
THIN        = Side(border_style='thin', color='BFBFBF')
BOX         = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left', vertical='center')
RIGHT       = Alignment(horizontal='right', vertical='center')

FMT_INT     = '#,##0'
FMT_BRL_M   = 'R$ #,##0.0,," mi"'   # valor em milhões de R$

MESES = ['jan','fev','mar','abr','mai','jun','jul','ago','set','out','nov','dez']

def fmt_periodo_humano(p):
    if not p or len(p) != 6: return p
    return f'{MESES[int(p[4:6])-1]}/{p[:4]}'

def conectar():
    if not os.path.exists(DB):
        raise SystemExit(f'ERRO: banco não encontrado em {DB}')
    return sqlite3.connect(DB)

def periodos_disponiveis(conn):
    cur = conn.cursor()
    cur.execute('SELECT DISTINCT periodo FROM posicoes_acoes ORDER BY periodo')
    return [r[0] for r in cur.fetchall()]

def buscar_matriz(conn):
    """Retorna {ticker: {periodo: (valor_mercado, n_fundos)}} agregado."""
    cur = conn.cursor()
    cur.execute('''
        SELECT cd_ativo, periodo,
               SUM(vl_mercado) AS vl,
               COUNT(DISTINCT cnpj_fundo) AS nf
        FROM posicoes_acoes
        WHERE cd_ativo IS NOT NULL AND cd_ativo <> ''
        GROUP BY cd_ativo, periodo
    ''')
    matriz = {}
    for ticker, periodo, vl, nf in cur.fetchall():
        matriz.setdefault(ticker, {})[periodo] = (vl or 0, nf or 0)
    return matriz

def ajustar_largura(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def main():
    conn = conectar()
    periodos = periodos_disponiveis(conn)
    if not periodos:
        raise SystemExit('Banco vazio. Rode: python carga_cvm_cda.py --ultimos 12')

    matriz = buscar_matriz(conn)
    ult = periodos[-1]
    # remove tickers que nunca tiveram valor de mercado em nenhum mês
    matriz = {t: v for t, v in matriz.items()
              if any(vl for vl, _ in v.values())}
    # ordena tickers pelo valor de mercado do último período (desc); sem posição = 0
    tickers = sorted(matriz.keys(),
                     key=lambda t: matriz[t].get(ult, (0, 0))[0],
                     reverse=True)

    print(f'Períodos: {", ".join(periodos)}')
    print(f'Tickers no banco: {len(tickers)}')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Matriz ticker x mês'

    # título
    ws.cell(1, 1, 'Valor de mercado agregado por ação — posições de fundos (CVM CDA)').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(periodos))
    ws.cell(2, 1, f'Soma do valor de mercado (R$, em milhões) de todos os fundos detentores, por mês. '
                  f'{len(tickers)} ações | {len(periodos)} períodos | ordenado por {fmt_periodo_humano(ult)}.'
            ).font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2 + len(periodos))

    # cabeçalho: Ticker | <meses...> | Variação último vs primeiro
    hrow = 4
    ws.cell(hrow, 1, 'Ticker')
    for j, p in enumerate(periodos, start=2):
        ws.cell(hrow, j, fmt_periodo_humano(p))
    col_var = 2 + len(periodos)
    ws.cell(hrow, col_var, 'Δ% (1º→últ)')
    for c in range(1, col_var + 1):
        cell = ws.cell(hrow, c)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[hrow].height = 22

    # corpo
    primeiro = periodos[0]
    for i, tk in enumerate(tickers, 1):
        rowi = hrow + i
        ws.cell(rowi, 1, tk).alignment = LEFT
        ws.cell(rowi, 1).border = BOX
        ws.cell(rowi, 1).font = Font(bold=True)
        for j, p in enumerate(periodos, start=2):
            vl = matriz[tk].get(p, (0, 0))[0]
            cell = ws.cell(rowi, j, (vl if vl else None))
            cell.number_format = FMT_BRL_M
            cell.alignment = RIGHT; cell.border = BOX
        # variação primeiro -> último (em %)
        v0 = matriz[tk].get(primeiro, (0, 0))[0]
        v1 = matriz[tk].get(ult, (0, 0))[0]
        var = ((v1 - v0) / v0) if v0 else None
        cv = ws.cell(rowi, col_var, var)
        cv.number_format = '0.0%'; cv.alignment = RIGHT; cv.border = BOX
        if i % 2 == 0:
            for c in range(1, col_var + 1):
                if ws.cell(rowi, c).fill.fgColor.rgb in (None, '00000000'):
                    ws.cell(rowi, c).fill = PatternFill('solid', fgColor='F2F2F2')

    # linha TOTAL (mercado inteiro por mês)
    trow = hrow + len(tickers) + 1
    ws.cell(trow, 1, 'TOTAL').font = Font(bold=True)
    for j, p in enumerate(periodos, start=2):
        total_p = sum(matriz[t].get(p, (0, 0))[0] for t in tickers)
        cell = ws.cell(trow, j, total_p if total_p else None)
        cell.number_format = FMT_BRL_M
    for c in range(1, col_var + 1):
        ws.cell(trow, c).font = Font(bold=True)
        ws.cell(trow, c).fill = PatternFill('solid', fgColor='DDEBF7')
        ws.cell(trow, c).border = BOX
        ws.cell(trow, c).alignment = RIGHT if c > 1 else LEFT

    # heatmap nas colunas de meses
    rng = f'{get_column_letter(2)}{hrow+1}:{get_column_letter(1+len(periodos))}{hrow+len(tickers)}'
    rule = ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',
                          mid_type='percentile', mid_value=80, mid_color='FFE699',
                          end_type='percentile', end_value=99, end_color='F8696B')
    ws.conditional_formatting.add(rng, rule)

    ajustar_largura(ws, [12] + [13] * len(periodos) + [13])
    ws.freeze_panes = 'B5'   # trava ticker + cabeçalho

    out = os.path.join(OUT_DIR, f'matriz_ticker_mes_{datetime.now():%Y%m%d_%H%M}.xlsx')
    wb.save(out)
    conn.close()
    print(f'\nRelatório salvo: {out}')
    print(f'Tamanho: {os.path.getsize(out)/1024:.1f} KB')

if __name__ == '__main__':
    main()
