"""
Gera Excel com o histórico trimestral COMPLETO de 1 empresa americana.
Espelha o estilo do financeiro/relatorio.py do projeto BR, com:
  - Aba RESUMO (cabeçalho, KPIs principais)
  - Aba DRE (todas as linhas, períodos como colunas)
  - Aba BALANÇO
  - Aba DFC (Fluxo de Caixa)
  - Aba INDICADORES (margens, ROE, alavancagem, etc.)

Uso:
    python historico_empresa.py AAPL
    python historico_empresa.py 0000320193    # CIK
    python historico_empresa.py               # interativo
"""
import os, sys, sqlite3, re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

BASE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE, 'empresas_us.db')
OUT_DIR = os.path.join(BASE, 'outputs')
os.makedirs(OUT_DIR, exist_ok=True)

# estilos
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
SUBHDR_FILL = PatternFill('solid', fgColor='2E75B6')
SUBHDR_FONT = Font(color='FFFFFF', bold=True, size=10)
TITLE_FONT  = Font(bold=True, size=14, color='1F4E79')
SUB_FONT    = Font(italic=True, size=9, color='595959')
TOTAL_FILL  = PatternFill('solid', fgColor='DDEBF7')
SECTION_FILL= PatternFill('solid', fgColor='E7E6E6')
SECTION_FONT= Font(bold=True, size=10, color='1F4E79')
THIN        = Side(border_style='thin', color='BFBFBF')
BOX         = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left', vertical='center')
RIGHT       = Alignment(horizontal='right', vertical='center')
FMT_USD_M   = '"US$ "#,##0.00,," mi"'    # exibe em milhões
FMT_USD_B   = '"US$ "#,##0.00,,,"B"'      # bilhões com letra B
FMT_PCT     = '0.00%'
FMT_NUM     = '0.00'

def resolver(conn, termo):
    cur = conn.cursor()
    s = str(termo).strip().upper()
    if s.isdigit() and len(s) >= 7:
        cik = f'{int(s):010d}'
        cur.execute('SELECT cik FROM empresas WHERE cik=?', (cik,))
        if cur.fetchone(): return cik
    cur.execute('SELECT cik FROM empresas WHERE UPPER(ticker)=? OR UPPER(ticker_alt) LIKE ?',
                (s, f'%{s}%'))
    r = cur.fetchone()
    if r: return r[0]
    cur.execute('SELECT cik, nome FROM empresas WHERE UPPER(nome) LIKE ?', (f'%{s}%',))
    rs = cur.fetchall()
    if len(rs) == 1: return rs[0][0]
    if len(rs) > 1:
        print(f'Ambíguo. {len(rs)} matches:')
        for c, n in rs[:10]: print(f'  CIK {c}  {n}')
        return None
    return None

def info_empresa(conn, cik):
    return conn.cursor().execute('''SELECT ticker, nome, sic_descricao, setor,
                                       exchange, fiscal_year_end FROM empresas
                                    WHERE cik=?''', (cik,)).fetchone()

def buscar_trimestres(conn, cik):
    """Retorna todos os trimestres+FY do CIK, ordenados cronologicamente."""
    return conn.cursor().execute('''
        SELECT ano, trimestre, tipo_periodo, data_inicio, data_fim, form,
               -- DRE
               receita_liquida, custo_receita, lucro_bruto, sg_a, r_e_d,
               despesas_operacionais, depreciacao_amortizacao, ebit,
               receitas_financeiras, despesas_financeiras, ir_csll, lucro_liquido,
               ebitda,
               -- Balanço Ativo
               ativo_total, ativo_circulante, caixa, investimentos_cp,
               contas_receber, estoques, ativo_nao_circulante, investimentos,
               imobilizado, intangivel, goodwill,
               -- Balanço Passivo
               passivo_total, passivo_circulante, fornecedores, emprestimos_cp,
               passivo_nao_circulante, emprestimos_lp, capital_social,
               lucros_acumulados, patrimonio_liquido, divida_bruta, divida_liquida,
               -- DFC
               fco, fci, fcf_financiamento, capex, aquisicoes, venda_ativos,
               captacoes, pagamento_dividas, recompra_acoes, dividendos_pagos,
               caixa_final, fcl
        FROM financeiros_trimestrais WHERE cik=?
        ORDER BY ano, CASE trimestre WHEN 'Q1' THEN 1 WHEN 'Q2' THEN 2
                                      WHEN 'Q3' THEN 3 WHEN 'Q4' THEN 4
                                      WHEN 'FY' THEN 9 END
    ''', (cik,)).fetchall()

# Colunas do SELECT na mesma ordem
COLS = ['ano','trimestre','tipo','data_inicio','data_fim','form',
        # DRE
        'receita_liquida','custo_receita','lucro_bruto','sg_a','r_e_d',
        'despesas_operacionais','depreciacao_amortizacao','ebit',
        'receitas_financeiras','despesas_financeiras','ir_csll','lucro_liquido','ebitda',
        # Ativo
        'ativo_total','ativo_circulante','caixa','investimentos_cp',
        'contas_receber','estoques','ativo_nao_circulante','investimentos',
        'imobilizado','intangivel','goodwill',
        # Passivo
        'passivo_total','passivo_circulante','fornecedores','emprestimos_cp',
        'passivo_nao_circulante','emprestimos_lp','capital_social',
        'lucros_acumulados','patrimonio_liquido','divida_bruta','divida_liquida',
        # DFC
        'fco','fci','fcf_financiamento','capex','aquisicoes','venda_ativos',
        'captacoes','pagamento_dividas','recompra_acoes','dividendos_pagos',
        'caixa_final','fcl']

def to_dict(row): return dict(zip(COLS, row))

# Seções estruturadas
DRE = [
    ('=== DRE ===', None),
    ('Receita Líquida',             'receita_liquida'),
    ('Custo da Receita (CMV/CSV)',  'custo_receita'),
    ('Lucro Bruto',                 'lucro_bruto'),
    ('SG&A',                        'sg_a'),
    ('Pesquisa & Desenvolvimento',  'r_e_d'),
    ('Despesas Operacionais (total)','despesas_operacionais'),
    ('Depreciação & Amortização',   'depreciacao_amortizacao'),
    ('EBIT (Op. Income)',           'ebit'),
    ('EBITDA',                      'ebitda'),
    ('Receitas Financeiras',        'receitas_financeiras'),
    ('Despesas Financeiras (Juros)','despesas_financeiras'),
    ('IR & Tributos',               'ir_csll'),
    ('Lucro Líquido',               'lucro_liquido'),
]

BP = [
    ('=== ATIVO ===', None),
    ('Ativo Total',                'ativo_total'),
    ('Ativo Circulante',           'ativo_circulante'),
    ('  Caixa & Equivalentes',     'caixa'),
    ('  Investimentos CP',         'investimentos_cp'),
    ('  Contas a Receber',         'contas_receber'),
    ('  Estoques',                 'estoques'),
    ('Ativo Não-Circulante',       'ativo_nao_circulante'),
    ('  Investimentos LP',         'investimentos'),
    ('  Imobilizado (PP&E)',       'imobilizado'),
    ('  Intangível',               'intangivel'),
    ('  Goodwill',                 'goodwill'),
    ('=== PASSIVO + PL ===', None),
    ('Passivo Total',              'passivo_total'),
    ('Passivo Circulante',         'passivo_circulante'),
    ('  Fornecedores',             'fornecedores'),
    ('  Empréstimos CP',           'emprestimos_cp'),
    ('Passivo Não-Circulante',     'passivo_nao_circulante'),
    ('  Empréstimos LP',           'emprestimos_lp'),
    ('Patrimônio Líquido',         'patrimonio_liquido'),
    ('  Capital Social',           'capital_social'),
    ('  Lucros Acumulados',        'lucros_acumulados'),
    ('Dívida Bruta (calc.)',       'divida_bruta'),
    ('Dívida Líquida (calc.)',     'divida_liquida'),
]

DFC = [
    ('=== FLUXO DE CAIXA ===', None),
    ('Fluxo de Caixa Operacional (FCO)', 'fco'),
    ('Fluxo de Caixa de Investimento (FCI)', 'fci'),
    ('  Capex',                     'capex'),
    ('  Aquisições',                'aquisicoes'),
    ('  Venda de Ativos',           'venda_ativos'),
    ('Fluxo de Caixa de Financiamento', 'fcf_financiamento'),
    ('  Captações',                 'captacoes'),
    ('  Pagamento de Dívidas',      'pagamento_dividas'),
    ('  Recompra de Ações',         'recompra_acoes'),
    ('  Dividendos Pagos',          'dividendos_pagos'),
    ('Caixa Final',                 'caixa_final'),
    ('Fluxo de Caixa Livre (FCL)',  'fcl'),
]

def montar_aba_secao(wb, nome, secoes, dados, periodos_label):
    ws = wb.create_sheet(nome)
    ws.cell(1, 1, nome).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(periodos_label)+1)
    ws.cell(2, 1, 'Valores em US$ milhões. Trimestres isolados (Q4 = FY - YTD9).').font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(periodos_label)+1)

    # Cabeçalho: Métrica | períodos
    ws.cell(4, 1, 'Métrica').fill = HEADER_FILL
    ws.cell(4, 1).font = HEADER_FONT; ws.cell(4, 1).alignment = CENTER
    ws.cell(4, 1).border = BOX
    for j, p in enumerate(periodos_label, 2):
        cell = ws.cell(4, j, p); cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[4].height = 22

    row_atual = 5
    for label, key in secoes:
        if key is None:  # seção
            ws.cell(row_atual, 1, label).fill = SECTION_FILL
            ws.cell(row_atual, 1).font = SECTION_FONT
            ws.merge_cells(start_row=row_atual, start_column=1,
                           end_row=row_atual, end_column=len(periodos_label)+1)
            row_atual += 1
            continue
        ws.cell(row_atual, 1, label).alignment = LEFT
        ws.cell(row_atual, 1).border = BOX
        for j, d in enumerate(dados, 2):
            v = d.get(key)
            cell = ws.cell(row_atual, j, (v/1e6) if v is not None else None)
            cell.number_format = '#,##0'   # já dividido por 1M, exibe inteiro
            cell.alignment = RIGHT; cell.border = BOX
        row_atual += 1

    ws.column_dimensions['A'].width = 38
    for j in range(2, len(periodos_label)+2):
        ws.column_dimensions[get_column_letter(j)].width = 14
    ws.freeze_panes = 'B5'

def montar_indicadores(wb, dados, periodos_label):
    ws = wb.create_sheet('Indicadores')
    ws.cell(1, 1, 'Indicadores derivados').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(periodos_label)+1)

    metricas = [
        ('=== Margens ===', None, None),
        ('Margem Bruta',        lambda d: pct(d.get('lucro_bruto'), d.get('receita_liquida')), FMT_PCT),
        ('Margem EBITDA',       lambda d: pct(d.get('ebitda'),      d.get('receita_liquida')), FMT_PCT),
        ('Margem EBIT',         lambda d: pct(d.get('ebit'),        d.get('receita_liquida')), FMT_PCT),
        ('Margem Líquida',      lambda d: pct(d.get('lucro_liquido'),d.get('receita_liquida')), FMT_PCT),
        ('=== Retorno ===', None, None),
        ('ROE (anualizado)',    lambda d: pct(d.get('lucro_liquido') and d['lucro_liquido']*4,
                                                d.get('patrimonio_liquido')) if d.get('tipo')=='Q'
                                            else pct(d.get('lucro_liquido'), d.get('patrimonio_liquido')), FMT_PCT),
        ('=== Endividamento ===', None, None),
        ('Dívida Líquida / EBITDA (Q)', lambda d: ratio(d.get('divida_liquida'),
                                                          d.get('ebitda') and d['ebitda']*4
                                                          if d.get('tipo')=='Q' else d.get('ebitda')), FMT_NUM),
        ('Caixa / Dívida Bruta',lambda d: pct(d.get('caixa'),       d.get('divida_bruta')), FMT_PCT),
        ('=== Fluxo de Caixa ===', None, None),
        ('FCL / Receita',       lambda d: pct(d.get('fcl'),         d.get('receita_liquida')), FMT_PCT),
        ('Capex / Receita',     lambda d: pct(d.get('capex'),       d.get('receita_liquida')), FMT_PCT),
        ('Dividendos / FCL',    lambda d: pct(d.get('dividendos_pagos') and -d['dividendos_pagos'] or None,
                                                d.get('fcl')), FMT_PCT),
    ]

    # cabeçalho
    ws.cell(4, 1, 'Indicador').fill = HEADER_FILL; ws.cell(4,1).font = HEADER_FONT
    ws.cell(4,1).alignment = CENTER; ws.cell(4,1).border = BOX
    for j, p in enumerate(periodos_label, 2):
        cell = ws.cell(4, j, p); cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BOX
    ws.row_dimensions[4].height = 22

    row_atual = 5
    for label, fn, fmt in metricas:
        if fn is None:
            ws.cell(row_atual, 1, label).fill = SECTION_FILL
            ws.cell(row_atual, 1).font = SECTION_FONT
            ws.merge_cells(start_row=row_atual, start_column=1,
                           end_row=row_atual, end_column=len(periodos_label)+1)
            row_atual += 1; continue
        ws.cell(row_atual, 1, label).alignment = LEFT
        ws.cell(row_atual, 1).border = BOX
        for j, d in enumerate(dados, 2):
            try: v = fn(d)
            except: v = None
            cell = ws.cell(row_atual, j, v)
            cell.number_format = fmt; cell.alignment = RIGHT; cell.border = BOX
        row_atual += 1

    ws.column_dimensions['A'].width = 32
    for j in range(2, len(periodos_label)+2):
        ws.column_dimensions[get_column_letter(j)].width = 12
    ws.freeze_panes = 'B5'

def pct(a, b):
    if a is None or b in (None, 0): return None
    try: return a/b
    except: return None

def ratio(a, b):
    if a is None or b in (None, 0): return None
    try: return a/b
    except: return None

def montar_resumo(wb, ticker, nome, info_extra, dados, periodos_label):
    ws = wb.create_sheet('Resumo', 0)
    sic_desc, setor, exch, fye = info_extra
    ws.cell(1,1, f'{nome} — {ticker}').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(2,1, f'Setor: {setor or "—"}  |  SIC: {sic_desc or "—"}  |  Bolsa: {exch or "—"}  |  Fim do FY: {fye}').font = SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.cell(3,1, f'Trimestres no banco: {len(periodos_label)}  |  Período: {periodos_label[0]} a {periodos_label[-1]}').font = SUB_FONT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)

    # KPIs últimos 4 trimestres (LTM)
    q_only = [d for d in dados if d.get('tipo') == 'Q']
    last4 = q_only[-4:] if len(q_only) >= 4 else q_only
    def soma(key): return sum((d.get(key) or 0) for d in last4) if last4 else None
    def media(key):
        vals = [d.get(key) for d in last4 if d.get(key) is not None]
        return sum(vals)/len(vals) if vals else None
    ws.cell(5,1, 'KPIs últimos 12 meses (LTM)').font = TITLE_FONT
    kpis = [
        ('Receita (LTM)',           soma('receita_liquida'),  FMT_USD_B),
        ('EBITDA (LTM)',            soma('ebitda'),           FMT_USD_B),
        ('Margem EBITDA (LTM)',     pct(soma('ebitda'), soma('receita_liquida')), FMT_PCT),
        ('Lucro Líquido (LTM)',     soma('lucro_liquido'),    FMT_USD_B),
        ('Margem Líq. (LTM)',       pct(soma('lucro_liquido'), soma('receita_liquida')), FMT_PCT),
        ('FCO (LTM)',               soma('fco'),              FMT_USD_B),
        ('Capex (LTM)',             soma('capex'),            FMT_USD_B),
        ('FCL (LTM)',               soma('fcl'),              FMT_USD_B),
        ('Ativo Total (último)',    last4[-1].get('ativo_total') if last4 else None, FMT_USD_B),
        ('Patrimônio Líquido (último)', last4[-1].get('patrimonio_liquido') if last4 else None, FMT_USD_B),
        ('Dívida Líquida (último)', last4[-1].get('divida_liquida') if last4 else None, FMT_USD_B),
        ('Div.Líq / EBITDA (LTM)',  ratio(last4[-1].get('divida_liquida') if last4 else None,
                                            soma('ebitda')), FMT_NUM),
    ]
    for i, (label, val, fmt) in enumerate(kpis, 6):
        ws.cell(i, 1, label).alignment = LEFT
        c = ws.cell(i, 2, val); c.alignment = RIGHT; c.number_format = fmt
        ws.cell(i, 1).border = BOX; c.border = BOX
        if i % 2 == 0:
            ws.cell(i,1).fill = PatternFill('solid', fgColor='F2F2F2')
            c.fill = PatternFill('solid', fgColor='F2F2F2')

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 22

def main():
    if not os.path.exists(DB):
        print(f'Banco não encontrado: {DB}\nRode antes: python carga_sec_companyfacts.py'); return
    conn = sqlite3.connect(DB)
    termo = sys.argv[1] if len(sys.argv)>1 else input('> Ticker, CIK ou nome: ').strip()
    if not termo: print('Nada informado.'); return
    cik = resolver(conn, termo)
    if not cik:
        print(f'Não encontrei empresa para "{termo}".'); return

    info = info_empresa(conn, cik)
    if not info: print(f'CIK {cik} sem dados.'); return
    ticker, nome, sic_d, setor, exch, fye = info

    rows = buscar_trimestres(conn, cik)
    if not rows: print(f'Empresa {nome} sem registros financeiros.'); return
    dados = [to_dict(r) for r in rows]
    periodos_label = [f"{d['ano']} {d['trimestre']}" for d in dados]

    print(f'Empresa: {nome} ({ticker}) | CIK {cik}')
    print(f'  {len(dados)} períodos | {periodos_label[0]} -> {periodos_label[-1]}')

    wb = Workbook(); wb.remove(wb.active)
    montar_aba_secao(wb, 'DRE', DRE, dados, periodos_label)
    montar_aba_secao(wb, 'Balanço', BP, dados, periodos_label)
    montar_aba_secao(wb, 'DFC', DFC, dados, periodos_label)
    montar_indicadores(wb, dados, periodos_label)
    montar_resumo(wb, ticker, nome, (sic_d, setor, exch, fye), dados, periodos_label)

    safe = re.sub(r'[^A-Za-z0-9._-]','_', f'{ticker or "company"}_{nome[:40]}')
    out = os.path.join(OUT_DIR, f'historico_{safe}_{datetime.now():%Y%m%d_%H%M}.xlsx')
    wb.save(out)
    print(f'Excel salvo: {out}  ({os.path.getsize(out)/1024:.1f} KB)')

if __name__ == '__main__':
    main()
