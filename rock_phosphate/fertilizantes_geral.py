"""Coletor + dashboard generico para fertilizantes - importacoes BR.

Pega varios NCMs (rocha fosfatica, KCl, ureia, MAP, DAP, sulfato amonio, NPK)
a partir dos CSVs anuais ja baixados no _cache/ do MDIC.

Gera Excel unico: data/dashboard_fertilizantes.xlsx
  - Aba 'Resumo Geral' - total BR por fertilizante x ano
  - Aba por fertilizante (rocha, KCl, ureia, MAP, DAP, etc.) com mesmo formato
"""
from __future__ import annotations
import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
CACHE = ROOT / "_cache"
OUT_XLSX = ROOT / "dashboard_fertilizantes.xlsx"

ANOS = [2020, 2021, 2022, 2023, 2024, 2025, 2026]

# ===== Familias de fertilizantes =====
# (label, prefixos NCM, descricao)
FAMILIAS = [
    ("Rocha Fosfática",  ["2510"],          "Fosfatos naturais (bruto e moído) - matéria-prima"),
    ("Cloreto de Potássio (KCl)", ["31042010", "31042090", "310420"], "Cloreto de potássio - o K do NPK"),
    ("Ureia",            ["31021010", "31021090", "310210"], "Ureia - principal fonte de N"),
    ("Sulfato de Amônio", ["31022100", "31022110", "31022190", "310221"], "Sulfato de amônio (N+S)"),
    ("Nitrato de Amônio", ["31023000", "310230"], "Nitrato de amônio"),
    ("MAP",              ["31054000", "310540"], "Fosfato monoamônico"),
    ("DAP",              ["31053000", "310530"], "Fosfato diamônico"),
    ("NPK formulados",   ["31052000", "310520"], "Adubos NPK formulados"),
    ("Sulfato de Potássio", ["31043000", "310430"], "Sulfato de potássio (K+S)"),
    ("Superfosfato Simples (SSP)", ["31031100", "310311"], "SSP - P+Ca+S"),
    ("Superfosfato Triplo (TSP)", ["31031900", "310319"], "TSP - alta concentração de P"),
]

# ===== Estilos =====
fill_header = PatternFill("solid", fgColor="305496")
fill_subheader = PatternFill("solid", fgColor="8EA9DB")
fill_alt = PatternFill("solid", fgColor="F2F2F2")
fill_egito = PatternFill("solid", fgColor="FFE699")
fill_destaque = PatternFill("solid", fgColor="C6EFCE")
f_title = Font(bold=True, color="FFFFFF", size=14)
f_header = Font(bold=True, color="FFFFFF", size=10)
f_subheader = Font(bold=True, size=10)
f_data = Font(size=10)
al_c = Alignment(horizontal="center", vertical="center")
al_l = Alignment(horizontal="left", vertical="center", indent=1)
al_r = Alignment(horizontal="right", vertical="center", indent=1)
s = Side(style="thin", color="BFBFBF")
brd = Border(left=s, right=s, top=s, bottom=s)


def carregar_paises():
    df = pd.read_csv(CACHE / "PAIS.csv", sep=";", encoding="latin-1", dtype=str)
    return df[["CO_PAIS", "NO_PAIS"]].copy()


def carregar_tabela_aux(nome: str, url: str, cols_keep: list[str]) -> pd.DataFrame:
    """Carrega tabela auxiliar (URF, VIA, etc) - baixa via curl se nao em cache."""
    import subprocess
    f = CACHE / f"{nome}.csv"
    if not f.exists() or f.stat().st_size < 100:
        print(f"  baixando {nome}.csv...")
        subprocess.run([
            "curl", "-k", "-L", "-s",
            "-A", "Mozilla/5.0 Chrome/120",
            "-o", str(f), url,
        ], timeout=60, check=True)
    df = pd.read_csv(f, sep=";", encoding="latin-1", dtype=str)
    keep = [c for c in cols_keep if c in df.columns]
    return df[keep].copy()


def carregar_csv_ano(ano: int) -> pd.DataFrame:
    """Carrega CSV anual completo (todos NCMs)."""
    f = CACHE / f"IMP_{ano}.csv"
    if not f.exists():
        print(f"  [skip] {ano}: arquivo nao em cache")
        return pd.DataFrame()
    print(f"  carregando IMP_{ano}.csv ...")
    df = pd.read_csv(f, sep=";", encoding="latin-1",
                     dtype={"CO_NCM": str, "CO_PAIS": str, "CO_ANO": str,
                            "CO_MES": str, "CO_URF": str, "CO_VIA": str,
                            "SG_UF_NCM": str})
    return df


def filtrar_familia(df_all_anos: pd.DataFrame, prefixos: list[str]) -> pd.DataFrame:
    """Filtra so linhas que comecam com algum dos prefixos NCM dados."""
    mask = pd.Series(False, index=df_all_anos.index)
    for p in prefixos:
        mask = mask | df_all_anos["CO_NCM"].str.startswith(p)
    return df_all_anos[mask].copy()


def construir_resumo(df: pd.DataFrame, paises: pd.DataFrame) -> pd.DataFrame:
    """Agrega por ano + pais. Retorna DF com: CO_ANO, NO_PAIS, TON, USD_FOB."""
    if df.empty:
        return pd.DataFrame()
    df = df.copy()
    df["KG_LIQUIDO"] = pd.to_numeric(df["KG_LIQUIDO"], errors="coerce").fillna(0)
    df["VL_FOB"] = pd.to_numeric(df["VL_FOB"], errors="coerce").fillna(0)
    df = df.merge(paises, on="CO_PAIS", how="left")
    res = (df.groupby(["CO_ANO", "NO_PAIS"], as_index=False)
             .agg(KG=("KG_LIQUIDO", "sum"),
                  VL_FOB_USD=("VL_FOB", "sum")))
    res["TON"] = (res["KG"] / 1000).round(0)
    res = res.sort_values(["CO_ANO", "TON"], ascending=[False, False])
    return res


def aplicar_header_row(ws, row, cols, valores, fill=fill_header, font=f_header):
    for c, v in zip(cols, valores):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = fill; cell.font = font; cell.alignment = al_c; cell.border = brd


def aba_resumo_geral(wb, resumos_por_familia: dict[str, pd.DataFrame]):
    """Aba 1: total Brasil por fertilizante x ano (em TON e USD)."""
    ws = wb.create_sheet("Resumo Geral", 0)
    ws.sheet_view.showGridLines = False

    ws["B2"] = "Importações BR de Fertilizantes — Resumo por Família"
    ws["B2"].font = f_title; ws["B2"].fill = fill_header; ws["B2"].alignment = al_c
    ws.merge_cells("B2:K2")

    # --- BLOCO 1: TON por familia x ano ---
    ws["B4"] = "TONELADAS por Família x Ano"
    ws["B4"].font = f_subheader; ws["B4"].fill = fill_subheader; ws["B4"].alignment = al_c
    ws.merge_cells("B4:K4")

    aplicar_header_row(ws, 5, [2] + list(range(3, 3 + len(ANOS))) + [3 + len(ANOS)],
                       ["Família"] + [str(a) for a in ANOS] + ["Total"])

    row = 6
    totais_por_ano = {a: 0 for a in ANOS}
    for label, df_res in resumos_por_familia.items():
        ws.cell(row=row, column=2, value=label).alignment = al_l
        total_fam = 0
        for i, ano in enumerate(ANOS, start=3):
            sub = df_res[df_res["CO_ANO"].astype(int) == ano] if not df_res.empty else pd.DataFrame()
            ton = float(sub["TON"].sum()) if not sub.empty else 0
            ws.cell(row=row, column=i, value=ton).number_format = '#,##0'
            ws.cell(row=row, column=i).alignment = al_r
            total_fam += ton
            totais_por_ano[ano] += ton
        # Total familia
        cell = ws.cell(row=row, column=3 + len(ANOS), value=total_fam)
        cell.number_format = '#,##0'; cell.alignment = al_r; cell.font = Font(bold=True, size=10)
        for c in range(2, 4 + len(ANOS)):
            cc = ws.cell(row=row, column=c)
            cc.border = brd
            if not cc.font.bold:
                cc.font = f_data
            if row % 2 == 0:
                cc.fill = fill_alt
        row += 1

    # linha total
    ws.cell(row=row, column=2, value="TOTAL FERTILIZANTES").font = Font(bold=True, color="FFFFFF", size=10)
    ws.cell(row=row, column=2).fill = fill_header
    ws.cell(row=row, column=2).alignment = al_l
    grand_total = 0
    for i, ano in enumerate(ANOS, start=3):
        cell = ws.cell(row=row, column=i, value=totais_por_ano[ano])
        cell.number_format = '#,##0'; cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = fill_header; cell.alignment = al_r
        grand_total += totais_por_ano[ano]
    cell = ws.cell(row=row, column=3 + len(ANOS), value=grand_total)
    cell.number_format = '#,##0'; cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = fill_header; cell.alignment = al_r
    for c in range(2, 4 + len(ANOS)):
        ws.cell(row=row, column=c).border = brd

    # --- BLOCO 2: USD FOB por familia x ano ---
    row += 3
    ws.cell(row=row, column=2, value="USD FOB por Família x Ano").font = f_subheader
    ws.cell(row=row, column=2).fill = fill_subheader
    ws.cell(row=row, column=2).alignment = al_c
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3 + len(ANOS))
    row += 1
    aplicar_header_row(ws, row, [2] + list(range(3, 3 + len(ANOS))) + [3 + len(ANOS)],
                       ["Família"] + [str(a) for a in ANOS] + ["Total"])
    row += 1

    totais_usd_ano = {a: 0 for a in ANOS}
    for label, df_res in resumos_por_familia.items():
        ws.cell(row=row, column=2, value=label).alignment = al_l
        total_fam = 0
        for i, ano in enumerate(ANOS, start=3):
            sub = df_res[df_res["CO_ANO"].astype(int) == ano] if not df_res.empty else pd.DataFrame()
            usd = float(sub["VL_FOB_USD"].sum()) if not sub.empty else 0
            ws.cell(row=row, column=i, value=usd).number_format = '"US$ "#,##0'
            ws.cell(row=row, column=i).alignment = al_r
            total_fam += usd
            totais_usd_ano[ano] += usd
        cell = ws.cell(row=row, column=3 + len(ANOS), value=total_fam)
        cell.number_format = '"US$ "#,##0'; cell.alignment = al_r; cell.font = Font(bold=True, size=10)
        for c in range(2, 4 + len(ANOS)):
            cc = ws.cell(row=row, column=c)
            cc.border = brd
            if not cc.font.bold:
                cc.font = f_data
            if row % 2 == 0:
                cc.fill = fill_alt
        row += 1

    # Larguras
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    for i in range(3, 4 + len(ANOS)):
        ws.column_dimensions[get_column_letter(i)].width = 16


def aba_familia(wb, label, df_res, df_raw, descricao):
    """1 aba por fertilizante com pivot pais x ano (TON e USD) + top 2026."""
    # sanitize nome aba (max 31 chars, sem :/*?[]\\
    nome = label[:31].replace("/", "-").replace(":", "")
    ws = wb.create_sheet(nome)
    ws.sheet_view.showGridLines = False

    ws["B2"] = f"{label} — Importações BR"
    ws["B2"].font = f_title; ws["B2"].fill = fill_header; ws["B2"].alignment = al_c
    ws.merge_cells("B2:K2")

    ws["B3"] = descricao
    ws["B3"].font = Font(italic=True, size=10, color="595959")
    ws["B3"].alignment = al_l
    ws.merge_cells("B3:K3")

    if df_res.empty:
        ws["B5"] = "Sem dados nesse NCM no período."
        ws["B5"].font = Font(italic=True, size=10)
        return

    # ===== Pivot pais x ano (TON) =====
    ws["B5"] = "Toneladas por País x Ano"
    ws["B5"].font = f_subheader; ws["B5"].fill = fill_subheader; ws["B5"].alignment = al_c
    ws.merge_cells("B5:K5")

    pivot = df_res.pivot_table(index="NO_PAIS", columns="CO_ANO",
                                values="TON", aggfunc="sum", fill_value=0)
    pivot.columns = [int(c) for c in pivot.columns]
    pivot = pivot.sort_index(axis=1)
    anos_pivot = list(pivot.columns)
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False)

    # Header
    headers_t = ["País"] + [str(a) for a in anos_pivot] + ["Total"]
    aplicar_header_row(ws, 6, list(range(2, 2 + len(headers_t))), headers_t)

    row = 7
    for pais, line in pivot.iterrows():
        ws.cell(row=row, column=2, value=str(pais)).alignment = al_l
        for i, ano in enumerate(anos_pivot, start=3):
            cell = ws.cell(row=row, column=i, value=float(line[ano]))
            cell.number_format = '#,##0'; cell.alignment = al_r
        # Total
        cell = ws.cell(row=row, column=3 + len(anos_pivot), value=float(line["Total"]))
        cell.number_format = '#,##0'; cell.alignment = al_r; cell.font = Font(bold=True, size=10)

        for c in range(2, 4 + len(anos_pivot)):
            cc = ws.cell(row=row, column=c)
            cc.border = brd
            if not cc.font.bold:
                cc.font = f_data
            if pais == "Egito":
                cc.fill = fill_egito
            elif row % 2 == 0:
                cc.fill = fill_alt
        row += 1

    # ===== Top 10 origens ano mais recente (USD) =====
    row += 2
    ano_max = int(df_res["CO_ANO"].max())
    ws.cell(row=row, column=2, value=f"Top Origens em {ano_max} (USD FOB)").font = f_subheader
    ws.cell(row=row, column=2).fill = fill_subheader
    ws.cell(row=row, column=2).alignment = al_c
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    row += 1
    aplicar_header_row(ws, row, [2, 3, 4, 5, 6, 7],
                       ["#", "País", "TON", "USD FOB", "USD/ton", "Share TON %"])
    row += 1

    top = df_res[df_res["CO_ANO"].astype(int) == ano_max].sort_values("TON", ascending=False).head(15).copy()
    total_ton_ano = float(top["TON"].sum())
    pos = 1
    for _, r in top.iterrows():
        ws.cell(row=row, column=2, value=pos).alignment = al_c
        ws.cell(row=row, column=3, value=str(r["NO_PAIS"])).alignment = al_l
        ws.cell(row=row, column=4, value=float(r["TON"])).number_format = '#,##0'
        ws.cell(row=row, column=5, value=float(r["VL_FOB_USD"])).number_format = '"US$ "#,##0'
        ws.cell(row=row, column=6, value=(float(r["VL_FOB_USD"]) / float(r["TON"])) if r["TON"] else 0).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=(float(r["TON"]) / total_ton_ano) if total_ton_ano else 0).number_format = '0.0%'
        for c in range(2, 8):
            cell = ws.cell(row=row, column=c)
            cell.border = brd; cell.font = f_data
            if r["NO_PAIS"] == "Egito":
                cell.fill = fill_egito
            elif pos % 2 == 0:
                cell.fill = fill_alt
        pos += 1
        row += 1

    # Larguras
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    for i in range(3, 4 + len(anos_pivot)):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.freeze_panes = "C7"


def _juntar_full(raws: dict, paises, urfs, vias, ncms_tab=None) -> pd.DataFrame:
    """Concatena raws de todas as familias e decodifica."""
    dfs = []
    for label, df in raws.items():
        if df.empty:
            continue
        d = df.copy()
        d["FAMILIA"] = label
        dfs.append(d)
    if not dfs:
        return pd.DataFrame()
    full = pd.concat(dfs, ignore_index=True)
    # Strip whitespace (codigos ja vem como str com zeros a esquerda)
    for col in ["CO_PAIS", "CO_URF", "CO_VIA"]:
        if col in full.columns:
            full[col] = full[col].str.strip()
    full = full.merge(paises, on="CO_PAIS", how="left")
    full = full.merge(urfs, on="CO_URF", how="left")
    full = full.merge(vias, on="CO_VIA", how="left")
    if ncms_tab is not None:
        full = full.merge(ncms_tab, on="CO_NCM", how="left")
    for col in ["KG_LIQUIDO", "VL_FOB", "VL_FRETE", "VL_SEGURO", "QT_ESTAT"]:
        if col in full.columns:
            full[col] = pd.to_numeric(full[col], errors="coerce").fillna(0)
    return full


def aba_banco_completo(wb, full: pd.DataFrame):
    """Aba final com TODAS as linhas raw de todos os fertilizantes, decodificado."""
    ws = wb.create_sheet("Banco Completo")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Banco Completo — Importações BR Cap.25/31 (todas as famílias)"
    ws["A1"].font = f_title; ws["A1"].fill = fill_header; ws["A1"].alignment = al_c

    if full.empty:
        return

    # Ordena: ano desc, valor desc
    full = full.sort_values(["CO_ANO", "VL_FOB"], ascending=[False, False])

    # Reordena colunas
    col_order = ["FAMILIA", "CO_ANO", "CO_MES", "CO_NCM", "DESCR_NCM",
                 "NO_PAIS", "CO_PAIS",
                 "SG_UF_NCM", "NO_VIA", "CO_VIA", "NO_URF", "CO_URF",
                 "KG_LIQUIDO", "VL_FOB", "VL_FRETE", "VL_SEGURO", "QT_ESTAT"]
    col_order = [c for c in col_order if c in full.columns]
    full = full[col_order]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_order))

    # Header
    for c, h in enumerate(col_order, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = fill_header; cell.font = f_header; cell.alignment = al_c; cell.border = brd

    # Linhas
    print(f"  escrevendo {len(full):,} linhas no banco completo (pode demorar 30-60s)...")
    for r_idx, row in enumerate(full.itertuples(index=False), start=4):
        for c_idx, v in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx,
                            value=v if not (isinstance(v, float) and pd.isna(v)) else None)
            cell.border = brd
            cell.font = f_data
            col_name = col_order[c_idx - 1]
            if col_name in ("KG_LIQUIDO", "VL_FOB", "VL_FRETE", "VL_SEGURO", "QT_ESTAT"):
                cell.number_format = '#,##0'
                cell.alignment = al_r
            elif col_name in ("NO_PAIS", "NO_URF", "NO_VIA", "FAMILIA", "DESCR_NCM"):
                cell.alignment = al_l
            else:
                cell.alignment = al_c
            if isinstance(v, str) and v == "Egito":
                cell.fill = fill_egito
            elif r_idx % 2 == 0:
                cell.fill = fill_alt

    # Larguras
    larguras = {"FAMILIA": 28, "CO_ANO": 8, "CO_MES": 7, "CO_NCM": 12,
                "DESCR_NCM": 60,
                "NO_PAIS": 22, "CO_PAIS": 8, "SG_UF_NCM": 8,
                "NO_VIA": 14, "CO_VIA": 7,
                "NO_URF": 36, "CO_URF": 10,
                "KG_LIQUIDO": 16, "VL_FOB": 18,
                "VL_FRETE": 12, "VL_SEGURO": 12, "QT_ESTAT": 14}
    for c, h in enumerate(col_order, start=1):
        ws.column_dimensions[get_column_letter(c)].width = larguras.get(h, 12)

    ws.auto_filter.ref = f"A3:{get_column_letter(len(col_order))}{len(full)+3}"
    ws.freeze_panes = "A4"


def aba_top_ncm_brasil(wb, anos_carregados: list, paises: pd.DataFrame,
                         ncms_tab: pd.DataFrame):
    """Top 50 NCMs mais importados pelo BR no ano mais recente disponivel.

    Vai alem dos fertilizantes - usa o universo COMPLETO do CSV.
    """
    ws = wb.create_sheet("Top 50 NCM BR")
    ws.sheet_view.showGridLines = False

    ws["B2"] = "Top 50 NCMs mais Importados pelo Brasil"
    ws["B2"].font = f_title; ws["B2"].fill = fill_header; ws["B2"].alignment = al_c
    ws.merge_cells("B2:J2")

    # Le o ano mais recente disponivel
    ano_max = max(anos_carregados)
    print(f"  carregando IMP_{ano_max}.csv completo (todos NCMs) para ranking...")
    f = CACHE / f"IMP_{ano_max}.csv"
    df = pd.read_csv(f, sep=";", encoding="latin-1",
                     dtype={"CO_NCM": str, "CO_PAIS": str, "CO_ANO": str,
                            "CO_MES": str, "CO_URF": str, "CO_VIA": str})
    df["KG_LIQUIDO"] = pd.to_numeric(df["KG_LIQUIDO"], errors="coerce").fillna(0)
    df["VL_FOB"] = pd.to_numeric(df["VL_FOB"], errors="coerce").fillna(0)
    df = df.merge(paises, on="CO_PAIS", how="left")

    # Top 50 NCMs por VL_FOB total
    top_ncm = (df.groupby("CO_NCM").agg(
                  KG=("KG_LIQUIDO", "sum"),
                  USD=("VL_FOB", "sum"),
                  N_LINHAS=("CO_NCM", "count"),
              ).reset_index())
    top_ncm["TON"] = (top_ncm["KG"] / 1000).round(0)
    top_ncm = top_ncm.sort_values("USD", ascending=False).head(50)

    # Junta descricao NCM
    top_ncm = top_ncm.merge(ncms_tab, on="CO_NCM", how="left")

    # Top pais por NCM (origem dominante)
    pais_por_ncm = (df.groupby(["CO_NCM", "NO_PAIS"])["VL_FOB"].sum()
                       .reset_index()
                       .sort_values(["CO_NCM", "VL_FOB"], ascending=[True, False])
                       .drop_duplicates("CO_NCM"))
    pais_dom = dict(zip(pais_por_ncm["CO_NCM"], pais_por_ncm["NO_PAIS"]))

    total_usd_br = df["VL_FOB"].sum()
    total_ton_br = df["KG_LIQUIDO"].sum() / 1000

    # Subtitulo com totais
    ws["B4"] = f"Ano: {ano_max}  |  Total importado BR: US$ {total_usd_br:,.0f}  |  {total_ton_br:,.0f} ton  |  {df['CO_NCM'].nunique():,} NCMs distintos"
    ws["B4"].font = Font(italic=True, size=10, color="595959")
    ws["B4"].alignment = al_l
    ws.merge_cells("B4:J4")

    aplicar_header_row(ws, 5, [2, 3, 4, 5, 6, 7, 8, 9, 10],
                       ["#", "NCM", "Descrição", "TON", "USD FOB", "USD/ton",
                        "% do BR", "Top Origem", "N Linhas"])

    pos = 1
    row = 6
    for _, r in top_ncm.iterrows():
        descr = str(r.get("DESCR_NCM", "") or "")
        if len(descr) > 250:
            descr = descr[:247] + "..."
        ws.cell(row=row, column=2, value=pos).alignment = al_c
        ws.cell(row=row, column=3, value=str(r["CO_NCM"])).alignment = al_c
        ws.cell(row=row, column=4, value=descr).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws.cell(row=row, column=5, value=float(r["TON"])).number_format = '#,##0'
        ws.cell(row=row, column=6, value=float(r["USD"])).number_format = '"US$ "#,##0'
        usd_ton = (float(r["USD"]) / float(r["TON"])) if r["TON"] else 0
        ws.cell(row=row, column=7, value=usd_ton).number_format = '#,##0.00'
        ws.cell(row=row, column=8, value=float(r["USD"]) / total_usd_br if total_usd_br else 0).number_format = '0.00%'
        ws.cell(row=row, column=9, value=str(pais_dom.get(r["CO_NCM"], "—"))).alignment = al_l
        ws.cell(row=row, column=10, value=int(r["N_LINHAS"])).number_format = '#,##0'

        for c in range(2, 11):
            cell = ws.cell(row=row, column=c)
            cell.border = brd; cell.font = f_data
            if pos % 2 == 0:
                cell.fill = fill_alt

        # altura maior por causa do wrap da descricao
        ws.row_dimensions[row].height = 30
        pos += 1
        row += 1

    # Larguras
    ws.column_dimensions["A"].width = 2
    widths = {2: 6, 3: 12, 4: 60, 5: 14, 6: 18, 7: 12, 8: 11, 9: 26, 10: 12}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "B6"


def aba_por_porto(wb, full: pd.DataFrame):
    """Aba com pivot Porto x Familia (TON) e ranking dos portos."""
    ws = wb.create_sheet("Por Porto", -1)  # antes do Banco Completo
    ws.sheet_view.showGridLines = False

    ws["B2"] = "Importações de Fertilizantes por Porto/URF de Entrada"
    ws["B2"].font = f_title; ws["B2"].fill = fill_header; ws["B2"].alignment = al_c
    ws.merge_cells("B2:L2")

    if full.empty:
        return

    # ===== BLOCO 1: Top 20 portos por toneladas (todos anos) =====
    ws["B4"] = "Top Portos por TON Total (todas as famílias, todos os anos)"
    ws["B4"].font = f_subheader; ws["B4"].fill = fill_subheader; ws["B4"].alignment = al_c
    ws.merge_cells("B4:G4")

    aplicar_header_row(ws, 5, [2, 3, 4, 5, 6, 7],
                       ["#", "Porto/URF", "TON", "USD FOB", "Modal Principal", "Share %"])
    top_porto = (full.groupby("NO_URF").agg(
                    KG=("KG_LIQUIDO", "sum"),
                    USD=("VL_FOB", "sum"),
                ).reset_index())
    top_porto["TON"] = top_porto["KG"] / 1000
    top_porto = top_porto.sort_values("TON", ascending=False).head(20)
    total_ton = top_porto["TON"].sum()

    # Modal principal por porto
    modal_por_porto = (full.groupby(["NO_URF", "NO_VIA"])["KG_LIQUIDO"].sum().reset_index()
                          .sort_values(["NO_URF", "KG_LIQUIDO"], ascending=[True, False])
                          .drop_duplicates("NO_URF"))
    mp = dict(zip(modal_por_porto["NO_URF"], modal_por_porto["NO_VIA"]))

    row = 6
    pos = 1
    for _, r in top_porto.iterrows():
        ws.cell(row=row, column=2, value=pos).alignment = al_c
        ws.cell(row=row, column=3, value=str(r["NO_URF"]) if pd.notna(r["NO_URF"]) else "(sem URF)").alignment = al_l
        ws.cell(row=row, column=4, value=float(r["TON"])).number_format = '#,##0'
        ws.cell(row=row, column=5, value=float(r["USD"])).number_format = '"US$ "#,##0'
        ws.cell(row=row, column=6, value=str(mp.get(r["NO_URF"], ""))).alignment = al_c
        ws.cell(row=row, column=7, value=float(r["TON"]) / total_ton if total_ton else 0).number_format = '0.0%'
        for c in range(2, 8):
            cell = ws.cell(row=row, column=c)
            cell.border = brd; cell.font = f_data
            if pos % 2 == 0: cell.fill = fill_alt
        pos += 1
        row += 1

    # ===== BLOCO 2: Pivot Porto x Familia =====
    row += 2
    ws.cell(row=row, column=2, value="Pivot: TON por Porto x Família (Top 15 portos)").font = f_subheader
    ws.cell(row=row, column=2).fill = fill_subheader
    ws.cell(row=row, column=2).alignment = al_c
    row += 1

    full_t = full.copy()
    full_t["TON"] = full_t["KG_LIQUIDO"] / 1000
    pivot = full_t.pivot_table(index="NO_URF", columns="FAMILIA",
                                values="TON", aggfunc="sum", fill_value=0)
    pivot["TOTAL"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("TOTAL", ascending=False).head(15)
    familias = [c for c in pivot.columns if c != "TOTAL"]
    headers_p = ["Porto/URF"] + familias + ["TOTAL"]
    aplicar_header_row(ws, row, list(range(2, 2 + len(headers_p))), headers_p)
    row += 1
    for porto, line in pivot.iterrows():
        ws.cell(row=row, column=2, value=str(porto)).alignment = al_l
        for i, fam in enumerate(familias, start=3):
            v = float(line[fam])
            cell = ws.cell(row=row, column=i, value=v if v != 0 else None)
            cell.number_format = '#,##0'; cell.alignment = al_r
        cell = ws.cell(row=row, column=2 + len(familias) + 1, value=float(line["TOTAL"]))
        cell.number_format = '#,##0'; cell.alignment = al_r; cell.font = Font(bold=True, size=10)
        for c in range(2, 3 + len(familias) + 1):
            cc = ws.cell(row=row, column=c)
            cc.border = brd
            if not cc.font.bold:
                cc.font = f_data
            if row % 2 == 0: cc.fill = fill_alt
        row += 1

    # ===== BLOCO 3: Modal Geral =====
    row += 2
    ws.cell(row=row, column=2, value="Modal de Transporte (todos os anos)").font = f_subheader
    ws.cell(row=row, column=2).fill = fill_subheader
    ws.cell(row=row, column=2).alignment = al_c
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    row += 1
    aplicar_header_row(ws, row, [2, 3, 4, 5], ["Modal", "TON", "USD FOB", "Share TON %"])
    row += 1
    modal_resumo = (full.groupby("NO_VIA").agg(
                        KG=("KG_LIQUIDO", "sum"),
                        USD=("VL_FOB", "sum"),
                    ).reset_index())
    modal_resumo["TON"] = modal_resumo["KG"] / 1000
    modal_resumo = modal_resumo.sort_values("TON", ascending=False)
    total_modal = modal_resumo["TON"].sum()
    for _, r in modal_resumo.iterrows():
        ws.cell(row=row, column=2, value=str(r["NO_VIA"]) if pd.notna(r["NO_VIA"]) else "(?)").alignment = al_l
        ws.cell(row=row, column=3, value=float(r["TON"])).number_format = '#,##0'
        ws.cell(row=row, column=4, value=float(r["USD"])).number_format = '"US$ "#,##0'
        ws.cell(row=row, column=5, value=float(r["TON"]) / total_modal if total_modal else 0).number_format = '0.0%'
        for c in range(2, 6):
            cell = ws.cell(row=row, column=c)
            cell.border = brd; cell.font = f_data
            if row % 2 == 0: cell.fill = fill_alt
        row += 1

    # Larguras
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    for i in range(3, 15):
        ws.column_dimensions[get_column_letter(i)].width = 14


def main():
    print("=== Coletor + Dashboard Fertilizantes (BR) ===\n")

    paises = carregar_paises()
    print(f"  paises: {len(paises)}")

    urfs = carregar_tabela_aux(
        "URF", "https://balanca.economia.gov.br/balanca/bd/tabelas/URF.csv",
        ["CO_URF", "NO_URF"]
    )
    vias = carregar_tabela_aux(
        "VIA", "https://balanca.economia.gov.br/balanca/bd/tabelas/VIA.csv",
        ["CO_VIA", "NO_VIA"]
    )
    ncms_tab = carregar_tabela_aux(
        "NCM", "https://balanca.economia.gov.br/balanca/bd/tabelas/NCM.csv",
        ["CO_NCM", "NO_NCM_POR"]
    )
    # Renomeia pra DESCR_NCM (mais curto)
    ncms_tab = ncms_tab.rename(columns={"NO_NCM_POR": "DESCR_NCM"})
    print(f"  URFs: {len(urfs)} | VIAs: {len(vias)} | NCMs: {len(ncms_tab)}\n")

    # 1. Carrega todos os anos UMA VEZ (varios GB em RAM mas evita reler)
    print("Carregando CSVs anuais (cache local)...")
    df_total = pd.concat([carregar_csv_ano(a) for a in ANOS if (CACHE / f"IMP_{a}.csv").exists()],
                         ignore_index=True)
    print(f"  total: {len(df_total):,} linhas em todos os anos\n")

    # Filtra apenas linhas que comecam com 25 ou 31 (otimizacao)
    df_total = df_total[df_total["CO_NCM"].str.startswith(("25", "31"))].copy()
    print(f"  filtrado (cap.25+31): {len(df_total):,} linhas\n")

    # 2. Para cada familia, agrega e guarda
    resumos = {}
    raws = {}
    for label, prefixos, descr in FAMILIAS:
        df_fam = filtrar_familia(df_total, prefixos)
        if df_fam.empty:
            print(f"  [vazio] {label}")
            resumos[label] = pd.DataFrame()
            raws[label] = pd.DataFrame()
            continue
        res = construir_resumo(df_fam, paises)
        resumos[label] = res
        raws[label] = df_fam
        total_ton = res["TON"].sum()
        total_usd = res["VL_FOB_USD"].sum()
        print(f"  {label:<35} {len(df_fam):>6,} linhas | TON {total_ton:>12,.0f} | USD {total_usd:>15,.0f}")

    # 3. Gera Excel
    print("\nGerando dashboard...")
    wb = Workbook()
    wb.remove(wb.active)

    aba_resumo_geral(wb, resumos)
    for label, _, descr in FAMILIAS:
        if not resumos[label].empty:
            aba_familia(wb, label, resumos[label], raws[label], descr)

    # Junta todos os raws decodificados (pais + porto + modal + ncm)
    full = _juntar_full(raws, paises, urfs, vias, ncms_tab)
    aba_por_porto(wb, full)
    aba_top_ncm_brasil(wb, [a for a in ANOS if (CACHE / f"IMP_{a}.csv").exists()],
                         paises, ncms_tab)
    aba_banco_completo(wb, full)

    wb.save(OUT_XLSX)
    print(f"\n✅ Salvo: {OUT_XLSX}")
    print(f"   Abas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
