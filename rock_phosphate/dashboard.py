"""Gera dashboard Excel com todos dados coletados de rock phosphate (NCM 2510).

Abas:
  1. Resumo BR     - KPIs por ano + ranking paises 2026
  2. Pivot Pais x Ano (TON)  - matriz visual
  3. Pivot Pais x Ano (USD)  - matriz valor
  4. Egito Detalhe  - foco Egito-Brasil (serie temporal + breakdown por NCM)
  5. Banco Completo - todas as 614 linhas raw
  6. Por NCM       - breakdown 2510.10 (bruto) vs 2510.20 (moido)
"""
from __future__ import annotations
import os
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
OUT = ROOT / "dashboard_rock_phosphate.xlsx"

# ===== Estilos =====
fill_header = PatternFill("solid", fgColor="305496")
fill_subheader = PatternFill("solid", fgColor="8EA9DB")
fill_destaque = PatternFill("solid", fgColor="C6EFCE")
fill_alt = PatternFill("solid", fgColor="F2F2F2")
fill_egito = PatternFill("solid", fgColor="FFE699")

f_title = Font(bold=True, color="FFFFFF", size=14)
f_header = Font(bold=True, color="FFFFFF", size=10)
f_subheader = Font(bold=True, size=10)
f_data = Font(size=10)
f_destaque = Font(bold=True, size=11)

al_c = Alignment(horizontal="center", vertical="center")
al_l = Alignment(horizontal="left", vertical="center", indent=1)
al_r = Alignment(horizontal="right", vertical="center", indent=1)

s = Side(style="thin", color="BFBFBF")
brd = Border(left=s, right=s, top=s, bottom=s)


def aplicar_header(ws, row, cols, valores, fill=fill_header, font=f_header):
    for c, v in zip(cols, valores):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = fill; cell.font = font; cell.alignment = al_c; cell.border = brd


def aba_resumo(wb, resumo):
    ws = wb.create_sheet("Resumo BR")
    ws.sheet_view.showGridLines = False

    # Titulo
    ws["B2"] = "Importações Brasileiras de Rocha Fosfática (NCM 2510)"
    ws["B2"].font = f_title; ws["B2"].fill = fill_header; ws["B2"].alignment = al_c
    ws.merge_cells("B2:G2")

    # ===== KPIs por ano =====
    ws["B4"] = "Total Importado por Ano"
    ws["B4"].font = f_subheader; ws["B4"].fill = fill_subheader; ws["B4"].alignment = al_c
    ws.merge_cells("B4:G4")

    aplicar_header(ws, 5, [2, 3, 4, 5, 6, 7],
                   ["Ano", "TON Totais", "USD FOB Total", "USD/ton Médio", "Países Origem", "Top Origem"])

    totais = (resumo.groupby("CO_ANO", as_index=False)
                    .agg(TON=("TON", "sum"),
                         USD=("VL_FOB_USD", "sum"),
                         N_PAISES=("NO_PAIS", "nunique")))
    totais["USD_TON"] = (totais["USD"] / totais["TON"]).round(2)
    # top origem por ano
    top_por_ano = (resumo.sort_values(["CO_ANO", "TON"], ascending=[True, False])
                          .groupby("CO_ANO").first().reset_index()
                          [["CO_ANO", "NO_PAIS"]])
    totais = totais.merge(top_por_ano, on="CO_ANO")
    totais = totais.sort_values("CO_ANO")

    row = 6
    for _, r in totais.iterrows():
        ws.cell(row=row, column=2, value=int(r["CO_ANO"])).alignment = al_c
        ws.cell(row=row, column=3, value=float(r["TON"])).number_format = '#,##0'
        ws.cell(row=row, column=4, value=float(r["USD"])).number_format = '"US$ "#,##0'
        ws.cell(row=row, column=5, value=float(r["USD_TON"])).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=int(r["N_PAISES"])).alignment = al_c
        ws.cell(row=row, column=7, value=str(r["NO_PAIS"])).alignment = al_l
        for c in range(2, 8):
            cell = ws.cell(row=row, column=c)
            cell.border = brd
            cell.font = f_data
            if row % 2 == 0:
                cell.fill = fill_alt
        row += 1

    # ===== Ranking 2026 (ano mais recente) =====
    row += 2
    ws.cell(row=row, column=2, value="Top Origens (2026)").font = f_subheader
    ws.cell(row=row, column=2).fill = fill_subheader
    ws.cell(row=row, column=2).alignment = al_c
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    row += 1
    aplicar_header(ws, row, [2, 3, 4, 5, 6, 7],
                   ["#", "País", "TON", "USD FOB", "USD/ton", "Share TON %"])
    row += 1
    ano_max = int(resumo["CO_ANO"].max())
    top = resumo[resumo["CO_ANO"] == ano_max].sort_values("TON", ascending=False).copy()
    total_ton = top["TON"].sum()
    pos = 1
    for _, r in top.iterrows():
        ws.cell(row=row, column=2, value=pos).alignment = al_c
        ws.cell(row=row, column=3, value=str(r["NO_PAIS"])).alignment = al_l
        ws.cell(row=row, column=4, value=float(r["TON"])).number_format = '#,##0'
        ws.cell(row=row, column=5, value=float(r["VL_FOB_USD"])).number_format = '"US$ "#,##0'
        usd_ton = float(r["VL_FOB_USD"]) / float(r["TON"]) if r["TON"] else 0
        ws.cell(row=row, column=6, value=usd_ton).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=float(r["TON"]) / total_ton if total_ton else 0).number_format = '0.0%'
        for c in range(2, 8):
            cell = ws.cell(row=row, column=c)
            cell.border = brd; cell.font = f_data
            if r["NO_PAIS"] == "Egito":
                cell.fill = fill_egito
            elif pos % 2 == 0:
                cell.fill = fill_alt
        pos += 1
        row += 1

    # Larguras
    widths = {"B": 8, "C": 28, "D": 16, "E": 18, "F": 14, "G": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions["A"].width = 2


def aba_pivot(wb, resumo, valor_col, title, num_fmt, sheet_name):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    ws["B2"] = title
    ws["B2"].font = f_title; ws["B2"].fill = fill_header; ws["B2"].alignment = al_c

    # Pivot
    pivot = resumo.pivot_table(index="NO_PAIS", columns="CO_ANO",
                                values=valor_col, aggfunc="sum", fill_value=0)
    pivot = pivot.sort_index(axis=1)  # anos crescentes
    # Adiciona total + ordena por total desc
    pivot["Total"] = pivot.sum(axis=1)
    pivot = pivot.sort_values("Total", ascending=False)

    anos = [c for c in pivot.columns if c != "Total"]

    # Header
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3 + len(anos) + 1)
    headers = ["País"] + [str(int(a)) for a in anos] + ["Total"]
    aplicar_header(ws, 4, list(range(2, 2 + len(headers))), headers)

    row = 5
    for pais, line in pivot.iterrows():
        ws.cell(row=row, column=2, value=str(pais)).alignment = al_l
        for i, ano in enumerate(anos, start=3):
            v = float(line[ano])
            cell = ws.cell(row=row, column=i, value=v)
            cell.number_format = num_fmt; cell.alignment = al_r
        # Total
        cell = ws.cell(row=row, column=3 + len(anos), value=float(line["Total"]))
        cell.number_format = num_fmt; cell.alignment = al_r; cell.font = f_destaque

        for c in range(2, 4 + len(anos)):
            cell = ws.cell(row=row, column=c)
            cell.border = brd
            if not cell.font.bold:
                cell.font = f_data
            if pais == "Egito":
                cell.fill = fill_egito
            elif row % 2 == 0:
                cell.fill = fill_alt
        row += 1

    # Larguras
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    for i in range(3, 4 + len(anos)):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.freeze_panes = "C5"


def aba_egito(wb, raw, resumo):
    ws = wb.create_sheet("Egito Detalhe")
    ws.sheet_view.showGridLines = False

    ws["B2"] = "Importações BR de Rocha Fosfática do Egito (NCM 2510)"
    ws["B2"].font = f_title; ws["B2"].fill = fill_header; ws["B2"].alignment = al_c
    ws.merge_cells("B2:H2")

    # ===== Série temporal =====
    egito_res = resumo[resumo["NO_PAIS"] == "Egito"].sort_values("CO_ANO")

    ws["B4"] = "Série Temporal Anual"
    ws["B4"].font = f_subheader; ws["B4"].fill = fill_subheader; ws["B4"].alignment = al_c
    ws.merge_cells("B4:H4")

    aplicar_header(ws, 5, [2, 3, 4, 5, 6, 7, 8],
                   ["Ano", "TON", "USD FOB", "USD/ton", "% do BR", "Δ TON YoY", "Δ USD YoY"])

    # totais BR por ano
    tot_br = resumo.groupby("CO_ANO")["TON"].sum()

    row = 6
    prev_ton, prev_usd = None, None
    for _, r in egito_res.iterrows():
        ano = int(r["CO_ANO"])
        ton = float(r["TON"])
        usd = float(r["VL_FOB_USD"])
        share = ton / tot_br.get(ano, 1) if tot_br.get(ano, 1) else 0
        d_ton = (ton / prev_ton - 1) if prev_ton else None
        d_usd = (usd / prev_usd - 1) if prev_usd else None

        ws.cell(row=row, column=2, value=ano).alignment = al_c
        ws.cell(row=row, column=3, value=ton).number_format = '#,##0'
        ws.cell(row=row, column=4, value=usd).number_format = '"US$ "#,##0'
        ws.cell(row=row, column=5, value=usd / ton if ton else 0).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=share).number_format = '0.0%'
        if d_ton is not None:
            ws.cell(row=row, column=7, value=d_ton).number_format = '+0.0%;-0.0%'
        if d_usd is not None:
            ws.cell(row=row, column=8, value=d_usd).number_format = '+0.0%;-0.0%'

        for c in range(2, 9):
            cell = ws.cell(row=row, column=c)
            cell.border = brd; cell.font = f_data
            cell.fill = fill_egito
        prev_ton, prev_usd = ton, usd
        row += 1

    # ===== Breakdown por NCM (Egito) =====
    row += 2
    ws.cell(row=row, column=2, value="Breakdown por NCM (Egito, todos anos)").font = f_subheader
    ws.cell(row=row, column=2).fill = fill_subheader
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    aplicar_header(ws, row, [2, 3, 4, 5, 6], ["NCM", "Descrição", "TON", "USD FOB", "USD/ton"])
    row += 1

    NCM_DESC = {
        "25101010": "Fosfato natural NÃO moído",
        "25102010": "Fosfato natural MOÍDO",
        "25101090": "Outros - não moídos",
        "25102090": "Outros - moídos",
    }

    egito_raw = raw[raw["NO_PAIS"] == "Egito"]
    by_ncm = egito_raw.groupby("CO_NCM").agg(
        TON=("KG_LIQUIDO", lambda x: x.sum() / 1000),
        USD=("VL_FOB", "sum"),
    ).sort_values("TON", ascending=False)

    for ncm, line in by_ncm.iterrows():
        ws.cell(row=row, column=2, value=str(ncm)).alignment = al_c
        ws.cell(row=row, column=3, value=NCM_DESC.get(str(ncm), "—")).alignment = al_l
        ws.cell(row=row, column=4, value=float(line["TON"])).number_format = '#,##0'
        ws.cell(row=row, column=5, value=float(line["USD"])).number_format = '"US$ "#,##0'
        usd_ton = float(line["USD"]) / float(line["TON"]) if line["TON"] else 0
        ws.cell(row=row, column=6, value=usd_ton).number_format = '#,##0.00'
        for c in range(2, 7):
            cell = ws.cell(row=row, column=c)
            cell.border = brd; cell.font = f_data
            if row % 2 == 0: cell.fill = fill_alt
        row += 1

    # ===== Mensal 2025-2026 =====
    row += 2
    ws.cell(row=row, column=2, value="Detalhamento Mensal (Egito, 2025-2026)").font = f_subheader
    ws.cell(row=row, column=2).fill = fill_subheader
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    aplicar_header(ws, row, [2, 3, 4, 5, 6], ["Ano", "Mês", "TON", "USD FOB", "USD/ton"])
    row += 1

    mensal = egito_raw[egito_raw["CO_ANO"].astype(int) >= 2025].copy()
    mensal["KG_LIQUIDO"] = pd.to_numeric(mensal["KG_LIQUIDO"], errors="coerce").fillna(0)
    mensal["VL_FOB"] = pd.to_numeric(mensal["VL_FOB"], errors="coerce").fillna(0)
    mensal = (mensal.groupby(["CO_ANO", "CO_MES"]).agg(
                  TON=("KG_LIQUIDO", lambda x: x.sum() / 1000),
                  USD=("VL_FOB", "sum"),
              ).reset_index().sort_values(["CO_ANO", "CO_MES"]))

    for _, m in mensal.iterrows():
        ws.cell(row=row, column=2, value=int(m["CO_ANO"])).alignment = al_c
        ws.cell(row=row, column=3, value=int(m["CO_MES"])).alignment = al_c
        ws.cell(row=row, column=4, value=float(m["TON"])).number_format = '#,##0'
        ws.cell(row=row, column=5, value=float(m["USD"])).number_format = '"US$ "#,##0'
        usd_ton = float(m["USD"]) / float(m["TON"]) if m["TON"] else 0
        ws.cell(row=row, column=6, value=usd_ton).number_format = '#,##0.00'
        for c in range(2, 7):
            cell = ws.cell(row=row, column=c)
            cell.border = brd; cell.font = f_data
            if row % 2 == 0: cell.fill = fill_alt
        row += 1

    # Larguras
    widths = {"B": 10, "C": 30, "D": 14, "E": 18, "F": 14, "G": 14, "H": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions["A"].width = 2


def aba_por_ncm(wb, raw):
    ws = wb.create_sheet("Por NCM")
    ws.sheet_view.showGridLines = False

    ws["B2"] = "Breakdown por NCM (bruto vs moído)"
    ws["B2"].font = f_title; ws["B2"].fill = fill_header; ws["B2"].alignment = al_c
    ws.merge_cells("B2:F2")

    NCM_DESC = {
        "25101010": "Fosfato NÃO moído (bruto)",
        "25102010": "Fosfato MOÍDO",
        "25101090": "Outros bruto",
        "25102090": "Outros moído",
    }

    raw2 = raw.copy()
    raw2["KG_LIQUIDO"] = pd.to_numeric(raw2["KG_LIQUIDO"], errors="coerce").fillna(0)
    raw2["VL_FOB"] = pd.to_numeric(raw2["VL_FOB"], errors="coerce").fillna(0)
    by_ncm_ano = (raw2.groupby(["CO_NCM", "CO_ANO"]).agg(
                    TON=("KG_LIQUIDO", lambda x: x.sum() / 1000),
                    USD=("VL_FOB", "sum"),
                ).reset_index().sort_values(["CO_NCM", "CO_ANO"]))

    aplicar_header(ws, 4, [2, 3, 4, 5, 6], ["NCM", "Descrição", "Ano", "TON", "USD FOB"])

    row = 5
    for _, r in by_ncm_ano.iterrows():
        ncm = str(r["CO_NCM"])
        ws.cell(row=row, column=2, value=ncm).alignment = al_c
        ws.cell(row=row, column=3, value=NCM_DESC.get(ncm, "—")).alignment = al_l
        ws.cell(row=row, column=4, value=int(r["CO_ANO"])).alignment = al_c
        ws.cell(row=row, column=5, value=float(r["TON"])).number_format = '#,##0'
        ws.cell(row=row, column=6, value=float(r["USD"])).number_format = '"US$ "#,##0'
        for c in range(2, 7):
            cell = ws.cell(row=row, column=c)
            cell.border = brd; cell.font = f_data
            if row % 2 == 0: cell.fill = fill_alt
        row += 1

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 18


def aba_raw(wb, raw):
    ws = wb.create_sheet("Banco Completo")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Banco Completo — Todas as importações BR NCM 2510 (raw)"
    ws["A1"].font = f_title; ws["A1"].fill = fill_header; ws["A1"].alignment = al_c
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(len(raw.columns), 14))

    # Header
    headers = list(raw.columns)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=str(h))
        cell.fill = fill_header; cell.font = f_header; cell.alignment = al_c; cell.border = brd

    # Dados
    for r_idx, (_, row) in enumerate(raw.iterrows(), start=4):
        for c_idx, col in enumerate(headers, start=1):
            v = row[col]
            cell = ws.cell(row=r_idx, column=c_idx, value=v if not pd.isna(v) else None)
            cell.border = brd; cell.font = f_data
            if isinstance(v, (int, float)) and col in ["KG_LIQUIDO", "VL_FOB", "VL_FRETE", "VL_SEGURO", "QT_ESTAT"]:
                cell.number_format = '#,##0'
            if r_idx % 2 == 0:
                cell.fill = fill_alt
            if row.get("NO_PAIS") == "Egito":
                cell.fill = fill_egito

    # Larguras
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{r_idx}"
    ws.freeze_panes = "A4"


def main():
    print("=== Gerando dashboard rock_phosphate ===\n")

    # Carrega dados
    raw_path = DATA / "comex_br_2510_raw.csv"
    resumo_path = DATA / "comex_br_2510_resumo.csv"
    if not raw_path.exists():
        print(f"❌ Arquivo nao encontrado: {raw_path}"); return

    raw = pd.read_csv(raw_path, sep=";", dtype={"CO_NCM": str, "CO_PAIS": str, "CO_ANO": str, "CO_MES": str})
    resumo = pd.read_csv(resumo_path, sep=";")
    print(f"  raw: {len(raw):,} linhas")
    print(f"  resumo: {len(resumo):,} linhas (pais x ano)\n")

    wb = Workbook()
    wb.remove(wb.active)

    print("  Gerando aba Resumo BR...")
    aba_resumo(wb, resumo)
    print("  Gerando aba Pivot TON...")
    aba_pivot(wb, resumo, "TON", "Toneladas Importadas (TON) — País x Ano", '#,##0', "Pivot TON")
    print("  Gerando aba Pivot USD...")
    aba_pivot(wb, resumo, "VL_FOB_USD", "Valor Importado (USD FOB) — País x Ano", '"US$ "#,##0', "Pivot USD")
    print("  Gerando aba Egito Detalhe...")
    aba_egito(wb, raw, resumo)
    print("  Gerando aba Por NCM...")
    aba_por_ncm(wb, raw)
    print("  Gerando aba Banco Completo...")
    aba_raw(wb, raw)

    wb.save(OUT)
    print(f"\n✅ Salvo: {OUT}")
    print(f"   Abas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
