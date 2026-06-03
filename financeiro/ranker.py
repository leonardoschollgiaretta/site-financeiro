"""
ranker.py — Ranking de ações a partir do banco financeiro.db
Lê configuração de tabela_de_pesos.xlsx
Gera ranking_resultado.xlsx com duas abas:
  - Ranking Completas : só tickers com todos os indicadores ativos preenchidos (≥4 anos nos multi-ano)
  - Ranking Geral     : completos + parciais (sem_dados excluídos)
"""
import sqlite3
import os
import sys
from datetime import date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "financeiro.db")
PESOS  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tabela_de_pesos.xlsx")
OUT    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ranking_resultado.xlsx")

ANOS        = [2021, 2022, 2023, 2024, 2025]
ANO_REF     = 2025
N_ANOS      = 5
MIN_ANOS    = 5   # mínimo de anos para indicadores multi-ano serem "completo"
FONTE_PRIO  = ["manual", "investsite", "statusinvest", "yfinance"]

# Indicadores que dependem de múltiplos anos
IND_MULTI = {
    "Margem Bruta Méd 5 anos",
    "Lucro Líquido Méd 5 anos %",
    "ROE Médio 5 anos",
    "FCO Médio 5 anos / Receita Líquida",
    "FCF Médio / Receita Líquida Média",
    "FCF Médio / Market Cap",
    "Capex / Margem Bruta (Méd 5 anos)",
    "Capex % Preço (Méd 5 anos)",
    "Dívida Líquida / Receita Méd 21-25",
    "Variação Receita 2025 / 2021",
    "Variação Receita 2025 / Méd 21-24",
    "Variação Lucro 2025 / 2021",
    "Variação Lucro 2025 / Méd 21-24",
}

FILL_CAT = {
    "PRECO":     "D9E1F2",
    "RECEITA":   "E2EFDA",
    "MARGEM":    "FFF2CC",
    "LUCRO":     "FCE4D6",
    "ROE":       "DDEBF7",
    "CAIXA":     "E2EFDA",
    "CAPEX":     "FFF2CC",
    "DIVIDENDO": "D9E1F2",
    "DIVIDA":    "FCE4D6",
}


# ── DB helpers ────────────────────────────────────────────────────────────────

def _fin_cols(conn):
    return [c[1] for c in conn.execute("PRAGMA table_info(financeiros_anuais)").fetchall()]

def _get_fin(conn, cols, ticker, ano):
    for fonte in FONTE_PRIO:
        r = conn.execute(
            "SELECT * FROM financeiros_anuais WHERE ticker=? AND ano=? AND fonte=?",
            (ticker, ano, fonte)
        ).fetchone()
        if r:
            return dict(zip(cols, r))
    return {}

def _v(d, k):
    v = d.get(k)
    return float(v) if v is not None else None

def _avg5(vals):
    """Divide por 5 sempre — mas retorna None se não houver nenhum valor."""
    non_none = [v for v in vals if v is not None]
    if not non_none:
        return None
    return sum(non_none) / N_ANOS

def _safe(a, b):
    if a is None or b is None or b == 0:
        return None
    return a / b

def _var(novo, antigo):
    if novo is None or antigo is None or antigo == 0:
        return None
    return (novo - antigo) / abs(antigo)


# ── Carregar pesos ────────────────────────────────────────────────────────────

def carregar_pesos():
    df = pd.read_excel(PESOS, sheet_name="Configuracao", header=None)
    data = df.iloc[3:].reset_index(drop=True)  # pula título, subtítulo, cabeçalho

    pesos = []
    cat_atual = None
    for _, row in data.iterrows():
        cat = str(row[0]).strip()
        if cat and cat not in ("nan", ""):
            cat_atual = cat
        ind   = str(row[2]).strip()
        ativo = str(row[8]).strip()
        if ind in ("nan", "") or ativo != "S":
            continue
        try:
            lim_inf = float(row[4]) if str(row[4]) not in ("nan", "") else None
            lim_sup = float(row[5]) if str(row[5]) not in ("nan", "") else None
            peso    = float(row[6]) if str(row[6]) not in ("nan", "") else 0.0
            se_neg  = float(row[7]) if str(row[7]) not in ("nan", "") else 0.0
            pesos.append({
                "categoria": cat_atual or "",
                "indicador": ind,
                "melhor":    str(row[3]).strip(),
                "lim_inf":   lim_inf,
                "lim_sup":   lim_sup,
                "peso":      peso,
                "se_neg":    se_neg,
            })
        except Exception:
            pass
    return pesos


# ── Calcular indicadores ──────────────────────────────────────────────────────

def calcular_indicadores(conn, cols, ticker):
    fin = {ano: _get_fin(conn, cols, ticker, ano) for ano in ANOS}

    emp = conn.execute(
        "SELECT acoes_free, acoes_total, moeda FROM empresas WHERE ticker=?", (ticker,)
    ).fetchone()
    acoes = float(emp[0]) if emp and emp[0] else None

    pa = conn.execute(
        "SELECT preco FROM preco_atual WHERE ticker=?", (ticker,)
    ).fetchone()
    preco = float(pa[0]) if pa and pa[0] else None

    divs = {}
    for r in conn.execute(
        "SELECT ano, dividendo_por_acao FROM dividendos_anuais WHERE ticker=? AND ano>=2021",
        (ticker,)
    ).fetchall():
        divs[int(r[0])] = float(r[1]) if r[1] else 0.0

    mkt = (preco * acoes) if (preco and acoes) else None

    def lst(campo):
        return [_v(fin[a], campo) for a in ANOS]

    def cnt(campo):
        return sum(1 for a in ANOS if _v(fin[a], campo) is not None)

    def cnt2(c1, c2):
        return sum(1 for a in ANOS if _v(fin[a], c1) is not None and _v(fin[a], c2) is not None)

    rec_l  = lst("receita_liquida")
    ll_l   = lst("lucro_liquido")
    lb_l   = lst("lucro_bruto")
    pl_l   = lst("patrimonio_liquido")
    fco_l  = lst("fco")
    fcl_l  = lst("fcl")
    cap_l  = lst("capex")

    rec25  = _v(fin[ANO_REF], "receita_liquida")
    rec21  = _v(fin[2021],    "receita_liquida")
    ll25   = _v(fin[ANO_REF], "lucro_liquido")
    ll21   = _v(fin[2021],    "lucro_liquido")
    pl25   = _v(fin[ANO_REF], "patrimonio_liquido")
    cx25   = _v(fin[ANO_REF], "caixa")
    dl25   = _v(fin[ANO_REF], "divida_liquida")
    fcl25  = _v(fin[ANO_REF], "fcl")

    # Médias 2021-2024 (para variação vs. méd)
    rec2124 = _safe(sum(v for v in [_v(fin[a], "receita_liquida") for a in [2021,2022,2023,2024]] if v),
                    sum(1 for a in [2021,2022,2023,2024] if _v(fin[a], "receita_liquida") is not None) or None)
    ll2124  = _safe(sum(v for v in [_v(fin[a], "lucro_liquido")   for a in [2021,2022,2023,2024]] if v),
                    sum(1 for a in [2021,2022,2023,2024] if _v(fin[a], "lucro_liquido")   is not None) or None)

    # Ratios anuais
    mb_a   = [_safe(_v(fin[a], "lucro_bruto"),   _v(fin[a], "receita_liquida"))  for a in ANOS]
    ml_a   = [_safe(_v(fin[a], "lucro_liquido"), _v(fin[a], "receita_liquida"))  for a in ANOS]
    roe_a  = [_safe(_v(fin[a], "lucro_liquido"), _v(fin[a], "patrimonio_liquido")) for a in ANOS]
    fr_a   = [_safe(_v(fin[a], "fco"),           _v(fin[a], "receita_liquida"))  for a in ANOS]
    # capex/margem_bruta anual (capex vem negativo → abs)
    cap_mb_a = []
    for a in ANOS:
        cap_v = _v(fin[a], "capex")
        mb_v  = _safe(_v(fin[a], "lucro_bruto"), _v(fin[a], "receita_liquida"))
        cap_mb_a.append(_safe(abs(cap_v) if cap_v is not None else None, mb_v))

    rec_med  = _avg5(rec_l)
    fcl_med  = _avg5(fcl_l)
    cap_abs_med = _avg5([abs(v) if v is not None else None for v in cap_l])

    # Dividendo: usa 0 para anos sem pagamento (correto — divide sempre por 5)
    div_vals = [divs.get(a, 0.0) for a in ANOS]
    div_med  = sum(div_vals) / N_ANOS
    div_yield = _safe(div_med, preco)

    # Anos disponíveis por campo (para check de completude)
    anys = {
        "receita": cnt("receita_liquida"),
        "ll":      cnt("lucro_liquido"),
        "lb_rec":  cnt2("lucro_bruto", "receita_liquida"),
        "pl":      cnt("patrimonio_liquido"),
        "fco":     cnt("fco"),
        "fcl":     cnt("fcl"),
        "capex":   cnt("capex"),
    }

    # (valor, anos_disponiveis)  — anos_disp usado só em IND_MULTI para check de completude
    ind = {
        "P/VP":                                  (_safe(mkt, pl25),                         1),
        "P/L":                                   (_safe(mkt, ll25),                         1),
        "Variação Receita 2025 / 2021":          (_var(rec25, rec21),                       anys["receita"]),
        "Variação Receita 2025 / Méd 21-24":     (_var(rec25, rec2124),                     anys["receita"]),
        "Margem Bruta Méd 5 anos":               (_avg5(mb_a),                              anys["lb_rec"]),
        "Lucro Líquido 2025 %":                  (_safe(ll25, rec25),                       1),
        "Lucro Líquido Méd 5 anos %":            (_avg5(ml_a),                              anys["ll"]),
        "Variação Lucro 2025 / 2021":            (_var(ll25, ll21),                         anys["ll"]),
        "Variação Lucro 2025 / Méd 21-24":       (_var(ll25, ll2124),                       anys["ll"]),
        "ROE 2025":                              (_safe(ll25, pl25),                        1),
        "ROE Médio 5 anos":                      (_avg5(roe_a),                             anys["pl"]),
        "FCO Médio 5 anos / Receita Líquida":    (_safe(_avg5(fco_l), rec_med or None),     anys["fco"]),
        "FCF / Receita Líquida 2025":            (_safe(fcl25, rec25),                      1),
        "FCF Médio / Receita Líquida Média":     (_safe(fcl_med, rec_med or None),          anys["fcl"]),
        "FCF Médio / Market Cap":                (_safe(mkt, abs(fcl_med) if fcl_med else None), anys["fcl"]),
        "Disponibilidade (Caixa / Mkt Cap)":     (_safe(cx25, mkt),                         1),
        "Capex / Margem Bruta (Méd 5 anos)":     (_avg5(cap_mb_a),                          anys["capex"]),
        "Capex % Preço (Méd 5 anos)":            (_safe(cap_abs_med, mkt),                  anys["capex"]),
        "Capex por Ação (Méd 5 anos)":           (None,                                     0),
        "Dividendo % Anualizado":                (div_yield,                                len([a for a in ANOS if a in divs])),
        "Dívida Líquida / Receita Méd 21-25":    (_safe(dl25, rec_med or None),             anys["receita"]),
        "Dívida Líquida / FCF Médio":            (_safe(dl25, fcl_med or None),             anys["fcl"]),
        "Dívida Líquida / Patrimônio Líquido":   (_safe(dl25, pl25),                        1),
    }
    return ind


# ── Classificar completude ────────────────────────────────────────────────────

def classificar(ind, pesos_ativos):
    algum_dado   = False
    faltando     = 0

    for p in pesos_ativos:
        nome = p["indicador"]
        if nome not in ind:
            faltando += 1
            continue
        valor, anos_disp = ind[nome]

        if valor is not None:
            algum_dado = True

        if nome in IND_MULTI:
            if nome == "Dividendo % Anualizado":
                # 0 dividendos é dado válido — não penaliza completude
                if valor is None:
                    faltando += 1
            elif anos_disp < MIN_ANOS:
                faltando += 1
        else:
            if valor is None:
                faltando += 1

    if not algum_dado:
        return "sem_dados"
    if faltando == 0:
        return "completo"
    return "parcial"


# ── Score ─────────────────────────────────────────────────────────────────────

def _score_hex(score, peso_max):
    """Retorna cor hex interpolada: verde (score alto) → amarelo → laranja (score baixo).
    Normaliza score pelo peso máximo do indicador."""
    if peso_max is None or peso_max == 0:
        return "FFFFFF"
    ratio = max(0.0, min(1.0, score / peso_max))
    # Verde  #70AD47  (112,173,71)
    # Amarelo #FFD966 (255,217,102)
    # Laranja #FF6B35 (255,107,53)
    if ratio >= 0.5:
        # Verde → Amarelo  (ratio 1.0→0.5)
        t  = (ratio - 0.5) / 0.5   # 1.0 em ratio=1.0, 0.0 em ratio=0.5
        r  = int(112 + (1 - t) * (255 - 112))
        g  = int(173 + (1 - t) * (217 - 173))
        b  = int(71  + (1 - t) * (102 - 71))
    else:
        # Amarelo → Laranja (ratio 0.5→0.0)
        t  = ratio / 0.5            # 1.0 em ratio=0.5, 0.0 em ratio=0.0
        r  = 255
        g  = int(107 + t * (217 - 107))
        b  = int(53  + t * (102 - 53))
    return f"{r:02X}{g:02X}{b:02X}"


def _total_hex(total, max_total):
    """Mesma lógica mas para as colunas de total."""
    return _score_hex(total, max_total)


def calcular_score(valor, melhor, lim_inf, lim_sup, peso, se_neg):
    if valor is None:
        return 0.0
    if valor < 0:
        return float(se_neg)
    if lim_inf is None or lim_sup is None or lim_sup == lim_inf:
        return 0.0
    if melhor.lower() == "maior":
        ratio = (valor - lim_inf) / (lim_sup - lim_inf)
    else:
        ratio = (lim_sup - valor) / (lim_sup - lim_inf)
    return round(max(0.0, min(1.0, ratio)) * peso, 4)


# ── Excel ─────────────────────────────────────────────────────────────────────

def _brd():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left",   vertical="center")

def gerar_aba(wb, nome_aba, resultados, pesos, titulo):
    ws = wb.create_sheet(nome_aba)

    n_ind    = len(pesos)
    # Colunas: Rank(1) Ticker(2) Status(3) Receita(4) PL(5) | Indicadores(6..5+n_ind) | TotalS TotalC Filtro
    COL_REC  = 4
    COL_PL   = 5
    IND_OFF  = 6   # primeira coluna de indicadores
    last_col = 5 + n_ind + 2          # totais
    COL_FILT = last_col + 1           # coluna auxiliar de filtro

    # Células de filtro (row 2)
    CEL_REC_MIN = "C2"
    CEL_REC_MAX = "E2"
    CEL_PL_MIN  = "H2"
    CEL_PL_MAX  = "J2"
    COL_REC_L   = get_column_letter(COL_REC)
    COL_PL_L    = get_column_letter(COL_PL)

    P_AMAREL = PatternFill("solid", fgColor="FFFF00")
    P_DARK   = PatternFill("solid", fgColor="1F3864")

    # ── Row 1: Título ─────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(last_col)}1")
    c = ws["A1"]
    c.value = titulo
    c.font  = Font(name="Arial", color="FFFFFF", bold=True, size=12)
    c.fill  = P_DARK
    c.alignment = CTR
    ws.row_dimensions[1].height = 26

    # ── Row 2: Painel de filtros ───────────────────────────────────────────────
    def lbl(cell, text):
        c = ws[cell]
        c.value = text
        c.font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill  = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="right", vertical="center")

    def inp(cell):
        c = ws[cell]
        c.fill      = P_AMAREL
        c.font      = Font(name="Arial", size=9)
        c.alignment = CTR
        c.border    = _brd()

    lbl("A2", "FILTROS →")
    lbl("B2", "Receita 2025 (R$ bi)  De:")
    inp(CEL_REC_MIN)
    lbl("D2", "Até:")
    inp(CEL_REC_MAX)
    lbl("F2", "")
    lbl("G2", "PL 2025 (R$ bi)  De:")
    inp(CEL_PL_MIN)
    lbl("I2", "Até:")
    inp(CEL_PL_MAX)
    ws.merge_cells("B2:B2")
    hint = ws.cell(row=2, column=COL_FILT,
                   value='← filtre esta col por "✓"')
    hint.font      = Font(name="Arial", size=8, italic=True, color="666666")
    hint.alignment = LFT
    ws.row_dimensions[2].height = 20

    # ── Row 3: Categorias ────────────────────────────────────────────────────
    for col in range(1, last_col + 1):
        ws.cell(row=3, column=col).fill = PatternFill("solid", fgColor="2F4F8F")

    cat_grupos = {}
    for i, p in enumerate(pesos):
        cat_grupos.setdefault(p["categoria"], []).append(IND_OFF + i)

    # Receita e PL sem categoria — deixa escuro
    for col in [COL_REC, COL_PL]:
        ws.cell(row=3, column=col).fill = P_DARK

    for cat, cols in cat_grupos.items():
        c1, c2 = cols[0], cols[-1]
        if c1 != c2:
            ws.merge_cells(f"{get_column_letter(c1)}3:{get_column_letter(c2)}3")
        cor_txt = "FFFFFF" if cat in ("PRECO", "ROE", "DIVIDA") else "000000"
        c = ws.cell(row=3, column=c1, value=cat)
        c.font      = Font(name="Arial", bold=True, size=9, color=cor_txt)
        c.fill      = PatternFill("solid", fgColor=FILL_CAT.get(cat, "999999"))
        c.alignment = CTR
        c.border    = _brd()

    for col, label in [(last_col - 1, "TOTAL S/PREÇO"), (last_col, "TOTAL C/PREÇO")]:
        c = ws.cell(row=3, column=col, value=label)
        c.font = Font(name="Arial", color="FFFFFF", bold=True, size=9)
        c.fill = P_DARK
        c.alignment = CTR
        c.border = _brd()

    # ── Row 4: Cabeçalhos ────────────────────────────────────────────────────
    HDR_ROW = 4
    for col, (label, w) in enumerate(
        [("Rank", 6), ("Ticker", 10), ("Status", 9)], 1
    ):
        c = ws.cell(row=HDR_ROW, column=col, value=label)
        c.font = Font(name="Arial", color="FFFFFF", bold=True, size=9)
        c.fill = P_DARK
        c.alignment = CTR
        c.border = _brd()
        ws.column_dimensions[get_column_letter(col)].width = w

    for col, label, w in [(COL_REC, "Receita 2025\n(R$ bi)", 12), (COL_PL, "PL 2025\n(R$ bi)", 12)]:
        c = ws.cell(row=HDR_ROW, column=col, value=label)
        c.font = Font(name="Arial", color="FFFFFF", bold=True, size=9)
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.alignment = CTR
        c.border = _brd()
        ws.column_dimensions[get_column_letter(col)].width = w

    for i, p in enumerate(pesos):
        col = IND_OFF + i
        label = f"{p['indicador']}\n({p['peso']:.0f}pt)"
        c = ws.cell(row=HDR_ROW, column=col, value=label)
        c.font      = Font(name="Arial", bold=True, size=8,
                           color="FFFFFF" if p["categoria"] in ("PRECO","ROE","DIVIDA") else "000000")
        c.fill      = PatternFill("solid", fgColor=FILL_CAT.get(p["categoria"], "CCCCCC"))
        c.alignment = CTR
        c.border    = _brd()
        ws.column_dimensions[get_column_letter(col)].width = 13

    for col, label in [(last_col-1, "Total\nS/Preço"), (last_col, "Total\nC/Preço")]:
        c = ws.cell(row=HDR_ROW, column=col, value=label)
        c.font = Font(name="Arial", color="FFFFFF", bold=True, size=9)
        c.fill = P_DARK
        c.alignment = CTR
        c.border = _brd()
        ws.column_dimensions[get_column_letter(col)].width = 10

    # Coluna FILTRO
    c = ws.cell(row=HDR_ROW, column=COL_FILT, value="FILTRO")
    c.font = Font(name="Arial", color="FFFFFF", bold=True, size=9)
    c.fill = PatternFill("solid", fgColor="2E75B6")
    c.alignment = CTR
    c.border = _brd()
    ws.column_dimensions[get_column_letter(COL_FILT)].width = 8

    ws.row_dimensions[HDR_ROW].height = 38

    # ── Dados ────────────────────────────────────────────────────────────────
    preco_inds  = {p["indicador"] for p in pesos if p["categoria"] == "PRECO"}
    FILL_COMP   = PatternFill("solid", fgColor="E2EFDA")
    FILL_PARC   = PatternFill("solid", fgColor="FFF2CC")
    max_total_c = sum(p["peso"] for p in pesos)
    max_total_s = sum(p["peso"] for p in pesos if p["indicador"] not in preco_inds)

    DATA_START = HDR_ROW + 1

    for rank, r in enumerate(resultados, 1):
        row = HDR_ROW + rank
        fill_st = FILL_COMP if r["status"] == "completo" else FILL_PARC

        ws.cell(row=row, column=1, value=rank).alignment = CTR
        ws.cell(row=row, column=1).font = Font(name="Arial", size=9)

        c = ws.cell(row=row, column=2, value=r["ticker"])
        c.font = Font(name="Arial", bold=True, size=9)
        c.alignment = LFT

        c = ws.cell(row=row, column=3, value="✅ Completo" if r["status"] == "completo" else "⚠️ Parcial")
        c.font = Font(name="Arial", size=8)
        c.fill = fill_st
        c.alignment = CTR

        # Receita e PL em R$ bi
        rec_bi = round(r["rec_2025"] / 1e9, 2) if r.get("rec_2025") else None
        pl_bi  = round(r["pl_2025"]  / 1e9, 2) if r.get("pl_2025")  else None
        for col, val in [(COL_REC, rec_bi), (COL_PL, pl_bi)]:
            c = ws.cell(row=row, column=col, value=val)
            c.font          = Font(name="Arial", size=9)
            c.fill          = PatternFill("solid", fgColor="DDEEFF")
            c.alignment     = CTR
            c.number_format = "#,##0.00"
            c.border        = _brd()

        total_sem = 0.0
        total_com = 0.0
        for i, p in enumerate(pesos):
            col   = IND_OFF + i
            score = r["scores"].get(p["indicador"], 0.0)
            hex_c = _score_hex(score, p["peso"])
            c = ws.cell(row=row, column=col, value=round(score, 2))
            c.font          = Font(name="Arial", size=9)
            c.fill          = PatternFill("solid", fgColor=hex_c)
            c.alignment     = CTR
            c.number_format = "0.00"
            c.border        = _brd()
            total_com += score
            if p["indicador"] not in preco_inds:
                total_sem += score

        for col, val_t, max_t in [(last_col-1, total_sem, max_total_s), (last_col, total_com, max_total_c)]:
            hex_t = _total_hex(val_t, max_t)
            c = ws.cell(row=row, column=col, value=round(val_t, 2))
            c.font          = Font(name="Arial", bold=True, size=9)
            c.fill          = PatternFill("solid", fgColor=hex_t)
            c.alignment     = CTR
            c.number_format = "0.00"
            c.border        = _brd()

        # Fórmula FILTRO
        rec_cell = f"{COL_REC_L}{row}"
        pl_cell  = f"{COL_PL_L}{row}"
        formula  = (
            f'=IF(AND(' 
            f'OR({CEL_REC_MIN}="",{CEL_REC_MIN}=0,{rec_cell}>={CEL_REC_MIN}*1000000000),'
            f'OR({CEL_REC_MAX}="",{CEL_REC_MAX}=0,{rec_cell}<={CEL_REC_MAX}*1000000000),'
            f'OR({CEL_PL_MIN}="",{CEL_PL_MIN}=0,{pl_cell}>={CEL_PL_MIN}*1000000000),'
            f'OR({CEL_PL_MAX}="",{CEL_PL_MAX}=0,{pl_cell}<={CEL_PL_MAX}*1000000000)'
            f'),\"✓\",\"\")' 
        )
        c = ws.cell(row=row, column=COL_FILT, value=formula)
        c.font      = Font(name="Arial", size=9)
        c.alignment = CTR
        c.border    = _brd()

        ws.row_dimensions[row].height = 15

    # AutoFilter no cabeçalho
    ws.auto_filter.ref = f"A{HDR_ROW}:{get_column_letter(COL_FILT)}{HDR_ROW + len(resultados)}"
    ws.freeze_panes = f"F{DATA_START}"


def gerar_excel(completos, todos, pesos):
    wb = Workbook()
    wb.remove(wb.active)
    hoje = date.today().strftime("%d/%m/%Y")
    gerar_aba(wb, "Ranking Completas",
              completos, pesos,
              f"🏆  RANKING — EMPRESAS COMPLETAS  |  Base: {hoje}")
    gerar_aba(wb, "Ranking Geral",
              todos, pesos,
              f"📊  RANKING GERAL (Completas + Parciais)  |  Base: {hoje}")
    wb.save(OUT)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("📥 Carregando pesos de tabela_de_pesos.xlsx...")
    pesos = carregar_pesos()
    print(f"   {len(pesos)} indicadores ativos\n")

    conn = sqlite3.connect(DB)
    cols = _fin_cols(conn)

    # Filtra tickers marcados como DESCONSIDERAR no validador
    tickers = [r[0] for r in conn.execute(
        "SELECT ticker FROM empresas "
        "WHERE considerar IS NULL OR considerar != 'DESCONSIDERAR' "
        "ORDER BY ticker"
    ).fetchall()]
    descartados = conn.execute(
        "SELECT COUNT(*) FROM empresas WHERE considerar = 'DESCONSIDERAR'"
    ).fetchone()[0]
    if descartados:
        print(f"   {descartados} ticker(s) ignorado(s) por estarem como DESCONSIDERAR")
    print(f"   {len(tickers)} tickers no banco\n")

    resultados = []
    sem_dados  = 0

    for i, ticker in enumerate(tickers, 1):
        if i % 100 == 0:
            print(f"   {i}/{len(tickers)}...")
        try:
            ind    = calcular_indicadores(conn, cols, ticker)
            status = classificar(ind, pesos)
            if status == "sem_dados":
                sem_dados += 1
                continue

            scores = {}
            for p in pesos:
                nome = p["indicador"]
                val_raw, _ = ind.get(nome, (None, 0))
                scores[nome] = calcular_score(
                    val_raw, p["melhor"], p["lim_inf"], p["lim_sup"],
                    p["peso"], p["se_neg"]
                )

            total_sem = sum(v for k, v in scores.items()
                            if next((p for p in pesos if p["indicador"] == k), {}).get("categoria") != "PRECO")
            total_com = sum(scores.values())

            # Receita e PL 2025 para filtros
            row_fin = None
            for fonte in FONTE_PRIO:
                row_fin = conn.execute(
                    "SELECT receita_liquida, patrimonio_liquido FROM financeiros_anuais WHERE ticker=? AND ano=2025 AND fonte=?",
                    (ticker, fonte)
                ).fetchone()
                if row_fin and (row_fin[0] or row_fin[1]):
                    break
            rec_2025 = float(row_fin[0]) if row_fin and row_fin[0] else None
            pl_2025  = float(row_fin[1]) if row_fin and row_fin[1] else None

            resultados.append({
                "ticker":    ticker,
                "status":    status,
                "scores":    scores,
                "total_sem": total_sem,
                "total_com": total_com,
                "rec_2025":  rec_2025,
                "pl_2025":   pl_2025,
            })
        except Exception as e:
            print(f"   ⚠️  {ticker}: {e}")

    conn.close()

    completos = sorted([r for r in resultados if r["status"] == "completo"],
                       key=lambda x: x["total_sem"], reverse=True)
    todos     = sorted(resultados, key=lambda x: x["total_sem"], reverse=True)

    print(f"\n   ✅ Completos : {len(completos)}")
    print(f"   ⚠️  Parciais  : {len(todos) - len(completos)}")
    print(f"   ❌ Sem dados  : {sem_dados}")

    gerar_excel(completos, todos, pesos)
    print(f"\n✅ Ranking salvo em: ranking_resultado.xlsx")


if __name__ == "__main__":
    main()
