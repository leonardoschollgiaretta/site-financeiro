"""
banco_resumido.py — Relatorio resumido por ano
Gera outputs/banco_resumido_YYYY.xlsx com um ticker por linha e indicadores por coluna.

Uso: python banco_resumido.py
"""
import sqlite3
import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

DB   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "financeiro.db")
DIR  = os.path.dirname(os.path.abspath(__file__))

PRIORIDADE = {"investsite": 0, "statusinvest": 1, "yfinance": 2, "fundamentus": 3, "manual": 4}
NI = "N/I"   # nao informado (NULL no banco)


# ─────────────────────────────────────────────────────────────────────────────
#  Carga do banco
# ─────────────────────────────────────────────────────────────────────────────

def carregar(ano):
    conn = sqlite3.connect(DB)

    empresas = pd.read_sql(
        "SELECT ticker, nome, moeda FROM empresas ORDER BY ticker", conn)

    fin_raw = pd.read_sql(f"SELECT * FROM financeiros_anuais WHERE ano={ano}", conn)

    precos = pd.read_sql(
        f"SELECT ticker, preco_medio FROM precos_anuais WHERE ano={ano}", conn)

    acoes = pd.read_sql(
        f"SELECT ticker, acoes_on, acoes_pn, acoes_total, acoes_free FROM acoes_anuais WHERE ano={ano}", conn)

    conn.close()

    # Melhor fonte por ticker
    if not fin_raw.empty:
        fin_raw["_p"] = fin_raw["fonte"].map(lambda f: PRIORIDADE.get(f, 99))
        fin = fin_raw.sort_values(["ticker","_p"]).groupby("ticker").first().reset_index()
    else:
        fin = pd.DataFrame()

    return empresas, fin, precos, acoes


# ─────────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _v(df, ticker, campo):
    """Retorna float ou None. None = dado ausente (NAO e zero)."""
    if df.empty or campo not in df.columns:
        return None
    sub = df[df["ticker"] == ticker]
    if sub.empty:
        return None
    v = sub[campo].values[0]
    return None if pd.isna(v) else float(v)


def _rmil(v):
    """R$ → R$ mil arredondado. Retorna NI se None."""
    if v is None:
        return NI
    return round(v / 1000, 0)


def _pct(num, den):
    """Percentual com 1 casa. Retorna NI se impossivel calcular."""
    if num is None or den is None or den == 0:
        return NI
    return round(abs(num) / abs(den) * 100, 1)


def _ratio(num, den):
    """Razao com 2 casas. Retorna NI se impossivel calcular."""
    if num is None or den is None or den == 0:
        return NI
    r = num / den
    return round(r, 2)


# ─────────────────────────────────────────────────────────────────────────────
#  Calculo dos indicadores por ticker
# ─────────────────────────────────────────────────────────────────────────────

def calcular(ticker, fin, precos, acoes):
    def v(campo): return _v(fin,    ticker, campo)
    def p(campo): return _v(precos, ticker, campo)
    def a(campo): return _v(acoes,  ticker, campo)

    # --- Financeiros base ---
    rl     = v("receita_liquida")
    lb     = v("lucro_bruto")
    sga    = v("despesas_operacionais")
    ll     = v("lucro_liquido")
    cx     = v("caixa")
    fco    = v("fco")
    fci    = v("fci")
    fcf    = v("fcf_financiamento")
    capex  = v("capex")
    pl_val = v("patrimonio_liquido")
    div_pg = v("dividendos_pagos")
    emp_cp = v("emprestimos_cp") or 0
    emp_lp = v("emprestimos_lp") or 0
    deb    = v("debentures")    or 0

    # --- Acoes e preco ---
    p_medio = p("preco_medio")
    ac_use  = a("acoes_free") or a("acoes_total") or (
              (a("acoes_on") or 0) + (a("acoes_pn") or 0)) or None

    # --- Calculados ---
    mkt_cap  = (p_medio * ac_use)          if (p_medio and ac_use) else None
    div_bruta = emp_cp + emp_lp + deb      if (emp_cp or emp_lp or deb) else None
    div_liq  = (div_bruta - (cx or 0))     if div_bruta is not None else None
    fcl      = (fco + capex)               if (fco is not None and capex is not None) else None

    # ROE: LL / PL
    roe = _pct(ll, pl_val)

    # LPA e P/L
    lpa = _ratio(ll, ac_use)
    pl_ratio = _ratio(p_medio, lpa) if (lpa not in (NI, None, 0) and p_medio) else NI

    # VPA e P/VP
    vpa = _ratio(pl_val, ac_use)
    pvp_ratio = _ratio(p_medio, vpa) if (vpa not in (NI, None, 0) and p_medio) else NI

    return [
        _rmil(mkt_cap),                 # Market Cap Medio
        _rmil(pl_val),                  # Valor Patrimonial
        _rmil(rl),                      # Receita Liquida
        _rmil(lb),                      # Lucro Bruto
        _pct(lb, rl),                   # Margem Bruta %
        _rmil(sga),                     # SG&A
        _pct(sga, rl),                  # SG&A % Receita
        _rmil(ll),                      # Lucro Liquido
        _pct(ll, rl),                   # Margem Liquida %
        _rmil(cx),                      # Caixa
        _rmil(fco),                     # FCO
        _rmil(fci),                     # FCI
        _rmil(fcf),                     # FCF
        _rmil(fcl),                     # Free Cash Flow
        _rmil(capex),                   # CAPEX
        _rmil(div_pg),                  # Dividendos Pagos
        _rmil(div_liq),                 # Divida Liquida
        roe,                            # ROE %
        pl_ratio,                       # P/L
        pvp_ratio,                      # P/VP
    ]


# ─────────────────────────────────────────────────────────────────────────────
#  Gera Excel
# ─────────────────────────────────────────────────────────────────────────────

COLUNAS = [
    "Ticker",
    "Market Cap\nMédio (R$mil)",
    "Valor\nPatrimonial (R$mil)",
    "Receita\nLíquida (R$mil)",
    "Lucro\nBruto (R$mil)",
    "Margem\nBruta (%)",
    "SG&A\n(R$mil)",
    "SG&A %\nReceita",
    "Lucro\nLíquido (R$mil)",
    "Margem\nLíquida (%)",
    "Caixa\n(R$mil)",
    "FCO\n(R$mil)",
    "FCI\n(R$mil)",
    "FCF\n(R$mil)",
    "Free Cash\nFlow (R$mil)",
    "CAPEX\n(R$mil)",
    "Dividendos\nPagos (R$mil)",
    "Dívida\nLíquida (R$mil)",
    "ROE\n(%)",
    "P/L",
    "P/VP",
]

# Colunas que sao percentuais ou ratios (nao exibem como inteiro)
COLS_PCT   = {5, 7, 9, 18}        # indices 1-based nas colunas de dados (apos ticker)
COLS_RATIO = {19, 20}


def gerar(ano):
    print(f"\nCarregando dados de {ano}...")
    empresas, fin, precos, acoes = carregar(ano)

    tickers = sorted(empresas[empresas["moeda"] == "BRL"]["ticker"].tolist())
    if not tickers:
        print("Nenhum ticker BR no banco.")
        return

    # ── Estilos ───────────────────────────────────────────────────────────────
    FILL_HDR   = PatternFill("solid", fgColor="1F3864")
    FILL_TICK  = PatternFill("solid", fgColor="2E75B6")
    FILL_PAR   = PatternFill("solid", fgColor="D6E4F0")
    FILL_IMPAR = PatternFill("solid", fgColor="F5F9FD")
    FILL_NI    = PatternFill("solid", fgColor="EFEFEF")

    BRD = Border(
        left=Side(style="thin",   color="CCCCCC"),
        right=Side(style="thin",  color="CCCCCC"),
        top=Side(style="thin",    color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    AL_C  = Alignment(horizontal="center", vertical="center")
    AL_CW = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_L  = Alignment(horizontal="left",   vertical="center", indent=1)

    wb = Workbook()
    ws = wb.active
    ws.title = str(ano)

    n_cols = len(COLUNAS)

    # ── Titulo ────────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1, f"  BANCO RESUMIDO — EXERCÍCIO {ano}")
    c.fill = FILL_HDR
    c.font = Font(bold=True, color="FFFFFF", size=13)
    c.alignment = AL_L
    ws.row_dimensions[1].height = 30

    # ── Cabecalho ─────────────────────────────────────────────────────────────
    for col, txt in enumerate(COLUNAS, start=1):
        c = ws.cell(2, col, txt)
        c.fill      = FILL_HDR
        c.font      = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = AL_CW
        c.border    = BRD
    ws.row_dimensions[2].height = 40

    # Larguras
    ws.column_dimensions["A"].width = 9
    for i in range(1, n_cols):
        ws.column_dimensions[ws.cell(2, i + 1).column_letter].width = 13

    # ── Dados ─────────────────────────────────────────────────────────────────
    for i, ticker in enumerate(tickers):
        row = 3 + i
        bg  = FILL_PAR if i % 2 == 0 else FILL_IMPAR

        indicadores = calcular(ticker, fin, precos, acoes)
        valores = [ticker] + indicadores

        for col, v in enumerate(valores, start=1):
            c = ws.cell(row, col, v)
            c.border    = BRD
            c.alignment = AL_C

            if col == 1:                      # Ticker
                c.fill = FILL_TICK
                c.font = Font(bold=True, color="FFFFFF", size=10)
            elif v == NI:                     # N/I
                c.fill = FILL_NI
                c.font = Font(size=9, color="999999", italic=True)
            else:
                c.fill = bg
                c.font = Font(size=9)
                # Formato de numero
                if col - 1 in COLS_PCT:
                    c.number_format = '#,##0.0"%"'
                elif col - 1 in COLS_RATIO:
                    c.number_format = '#,##0.00'
                else:
                    c.number_format = '#,##0'

        ws.row_dimensions[row].height = 18

    # ── Rodape ────────────────────────────────────────────────────────────────
    r_rod = 3 + len(tickers) + 1
    ws.merge_cells(start_row=r_rod, start_column=1, end_row=r_rod, end_column=n_cols)
    c = ws.cell(r_rod, 1,
        "  Valores em R$ mil  |  N/I = dado ausente no banco  |  Percentuais calculados sobre valores absolutos")
    c.font      = Font(italic=True, size=8, color="666666")
    c.alignment = AL_L
    ws.row_dimensions[r_rod].height = 14

    # ── Salva ─────────────────────────────────────────────────────────────────
    saida = os.path.join(DIR, "outputs", f"banco_resumido_{ano}.xlsx")
    os.makedirs(os.path.dirname(saida), exist_ok=True)
    wb.save(saida)
    print(f"Salvo em: outputs/banco_resumido_{ano}.xlsx")
    return saida


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import banco
    banco.criar_banco()

    entrada = input("\nDigite o ano (ex: 2025): ").strip()
    if not entrada.isdigit() or len(entrada) != 4:
        print("Ano invalido.")
        sys.exit()

    gerar(int(entrada))
