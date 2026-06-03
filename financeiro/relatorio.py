"""
relatorio.py — Gera relatório financeiro por empresa
Baseado no template novomodelo.xlsx
Preenche campos de input com dados do banco; mantém todas as fórmulas intactas.
"""
import sqlite3
import os
import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from copy import copy as _copy

DB        = os.path.join(os.path.dirname(os.path.abspath(__file__)), "financeiro.db")
TEMPLATE  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "novomodelo.xlsx")
TEMPLATE2 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "novomodelo2.xlsx")
SAIDA     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs", "relatorio.xlsx")

def _copiar_aba(ws_orig, wb_dest, nome_dest):
    """Copia uma aba de outro workbook para wb_dest (celula a celula)."""
    ws = wb_dest.create_sheet(nome_dest)

    # Dimensoes de colunas e linhas
    for col, dim in ws_orig.column_dimensions.items():
        ws.column_dimensions[col].width  = dim.width
        ws.column_dimensions[col].hidden = dim.hidden
    for row, dim in ws_orig.row_dimensions.items():
        ws.row_dimensions[row].height = dim.height
        ws.row_dimensions[row].hidden = dim.hidden

    # Celulas mescladas
    for merge in ws_orig.merged_cells.ranges:
        ws.merge_cells(str(merge))

    # Valores e estilos
    for row in ws_orig.iter_rows():
        for cell in row:
            nc = ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                nc.font         = _copy(cell.font)
                nc.border       = _copy(cell.border)
                nc.fill         = _copy(cell.fill)
                nc.alignment    = _copy(cell.alignment)
                nc.number_format = cell.number_format
    return ws


PRIORIDADE = {"investsite": 0, "statusinvest": 1, "yfinance": 2, "fundamentus": 3, "manual": 4}

# Ano → coluna Excel (B=TTM, C=2025 ... H=2020)
ANO_COL = {2025: 3, 2024: 4, 2023: 5, 2022: 6, 2021: 7, 2020: 8}

# Campos do banco → linha do template (valores em R$, serão divididos por 1000)
MAPA_FINANCEIRO = {
    # ── DRE ──────────────────────────────────────────────────────────────────
    "receita_liquida":          21,
    "custo_receita":            24,
    "lucro_bruto":              25,
    "despesas_operacionais":    27,   # SG&A
    "depreciacao_amortizacao":  29,   # D&A (também espelhado em linha 78 via código especial)
    "ebit":                     32,
    "receitas_financeiras":     35,
    "despesas_financeiras":     36,
    "resultado_financeiro":     37,
    "ebt":                      38,
    "ir_csll":                  39,
    "lucro_liquido":            41,
    # ── Balanço Ativo ─────────────────────────────────────────────────────────
    "caixa":                    48,
    "contas_receber":           49,
    "estoques":                 50,
    # 51: Outros Ativos Circulantes  → calculado (residual)
    "ativo_circulante":         52,
    "imobilizado":              53,
    "intangivel":               54,
    "investimentos":            55,
    "outros_ativos_nc":         56,
    "ativo_nao_circulante":     57,   # total ANC da CVM (1.02)
    "ativo_total":              58,
    # ── Balanço Passivo ───────────────────────────────────────────────────────
    "emprestimos_cp":           60,
    "fornecedores":             61,
    # 62: Outros Passivos CP        → calculado (residual)
    "passivo_circulante":       63,
    "emprestimos_lp":           64,
    "debentures":               65,
    # 66: Outros Passivos NC        → calculado (residual)
    "passivo_nao_circulante":   67,   # total PNC da CVM (2.02)
    "capital_social":           68,
    "reservas_lucro":           69,
    "lucros_acumulados":        70,
    "patrimonio_liquido":       71,
    # ── Fluxo de Caixa ────────────────────────────────────────────────────────
    # 77: Lucro Líquido repetido    → espelhado via código especial
    # 78: D&A no FC                 → espelhado de depreciacao_amortizacao
    # 79: Variação Capital de Giro  → zeramos (código CVM varia muito)
    "fco":                      81,
    "capex":                    83,
    "venda_ativos":             84,
    "aquisicoes":               85,
    "fci":                      86,
    "captacoes":                88,
    "pagamento_dividas":        89,
    "recompra_acoes":           90,
    "dividendos_pagos":         91,
    "fcf_financiamento":        92,
    "variacao_caixa":           94,
    "caixa_inicial":            95,
    "caixa_final":              96,
}

# Linha 77 = lucro_liquido repetido na seção FC
LINHA_LUCRO_FC = 77
# Linha 78 = D&A repetido na seção FC
LINHA_DA_FC    = 78

# Linhas de input sem dados → zerar para não herdar valores do MODELO
LINHAS_ZERAR = [
    51,     # Outros Ativos Circulantes (calculado)
    62,     # Outros Passivos CP (calculado)
    66,     # Outros Passivos NC (calculado)
    72,     # Total Passivo + PL (= ativo total, calculado)
    79,     # Variação Capital de Giro (código varia, não coletado)
    108,    # Dividendos por Ação
    109,    # JCP por Ação
    116,    # Nº de Pagamentos
]

LINHA_ACOES     = 9    # nº de ações (em milhares)
LINHA_MKT_MEDIO = 10   # market cap médio  (preco_medio × acoes)
LINHA_MKT_MAX   = 11   # market cap máximo (preco_max   × acoes)
LINHA_MKT_MIN   = 12   # market cap mínimo (preco_min   × acoes)

# Bloco secundário do par ON/PN — detectado dinamicamente do template
# (valores default usados se deteccao falhar)
PAR_COL_BASE = 14
PAR_TTM_COL  = 15
PAR_ANO_COLS = {2025: 16, 2024: 17, 2023: 18, 2022: 19, 2021: 20, 2020: 21}


def detectar_colunas_par(ws, header_row=5):
    """
    Le a linha header_row do template e detecta onde estao TTM e cada ano.
    Atualiza PAR_COL_BASE, PAR_TTM_COL e PAR_ANO_COLS automaticamente.
    """
    global PAR_COL_BASE, PAR_TTM_COL, PAR_ANO_COLS

    ttm_col  = None
    ano_cols = {}

    for cell in ws[header_row]:
        if cell.value is None:
            continue
        v = str(cell.value).strip()
        if v == "TTM":
            ttm_col = cell.column
        elif v.isdigit() and int(v) in ANO_COL:
            ano_cols[int(v)] = cell.column

    if ttm_col and ano_cols:
        PAR_TTM_COL  = ttm_col
        PAR_COL_BASE = ttm_col - 1   # coluna de rotulos = uma antes do TTM
        PAR_ANO_COLS = ano_cols
        print(f"  Bloco par detectado: rotulos=col {PAR_COL_BASE}, TTM=col {PAR_TTM_COL}, anos={ano_cols}")
        return True

    print("  Aviso: nao foi possivel detectar colunas do bloco par — usando defaults")
    return False


# ─────────────────────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────────────────────

def is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")


def set_val(ws, row, col, value):
    """Preenche célula apenas se NÃO for fórmula."""
    cell = ws.cell(row=row, column=col)
    if not is_formula(cell):
        cell.value = value


def r_mil(val):
    """Converte reais → R$ mil (arredonda para inteiro)."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0
    return round(float(val) / 1000, 0)


# ─────────────────────────────────────────────────────────────────────────────
#  Carga do banco
# ─────────────────────────────────────────────────────────────────────────────

def carregar_preco_atual(ticker):
    """Retorna (preco, data_fechamento) da tabela preco_atual."""
    conn = sqlite3.connect(DB)
    row = conn.execute(
        "SELECT preco, data_fechamento FROM preco_atual WHERE ticker=?", (ticker,)
    ).fetchone()
    conn.close()
    return row if row else (None, None)


def carregar_dados(ticker):
    conn = sqlite3.connect(DB)
    fin_raw = pd.read_sql(f"SELECT * FROM financeiros_anuais WHERE ticker='{ticker}'", conn)
    precos  = pd.read_sql(f"SELECT * FROM precos_anuais    WHERE ticker='{ticker}'", conn)
    emp     = pd.read_sql(f"SELECT * FROM empresas          WHERE ticker='{ticker}'", conn)

    # Dividendos: agrega de dividendos_pagamentos por ano (total histórico sem filtro de data)
    divs = pd.read_sql(f"""
        SELECT CAST(substr(data_com,1,4) AS INTEGER) AS ano,
               SUM(valor) AS dividendo_por_acao
        FROM dividendos_pagamentos
        WHERE ticker='{ticker}'
        GROUP BY ano
        ORDER BY ano
    """, conn)

    conn.close()

    if fin_raw.empty:
        return pd.DataFrame(), precos, divs, emp

    fin_raw["_prio"] = fin_raw["fonte"].map(lambda f: PRIORIDADE.get(f, 99))
    fin_raw = fin_raw.sort_values(["ano", "_prio"])
    cols = [c for c in fin_raw.columns if c not in ["ticker", "ano", "fonte", "moeda", "_prio"]]
    fin  = fin_raw.groupby("ano")[cols].first().reset_index()
    return fin, precos, divs, emp


def carregar_dividendos_ttm(ticker):
    """
    Soma dividendos do ano corrente com Data Com <= hoje.
    Regra: só conta dividendos que o investidor já tem direito (Data Com passada).
    """
    hoje      = date.today().isoformat()          # 'YYYY-MM-DD'
    ano_atual = str(date.today().year)

    conn = sqlite3.connect(DB)
    row = conn.execute("""
        SELECT SUM(valor)
        FROM dividendos_pagamentos
        WHERE ticker=?
          AND substr(data_com,1,4) = ?
          AND data_com <= ?
    """, (ticker, ano_atual, hoje)).fetchone()
    conn.close()
    return round(float(row[0]), 6) if row and row[0] else 0


def carregar_acoes(ticker):
    """
    Retorna:
      - acoes_por_ano : dict {ano: acoes_free} para preencher linha 9 por coluna
      - acoes_info    : dict com snapshot atual + ticker_on/ticker_pn para mkt cap TTM
    """
    conn = sqlite3.connect(DB)

    # Historico anual
    rows_ano = conn.execute("""
        SELECT ano, acoes_on, acoes_pn, acoes_free
        FROM acoes_anuais WHERE ticker=? ORDER BY ano
    """, (ticker,)).fetchall()

    # Snapshot + par da tabela empresas
    emp_row = conn.execute("""
        SELECT acoes_free, acoes_on, acoes_pn, ticker_on, ticker_pn
        FROM empresas WHERE ticker=?
    """, (ticker,)).fetchone()
    conn.close()

    acoes_por_ano = {r[0]: {"acoes_free": r[1] or 0 + r[2] or 0,
                             "acoes_on": r[1] or 0,
                             "acoes_pn": r[2] or 0,
                             "acoes_free_real": r[3] or 0}
                     for r in rows_ano}
    # usa acoes_free_real como campo principal
    acoes_por_ano = {r[0]: {"acoes_on": r[1] or 0,
                             "acoes_pn": r[2] or 0,
                             "acoes_free": r[3] or 0}
                     for r in rows_ano}

    acoes_info = None
    if emp_row and emp_row[0] is not None:
        acoes_info = {
            "acoes_free": emp_row[0] or 0,
            "acoes_on":   emp_row[1] or 0,
            "acoes_pn":   emp_row[2] or 0,
            "ticker_on":  emp_row[3],
            "ticker_pn":  emp_row[4],
        }

    return acoes_por_ano, acoes_info


def carregar_precos_par(ticker_par):
    """Carrega precos_anuais e preco_atual do ticker parceiro (ON ou PN)."""
    conn = sqlite3.connect(DB)
    precos = pd.read_sql(f"SELECT * FROM precos_anuais WHERE ticker='{ticker_par}'", conn)
    row = conn.execute("SELECT preco FROM preco_atual WHERE ticker=?", (ticker_par,)).fetchone()
    conn.close()
    preco_atual = float(row[0]) if row and row[0] else None
    return precos, preco_atual


def calcular_mkt_cap(acoes_info, preco_ticker):
    """
    Calcula market cap considerando par ON+PN quando existir.

    - Empresa somente ON (ou units): preco_ticker x acoes_free
    - Empresa com ON + PN:
        (preco_on x acoes_on) + (preco_pn x acoes_pn)
      onde preco_on e preco_pn sao buscados do preco_atual de cada ticker.
    """
    if acoes_info is None:
        return None

    ticker_on = acoes_info["ticker_on"]
    ticker_pn = acoes_info["ticker_pn"]

    if ticker_pn is None:
        # Somente uma classe de acoes
        if preco_ticker is None:
            return None
        return preco_ticker * acoes_info["acoes_free"]

    # Empresa com ON + PN — busca preco de cada ticker
    conn = sqlite3.connect(DB)
    def preco(t):
        r = conn.execute("SELECT preco FROM preco_atual WHERE ticker=?", (t,)).fetchone()
        return float(r[0]) if r and r[0] else None
    p_on = preco(ticker_on)
    p_pn = preco(ticker_pn)
    conn.close()

    mkt = 0
    if p_on: mkt += p_on * acoes_info["acoes_on"]
    if p_pn: mkt += p_pn * acoes_info["acoes_pn"]
    return mkt if mkt else None


def carregar_meta(ticker):
    """Retorna data de atualização e fonte por seção."""
    conn = sqlite3.connect(DB)
    fin_raw = pd.read_sql(
        f"SELECT fonte, atualizado_em FROM financeiros_anuais WHERE ticker='{ticker}'", conn)
    pre_raw = pd.read_sql(
        f"SELECT atualizado_em FROM precos_anuais WHERE ticker='{ticker}'", conn)
    div_raw = pd.read_sql(
        f"SELECT atualizado_em FROM dividendos_anuais WHERE ticker='{ticker}'", conn)
    conn.close()

    fonte_fin = "-"
    if not fin_raw.empty:
        contagem = fin_raw["fonte"].value_counts()
        fonte_fin = contagem.index[0] if not contagem.empty else "-"

    datas = []
    for df in [fin_raw, pre_raw, div_raw]:
        if not df.empty and "atualizado_em" in df.columns:
            vals = df["atualizado_em"].dropna()
            if not vals.empty:
                datas.append(vals.max())

    data_max = max(datas)[:10] if datas else "-"

    return {
        "data":       data_max,
        "dre":        fonte_fin,
        "balanco":    fonte_fin,
        "fluxo":      fonte_fin,
        "dividendos": "statusinvest",
        "precos":     "yfinance",
    }


def row_val(df, ano, campo):
    """Retorna valor para (ano, campo) ou None."""
    sub = df[df["ano"] == ano]
    if sub.empty or campo not in df.columns:
        return None
    v = sub[campo].values[0]
    return None if pd.isna(v) else float(v)


# ─────────────────────────────────────────────────────────────────────────────
#  Bloco de fonte / data (colunas I-K)
#  Linha inicial: 11 para nao colidir com o bloco do par ON+PN (linhas 1-9)
# ─────────────────────────────────────────────────────────────────────────────

META_ROW_START = 11   # primeira linha do bloco FONTES DOS DADOS

def escrever_meta(ws, meta):
    f_tit  = Font(bold=True, color="FFFFFF", size=9)
    f_lbl  = Font(bold=True, size=9)
    f_val  = Font(size=9)
    al_l   = Alignment(horizontal="left", vertical="center", indent=1)
    al_c   = Alignment(horizontal="center", vertical="center")
    fill_h = PatternFill("solid", fgColor="1F3864")
    fill_s = PatternFill("solid", fgColor="DAEEF3")

    r0 = META_ROW_START
    ws.merge_cells(start_row=r0, start_column=9, end_row=r0, end_column=11)
    c = ws.cell(r0, 9, "  FONTES DOS DADOS")
    c.font = f_tit; c.fill = fill_h; c.alignment = al_l

    linhas = [
        ("Atualizado em",  meta["data"]),
        ("DRE",            meta["dre"]),
        ("Balanco",        meta["balanco"]),
        ("Fluxo de Caixa", meta["fluxo"]),
        ("Dividendos",     meta["dividendos"]),
        ("Precos",         meta["precos"]),
    ]
    for i, (label, valor) in enumerate(linhas, start=r0 + 1):
        row_fill = fill_s if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        c_lbl = ws.cell(i, 9, label)
        c_lbl.font = f_lbl; c_lbl.alignment = al_l; c_lbl.fill = row_fill
        c_val = ws.cell(i, 10, valor)
        c_val.font = f_val; c_val.alignment = al_c; c_val.fill = row_fill
        ws.merge_cells(start_row=i, start_column=10, end_row=i, end_column=11)
        ws.cell(i, 11).fill = PatternFill("solid", fgColor="FFFFFF")

    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 4


# ─────────────────────────────────────────────────────────────────────────────
#  Preenchimento do template
# ─────────────────────────────────────────────────────────────────────────────

def _preencher_bloco_par_valores(ws, ticker_par, acoes_por_ano, precos_par, preco_atual_par, campo_par="acoes_on"):
    """
    Preenche apenas os VALORES do bloco par nas colunas PAR_* do template.
    O layout (labels, cores, bordas) ja existe no novomodelo2.xlsx.
    """
    # Ticker par no cabecalho (linha 3, coluna PAR_COL_BASE)
    set_val(ws, 3, PAR_COL_BASE, ticker_par)

    # Preco atual (TTM) — linha 6
    if preco_atual_par:
        set_val(ws, 6, PAR_TTM_COL, round(preco_atual_par, 2))

    # Preco min/max TTM (2026) — linhas 7 e 8
    if not precos_par.empty:
        sub_2026 = precos_par[precos_par["ano"] == 2026]
        if not sub_2026.empty:
            pmin_26 = sub_2026["preco_min"].values[0]
            pmax_26 = sub_2026["preco_max"].values[0]
            if pd.notna(pmin_26): set_val(ws, 7, PAR_TTM_COL, round(float(pmin_26), 2))
            if pd.notna(pmax_26): set_val(ws, 8, PAR_TTM_COL, round(float(pmax_26), 2))

    # Acoes TTM (snapshot mais recente) — linha 9
    if acoes_por_ano:
        d_atual = acoes_por_ano.get(max(acoes_por_ano.keys()))
        if d_atual and d_atual.get(campo_par, 0):
            set_val(ws, 9, PAR_TTM_COL, round(d_atual[campo_par] / 1000, 0))

    # Anos historicos
    for ano, col in PAR_ANO_COLS.items():
        sub = precos_par[precos_par["ano"] == ano]
        if not sub.empty:
            pmedio = sub["preco_medio"].values[0] if "preco_medio" in sub.columns else None
            pmin   = sub["preco_min"].values[0]
            pmax   = sub["preco_max"].values[0]
            if pmedio is not None and pd.notna(pmedio): set_val(ws, 6, col, round(float(pmedio), 2))
            if pd.notna(pmin):                          set_val(ws, 7, col, round(float(pmin),   2))
            if pd.notna(pmax):                          set_val(ws, 8, col, round(float(pmax),   2))

        d = acoes_por_ano.get(ano)
        if d and d.get(campo_par, 0):
            set_val(ws, 9, col, round(d[campo_par] / 1000, 0))


def escrever_bloco_par(ws, ticker_par, acoes_por_ano, precos_par, preco_atual_par):
    """
    Escreve o mini-bloco do ticker parceiro (ON ou PN) a direita do template.
    Mostra: preco medio/min/max por ano + numero de acoes do tipo correspondente.
    ticker_par = PETR3 se main eh PETR4 (e vice-versa).
    acoes_por_ano = dict {ano: {acoes_on, acoes_pn, acoes_free}} do ticker principal.
    """
    fill_amarelo = PatternFill("solid", fgColor="FFFF00")
    fill_header  = PatternFill("solid", fgColor="1F3864")
    fill_par     = PatternFill("solid", fgColor="F2F2F2")
    f_bold9  = Font(bold=True, size=9)
    f_white9 = Font(bold=True, color="FFFFFF", size=9)
    f_dat    = Font(size=9)
    al_c = Alignment(horizontal="center", vertical="center")
    al_l = Alignment(horizontal="left",   vertical="center", indent=1)
    s    = Side(style="thin", color="CCCCCC")
    brd  = Border(left=s, right=s, top=s, bottom=s)

    # Determina se o par e ON ou PN para saber qual campo de acoes usar
    par_eh_on = ticker_par.endswith("3") or ticker_par.endswith("11")
    campo_acoes_par = "acoes_on" if par_eh_on else "acoes_pn"

    # Row 3 — ticker name
    c = ws.cell(row=3, column=PAR_COL_BASE, value=ticker_par)
    c.font = Font(bold=True, size=11); c.fill = fill_amarelo; c.alignment = al_c

    # Row 5 — cabecalho do bloco (sem merge para nao bloquear colunas de dados)
    for col in [PAR_COL_BASE, PAR_TTM_COL] + list(PAR_ANO_COLS.values()):
        ws.cell(row=5, column=col).fill = fill_header

    c = ws.cell(row=5, column=PAR_COL_BASE, value="   DADOS DE MERCADO")
    c.font = f_white9; c.alignment = al_l

    c = ws.cell(row=5, column=PAR_TTM_COL, value="TTM")
    c.font = f_white9; c.alignment = al_c

    for ano, col in PAR_ANO_COLS.items():
        c = ws.cell(row=5, column=col, value=str(ano))
        c.font = f_white9; c.alignment = al_c

    # Linhas 6-9 — rotulos
    rotulos = {
        6: "Preco da Acao (R$)",
        7: "Preco Minimo da Acao (R$)",
        8: "Preco Maximo da Acao (R$)",
        9: "Numero de Acoes (mil)",
    }
    for row, label in rotulos.items():
        c = ws.cell(row=row, column=PAR_COL_BASE, value=label)
        c.font = f_bold9; c.alignment = al_l
        if row % 2 == 0: c.fill = fill_par

    # TTM — preco atual do par
    if preco_atual_par:
        c = ws.cell(row=6, column=PAR_TTM_COL, value=round(preco_atual_par, 2))
        c.font = f_dat; c.alignment = al_c; c.border = brd

    # TTM — acoes do par (snapshot mais recente)
    if acoes_por_ano:
        d_atual = acoes_por_ano.get(max(acoes_por_ano.keys()))
        if d_atual:
            val_acoes = d_atual.get(campo_acoes_par, 0)
            if val_acoes:
                c = ws.cell(row=9, column=PAR_TTM_COL, value=round(val_acoes / 1000, 0))
                c.font = f_dat; c.alignment = al_c; c.border = brd

    # Anos historicos
    for ano, col in PAR_ANO_COLS.items():
        sub = precos_par[precos_par["ano"] == ano] if not precos_par.empty else pd.DataFrame()
        if not sub.empty:
            pmedio = sub["preco_medio"].values[0] if "preco_medio" in sub.columns else None
            pmin   = sub["preco_min"].values[0]
            pmax   = sub["preco_max"].values[0]
            for row, v in [(6, pmedio), (7, pmin), (8, pmax)]:
                if v is not None and pd.notna(v):
                    c = ws.cell(row=row, column=col, value=round(float(v), 2))
                    c.font = f_dat; c.alignment = al_c; c.border = brd

        d = acoes_por_ano.get(ano)
        if d:
            val_acoes = d.get(campo_acoes_par, 0)
            if val_acoes:
                c = ws.cell(row=9, column=col, value=round(val_acoes / 1000, 0))
                c.font = f_dat; c.alignment = al_c; c.border = brd


TTM_COL = 2  # coluna B = TTM

def preencher_ttm(ws, ticker, fin, precos, divs, acoes_info=None):
    """
    Preenche coluna B (TTM):
      - DRE / Balanço / FC  → copia valores de 2025
      - Preço min/max        → dados de 2026 (ano corrente)
      - Preço da Ação        → fechamento do dia anterior (preco_atual)
      - Dividendos           → soma 2026
    """
    col_2025 = ANO_COL[2025]

    # Linhas do MAPA + espelhos + linhas calculadas (resíduos e totais)
    LINHAS_CALCULADAS = [
        51,   # Outros Ativos Circulantes
        62,   # Outros Passivos CP
        66,   # Outros Passivos NC
        72,   # Total Passivo + PL
        103,  # Working Capital
        104,  # Capital Investido
        LINHA_LUCRO_FC,  # 77 — Lucro Líquido repetido no FC
        LINHA_DA_FC,     # 78 — D&A repetido no FC
    ]

    todas_as_linhas = list(MAPA_FINANCEIRO.values()) + LINHAS_CALCULADAS
    for linha in todas_as_linhas:
        cell_2025 = ws.cell(row=linha, column=col_2025)
        cell_ttm  = ws.cell(row=linha, column=TTM_COL)
        if not is_formula(cell_ttm) and not is_formula(cell_2025):
            cell_ttm.value = cell_2025.value

    # Preço da Ação (B6): fechamento do dia anterior
    preco_fech, _ = carregar_preco_atual(ticker)
    if preco_fech:
        ws.cell(row=6, column=TTM_COL).value = round(preco_fech, 2)

    # Preço Mínimo (B7) e Máximo (B8): ano corrente 2026
    sub_2026 = precos[precos["ano"] == 2026]
    if not sub_2026.empty:
        pmin = sub_2026["preco_min"].values[0]
        pmax = sub_2026["preco_max"].values[0]
        if pd.notna(pmin): set_val(ws, 7, TTM_COL, round(float(pmin), 2))
        if pd.notna(pmax): set_val(ws, 8, TTM_COL, round(float(pmax), 2))

    # Dividendos TTM: soma do ano corrente com Data Com <= hoje
    total_div_ttm = carregar_dividendos_ttm(ticker)
    set_val(ws, 108, TTM_COL, round(total_div_ttm, 4) if total_div_ttm else 0)

    # Nº de ações (linha 9) e Market Cap TTM — com suporte a par ON+PN
    if acoes_info:
        tem_par_ttm = acoes_info.get("ticker_pn") is not None
        eh_pn_ttm   = ticker == acoes_info.get("ticker_pn")
        campo_ttm   = "acoes_pn" if eh_pn_ttm else ("acoes_on" if tem_par_ttm else "acoes_free")

        acoes_main_ttm = acoes_info.get(campo_ttm, 0) or 0
        if not acoes_main_ttm:
            if campo_ttm == "acoes_free":
                acoes_main_ttm = acoes_info.get("acoes_on", 0) or acoes_info.get("acoes_total", 0) or 0
            elif campo_ttm == "acoes_pn":
                total = acoes_info.get("acoes_total", 0) or 0
                on    = acoes_info.get("acoes_on",    0) or 0
                acoes_main_ttm = (total - on) if total > on else total
        if acoes_main_ttm:
            set_val(ws, LINHA_ACOES, TTM_COL, round(acoes_main_ttm / 1000, 0))

        preco_fech, _ = carregar_preco_atual(ticker)

        if tem_par_ttm:
            ticker_par_ttm  = acoes_info["ticker_on"] if eh_pn_ttm else acoes_info["ticker_pn"]
            campo_par_ttm   = "acoes_on" if eh_pn_ttm else "acoes_pn"
            acoes_par_ttm   = acoes_info.get(campo_par_ttm, 0)
            _, preco_par_ttm = carregar_precos_par(ticker_par_ttm)

            mkt_ttm = 0
            if preco_fech and acoes_main_ttm: mkt_ttm += preco_fech    * acoes_main_ttm
            if preco_par_ttm and acoes_par_ttm: mkt_ttm += preco_par_ttm * acoes_par_ttm
            if mkt_ttm:
                set_val(ws, LINHA_MKT_MEDIO, TTM_COL, round(mkt_ttm / 1000, 0))
        else:
            mkt_ttm = calcular_mkt_cap(acoes_info, preco_fech)
            if mkt_ttm:
                set_val(ws, LINHA_MKT_MEDIO, TTM_COL, round(mkt_ttm / 1000, 0))


def preencher_aba(ws, ticker, fin, precos, divs, emp, meta):
    nome = emp["nome"].values[0] if not emp.empty else ticker

    # Bloco de fonte/data
    escrever_meta(ws, meta)

    # Cabeçalho
    set_val(ws, 1, 1, f"  {nome}  ({ticker})")
    set_val(ws, 2, 1, "  Valores em R$ milhares  |  Dados coletados automaticamente")
    set_val(ws, 3, 1, ticker)

    # Zera inputs sem dados (para não herdar do MODELO)
    for row in LINHAS_ZERAR:
        for col in ANO_COL.values():
            set_val(ws, row, col, 0)

    # Nº de ações e Market Cap — por ano
    acoes_por_ano, acoes_info = carregar_acoes(ticker)

    # Determina se este ticker e o PN ou ON do par
    tem_par    = acoes_info and acoes_info.get("ticker_pn") is not None
    eh_pn      = acoes_info and ticker == acoes_info.get("ticker_pn")
    campo_main = "acoes_pn" if eh_pn else ("acoes_on" if tem_par else "acoes_free")

    # Carrega dados do par (se existir)
    ticker_par    = None
    precos_par    = pd.DataFrame()
    preco_atual_par = None
    if tem_par:
        ticker_par = acoes_info["ticker_on"] if eh_pn else acoes_info["ticker_pn"]
        precos_par, preco_atual_par = carregar_precos_par(ticker_par)

    for ano, col in ANO_COL.items():
        d = acoes_por_ano.get(ano)
        if not d:
            continue

        acoes_main = d.get(campo_main, 0) or 0
        # Fallback robusto: se campo preferido for 0, tenta alternativas
        if not acoes_main:
            if campo_main == "acoes_free":
                # Empresa sem par: total = ON (unica classe) ou ON+PN
                acoes_main = d.get("acoes_on", 0) or d.get("acoes_total", 0) or 0
            elif campo_main == "acoes_pn":
                # PN: tenta total - ON como proxy
                total = d.get("acoes_total", 0) or 0
                on    = d.get("acoes_on",    0) or 0
                acoes_main = (total - on) if total > on else total
        if acoes_main:
            set_val(ws, LINHA_ACOES, col, round(acoes_main / 1000, 0))

        sub_main = precos[precos["ano"] == ano]
        pmedio_main = float(sub_main["preco_medio"].values[0]) if not sub_main.empty and pd.notna(sub_main["preco_medio"].values[0]) else None
        pmin_main   = float(sub_main["preco_min"].values[0])   if not sub_main.empty and pd.notna(sub_main["preco_min"].values[0])   else None
        pmax_main   = float(sub_main["preco_max"].values[0])   if not sub_main.empty and pd.notna(sub_main["preco_max"].values[0])   else None

        if tem_par:
            # Market cap combinado: main + par
            campo_par = "acoes_on" if eh_pn else "acoes_pn"
            acoes_par_val = d.get(campo_par, 0)
            sub_par = precos_par[precos_par["ano"] == ano] if not precos_par.empty else pd.DataFrame()
            pmedio_par = float(sub_par["preco_medio"].values[0]) if not sub_par.empty and pd.notna(sub_par["preco_medio"].values[0]) else None
            pmin_par   = float(sub_par["preco_min"].values[0])   if not sub_par.empty and pd.notna(sub_par["preco_min"].values[0])   else None
            pmax_par   = float(sub_par["preco_max"].values[0])   if not sub_par.empty and pd.notna(sub_par["preco_max"].values[0])   else None

            def mkt(p_main, p_par):
                if p_main and acoes_main and p_par and acoes_par_val:
                    return round((p_main * acoes_main + p_par * acoes_par_val) / 1000, 0)
                elif p_main and acoes_main:
                    return round(p_main * acoes_main / 1000, 0)
                return None

            if mkt(pmedio_main, pmedio_par): set_val(ws, LINHA_MKT_MEDIO, col, mkt(pmedio_main, pmedio_par))
            if mkt(pmax_main,   pmax_par):   set_val(ws, LINHA_MKT_MAX,   col, mkt(pmax_main,   pmax_par))
            if mkt(pmin_main,   pmin_par):   set_val(ws, LINHA_MKT_MIN,   col, mkt(pmin_main,   pmin_par))
        else:
            # Empresa sem par — market cap simples
            if pmedio_main and acoes_main: set_val(ws, LINHA_MKT_MEDIO, col, round(pmedio_main * acoes_main / 1000, 0))
            if pmax_main   and acoes_main: set_val(ws, LINHA_MKT_MAX,   col, round(pmax_main   * acoes_main / 1000, 0))
            if pmin_main   and acoes_main: set_val(ws, LINHA_MKT_MIN,   col, round(pmin_main   * acoes_main / 1000, 0))

    # Bloco secundario do par — preenche valores nas celulas do MODELO2
    # (layout ja vem do template novomodelo2.xlsx)
    if tem_par and ticker_par and not precos_par.empty:
        _preencher_bloco_par_valores(ws, ticker_par, acoes_por_ano, precos_par, preco_atual_par, campo_par="acoes_on" if eh_pn else "acoes_pn")

    if fin.empty:
        return

    # ── Preenche todos os campos do MAPA (R$ → R$ mil) ──────────────────────
    for campo, linha in MAPA_FINANCEIRO.items():
        for ano, col in ANO_COL.items():
            v = row_val(fin, ano, campo)
            set_val(ws, linha, col, r_mil(v) if v is not None else 0)

    # ── Lucro Líquido repetido na seção FC (linha 77) ────────────────────────
    for ano, col in ANO_COL.items():
        v = row_val(fin, ano, "lucro_liquido")
        set_val(ws, LINHA_LUCRO_FC, col, r_mil(v) if v is not None else 0)

    # ── D&A repetido na seção FC (linha 78) ─────────────────────────────────
    for ano, col in ANO_COL.items():
        v = row_val(fin, ano, "depreciacao_amortizacao")
        set_val(ws, LINHA_DA_FC, col, r_mil(v) if v is not None else 0)

    # ── Campos calculados ─────────────────────────────────────────────────────
    for ano, col in ANO_COL.items():
        ac  = row_val(fin, ano, "ativo_circulante")    or 0
        at  = row_val(fin, ano, "ativo_total")          or 0
        cx  = row_val(fin, ano, "caixa")                or 0
        cr  = row_val(fin, ano, "contas_receber")       or 0
        est = row_val(fin, ano, "estoques")             or 0
        pl  = row_val(fin, ano, "patrimonio_liquido")   or 0
        pc  = row_val(fin, ano, "passivo_circulante")   or 0
        fnc = row_val(fin, ano, "fornecedores")         or 0
        ecp = row_val(fin, ano, "emprestimos_cp")       or 0
        pnc = row_val(fin, ano, "passivo_nao_circulante") or 0
        elp = row_val(fin, ano, "emprestimos_lp")       or 0
        deb = row_val(fin, ano, "debentures")           or 0

        # Outros Ativos Circulantes (residual: AC - caixa - CR - estoques)
        outros_ac = ac - cx - cr - est
        set_val(ws, 51, col, r_mil(outros_ac) if outros_ac > 0 else 0)

        # Outros Passivos CP (residual: PC - fornecedores - emprestimos_cp)
        outros_pc = pc - fnc - ecp
        set_val(ws, 62, col, r_mil(outros_pc) if outros_pc > 0 else 0)

        # Outros Passivos NC (residual: PNC - emprestimos_lp - debentures)
        outros_pnc = pnc - elp - deb
        set_val(ws, 66, col, r_mil(outros_pnc) if outros_pnc > 0 else 0)

        # Total Passivo + PL (= ativo total)
        set_val(ws, 72, col, r_mil(at) if at else 0)

        # Working Capital (linha 103) = AC - PC
        wc = ac - pc
        set_val(ws, 103, col, r_mil(wc))

        # Capital Investido (linha 104) = PL + dívida bruta
        div_bruta = ecp + elp + deb
        ci = pl + div_bruta
        set_val(ws, 104, col, r_mil(ci) if ci else 0)

    # ── Preços (em R$, sem divisão por 1000) ─────────────────────────────────
    for ano, col in ANO_COL.items():
        sub = precos[precos["ano"] == ano]
        if not sub.empty:
            pmin = sub["preco_min"].values[0]
            pmax = sub["preco_max"].values[0]
            if pd.notna(pmin): set_val(ws, 7, col, round(float(pmin), 2))
            if pd.notna(pmax): set_val(ws, 8, col, round(float(pmax), 2))

    # ── Dividendos por ação (anos históricos) ────────────────────────────────
    for ano, col in ANO_COL.items():
        sub = divs[divs["ano"] == ano]
        if not sub.empty and pd.notna(sub["dividendo_por_acao"].values[0]):
            set_val(ws, 108, col, round(float(sub["dividendo_por_acao"].values[0]), 4))

    # ── TTM (coluna B) — preenche por último para sobrescrever corretamente ──
    preencher_ttm(ws, ticker, fin, precos, divs, acoes_info)



# ─────────────────────────────────────────────────────────────────────────────
#  Seção detalhada (abaixo do template, linha 183+)
# ─────────────────────────────────────────────────────────────────────────────

def adicionar_detalhes(ws, fin, precos, divs):
    if fin.empty:
        return

    anos = sorted(ANO_COL.keys(), reverse=True)
    START = 183

    s_tit  = Side(style="thin", color="CCCCCC")
    brd    = Border(left=s_tit, right=s_tit, top=s_tit, bottom=s_tit)
    f_hdr  = Font(bold=True, color="FFFFFF", size=10)
    f_sub  = Font(bold=True, size=9)
    f_dat  = Font(size=9)
    al_c   = Alignment(horizontal="center", vertical="center")
    al_l   = Alignment(horizontal="left",   vertical="center", indent=1)
    fill_azul  = PatternFill("solid", fgColor="1F3864")
    fill_cinza = PatternFill("solid", fgColor="F2F2F2")

    def bloco(r, titulo, campos):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(r, 1, titulo)
        c.fill = fill_azul; c.font = f_hdr; c.alignment = al_l
        ws.row_dimensions[r].height = 20
        r += 1

        ws.cell(r, 1, "Campo").font = f_sub
        ws.cell(r, 1).alignment = al_l
        for i, ano in enumerate(anos):
            c = ws.cell(r, 3 + i, str(ano))
            c.font = f_sub; c.alignment = al_c; c.border = brd
        ws.row_dimensions[r].height = 16
        r += 1

        for label, campo, divisor in campos:
            # campo pode ser uma string (campo do DB) ou uma função (campo calculado)
            bg = fill_cinza if r % 2 == 0 else None
            c = ws.cell(r, 1, label)
            c.font = f_dat; c.alignment = al_l
            if bg: c.fill = bg

            for i, ano in enumerate(anos):
                if callable(campo):
                    v = campo(fin, ano)
                elif campo not in fin.columns:
                    v = None
                else:
                    v = row_val(fin, ano, campo)
                c = ws.cell(r, 3 + i)
                c.font = f_dat; c.alignment = al_c; c.border = brd
                if bg: c.fill = bg
                if v is not None:
                    c.value = round(v / divisor, 0) if divisor == 1000 else round(v / divisor, 4)
                    c.number_format = "#,##0" if divisor == 1000 else "#,##0.0000"
            ws.row_dimensions[r].height = 16
            r += 1
        return r + 1

    # Funções para campos calculados
    def calc_ebitda(df, ano):
        ebit = row_val(df, ano, "ebit")
        da   = row_val(df, ano, "depreciacao_amortizacao")
        if ebit is not None and da is not None:
            return ebit + da
        return ebit  # se não tiver D&A, retorna só EBIT

    def calc_divida_bruta(df, ano):
        ecp = row_val(df, ano, "emprestimos_cp") or 0
        elp = row_val(df, ano, "emprestimos_lp") or 0
        deb = row_val(df, ano, "debentures")     or 0
        total = ecp + elp + deb
        return total if total else None

    def calc_divida_liquida(df, ano):
        db = calc_divida_bruta(df, ano)
        cx = row_val(df, ano, "caixa")
        if db is not None and cx is not None:
            return db - cx
        return None

    def calc_fcl(df, ano):
        fco   = row_val(df, ano, "fco")
        capex = row_val(df, ano, "capex")
        if fco is not None and capex is not None:
            return fco + capex  # capex já é negativo
        return row_val(df, ano, "fcl")

    def calc_wc(df, ano):
        ac = row_val(df, ano, "ativo_circulante") or 0
        pc = row_val(df, ano, "passivo_circulante") or 0
        return ac - pc if (ac or pc) else None

    def calc_ci(df, ano):
        pl  = row_val(df, ano, "patrimonio_liquido") or 0
        db  = calc_divida_bruta(df, ano) or 0
        return pl + db if (pl or db) else None

    r = START

    r = bloco(r, "   📊  DRE — DEMONSTRAÇÃO DE RESULTADO  (R$ mil)", [
        ("Receita Líquida",             "receita_liquida",         1000),
        ("CPV / CMV",                   "custo_receita",           1000),
        ("Lucro Bruto",                 "lucro_bruto",             1000),
        ("Despesas Operacionais (SG&A)","despesas_operacionais",   1000),
        ("Depreciação & Amortização",   "depreciacao_amortizacao", 1000),
        ("EBIT",                        "ebit",                    1000),
        ("EBITDA",                      calc_ebitda,               1000),
        ("Receitas Financeiras",        "receitas_financeiras",    1000),
        ("Despesas Financeiras",        "despesas_financeiras",    1000),
        ("Resultado Financeiro Líq.",   "resultado_financeiro",    1000),
        ("EBT",                         "ebt",                     1000),
        ("IR e CSLL",                   "ir_csll",                 1000),
        ("Lucro Líquido",               "lucro_liquido",           1000),
    ])

    r = bloco(r, "   🏦  BALANÇO — ATIVO  (R$ mil)", [
        ("Caixa + Aplicações Fin.",     "caixa",                1000),
        ("Contas a Receber",            "contas_receber",       1000),
        ("Estoques",                    "estoques",             1000),
        ("TOTAL ATIVO CIRCULANTE",      "ativo_circulante",     1000),
        ("Imobilizado (líquido)",       "imobilizado",          1000),
        ("Intangíveis",                 "intangivel",           1000),
        ("Investimentos",               "investimentos",        1000),
        ("Outros Ativos NC",            "outros_ativos_nc",     1000),
        ("TOTAL ATIVO NÃO CIRC.",       "ativo_nao_circulante", 1000),
        ("TOTAL DO ATIVO",              "ativo_total",          1000),
    ])

    r = bloco(r, "   🏦  BALANÇO — PASSIVO  (R$ mil)", [
        ("Empréstimos CP",              "emprestimos_cp",           1000),
        ("Fornecedores",                "fornecedores",             1000),
        ("TOTAL PASSIVO CIRCULANTE",    "passivo_circulante",       1000),
        ("Empréstimos LP",              "emprestimos_lp",           1000),
        ("Debêntures",                  "debentures",               1000),
        ("TOTAL PASSIVO NÃO CIRC.",     "passivo_nao_circulante",   1000),
        ("Capital Social",              "capital_social",           1000),
        ("Reservas de Lucro",           "reservas_lucro",           1000),
        ("Lucros / Prejuízos Acum.",    "lucros_acumulados",        1000),
        ("TOTAL PATRIMÔNIO LÍQUIDO",    "patrimonio_liquido",       1000),
        ("Dívida Bruta",                calc_divida_bruta,          1000),
        ("Dívida Líquida",              calc_divida_liquida,        1000),
    ])

    r = bloco(r, "   💸  FLUXO DE CAIXA  (R$ mil)", [
        ("Lucro Líquido do Período",    "lucro_liquido",            1000),
        ("(+) D&A",                     "depreciacao_amortizacao",  1000),
        ("FLUXO CAIXA OPERACIONAL",     "fco",                      1000),
        ("CAPEX",                       "capex",                    1000),
        ("Venda de Ativos",             "venda_ativos",             1000),
        ("Aquisições / Participações",  "aquisicoes",               1000),
        ("FLUXO CAIXA INVESTIMENTOS",   "fci",                      1000),
        ("Captações",                   "captacoes",                1000),
        ("Pagamento de Dívidas",        "pagamento_dividas",        1000),
        ("Recompra de Ações",           "recompra_acoes",           1000),
        ("Dividendos / JCP Pagos",      "dividendos_pagos",         1000),
        ("FLUXO CAIXA FINANCIAMENTOS",  "fcf_financiamento",        1000),
        ("Variação Líquida de Caixa",   "variacao_caixa",           1000),
        ("Caixa Inicial",               "caixa_inicial",            1000),
        ("Caixa Final",                 "caixa_final",              1000),
        ("Free Cash Flow (FCO−CAPEX)",  calc_fcl,                   1000),
    ])

    r = bloco(r, "   📐  INDICADORES CALCULADOS  (R$ mil)", [
        ("Working Capital",             calc_wc,                    1000),
        ("Capital Investido",           calc_ci,                    1000),
        ("Dívida Bruta",                calc_divida_bruta,          1000),
        ("Dívida Líquida",              calc_divida_liquida,        1000),
    ])

    if not divs.empty:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(r, 1, "   💰  DIVIDENDOS  (R$ por ação)")
        c.fill = fill_azul; c.font = f_hdr; c.alignment = al_l
        ws.row_dimensions[r].height = 20
        r += 1

        ws.cell(r, 1, "Dividendos/JCP por Ação").font = f_sub
        for i, ano in enumerate(anos):
            sub = divs[divs["ano"] == ano]
            c = ws.cell(r, 3 + i)
            c.font = f_dat; c.alignment = al_c; c.border = brd
            if not sub.empty and pd.notna(sub["dividendo_por_acao"].values[0]):
                c.value = round(float(sub["dividendo_por_acao"].values[0]), 4)
                c.number_format = "#,##0.0000"
        ws.row_dimensions[r].height = 16


# ─────────────────────────────────────────────────────────────────────────────
#  Orquestrador principal
# ─────────────────────────────────────────────────────────────────────────────

def criar_aba_dcf(wb):
    """
    Cria aba 'DCF' (calculadora de valor presente / fluxo de caixa descontado de equity).

    Modelo de dividendo descontado (Gordon), nominal:
      - Inputs: Preco atual, N acoes, PL inicial 2025, premissas (inflacao, payouts,
        r real, g perpetuidade), ROE por ano (2026..2034 + perpetuidade)
      - Modelo: para cada ano: LL = PL_ini * ROE; Div = LL * payout;
                PL_fim = PL_ini + LL - Div
      - Valuation: VP(divs 2026..2034) + VP(Valor Terminal: Div_2035/(r - g))
      - Preco justo = Mkt cap justo / N acoes
    """
    ws = wb.create_sheet("DCF", 0)  # 0 = primeira posicao
    ws.sheet_view.showGridLines = False

    # ===== Estilos =====
    fill_header = PatternFill("solid", fgColor="1F3864")
    fill_input  = PatternFill("solid", fgColor="FFF2CC")  # amarelo claro (input)
    fill_calc   = PatternFill("solid", fgColor="E7E6E6")  # cinza claro (calculado)
    fill_titulo = PatternFill("solid", fgColor="305496")
    fill_destaque = PatternFill("solid", fgColor="C6EFCE")  # verde claro (resultado)

    f_titulo = Font(bold=True, color="FFFFFF", size=12)
    f_subtit = Font(bold=True, color="FFFFFF", size=10)
    f_label  = Font(bold=True, size=10)
    f_input  = Font(size=10)
    f_calc   = Font(size=10, italic=True)
    f_destaq = Font(bold=True, size=11)

    al_c = Alignment(horizontal="center", vertical="center")
    al_l = Alignment(horizontal="left",   vertical="center", indent=1)
    al_r = Alignment(horizontal="right",  vertical="center", indent=1)

    s = Side(style="thin", color="BFBFBF")
    brd = Border(left=s, right=s, top=s, bottom=s)

    # ===== Layout das colunas =====
    # B = rotulos / C = valores (inputs e resultados principais)
    # E = label premissas / F = valor premissas
    # B..N para o modelo nominal (anos 2025..2034 + perpetuidade)
    widths = {"A": 2, "B": 28, "C": 14, "D": 3, "E": 24, "F": 12, "G": 2}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    # Colunas dos anos no modelo nominal: C..M (2025..2034) + N (perpetuidade)
    for col_letter in ["H", "I", "J", "K", "L", "M", "N"]:
        ws.column_dimensions[col_letter].width = 11

    # ===== Bloco 1: VALUATION (B2:C8) =====
    ws["B2"] = "Valuation"
    ws["B2"].font = f_titulo; ws["B2"].fill = fill_titulo; ws["B2"].alignment = al_c
    ws.merge_cells("B2:C2")

    ws["B3"] = "Preco atual";        ws["C3"] = 20.76
    ws["B4"] = "N acoes (mil)";      ws["C4"] = 5728000      # padrao TAEE11
    ws["B5"] = "PL inicial 2025 (R$ mil)"; ws["C5"] = 8500000
    ws["B6"] = "Preco justo";        ws["C6"] = "=IFERROR(C28/C4,0)"
    ws["B7"] = "Upside/downside";    ws["C7"] = "=IFERROR(C6/C3-1,0)"
    ws["B8"] = "Mkt cap atual (R$ mil)"; ws["C8"] = "=C3*C4"
    ws["B9"] = "Mkt cap justo (R$ mil)"; ws["C9"] = "=C28"

    # Formatacoes
    for r in [3, 4, 5]:
        ws.cell(r, 3).fill = fill_input
    for r in [6, 7, 8, 9]:
        ws.cell(r, 3).fill = fill_calc; ws.cell(r, 3).font = f_calc
    ws["C3"].number_format = '"R$ "#,##0.00'
    ws["C4"].number_format = '#,##0'
    ws["C5"].number_format = '#,##0'
    ws["C6"].number_format = '"R$ "#,##0.00'
    ws["C7"].number_format = '0.0%'
    ws["C8"].number_format = '#,##0'
    ws["C9"].number_format = '#,##0'
    for r in range(3, 10):
        ws.cell(r, 2).font = f_label; ws.cell(r, 2).alignment = al_l
        ws.cell(r, 3).alignment = al_r; ws.cell(r, 3).border = brd
        ws.cell(r, 2).border = brd
    # Destaque preco justo + upside
    ws["C6"].fill = fill_destaque; ws["C6"].font = f_destaq
    ws["C7"].fill = fill_destaque; ws["C7"].font = f_destaq

    # ===== Bloco 2: PREMISSAS (E2:F8) =====
    ws["E2"] = "Premissas"
    ws["E2"].font = f_titulo; ws["E2"].fill = fill_titulo; ws["E2"].alignment = al_c
    ws.merge_cells("E2:F2")

    ws["E3"] = "Inflacao";                   ws["F3"] = 0.05
    ws["E4"] = "Payout (periodo projetado)"; ws["F4"] = 0.40
    ws["E5"] = "Payout (perpetuidade)";      ws["F5"] = 0.45
    ws["E6"] = "r = 10% real (nominal)";     ws["F6"] = "=(1+0.10)*(1+F3)-1"  # 10% real
    ws["E7"] = "g (perpetuidade nominal)";   ws["F7"] = "=(1+0.013)*(1+F3)-1"  # 1.3% real
    ws["E8"] = "TIR real (saida)";           ws["F8"] = ""  # informativo, sem calc por enquanto

    for r in [3, 4, 5]:
        ws.cell(r, 6).fill = fill_input
    for r in [6, 7]:
        ws.cell(r, 6).fill = fill_calc; ws.cell(r, 6).font = f_calc

    for r in range(3, 9):
        ws.cell(r, 5).font = f_label; ws.cell(r, 5).alignment = al_l
        ws.cell(r, 6).alignment = al_r; ws.cell(r, 6).border = brd
        ws.cell(r, 5).border = brd
        ws.cell(r, 6).number_format = '0.0%'

    # ===== Bloco 3: MODELO NOMINAL (linha 12+) =====
    ws["B11"] = "Modelo nominal (R$ mil)"
    ws["B11"].font = f_titulo; ws["B11"].fill = fill_titulo; ws["B11"].alignment = al_c
    ws.merge_cells("B11:N11")

    # Linha 12: anos (cabecalho)
    ws["B12"] = ""
    anos = [2025, 2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034]
    # B=label, C=2025, D=2026, ..., L=2034, M=PERP
    # Refazendo: B(rotulo), C..L (2025..2034), M (perpetuidade)
    cols_anos = {2025: 3, 2026: 4, 2027: 5, 2028: 6, 2029: 7,
                 2030: 8,  2031: 9, 2032: 10, 2033: 11, 2034: 12}
    COL_PERP = 13

    for ano, col in cols_anos.items():
        c = ws.cell(12, col, ano)
        c.font = f_subtit; c.fill = fill_header; c.alignment = al_c; c.border = brd
    c = ws.cell(12, COL_PERP, "Perpetuidade")
    c.font = f_subtit; c.fill = fill_header; c.alignment = al_c; c.border = brd

    # Linha 13: ROE (inputs - voce digita)
    ws["B13"] = "ROE"; ws["B13"].font = f_label; ws["B13"].alignment = al_l; ws["B13"].border = brd
    roe_defaults = {2026: 0.096, 2027: 0.120, 2028: 0.140, 2029: 0.150,
                    2030: 0.150, 2031: 0.150, 2032: 0.150, 2033: 0.150,
                    2034: 0.150}
    for ano, col in cols_anos.items():
        if ano in roe_defaults:
            c = ws.cell(13, col, roe_defaults[ano])
            c.fill = fill_input; c.number_format = '0.0%'; c.font = f_input
            c.alignment = al_c; c.border = brd
        else:
            # 2025 nao tem ROE projetado (linha base)
            c = ws.cell(13, col, "")
            c.fill = fill_calc; c.border = brd
    # Perpetuidade
    c = ws.cell(13, COL_PERP, 0.12)
    c.fill = fill_input; c.number_format = '0.0%'; c.font = f_input
    c.alignment = al_c; c.border = brd

    # Linha 14: ROAE (calculado) - ROE medio sobre PL medio
    ws["B14"] = "ROAE"; ws["B14"].font = f_label; ws["B14"].alignment = al_l; ws["B14"].border = brd
    for ano, col in cols_anos.items():
        if ano == 2025:
            c = ws.cell(14, col, "")
        else:
            col_letter = ws.cell(14, col).column_letter
            # ROAE = LL / media(PL_ini, PL_fim) - implementacao simples (mesmo que ROE * fator)
            c = ws.cell(14, col, f"=IFERROR(C15/((C16+{col_letter}16)/2),0)" if col != 4 else f"=IFERROR(D15/((C16+D16)/2),0)")
            # Forma generalizada: PL_ini = coluna anterior linha 16, PL_fim = coluna atual linha 16
            prev_col = ws.cell(14, col-1).column_letter
            c.value = f"=IFERROR({col_letter}15/(({prev_col}16+{col_letter}16)/2),0)"
        c.fill = fill_calc; c.font = f_calc; c.number_format = '0.0%'; c.alignment = al_c; c.border = brd
    # Perpetuidade ROAE = mesma logica
    col_letter = "M"; prev_col = "L"
    c = ws.cell(14, COL_PERP, f"=IFERROR({col_letter}15/(({prev_col}16+{col_letter}16)/2),0)")
    c.fill = fill_calc; c.font = f_calc; c.number_format = '0.0%'; c.alignment = al_c; c.border = brd

    # Linha 15: Lucro liquido = PL_ini * ROE
    ws["B15"] = "Lucro liquido"; ws["B15"].font = f_label; ws["B15"].alignment = al_l; ws["B15"].border = brd
    for ano, col in cols_anos.items():
        if ano == 2025:
            # Linha base - sem LL projetado, deixa vazio
            c = ws.cell(15, col, "")
        else:
            col_letter = ws.cell(15, col).column_letter
            prev_col = ws.cell(15, col-1).column_letter
            # LL_ano = PL_ini (do ano anterior, linha 16) * ROE_ano (linha 13)
            c = ws.cell(15, col, f"={prev_col}16*{col_letter}13")
        c.fill = fill_calc; c.font = f_calc; c.number_format = '#,##0'; c.alignment = al_c; c.border = brd
    # Perpetuidade LL = PL_2034 * (1+g) * ROE_perp (simplificacao: LL cresce a taxa g)
    c = ws.cell(15, COL_PERP, "=L16*M13")
    c.fill = fill_calc; c.font = f_calc; c.number_format = '#,##0'; c.alignment = al_c; c.border = brd

    # Linha 16: Patrimonio liquido (final do ano)
    ws["B16"] = "Patrimonio liquido"; ws["B16"].font = f_label; ws["B16"].alignment = al_l; ws["B16"].border = brd
    # 2025: linha base = $C$5 (input)
    c = ws.cell(16, 3, "=$C$5")
    c.fill = fill_calc; c.font = f_calc; c.number_format = '#,##0'; c.alignment = al_c; c.border = brd
    # 2026..2034: PL_fim = PL_ini + LL - Dividendos
    for ano, col in cols_anos.items():
        if ano == 2025:
            continue
        col_letter = ws.cell(16, col).column_letter
        prev_col = ws.cell(16, col-1).column_letter
        # PL_fim = PL_ini_anterior + LL_atual - Div_atual
        c = ws.cell(16, col, f"={prev_col}16+{col_letter}15-{col_letter}17")
        c.fill = fill_calc; c.font = f_calc; c.number_format = '#,##0'; c.alignment = al_c; c.border = brd
    # Perpetuidade: PL cresce a taxa g (simplificacao - na verdade na perpetuidade os fluxos crescem a g)
    c = ws.cell(16, COL_PERP, "=L16*(1+$F$7)")
    c.fill = fill_calc; c.font = f_calc; c.number_format = '#,##0'; c.alignment = al_c; c.border = brd

    # Linha 17: Dividendos = LL * payout
    ws["B17"] = "Dividendos"; ws["B17"].font = f_label; ws["B17"].alignment = al_l; ws["B17"].border = brd
    for ano, col in cols_anos.items():
        if ano == 2025:
            c = ws.cell(17, col, "")
        else:
            col_letter = ws.cell(17, col).column_letter
            # Div = LL * payout_periodo
            c = ws.cell(17, col, f"={col_letter}15*$F$4")
        c.fill = fill_calc; c.font = f_calc; c.number_format = '#,##0'; c.alignment = al_c; c.border = brd
    # Perpetuidade: Div = LL * payout_perpetuidade
    c = ws.cell(17, COL_PERP, "=M15*$F$5")
    c.fill = fill_calc; c.font = f_calc; c.number_format = '#,##0'; c.alignment = al_c; c.border = brd

    # Linha 18: Fator de desconto = 1/(1+r)^n
    ws["B18"] = "Fator desconto"; ws["B18"].font = f_label; ws["B18"].alignment = al_l; ws["B18"].border = brd
    for ano, col in cols_anos.items():
        n = ano - 2025  # 2026 = 1, 2027 = 2, ...
        if ano == 2025:
            c = ws.cell(18, col, "")
        else:
            c = ws.cell(18, col, f"=1/(1+$F$6)^{n}")
        c.fill = fill_calc; c.font = f_calc; c.number_format = '0.0000'; c.alignment = al_c; c.border = brd
    # Perpetuidade: usa o mesmo fator do ultimo ano (2034 = 9)
    c = ws.cell(18, COL_PERP, "=1/(1+$F$6)^9")
    c.fill = fill_calc; c.font = f_calc; c.number_format = '0.0000'; c.alignment = al_c; c.border = brd

    # Linha 19: VP dos dividendos = Div * Fator
    ws["B19"] = "VP dividendos"; ws["B19"].font = f_label; ws["B19"].alignment = al_l; ws["B19"].border = brd
    for ano, col in cols_anos.items():
        if ano == 2025:
            c = ws.cell(19, col, "")
        else:
            col_letter = ws.cell(19, col).column_letter
            c = ws.cell(19, col, f"={col_letter}17*{col_letter}18")
        c.fill = fill_calc; c.font = f_calc; c.number_format = '#,##0'; c.alignment = al_c; c.border = brd
    # Perpetuidade: Valor terminal = Div_perp / (r - g), descontado ao presente
    # VT = M17 / (F6 - F7); VP(VT) = VT * fator_2034 (M18)
    c = ws.cell(19, COL_PERP, "=IFERROR(M17/($F$6-$F$7)*M18,0)")
    c.fill = fill_destaque; c.font = f_destaq; c.number_format = '#,##0'; c.alignment = al_c; c.border = brd

    # ===== Bloco 4: RESULTADOS (linha 22+) =====
    ws["B22"] = "Resultados"
    ws["B22"].font = f_titulo; ws["B22"].fill = fill_titulo; ws["B22"].alignment = al_c
    ws.merge_cells("B22:C22")

    ws["B23"] = "Soma VP dividendos (2026-2034)"
    ws["C23"] = "=SUM(D19:L19)"
    ws["B24"] = "VP valor terminal (perpetuidade)"
    ws["C24"] = "=M19"
    ws["B25"] = "Total VP (Mkt Cap justo)"
    ws["C25"] = "=C23+C24"

    ws["B26"] = ""  # espacador
    ws["B27"] = "Verificacao: Preco justo (R$)"
    ws["C27"] = "=IFERROR(C25/C4,0)"
    ws["B28"] = "Mkt Cap justo (R$ mil)"
    ws["C28"] = "=C25"

    for r in range(23, 29):
        if r == 26: continue
        ws.cell(r, 2).font = f_label; ws.cell(r, 2).alignment = al_l; ws.cell(r, 2).border = brd
        ws.cell(r, 3).fill = fill_calc; ws.cell(r, 3).font = f_calc
        ws.cell(r, 3).alignment = al_r; ws.cell(r, 3).border = brd
    ws["C23"].number_format = '#,##0'
    ws["C24"].number_format = '#,##0'
    ws["C25"].number_format = '#,##0'
    ws["C27"].number_format = '"R$ "#,##0.00'
    ws["C28"].number_format = '#,##0'
    # Destaque preco justo final
    ws["C27"].fill = fill_destaque; ws["C27"].font = f_destaq
    ws["C25"].fill = fill_destaque; ws["C25"].font = f_destaq

    # ===== Legenda =====
    ws["B30"] = "Legenda:"
    ws["B30"].font = Font(bold=True, size=9)
    ws["B31"] = "Amarelo = input (voce digita)"
    ws["B31"].font = Font(size=9); ws["B31"].fill = fill_input
    ws["B32"] = "Cinza = calculado automatico"
    ws["B32"].font = Font(size=9, italic=True); ws["B32"].fill = fill_calc
    ws["B33"] = "Verde = resultado principal"
    ws["B33"].font = Font(bold=True, size=9); ws["B33"].fill = fill_destaque

    # ===== Notas =====
    ws["B35"] = "Notas:"
    ws["B35"].font = Font(bold=True, size=9)
    ws["B36"] = "* Modelo de Dividendos Descontados (DDM/Gordon) sobre equity."
    ws["B37"] = "* PL inicial = patrimonio liquido em 2025 (R$ mil)."
    ws["B38"] = "* Para cada ano: LL = PL_ini * ROE; Div = LL * Payout; PL_fim = PL_ini + LL - Div."
    ws["B39"] = "* Valor Terminal = Div_perpetuidade / (r - g), pelo modelo de Gordon."
    ws["B40"] = "* Preco justo = (Soma VP dividendos + VP do valor terminal) / N acoes."
    ws["B41"] = "* Ajuste r e g como NOMINAIS (ja incluem inflacao)."
    for r in range(35, 42):
        ws.cell(r, 2).font = Font(size=8, italic=(r != 35))

    # Congela paineis (B do topo)
    ws.freeze_panes = "B3"

    return ws


def gerar_relatorio(tickers):
    if not os.path.exists(TEMPLATE):
        print(f"❌ Template não encontrado: {TEMPLATE}")
        return

    wb = load_workbook(TEMPLATE)

    # Remove abas que nao sao MODELO
    for name in list(wb.sheetnames):
        if name != "MODELO":
            del wb[name]

    # Injeta MODELO2 (para ON+PN) se o arquivo existir
    if os.path.exists(TEMPLATE2):
        wb2 = load_workbook(TEMPLATE2)
        src = wb2["MODELO"] if "MODELO" in wb2.sheetnames else wb2.active
        _copiar_aba(src, wb, "MODELO2")
        detectar_colunas_par(src)   # auto-detecta colunas do bloco par a partir do template
        print("Template ON+PN (novomodelo2.xlsx) carregado.")
    else:
        print(f"Aviso: {TEMPLATE2} nao encontrado — ON+PN usara MODELO padrao.")

    modelo       = wb["MODELO"]
    modelo2      = wb["MODELO2"] if "MODELO2" in wb.sheetnames else None

    for ticker in tickers:
        ticker = ticker.upper()
        print(f"\n📋 {ticker}...")

        fin, precos, divs, emp = carregar_dados(ticker)
        meta = carregar_meta(ticker)

        # Detecta se ticker tem par ON+PN para escolher o template
        _, acoes_info = carregar_acoes(ticker)
        tem_par = acoes_info is not None and acoes_info.get("ticker_pn") is not None

        template_usar = (modelo2 if tem_par and modelo2 else modelo)
        ws       = wb.copy_worksheet(template_usar)
        ws.title = ticker

        preencher_aba(ws, ticker, fin, precos, divs, emp, meta)
        adicionar_detalhes(ws, fin, precos, divs)

        status = "✅" if not fin.empty else "⚠️  sem dados financeiros"
        tipo   = "ON+PN" if tem_par else "padrao"
        print(f"   {status}  [{tipo}]")

    # Remove abas de template do output
    for nome in ["MODELO", "MODELO2"]:
        if nome in wb.sheetnames:
            del wb[nome]

    # Cria aba DCF (calculadora de valor presente) como PRIMEIRA aba
    if "DCF" in wb.sheetnames:
        del wb["DCF"]
    criar_aba_dcf(wb)
    print("📐 Aba DCF (calculadora de valor presente) criada como primeira aba.")

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    print(f"\n✅ Relatório salvo em: outputs/relatorio.xlsx")


if __name__ == "__main__":
    import sys
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    import banco
    banco.criar_banco()

    entrada = input("\nDigite os tickers (separados por espaco ou virgula): ").strip()
    if not entrada:
        print("Nenhum ticker informado.")
        import sys; sys.exit()

    tickers = [t.strip().upper() for t in entrada.replace(",", " ").split() if t.strip()]

    conn = sqlite3.connect(DB)
    disponiveis = set(row[0] for row in conn.execute("SELECT DISTINCT ticker FROM financeiros_anuais"))
    desconsiderados = set(row[0] for row in conn.execute(
        "SELECT ticker FROM empresas WHERE considerar = 'DESCONSIDERAR'"
    ).fetchall())
    conn.close()

    com_dados = [t for t in tickers if t in disponiveis]
    sem_dados = [t for t in tickers if t not in disponiveis]

    if sem_dados:
        print(f"Sem dados no banco para: {sem_dados} - serao ignorados")

    # Avisa sobre tickers DESCONSIDERAR mas gera mesmo assim (intencao manual)
    flag_desc = [t for t in com_dados if t in desconsiderados]
    if flag_desc:
        print(f"⚠️  Atencao: {flag_desc} esta(o) marcado(s) como DESCONSIDERAR no validador")
        print(f"   (gerando relatorio assim mesmo, pois foi pedido explicitamente)")

    if not com_dados:
        print("Nenhum dos tickers informados tem dados no banco.")
        import sys; sys.exit()

    print(f"Gerando relatorio para: {com_dados}\n")
    gerar_relatorio(com_dados)
