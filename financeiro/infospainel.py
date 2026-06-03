"""
infospainel.py — Painel de status de dados por ticker
Gera outputs/infospainel.xlsx com semaforo de qualidade por ticker / tipo / ano.

Nao modifica nenhum outro arquivo — apenas le o banco.
Uso: python financeiro/infospainel.py
"""
import sqlite3
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

DB    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "financeiro.db")
SAIDA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs", "infospainel.xlsx")

ANOS = [2025, 2024, 2023, 2022, 2021, 2020]

# ── Cores ─────────────────────────────────────────────────────────────────────
FILL_OK     = PatternFill("solid", fgColor="70AD47")   # verde
FILL_AVISO  = PatternFill("solid", fgColor="FFD966")   # amarelo
FILL_ERRO   = PatternFill("solid", fgColor="FF6B6B")   # vermelho
FILL_ZERO   = PatternFill("solid", fgColor="C6EFCE")   # verde claro — coletado, confirmado zero
FILL_HEADER = PatternFill("solid", fgColor="1F3864")   # azul escuro
FILL_TICKER = PatternFill("solid", fgColor="2E75B6")   # azul medio
FILL_PAR    = PatternFill("solid", fgColor="D6E4F0")   # azul claro (linha par)
FILL_IMPAR  = PatternFill("solid", fgColor="F5F9FD")   # azul muito claro (linha impar)

BORDA = Border(
    left=Side(style="thin",   color="CCCCCC"),
    right=Side(style="thin",  color="CCCCCC"),
    top=Side(style="thin",    color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
AL_C = Alignment(horizontal="center", vertical="center")
AL_L = Alignment(horizontal="left",   vertical="center", indent=1)

TIPOS  = ["DRE", "Balanco", "Fluxo de Caixa", "Dividendos", "Precos"]
FILL_SEP = PatternFill("solid", fgColor="B8CCE4")   # azul medio claro (separador)

# Campos usados para determinar status de cada tipo
CAMPOS_DRE    = ["receita_liquida", "lucro_liquido", "ebit"]
CAMPOS_BAL    = ["ativo_total", "patrimonio_liquido", "caixa"]
CAMPOS_FLUXO  = ["fco", "fci", "fcf_financiamento"]


# ── Helpers de status ─────────────────────────────────────────────────────────

def _presentes(d, campos):
    return sum(
        1 for c in campos
        if c in d and d[c] is not None
        and not (isinstance(d[c], float) and pd.isna(d[c]))
    )

def status_dre(d):
    p = _presentes(d, CAMPOS_DRE)
    if p == len(CAMPOS_DRE): return "OK"
    if p >= 1:               return "AVISO"
    return "ERRO"

def status_balanco(d):
    p = _presentes(d, CAMPOS_BAL)
    if p == len(CAMPOS_BAL): return "OK"
    if p >= 1:               return "AVISO"
    return "ERRO"

def status_fluxo(d):
    p = _presentes(d, CAMPOS_FLUXO)
    if p == len(CAMPOS_FLUXO): return "OK"
    if p >= 1:                 return "AVISO"
    return "ERRO"

FN_STATUS = {
    "DRE":           status_dre,
    "Balanco":       status_balanco,
    "Fluxo de Caixa": status_fluxo,
}


# ── Carga do banco ────────────────────────────────────────────────────────────

def carregar_banco():
    if not os.path.exists(DB):
        print("Banco nao encontrado:", DB)
        return [], [], pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    conn = sqlite3.connect(DB)

    # Filtra tickers marcados como DESCONSIDERAR no validador
    FILTRO_ATIVO = "(considerar IS NULL OR considerar != 'DESCONSIDERAR')"

    tickers_br = [r[0] for r in conn.execute(
        f"SELECT ticker FROM empresas WHERE moeda='BRL' AND {FILTRO_ATIVO} ORDER BY ticker"
    ).fetchall()]

    tickers_us = [r[0] for r in conn.execute(
        f"SELECT ticker FROM empresas WHERE moeda!='BRL' AND {FILTRO_ATIVO} ORDER BY ticker"
    ).fetchall()]

    # Tickers cujo coletor de dividendos ja rodou pelo menos uma vez
    div_coletados = set(r[0] for r in conn.execute(
        "SELECT ticker FROM empresas WHERE dividendos_coletados_em IS NOT NULL"
    ).fetchall())

    fin = pd.read_sql("""
        SELECT ticker, ano, fonte, atualizado_em,
               receita_liquida, lucro_liquido, ebit,
               ativo_total, patrimonio_liquido, caixa,
               fco, fci, fcf_financiamento
        FROM financeiros_anuais
    """, conn)

    precos = pd.read_sql("""
        SELECT ticker, ano, preco_min, preco_max, atualizado_em
        FROM precos_anuais
    """, conn)

    # Verifica se dividendos_pagamentos existe
    tabelas = [r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'"
    ).fetchall()]
    if "dividendos_pagamentos" in tabelas:
        divs = pd.read_sql("""
            SELECT ticker,
                   CAST(substr(data_com,1,4) AS INTEGER) AS ano,
                   COUNT(*) AS n_pagamentos,
                   MAX(atualizado_em) AS atualizado_em
            FROM dividendos_pagamentos
            GROUP BY ticker, ano
        """, conn)
    else:
        # fallback para dividendos_anuais
        divs = pd.read_sql("""
            SELECT ticker, ano,
                   CASE WHEN dividendo_por_acao > 0 THEN 1 ELSE 0 END AS n_pagamentos,
                   atualizado_em
            FROM dividendos_anuais
        """, conn)

    conn.close()
    return tickers_br, tickers_us, fin, precos, divs, div_coletados


# ── Monta linhas de status ────────────────────────────────────────────────────

def montar_linhas(tickers, fin, precos, divs, div_coletados=None):
    PRIORIDADE = {"manual": 0, "investsite": 1, "statusinvest": 2, "yfinance": 3}

    # Melhor fonte por ticker/ano para financeiros
    if not fin.empty:
        fin["_p"] = fin["fonte"].map(lambda f: PRIORIDADE.get(f, 5))
        fin = fin.sort_values(["ticker", "ano", "_p"])
        fin_best = fin.groupby(["ticker", "ano"]).first().reset_index()
    else:
        fin_best = pd.DataFrame()

    linhas = []

    for ticker in tickers:
        fin_t    = fin_best[fin_best["ticker"] == ticker]   if not fin_best.empty else pd.DataFrame()
        precos_t = precos[precos["ticker"]   == ticker]
        divs_t   = divs[divs["ticker"]       == ticker]

        for tipo in TIPOS:
            linha = {"ticker": ticker, "tipo": tipo}

            # --- Financeiros (DRE / Balanco / Fluxo de Caixa) ---
            if tipo in FN_STATUS:
                fn = FN_STATUS[tipo]
                # Data de atualizacao: max entre todos os anos
                linha["dt"] = fin_t["atualizado_em"].max() if not fin_t.empty else None

                for ano in ANOS:
                    sub = fin_t[fin_t["ano"] == ano]
                    if sub.empty:
                        linha[ano] = "ERRO"
                    else:
                        linha[ano] = fn(sub.iloc[0].to_dict())

            # --- Dividendos ---
            elif tipo == "Dividendos":
                linha["dt"] = divs_t["atualizado_em"].max() if not divs_t.empty else None
                ja_coletado = div_coletados and (ticker in div_coletados)

                for ano in ANOS:
                    sub = divs_t[divs_t["ano"] == ano]
                    tem_pagamento = not sub.empty and int(sub["n_pagamentos"].values[0]) > 0
                    if tem_pagamento:
                        linha[ano] = "OK"
                    elif ja_coletado:
                        linha[ano] = "ZERO"   # coletado, confirmado sem pagamento
                    else:
                        linha[ano] = "ERRO"   # nunca coletamos

            # --- Precos ---
            elif tipo == "Precos":
                linha["dt"] = precos_t["atualizado_em"].max() if not precos_t.empty else None

                for ano in ANOS:
                    sub = precos_t[precos_t["ano"] == ano]
                    if sub.empty:
                        linha[ano] = "ERRO"
                    else:
                        pmin = sub["preco_min"].values[0]
                        pmax = sub["preco_max"].values[0]
                        if pd.notna(pmin) and pd.notna(pmax):
                            linha[ano] = "OK"
                        elif pd.notna(pmin) or pd.notna(pmax):
                            linha[ano] = "AVISO"
                        else:
                            linha[ano] = "ERRO"

            linhas.append(linha)

    return linhas


# ── Resumo por ticker ─────────────────────────────────────────────────────────

def _contar(statuses_por_ticker):
    """Recebe {ticker: [lista de statuses]} e devolve (total, ok, parcial, vazio)."""
    total   = len(statuses_por_ticker)
    ok      = sum(1 for s in statuses_por_ticker.values() if all(x in ("OK", "ZERO") for x in s))
    vazio   = sum(1 for s in statuses_por_ticker.values() if all(x == "ERRO" for x in s))
    parcial = total - ok - vazio
    return total, ok, parcial, vazio


def calcular_resumo(linhas):
    """
    Classifica cada ticker como:
      - OK      : todos os tipos em todos os anos = OK
      - VAZIO   : todos os tipos em todos os anos = ERRO (sem nada)
      - PARCIAL : alguma coisa tem mas nao esta completo
    """
    from collections import defaultdict
    por_ticker = defaultdict(list)
    for linha in linhas:
        for ano in ANOS:
            por_ticker[linha["ticker"]].append(linha.get(ano, "ERRO"))

    total   = len(por_ticker)
    ok      = 0
    vazio   = 0
    parcial = 0

    total, ok, parcial, vazio = _contar(por_ticker)
    return {"total": total, "ok": ok, "vazio": vazio, "parcial": parcial}


def calcular_resumo_por_tipo(linhas):
    """Para cada tipo de dado, quantos tickers estao completos/parciais/vazios."""
    from collections import defaultdict
    resultado = {}
    for tipo in TIPOS:
        por_ticker = defaultdict(list)
        for linha in linhas:
            if linha["tipo"] == tipo:
                for ano in ANOS:
                    por_ticker[linha["ticker"]].append(linha.get(ano, "ERRO"))
        total, ok, parcial, vazio = _contar(por_ticker)
        resultado[tipo] = {"total": total, "ok": ok, "parcial": parcial, "vazio": vazio}
    return resultado


def calcular_resumo_por_ano(linhas):
    """Para cada ano, quantos tickers tem todos os tipos OK/parcial/vazio."""
    from collections import defaultdict
    resultado = {}
    for ano in ANOS:
        por_ticker = defaultdict(list)
        for linha in linhas:
            por_ticker[linha["ticker"]].append(linha.get(ano, "ERRO"))
        total, ok, parcial, vazio = _contar(por_ticker)
        resultado[ano] = {"total": total, "ok": ok, "parcial": parcial, "vazio": vazio}
    return resultado


def _escrever_tabela_resumo(ws, titulo, itens, row_ini, n_cols):
    """
    Escreve um bloco de resumo em formato tabela a partir de row_ini.
    itens = lista de (label, total, ok, parcial, vazio)
    Retorna a proxima linha disponivel apos o bloco.
    """
    FILL_HDR  = PatternFill("solid", fgColor="1F3864")
    FILL_TOT  = PatternFill("solid", fgColor="2E75B6")
    FILL_OK_  = PatternFill("solid", fgColor="70AD47")
    FILL_PAR_ = PatternFill("solid", fgColor="FFD966")
    FILL_VAZ_ = PatternFill("solid", fgColor="FF6B6B")
    FILL_ROT  = PatternFill("solid", fgColor="D6E4F0")

    # Titulo
    ws.merge_cells(start_row=row_ini, start_column=1, end_row=row_ini, end_column=n_cols)
    c = ws.cell(row_ini, 1, f"  {titulo}")
    c.fill = FILL_HDR
    c.font = Font(bold=True, color="FFFFFF", size=10)
    c.alignment = AL_L
    ws.row_dimensions[row_ini].height = 20
    row_ini += 1

    # Cabecalho das colunas
    for col, (txt, fill, cor) in enumerate([
        ("",          FILL_HDR,  "FFFFFF"),
        ("Total",     FILL_TOT,  "FFFFFF"),
        ("Completos", FILL_OK_,  "FFFFFF"),
        ("Parciais",  FILL_PAR_, "000000"),
        ("Sem Dados", FILL_VAZ_, "FFFFFF"),
    ], start=1):
        c = ws.cell(row_ini, col, txt)
        c.fill = fill; c.font = Font(bold=True, size=9, color=cor)
        c.alignment = AL_C; c.border = BORDA
    ws.row_dimensions[row_ini].height = 16
    row_ini += 1

    # Linhas de dados
    for i, (label, total, ok, parcial, vazio) in enumerate(itens):
        bg = FILL_ROT if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        c = ws.cell(row_ini, 1, label)
        c.fill = bg; c.font = Font(bold=True, size=9); c.alignment = AL_L; c.border = BORDA

        for col, (val, fill, cor) in enumerate([
            (total,   FILL_TOT,  "FFFFFF"),
            (ok,      FILL_OK_,  "FFFFFF"),
            (parcial, FILL_PAR_, "000000"),
            (vazio,   FILL_VAZ_, "FFFFFF"),
        ], start=2):
            c = ws.cell(row_ini, col, val)
            c.fill = fill; c.font = Font(bold=True, size=10, color=cor)
            c.alignment = AL_C; c.border = BORDA

        ws.row_dimensions[row_ini].height = 16
        row_ini += 1

    # Linha em branco de separacao
    for col in range(1, n_cols + 1):
        ws.cell(row_ini, col).fill = PatternFill("solid", fgColor="FFFFFF")
    ws.row_dimensions[row_ini].height = 6
    return row_ini + 1


def _escrever_resumo(ws, resumo, n_cols):
    """Escreve o bloco de resumo nas primeiras 3 linhas da aba."""
    FILL_RESUMO_HDR = PatternFill("solid", fgColor="1F3864")
    FILL_OK_R   = PatternFill("solid", fgColor="70AD47")
    FILL_PAR_R  = PatternFill("solid", fgColor="FFD966")
    FILL_VAZ_R  = PatternFill("solid", fgColor="FF6B6B")
    FILL_TOT_R  = PatternFill("solid", fgColor="2E75B6")

    # Linha 1: titulo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(1, 1, "  RESUMO DO PAINEL DE QUALIDADE")
    c.fill = FILL_RESUMO_HDR
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = AL_L
    ws.row_dimensions[1].height = 22

    # Linha 2: labels
    blocos = [
        ("TOTAL", resumo["total"],   FILL_TOT_R, "FFFFFF"),
        ("COMPLETOS",  resumo["ok"],      FILL_OK_R,  "FFFFFF"),
        ("PARCIAIS",   resumo["parcial"], FILL_PAR_R, "000000"),
        ("SEM DADOS",  resumo["vazio"],   FILL_VAZ_R, "FFFFFF"),
    ]
    for i, (label, valor, fill, cor_txt) in enumerate(blocos):
        col_lbl = 1 + i * 2
        col_val = col_lbl + 1

        c_lbl = ws.cell(2, col_lbl, label)
        c_lbl.fill = fill
        c_lbl.font = Font(bold=True, size=10, color=cor_txt)
        c_lbl.alignment = AL_C
        c_lbl.border = BORDA

        c_val = ws.cell(2, col_val, valor)
        c_val.fill = fill
        c_val.font = Font(bold=True, size=14, color=cor_txt)
        c_val.alignment = AL_C
        c_val.border = BORDA

        ws.row_dimensions[2].height = 28

    # Linha 3: separador vazio
    for col in range(1, n_cols + 1):
        ws.cell(3, col).fill = PatternFill("solid", fgColor="FFFFFF")
    ws.row_dimensions[3].height = 6


# ── Gera Excel ────────────────────────────────────────────────────────────────

ROTULO  = {"OK": "OK",  "AVISO": "!",  "ERRO": "-",  "ZERO": "R$0"}
FILL_ST = {"OK": FILL_OK, "AVISO": FILL_AVISO, "ERRO": FILL_ERRO, "ZERO": FILL_ZERO}
COR_TXT = {"OK": "FFFFFF", "AVISO": "000000", "ERRO": "FFFFFF",  "ZERO": "375623"}

def _preencher_aba(ws, linhas):
    """Preenche uma aba com as linhas de status. ws ja deve existir."""

    n_cols = 3 + len(ANOS)

    # ── Bloco 1: resumo geral (linhas 1-3) ───────────────────────────────────
    resumo = calcular_resumo(linhas)
    _escrever_resumo(ws, resumo, n_cols)

    # ── Bloco 2: resumo por tipo de dado ─────────────────────────────────────
    res_tipo = calcular_resumo_por_tipo(linhas)
    itens_tipo = [
        (tipo, d["total"], d["ok"], d["parcial"], d["vazio"])
        for tipo, d in res_tipo.items()
    ]
    prox = _escrever_tabela_resumo(ws, "RESUMO POR TIPO DE DADO", itens_tipo, 4, n_cols)

    # ── Bloco 3: resumo por ano ───────────────────────────────────────────────
    res_ano = calcular_resumo_por_ano(linhas)
    itens_ano = [
        (str(ano), d["total"], d["ok"], d["parcial"], d["vazio"])
        for ano, d in res_ano.items()
    ]
    prox = _escrever_tabela_resumo(ws, "RESUMO POR ANO", itens_ano, prox, n_cols)

    # ── Cabecalho da tabela de tickers ────────────────────────────────────────
    ROW_HDR = prox
    cabecalho = ["Ticker", "Tipo de Dado", "Ultima Atualizacao"] + [str(a) for a in ANOS]
    for col, txt in enumerate(cabecalho, start=1):
        c = ws.cell(row=ROW_HDR, column=col, value=txt)
        c.fill   = FILL_HEADER
        c.font   = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = AL_C
        c.border = BORDA
    ws.row_dimensions[ROW_HDR].height = 24

    # ── Larguras ───────────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    for i, _ in enumerate(ANOS):
        ws.column_dimensions[chr(ord("D") + i)].width = 8

    # ── Linhas de dados ───────────────────────────────────────────────────────
    ticker_atual = None
    linha_par    = True   # alterna por grupo de ticker
    r_idx        = ROW_HDR  # sera incrementado antes de cada escrita

    for i_linha, linha in enumerate(linhas):
        ticker = linha["ticker"]
        tipo   = linha["tipo"]

        # Novo ticker: inserir linha separadora (exceto antes do primeiro)
        if ticker != ticker_atual:
            if ticker_atual is not None:
                r_idx += 1
                for col in range(1, n_cols + 1):
                    sc = ws.cell(row=r_idx, column=col, value="")
                    sc.fill = FILL_SEP
                ws.row_dimensions[r_idx].height = 5

            linha_par    = not linha_par
            ticker_atual = ticker

        bg = FILL_PAR if linha_par else FILL_IMPAR

        r_idx += 1

        # Data de atualizacao formatada
        dt_raw = linha.get("dt", "")
        dt_str = str(dt_raw)[:10] if dt_raw and str(dt_raw) not in ("", "None", "nan") else "-"

        # Col A — Ticker (sempre visivel em todas as linhas do grupo)
        ca = ws.cell(row=r_idx, column=1, value=ticker)
        ca.fill      = FILL_TICKER
        ca.font      = Font(bold=True, color="FFFFFF", size=10)
        ca.alignment = AL_C
        ca.border    = BORDA

        # Col B — Tipo
        cb = ws.cell(row=r_idx, column=2, value=tipo)
        cb.fill      = bg
        cb.font      = Font(size=9)
        cb.alignment = AL_L
        cb.border    = BORDA

        # Col C — Ultima atualizacao
        cc = ws.cell(row=r_idx, column=3, value=dt_str)
        cc.fill      = bg
        cc.font      = Font(size=9)
        cc.alignment = AL_C
        cc.border    = BORDA

        # Cols D+ — Status por ano
        for j, ano in enumerate(ANOS):
            st    = linha.get(ano, "ERRO")
            label = ROTULO.get(st, "-")
            fill  = FILL_ST.get(st, FILL_ERRO)
            cor   = COR_TXT.get(st, "FFFFFF")

            cx = ws.cell(row=r_idx, column=4 + j, value=label)
            cx.fill      = fill
            cx.font      = Font(bold=True, size=10, color=cor)
            cx.alignment = AL_C
            cx.border    = BORDA

        ws.row_dimensions[r_idx].height = 18

    # ── Legenda ───────────────────────────────────────────────────────────────
    r_leg = r_idx + 2
    ws.cell(r_leg, 1, "Legenda:").font = Font(bold=True, size=9)

    for col, (txt, fill, cor) in enumerate([
        ("OK  — dados completos", FILL_OK,    "FFFFFF"),
        ("!   — dados parciais",  FILL_AVISO, "000000"),
        ("-   — sem dados",        FILL_ERRO,  "FFFFFF"),
    ], start=2):
        c = ws.cell(r_leg, col, txt)
        c.fill      = fill
        c.font      = Font(bold=True, size=9, color=cor)
        c.alignment = AL_C
        c.border    = BORDA


def gerar_excel(linhas_br, linhas_us):
    wb = Workbook()

    # Aba BR
    ws_br = wb.active
    ws_br.title = "BR"
    _preencher_aba(ws_br, linhas_br)

    # Aba US (mesmo que vazia, cria a aba)
    ws_us = wb.create_sheet(title="US")
    _preencher_aba(ws_us, linhas_us)

    os.makedirs(os.path.dirname(SAIDA), exist_ok=True)
    wb.save(SAIDA)
    print(f"Painel salvo em: outputs/infospainel.xlsx")

# ── Main ─────────────────────────────────────────────

def main():
    print("Lendo banco de dados...")
    tickers_br, tickers_us, fin, precos, divs, div_coletados = carregar_banco()

    if not tickers_br and not tickers_us:
        print("Nenhum ticker no banco.")
        return

    print(f"BR: {len(tickers_br)} ticker(s)  |  US: {len(tickers_us)} ticker(s)")

    linhas_br = montar_linhas(tickers_br, fin, precos, divs, div_coletados)
    linhas_us = montar_linhas(tickers_us, fin, precos, divs, div_coletados)

    gerar_excel(linhas_br, linhas_us)

if __name__ == "__main__":
    main()
