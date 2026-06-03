"""
gerar_validador.py — Gera (ou atualiza) validador.xlsx

Colunas:
  Ticker | PL 2025 (R$ bi) | CONSIDERAR? | Status Geral | DRE | Balanco | Fluxo | Dividendos | Precos
  [separador] DRE | Balanco | Fluxo | Dividendos | Acoes | Precos  (tipo lock)
  [separador] 2025 | 2024 | 2023 | 2022 | 2021 | 2020              (ano lock)

Células amarelas = editáveis pelo usuário.
Células coloridas (status) = somente leitura, geradas pelo script.

Uso: python gerar_validador.py
"""
import sqlite3
import os
import sys
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

DB      = os.path.join(os.path.dirname(os.path.abspath(__file__)), "financeiro.db")
SAIDA   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "validador.xlsx")

ANOS       = [2025, 2024, 2023, 2022, 2021, 2020]   # BR (6 anos)
ANOS_US    = [2025, 2024, 2023, 2022]                # US (4 anos — limite do yfinance gratuito)
TIPOS   = ["DRE", "Balanco", "Fluxo", "Dividendos", "Acoes", "Precos"]
FONTE_PRIO = ["investsite", "statusinvest", "yfinance", "manual"]

CAMPOS_DRE   = ["receita_liquida", "lucro_liquido", "ebit"]
CAMPOS_BAL   = ["ativo_total", "patrimonio_liquido", "caixa"]
CAMPOS_FLUXO = ["fco", "fci", "fcf_financiamento"]

# ── Estilos ───────────────────────────────────────────────────────────────────
F_HDR    = Font(name="Arial", color="FFFFFF", bold=True, size=9)
F_BODY   = Font(name="Arial", size=9)
F_TICK   = Font(name="Arial", bold=True, size=9)
F_BLUE   = Font(name="Arial", color="0000FF", size=9)
F_GRAY   = Font(name="Arial", color="999999", italic=True, size=9)

P_DARK   = PatternFill("solid", fgColor="1F3864")
P_MED    = PatternFill("solid", fgColor="2E75B6")
P_AMAREL = PatternFill("solid", fgColor="FFFF00")
P_OK     = PatternFill("solid", fgColor="70AD47")
P_AVISO  = PatternFill("solid", fgColor="FFD966")
P_ERRO   = PatternFill("solid", fgColor="FF6B6B")
P_ZERO   = PatternFill("solid", fgColor="C6EFCE")
P_TIPO   = PatternFill("solid", fgColor="D9E1F2")
P_ANO    = PatternFill("solid", fgColor="E2EFDA")
P_PAR    = PatternFill("solid", fgColor="F5F9FD")
P_IMPAR  = PatternFill("solid", fgColor="FFFFFF")

CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left",   vertical="center", indent=1)

def brd():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Helpers de status (mesma lógica do infospainel) ───────────────────────────

def _presentes(d, campos):
    return sum(1 for c in campos if d.get(c) is not None)

def status_tipo(tipo, d):
    if tipo == "DRE":
        p = _presentes(d, CAMPOS_DRE)
        return "OK" if p == 3 else ("AVISO" if p >= 1 else "ERRO")
    if tipo == "Balanco":
        p = _presentes(d, CAMPOS_BAL)
        return "OK" if p == 3 else ("AVISO" if p >= 1 else "ERRO")
    if tipo == "Fluxo":
        p = _presentes(d, CAMPOS_FLUXO)
        return "OK" if p == 3 else ("AVISO" if p >= 1 else "ERRO")
    return "ERRO"

def status_geral(statuses):
    """statuses = lista de 'OK'|'AVISO'|'ERRO'|'ZERO'"""
    tudo_ok    = all(s in ("OK", "ZERO") for s in statuses)
    tudo_erro  = all(s == "ERRO" for s in statuses)
    if tudo_ok:   return "Completo"
    if tudo_erro: return "Sem Dados"
    return "Parcial"

# ── Carrega dados do banco ────────────────────────────────────────────────────

def carregar_dados(moeda="BRL"):
    """
    Carrega dados de tickers de uma moeda especifica (BRL ou USD).
    Retorna: tickers, fin_best, divs, div_coletados, precos, cons_map, val_df
    """
    conn = sqlite3.connect(DB)

    if moeda == "BRL":
        filtro_sql = "moeda='BRL'"
    else:
        filtro_sql = "moeda!='BRL'"

    tickers = [r[0] for r in conn.execute(
        f"SELECT ticker FROM empresas WHERE {filtro_sql} ORDER BY ticker"
    ).fetchall()]

    # PL 2025 e campo considerar
    pl_map = {}
    cons_map = {}
    for r in conn.execute(
        f"SELECT ticker, considerar FROM empresas WHERE {filtro_sql}"
    ).fetchall():
        cons_map[r[0]] = r[1] or ""

    # Financeiros — melhor fonte por ticker/ano
    fin_raw = pd.read_sql("""
        SELECT ticker, ano, fonte, receita_liquida, lucro_liquido, ebit,
               ativo_total, patrimonio_liquido, caixa, fco, fci, fcf_financiamento
        FROM financeiros_anuais
    """, conn)
    if not fin_raw.empty:
        fin_raw["_p"] = fin_raw["fonte"].map(lambda f: {"investsite":0,"statusinvest":1,"yfinance":2,"manual":9}.get(f,5))
        fin_raw = fin_raw.sort_values(["ticker","ano","_p"])
        fin_best = fin_raw.groupby(["ticker","ano"]).first().reset_index()
    else:
        fin_best = pd.DataFrame()

    # Dividendos
    tabelas = [r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'"
    ).fetchall()]
    div_coletados = set(r[0] for r in conn.execute(
        "SELECT ticker FROM empresas WHERE dividendos_coletados_em IS NOT NULL"
    ).fetchall())
    if "dividendos_pagamentos" in tabelas:
        divs = pd.read_sql("""
            SELECT ticker, CAST(substr(data_com,1,4) AS INTEGER) AS ano,
                   COUNT(*) AS n FROM dividendos_pagamentos GROUP BY ticker, ano
        """, conn)
    else:
        divs = pd.read_sql("""
            SELECT ticker, ano,
                   CASE WHEN dividendo_por_acao > 0 THEN 1 ELSE 0 END AS n
            FROM dividendos_anuais
        """, conn)

    # Preços
    precos = pd.read_sql("SELECT ticker, ano FROM precos_anuais WHERE preco_min IS NOT NULL", conn)

    # Validações atuais
    val_df = pd.read_sql("SELECT ticker, kind, valor FROM validacoes", conn)

    conn.close()
    return tickers, fin_best, divs, div_coletados, precos, cons_map, val_df


def montar_status(ticker, fin_best, divs, div_coletados, precos, anos=None):
    """Retorna dict {(tipo, ano): status_str} para um ticker."""
    if anos is None:
        anos = ANOS
    res = {}
    fin_t   = fin_best[fin_best["ticker"] == ticker]   if not fin_best.empty  else pd.DataFrame()
    div_t   = divs[divs["ticker"] == ticker]
    prec_t  = precos[precos["ticker"] == ticker]
    ja_col  = ticker in div_coletados

    for ano in anos:
        # DRE / Balanco / Fluxo
        sub = fin_t[fin_t["ano"] == ano]
        d   = sub.iloc[0].to_dict() if not sub.empty else {}
        for tipo in ["DRE", "Balanco", "Fluxo"]:
            res[(tipo, ano)] = status_tipo(tipo, d) if d else "ERRO"

        # Dividendos
        dsub = div_t[div_t["ano"] == ano]
        tem_pgto = not dsub.empty and int(dsub["n"].values[0]) > 0
        if tem_pgto:
            res[("Dividendos", ano)] = "OK"
        elif ja_col:
            res[("Dividendos", ano)] = "ZERO"
        else:
            res[("Dividendos", ano)] = "ERRO"

        # Precos
        res[("Precos", ano)] = "OK" if not prec_t[prec_t["ano"] == ano].empty else "ERRO"

        # Acoes — não tem status histórico por ano no infospainel, usa ERRO como default
        res[("Acoes", ano)] = "N/A"

    return res


# ── Gera o Excel ──────────────────────────────────────────────────────────────

def _preencher_aba(ws, tickers, fin_best, divs, div_coletados, precos, cons_map, val_set, anos, simbolo_moeda="R$"):
    """
    Preenche uma aba do validador com tickers de uma moeda especifica.
    `anos` define quais anos terao colunas de validacao (BR=6, US=4).
    `simbolo_moeda` muda apenas o rotulo do PL ('R$ bi' vs '$ bi').
    """
    # ── Estrutura de colunas ───────────────────────────────────────────────────
    # Col 1: Ticker
    # Col 2: PL 2025 (R$ bi)
    # Col 3: CONSIDERAR?
    # Col 4: Status Geral
    # Col 5-9: Status DRE | Balanco | Fluxo | Dividendos | Precos  (por tipo, todos anos)
    # Col 10: [sep]
    # Col 11-16: DRE | Balanco | Fluxo | Dividendos | Acoes | Precos  (tipo lock)
    # Col 17: [sep]
    # Col 18-23: 2025 | 2024 | 2023 | 2022 | 2021 | 2020  (ano lock)

    COL_TICKER   = 1
    COL_PL       = 2
    COL_CONS     = 3
    COL_STATUS   = 4
    # Status por tipo: cols 5-9 (DRE, Balanco, Fluxo, Dividendos, Precos)
    STATUS_TIPOS = ["DRE", "Balanco", "Fluxo", "Dividendos", "Precos"]
    COL_ST_FIRST = 5
    COL_ST_LAST  = COL_ST_FIRST + len(STATUS_TIPOS) - 1  # 9

    COL_SEP1     = COL_ST_LAST + 1   # 10
    COL_TL_FIRST = COL_SEP1 + 1      # 11  (tipo locks)
    COL_TL_LAST  = COL_TL_FIRST + len(TIPOS) - 1  # 16

    COL_SEP2     = COL_TL_LAST + 1   # 17
    COL_AL_FIRST = COL_SEP2 + 1      # 18  (ano locks)
    COL_AL_LAST  = COL_AL_FIRST + len(anos) - 1

    LAST_COL = COL_AL_LAST

    # ── Linha 1: Título ────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(LAST_COL)}1")
    c = ws["A1"]
    c.value = "🔒  VALIDADOR DE DADOS — Marque VALIDADO nas células amarelas para travar a recoleta"
    c.font  = Font(name="Arial", color="FFFFFF", bold=True, size=11)
    c.fill  = P_DARK
    c.alignment = LFT
    ws.row_dimensions[1].height = 24

    # ── Linha 2: Grupos de colunas ─────────────────────────────────────────────
    def grupo(ws, col1, col2, label, fill, font_color="FFFFFF"):
        if col1 != col2:
            ws.merge_cells(f"{get_column_letter(col1)}2:{get_column_letter(col2)}2")
        c = ws.cell(2, col1, label)
        c.font = Font(name="Arial", bold=True, size=9, color=font_color)
        c.fill = fill; c.alignment = CTR; c.border = brd()

    grupo(ws, COL_TICKER, COL_TICKER, "Ticker",         P_DARK)
    grupo(ws, COL_PL,     COL_PL,     f"PL 2025 ({simbolo_moeda}bi)", P_DARK)
    grupo(ws, COL_CONS,   COL_CONS,   "CONSIDERAR?",    P_DARK)
    grupo(ws, COL_STATUS, COL_STATUS, "Status Geral",   P_DARK)
    grupo(ws, COL_ST_FIRST, COL_ST_LAST, "STATUS POR TIPO (somente leitura)", P_MED)
    grupo(ws, COL_SEP1,   COL_SEP1,   "",               P_DARK)
    grupo(ws, COL_TL_FIRST, COL_TL_LAST, "🔒 VALIDAR TIPO (trava TODOS os anos)", P_TIPO, "000000")
    grupo(ws, COL_SEP2,   COL_SEP2,   "",               P_DARK)
    grupo(ws, COL_AL_FIRST, COL_AL_LAST, "🔒 VALIDAR ANO (trava TODOS os tipos)", P_ANO, "000000")
    ws.row_dimensions[2].height = 28

    # ── Linha 3: Cabeçalhos de colunas ────────────────────────────────────────
    headers = (
        [(COL_TICKER, "Ticker", 10),
         (COL_PL,     f"PL 2025\n({simbolo_moeda} bi)", 12),
         (COL_CONS,   "CONSIDERAR?", 18),
         (COL_STATUS, "Status\nGeral", 11)] +
        [(COL_ST_FIRST + i, t, 10) for i, t in enumerate(STATUS_TIPOS)] +
        [(COL_SEP1, "", 2)] +
        [(COL_TL_FIRST + i, t, 11) for i, t in enumerate(TIPOS)] +
        [(COL_SEP2, "", 2)] +
        [(COL_AL_FIRST + i, str(a), 10) for i, a in enumerate(anos)]
    )
    for col, label, width in headers:
        c = ws.cell(3, col, label)
        if col in (COL_SEP1, COL_SEP2):
            c.fill = P_DARK
        elif col >= COL_TL_FIRST and col <= COL_TL_LAST:
            c.font = Font(name="Arial", bold=True, size=9, color="000000")
            c.fill = P_TIPO; c.alignment = CTR; c.border = brd()
        elif col >= COL_AL_FIRST:
            c.font = Font(name="Arial", bold=True, size=9, color="000000")
            c.fill = P_ANO; c.alignment = CTR; c.border = brd()
        else:
            c.font = F_HDR; c.fill = P_DARK; c.alignment = CTR; c.border = brd()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 32

    # Data validation dropdown para CONSIDERAR?
    dv_cons = DataValidation(
        type="list",
        formula1='"100% VALIDADA,VALIDADA PARCIAL,DESCONSIDERAR"',
        allow_blank=True
    )
    dv_cons.error      = "Use: 100% VALIDADA, VALIDADA PARCIAL ou DESCONSIDERAR"
    dv_cons.errorTitle = "Valor inválido"
    ws.add_data_validation(dv_cons)

    dv_val = DataValidation(
        type="list",
        formula1='"VALIDADO"',
        allow_blank=True
    )
    ws.add_data_validation(dv_val)

    FILL_STATUS = {"OK": P_OK, "AVISO": P_AVISO, "ERRO": P_ERRO, "ZERO": P_ZERO, "N/A": P_IMPAR}
    COR_GERAL   = {"Completo": P_OK, "Parcial": P_AVISO, "Sem Dados": P_ERRO}

    # ── Linhas de dados ────────────────────────────────────────────────────────
    for i, ticker in enumerate(tickers):
        row = 4 + i
        fill_bg = P_PAR if i % 2 == 0 else P_IMPAR

        st_map = montar_status(ticker, fin_best, divs, div_coletados, precos, anos=anos)

        # Status geral: considera DRE + Balanco + Fluxo + Dividendos + Precos
        todos_st = [st_map.get((t, a), "ERRO") for t in STATUS_TIPOS for a in anos]
        sg = status_geral(todos_st)

        # PL 2025 — busca no fin_best
        pl_val = None
        if not fin_best.empty:
            sub = fin_best[(fin_best["ticker"] == ticker) & (fin_best["ano"] == 2025)]
            if not sub.empty and sub["patrimonio_liquido"].values[0] is not None:
                pl_val = round(float(sub["patrimonio_liquido"].values[0]) / 1e9, 2)

        # Col Ticker
        c = ws.cell(row, COL_TICKER, ticker)
        c.font = F_TICK; c.fill = fill_bg; c.alignment = LFT; c.border = brd()

        # Col PL
        c = ws.cell(row, COL_PL, pl_val)
        c.font = F_BODY; c.fill = fill_bg; c.alignment = CTR; c.border = brd()
        if pl_val is not None: c.number_format = "#,##0.00"

        # Col CONSIDERAR? — amarelo (editável), pré-preenchido se já existe
        c = ws.cell(row, COL_CONS, cons_map.get(ticker, ""))
        c.font = F_BLUE; c.fill = P_AMAREL; c.alignment = CTR; c.border = brd()
        dv_cons.add(c)

        # Col Status Geral
        c = ws.cell(row, COL_STATUS, sg)
        c.font = Font(name="Arial", bold=True, size=9,
                      color="FFFFFF" if sg != "Parcial" else "000000")
        c.fill = COR_GERAL.get(sg, P_AVISO); c.alignment = CTR; c.border = brd()

        # Cols status por tipo (somente leitura — status agregado de todos os anos)
        for j, tipo in enumerate(STATUS_TIPOS):
            col = COL_ST_FIRST + j
            sts = [st_map.get((tipo, a), "ERRO") for a in anos]
            # Resumo do tipo: OK se todos OK/ZERO, ERRO se todos ERRO, senão AVISO
            if all(s in ("OK", "ZERO") for s in sts): resumo = "OK"
            elif all(s == "ERRO" for s in sts):       resumo = "ERRO"
            else:                                       resumo = "AVISO"
            c = ws.cell(row, col, resumo)
            c.font = Font(name="Arial", bold=True, size=8,
                          color="FFFFFF" if resumo in ("OK","ERRO") else "000000")
            c.fill = FILL_STATUS.get(resumo, P_IMPAR)
            c.alignment = CTR; c.border = brd()

        # Separador 1
        ws.cell(row, COL_SEP1).fill = P_DARK

        # Cols tipo lock (amarelo editável — pré-preenchido se validado)
        for j, tipo in enumerate(TIPOS):
            col = COL_TL_FIRST + j
            tipo_lower = tipo.lower()
            ja_val = "VALIDADO" if (ticker, "tipo", tipo_lower) in val_set else ""
            c = ws.cell(row, col, ja_val)
            c.font = F_BLUE; c.fill = P_AMAREL; c.alignment = CTR; c.border = brd()
            dv_val.add(c)

        # Separador 2
        ws.cell(row, COL_SEP2).fill = P_DARK

        # Cols ano lock (amarelo editável — pré-preenchido se validado)
        for j, ano in enumerate(anos):
            col = COL_AL_FIRST + j
            ja_val = "VALIDADO" if (ticker, "ano", str(ano)) in val_set else ""
            c = ws.cell(row, col, ja_val)
            c.font = F_BLUE; c.fill = P_AMAREL; c.alignment = CTR; c.border = brd()
            dv_val.add(c)

        ws.row_dimensions[row].height = 16

    # Congela até col D (deixa Ticker..Status Geral fixos ao rolar)
    ws.freeze_panes = "D4"


def gerar():
    print("📥 Carregando dados do banco...")

    # ── Carrega dados BR e US ──────────────────────────────────────────────────
    tickers_br, fin_br, divs_br, divc_br, prec_br, cons_br, val_br = carregar_dados("BRL")
    tickers_us, fin_us, divs_us, divc_us, prec_us, cons_us, val_us = carregar_dados("USD")
    print(f"   {len(tickers_br)} tickers BR | {len(tickers_us)} tickers US")

    # Indice de validacoes existentes (banco e o mesmo para BR e US)
    val_set = set()
    for _, r in val_br.iterrows():
        val_set.add((r["ticker"], r["kind"], str(r["valor"])))
    # val_us == val_br (mesma tabela), nao precisa duplicar

    wb = Workbook()

    # Aba BR (sempre presente — mantem padrao do projeto)
    ws_br = wb.active
    ws_br.title = "Validador"
    _preencher_aba(ws_br, tickers_br, fin_br, divs_br, divc_br, prec_br, cons_br,
                   val_set, anos=ANOS, simbolo_moeda="R$")

    # Aba US (so cria se houver tickers US)
    if tickers_us:
        ws_us = wb.create_sheet("Validador US")
        _preencher_aba(ws_us, tickers_us, fin_us, divs_us, divc_us, prec_us, cons_us,
                       val_set, anos=ANOS_US, simbolo_moeda="$")

    wb.save(SAIDA)
    total = len(tickers_br) + len(tickers_us)
    print(f"✅ validador.xlsx gerado: {SAIDA}")
    print(f"   {total} tickers ({len(tickers_br)} BR + {len(tickers_us)} US)")
    print(f"   Edite CONSIDERAR? e colunas amarelas de validacao")
    print(f"   Depois rode sync_validador.py para gravar no banco")


if __name__ == "__main__":
    import banco
    banco.criar_banco()
    gerar()
