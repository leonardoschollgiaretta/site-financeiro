"""
excelindicadoresranking.py -- Gera Excel com ranking dos 8 indicadores
para todas as acoes BR e US (abas separadas).

Uma linha por acao. Para cada indicador: valor e posicao no ranking.

Uso: python financeiro/excelindicadoresranking.py
"""
import os
import sys
import sqlite3
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from insta_relatorio import calcular_indicadores_ticker, _direcoes

DB      = os.path.join(os.path.dirname(os.path.abspath(__file__)), "financeiro.db")
DIR_OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs")


def _ordenar(validos, nome_ind, maior_melhor):
    """Mesma logica de ordenacao do calcular_ranking_v2 (P/L e P/VP negativos vao pro fim)."""
    if not maior_melhor and nome_ind in ("P/L", "P/VP"):
        positivos = sorted([(t, v) for t, v in validos if v > 0], key=lambda x: x[1])
        negativos = sorted([(t, v) for t, v in validos if v <= 0],
                           key=lambda x: x[1], reverse=True)
        return positivos + negativos
    return sorted(validos, key=lambda x: x[1], reverse=maior_melhor)


def montar_dataframe(universo, ano_ref):
    """universo: 'BR' ou 'US'."""
    conn = sqlite3.connect(DB)
    filtro = "moeda='BRL'" if universo == "BR" else "moeda!='BRL'"
    tickers = [r[0] for r in conn.execute(
        f"SELECT ticker FROM empresas "
        f"WHERE {filtro} "
        f"AND (considerar IS NULL OR considerar != 'DESCONSIDERAR') "
        f"ORDER BY ticker"
    ).fetchall()]

    nomes = dict(conn.execute(
        f"SELECT ticker, nome FROM empresas WHERE {filtro}"
    ).fetchall())

    indicadores_por_ticker = {}
    for t in tickers:
        ind = calcular_indicadores_ticker(conn, t, ano_ref)
        if ind:
            indicadores_por_ticker[t] = ind
    conn.close()

    if not indicadores_por_ticker:
        return None

    direcoes = _direcoes(ano_ref)
    nomes_ind = list(direcoes.keys())

    # Calcula posicao de cada ticker em cada indicador
    posicoes = {ind: {} for ind in nomes_ind}
    totais = {}
    for ind, maior_melhor in direcoes.items():
        validos = [(t, v[ind]) for t, v in indicadores_por_ticker.items()
                   if v.get(ind) is not None]
        ordenado = _ordenar(validos, ind, maior_melhor)
        totais[ind] = len(ordenado)
        for pos, (t, _) in enumerate(ordenado, 1):
            posicoes[ind][t] = pos

    linhas = []
    for t, vals in indicadores_por_ticker.items():
        linha = {"Ticker": t, "Nome": nomes.get(t, "")}
        for ind in nomes_ind:
            linha[ind] = vals.get(ind)
            linha[f"Pos {ind}"] = posicoes[ind].get(t)
        linhas.append(linha)

    df = pd.DataFrame(linhas).sort_values("Ticker").reset_index(drop=True)

    # Adiciona linha com o total de empresas avaliadas em cada ranking
    total_row = {"Ticker": "_total_avaliadas", "Nome": ""}
    for ind in nomes_ind:
        total_row[ind] = None
        total_row[f"Pos {ind}"] = totais[ind]
    df.attrs["totais"] = totais
    return df


def aplicar_formatacao(writer, sheet_name, df):
    """Formata numeros e percentuais (engine openpyxl)."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = writer.sheets[sheet_name]

    header_fill = PatternFill("solid", fgColor="0B1929")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="1E3A4A")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = header_border

    indicadores_pct = {"Margem Bruta", "Margem Liquida", "ROE", "DY medio 5 anos"}

    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        if col_name == "Nome":
            ws.column_dimensions[letter].width = 28
            continue
        if col_name == "Ticker":
            ws.column_dimensions[letter].width = 12
            continue

        if col_name.startswith("Pos "):
            num_fmt = "0"
            width = 10
        elif col_name in indicadores_pct or col_name.startswith("DY "):
            num_fmt = "0.0%"
            width = 12
        else:
            num_fmt = "0.00"
            width = 10

        ws.column_dimensions[letter].width = width
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).number_format = num_fmt

    ws.freeze_panes = "C2"


def gerar_excel(ano_ref):
    os.makedirs(DIR_OUT, exist_ok=True)
    caminho = os.path.join(DIR_OUT, f"ranking_indicadores_{ano_ref}.xlsx")

    print(f"\nGerando ranking de indicadores para {ano_ref}...")

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        for universo in ("BR", "US"):
            print(f"  [{universo}] processando...")
            df = montar_dataframe(universo, ano_ref)
            if df is None or df.empty:
                print(f"  [{universo}] sem dados, pulando.")
                continue
            df.to_excel(writer, sheet_name=universo, index=False)
            aplicar_formatacao(writer, universo, df)
            print(f"  [{universo}] {len(df)} acoes.")

    print(f"\nArquivo salvo: {caminho}")


if __name__ == "__main__":
    print("=" * 50)
    print("  RANKING DE INDICADORES -> EXCEL")
    print("=" * 50)
    try:
        ano = int(input("\n  Ano-base (ex: 2025): ").strip())
    except ValueError:
        print("  Ano invalido.")
        sys.exit()
    gerar_excel(ano)
