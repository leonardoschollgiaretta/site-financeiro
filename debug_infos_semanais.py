"""
DEBUG isolado — inspeciona as planilhas de import/export em INFOS SEMANAIS.
Só leitura, não altera nada.
"""
import os, glob, sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
PASTA = os.path.join(BASE, "INFOS SEMANAIS")

xlsxs = glob.glob(os.path.join(PASTA, "**", "*.xlsx"), recursive=True)
print(f"Planilhas encontradas: {len(xlsxs)}\n")

for fp in xlsxs:
    print("=" * 70)
    print("ARQUIVO:", os.path.relpath(fp, BASE))
    try:
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    except Exception as e:
        print("  ERRO ao abrir:", e)
        continue
    for ws in wb.worksheets:
        dims = ws.calculate_dimension()
        print(f"\n  --- Aba: {ws.title!r} | dims {dims} | max_row={ws.max_row} max_col={ws.max_column}")
        # imprime as primeiras linhas pra entender estrutura
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            vals = ["" if v is None else str(v) for v in row]
            # corta linhas vazias
            if not any(vals):
                continue
            print("   r%-3d|" % i, " | ".join(v[:18] for v in vals[:12]))
            if i >= 12:
                print("    ... (mais linhas)")
                break
    wb.close()
