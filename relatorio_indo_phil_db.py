"""
Relatório de IMPORTAÇÕES de Indonésia e Filipinas (como destino),
histórico 2025 e 2026, a partir do banco oficial Weekly report/vessels.db
(carregado de '2025 2026 database.xlsx', abas data25/data26).

Data de referência: bl_date (Bill of Lading); quando ausente, usa eta.
2026 é ano corrente / parcial (vai até ~junho).

Gera: Weekly report/outputs/importacoes_indonesia_filipinas_2025_2026.xlsx
Só lê o banco; escreve apenas o relatório.
"""
import os, sqlite3, sys
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE, "Weekly report", "vessels.db")
OUT_DIR = os.path.join(BASE, "Weekly report", "outputs")
os.makedirs(OUT_DIR, exist_ok=True)
OUT = os.path.join(OUT_DIR, "importacoes_indonesia_filipinas_2025_2026.xlsx")

# data de referência: bl_date, senão eta
DATA = "COALESCE(NULLIF(bl_date,''), eta)"
FILTRO = "discharge IN ('Indonesia','Philippines')"
ANO = f"substr({DATA},1,4)"
MES = f"substr({DATA},6,2)"

conn = sqlite3.connect(DB)
c = conn.cursor()

# ---- estilos ----
HDR = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="1F4E78")
TOT = Font(bold=True)
TOTFILL = PatternFill("solid", fgColor="DDEBF7")
CEN = Alignment(horizontal="center")
THIN = Border(*([Side(style="thin", color="D9D9D9")] * 4))

def header(ws, ncols, row=1):
    for col in range(1, ncols + 1):
        cell = ws.cell(row, col)
        cell.font = HDR; cell.fill = FILL; cell.alignment = CEN
        cell.border = THIN

def autofit(ws):
    for cells in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in cells)
        ws.column_dimensions[get_column_letter(cells[0].column)].width = min(max(w + 2, 10), 42)

def fmt_milhares(ws, col_letter, start=2):
    for row in range(start, ws.max_row + 1):
        ws[f"{col_letter}{row}"].number_format = "#,##0"

wb = openpyxl.Workbook()

# ===== Aba 0: Capa / notas =====
ws0 = wb.active; ws0.title = "Leia-me"
notas = [
    ["RELATÓRIO — Importações de Indonésia e Filipinas (destino)", ""],
    ["", ""],
    ["Fonte", "Weekly report/vessels.db (de '2025 2026 database.xlsx', abas data25/data26)"],
    ["Países (discharge)", "Indonesia, Philippines"],
    ["Período", "2025 (ano completo) e 2026 (em andamento)"],
    ["Data de referência", "Bill of Lading (bl_date); quando ausente, usa ETA"],
    ["Unidade", "toneladas métricas (mt)"],
    ["", ""],
    ["⚠ AVISO 1", "2026 é ANO CORRENTE / PARCIAL — só vai até ~junho/2026. Não comparar"],
    ["", "total de 2026 com total de 2025 como se fossem anos fechados."],
    ["⚠ AVISO 2", "Junho/2026 tem pouquíssimos embarques (parcial dentro do mês)."],
    ["⚠ AVISO 3", "Tudo aqui é IMPORTAÇÃO destes países (eles como destino/discharge),"],
    ["", "somando todas as commodities (trigo, farelo de soja, milho, soja, cevada...)."],
]
for r in notas:
    ws0.append(r)
ws0["A1"].font = Font(bold=True, size=13)
for r in (9, 11, 12):
    ws0[f"A{r}"].font = Font(bold=True, color="C00000")
ws0.column_dimensions["A"].width = 20
ws0.column_dimensions["B"].width = 85

# ===== Aba 1: Resumo país x ano =====
ws1 = wb.create_sheet("Resumo país x ano")
ws1.append(["País", "Ano", "Embarques", "Volume (mt)", "Obs"])
header(ws1, 5)
for r in c.execute(f"""
    SELECT discharge, {ANO} ano, COUNT(*), SUM(quantity_mt)
    FROM embarques WHERE {FILTRO} AND {ANO} IN ('2025','2026')
    GROUP BY discharge, ano ORDER BY discharge, ano
"""):
    obs = "ano completo" if r[1] == "2025" else "PARCIAL (até ~jun/26)"
    ws1.append([r[0], r[1], r[2], r[3] or 0, obs])
fmt_milhares(ws1, "D"); autofit(ws1)

# ===== Aba 2: Por mês (matriz país/ano nas linhas, meses nas colunas) =====
ws2 = wb.create_sheet("Por mês")
MESES = ["01","02","03","04","05","06","07","08","09","10","11","12"]
NOMES = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
ws2.append(["País", "Ano"] + NOMES + ["Total"])
header(ws2, 15)
dados_mes = {}
for r in c.execute(f"""
    SELECT discharge, {ANO} ano, {MES} mes, SUM(quantity_mt)
    FROM embarques WHERE {FILTRO} AND {ANO} IN ('2025','2026')
    GROUP BY discharge, ano, mes
"""):
    dados_mes[(r[0], r[1], r[2])] = r[3] or 0
for pais in ("Indonesia", "Philippines"):
    for ano in ("2025", "2026"):
        linha = [pais, ano]
        tot = 0
        for m in MESES:
            v = dados_mes.get((pais, ano, m), 0)
            linha.append(v if v else None)
            tot += v
        linha.append(tot)
        ws2.append(linha)
for col in range(3, 16):
    fmt_milhares(ws2, get_column_letter(col))
autofit(ws2)

# ===== Aba 3: Por commodity (país x ano x commodity) =====
ws3 = wb.create_sheet("Por commodity")
ws3.append(["País", "Ano", "Commodity", "Embarques", "Volume (mt)"])
header(ws3, 5)
for r in c.execute(f"""
    SELECT discharge, {ANO} ano, commodity, COUNT(*), SUM(quantity_mt)
    FROM embarques WHERE {FILTRO} AND {ANO} IN ('2025','2026')
    GROUP BY discharge, ano, commodity ORDER BY discharge, ano, 5 DESC
"""):
    ws3.append([r[0], r[1], r[2], r[3], r[4] or 0])
fmt_milhares(ws3, "E"); autofit(ws3)

# ===== Aba 4: Por origem =====
ws4 = wb.create_sheet("Por origem")
ws4.append(["País", "Ano", "Origem (país carga)", "Embarques", "Volume (mt)"])
header(ws4, 5)
for r in c.execute(f"""
    SELECT discharge, {ANO} ano, country, COUNT(*), SUM(quantity_mt)
    FROM embarques WHERE {FILTRO} AND {ANO} IN ('2025','2026')
    GROUP BY discharge, ano, country ORDER BY discharge, ano, 5 DESC
"""):
    ws4.append([r[0], r[1], r[2] or "(sem origem)", r[3], r[4] or 0])
fmt_milhares(ws4, "E"); autofit(ws4)

# ===== Aba 5: Detalhe (todas as linhas) =====
ws5 = wb.create_sheet("Detalhe")
cols = ["País destino", "Ano", "Mês", "Commodity", "Origem", "Porto carga",
        "Navio", "BL/ETA", "Volume (mt)", "Status"]
ws5.append(cols); header(ws5, len(cols))
for r in c.execute(f"""
    SELECT discharge, {ANO} ano, {MES} mes, commodity, country, port, vessel,
           {DATA} dataref, quantity_mt, status
    FROM embarques WHERE {FILTRO} AND {ANO} IN ('2025','2026')
    ORDER BY discharge, dataref
"""):
    ws5.append([r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8] or 0, r[9]])
fmt_milhares(ws5, "I"); autofit(ws5)
ws5.freeze_panes = "A2"

conn.close()
wb.save(OUT)
print("Relatório salvo em:", OUT)
print("Abas:", wb.sheetnames)
